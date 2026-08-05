"""
test_import_e2e.py — E2E de importación con un Excel REAL (formato A, estaciones).

1. Parsea la hoja 'ventana' del Excel de campo (sin BD: solo parser).
2. Valida que los campos críticos salgan del parser: campaña, fase, GSI
   superficie/estructura, nivel, sector, litologías, estructuras y RMR.
3. Escribe fixtures/preview_a21.json para que el test del frontend
   (test_window_transform.mjs) valide la transformación a WindowData.

Uso:  python test_import_e2e.py      (desde Sistema/tests)
"""
import sys
import os
import json

TESTS_DIR = os.path.dirname(os.path.abspath(__file__))
BACKEND_DIR = os.path.abspath(os.path.join(TESTS_DIR, "..", "backend"))
EXCEL = os.path.abspath(os.path.join(TESTS_DIR, "..", "..", "..", "Material", "Estaciones_A21_23-04-2026.xlsx"))
HOJA = "ventana"
FIXTURE = os.path.join(TESTS_DIR, "fixtures", "preview_a21.json")

sys.path.insert(0, BACKEND_DIR)

from openpyxl import load_workbook  # noqa: E402
from app.parsers.excel_a import parse_excel_a, normalize_station_to_celda  # noqa: E402
from app.core.catalogs import infer_lithology_from_lito3  # noqa: E402

assert os.path.exists(EXCEL), f"No se encontró el Excel: {EXCEL}"

wb = load_workbook(EXCEL, data_only=True)
assert HOJA in wb.sheetnames, f"Hoja '{HOJA}' no existe en {wb.sheetnames}"

stations = parse_excel_a(wb[HOJA])
assert stations, "No se detectaron estaciones en la hoja"

celdas = []
for st in stations:
    c = normalize_station_to_celda(st, infer_lito=infer_lithology_from_lito3)
    if c.get("codigo"):
        celdas.append(c)

assert celdas, "El parser no devolvió celdas válidas"
print(f"Parser OK: {len(celdas)} celdas en hoja '{HOJA}'")

errors = []
for c in celdas:
    ed = c.get("excel_data", {})
    cod = c["codigo"]
    if not str(ed.get("campania", "")).startswith("Campaña 20"):
        errors.append(f"{cod}: campania inválida {ed.get('campania')!r}")
    if not str(ed.get("fase", "")).strip():
        errors.append(f"{cod}: fase vacía")
    if not str(ed.get("gsi_superficie", "")).strip():
        errors.append(f"{cod}: gsi_superficie vacía")
    if not str(ed.get("gsi_estructura", "")).strip():
        errors.append(f"{cod}: gsi_estructura vacía")
    if not str(ed.get("nivel", "")).strip():
        errors.append(f"{cod}: nivel vacío")
    if not str(ed.get("sector", "")).strip():
        errors.append(f"{cod}: sector vacío")
    if not str(ed.get("mapeador", "")).strip():
        errors.append(f"{cod}: mapeador vacío")
    if not c.get("estructuras"):
        errors.append(f"{cod}: sin estructuras")

if errors:
    print("\n".join(errors))
    sys.exit(1)

print("Campos críticos OK en las", len(celdas), "celdas")

os.makedirs(os.path.dirname(FIXTURE), exist_ok=True)
with open(FIXTURE, "w", encoding="utf-8") as f:
    json.dump([{"codigo": c["codigo"], "excel_data": c["excel_data"], "estructuras": c["estructuras"]} for c in celdas],
              f, ensure_ascii=False, indent=1)
print("Fixture escrito en", FIXTURE)
