"""
app/core/excel_styles.py
Paleta de estilos openpyxl compartida entre auditoria.py y comparativo.py.
Define una única función get_styles() que retorna el dict completo de estilos.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def get_styles() -> dict:
    """
    Retorna un diccionario con todos los estilos openpyxl reutilizables.
    Llamar una vez por workbook y pasar el dict a las funciones de escritura.
    """
    return {
        # --- Fuentes ---
        "font_title":       Font(name="Segoe UI", size=16, bold=True,  color="1B365D"),
        "font_subtitle":    Font(name="Segoe UI", size=10, italic=True, color="555555"),
        "font_section":     Font(name="Segoe UI", size=11, bold=True,  color="1B365D"),
        "font_header":      Font(name="Segoe UI", size=10, bold=True,  color="FFFFFF"),
        "font_bold":        Font(name="Segoe UI", size=10, bold=True,  color="000000"),
        "font_regular":     Font(name="Segoe UI", size=10,             color="000000"),
        "font_kpi_lbl":     Font(name="Segoe UI", size=9,  bold=True,  color="555555"),
        "font_kpi_blue":    Font(name="Segoe UI", size=18, bold=True,  color="1B365D"),
        "font_kpi_green":   Font(name="Segoe UI", size=18, bold=True,  color="375623"),
        "font_kpi_red":     Font(name="Segoe UI", size=18, bold=True,  color="C00000"),
        "font_kpi_orange":  Font(name="Segoe UI", size=18, bold=True,  color="C65911"),
        "font_delta_green": Font(name="Segoe UI", size=10, bold=True,  color="375623"),
        "font_delta_red":   Font(name="Segoe UI", size=10, bold=True,  color="C00000"),
        "font_delta_gray":  Font(name="Segoe UI", size=10,             color="555555"),
        "font_link":        Font(name="Segoe UI", size=10, bold=True,  color="1B365D", underline="single"),
        # --- Rellenos ---
        "fill_primary":     PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid"),
        "fill_a":           PatternFill(start_color="2B4F8C", end_color="2B4F8C", fill_type="solid"),
        "fill_b":           PatternFill(start_color="14532D", end_color="14532D", fill_type="solid"),
        "fill_green":       PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "fill_yellow":      PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "fill_orange":      PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "fill_red":         PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),
        "fill_zebra":       PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid"),
        "fill_kpi_gray":    PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid"),
        "fill_new":         PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid"),
        "fill_resolved":    PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        # --- Bordes ---
        "border_thin": Border(
            left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin",  color="E2E8F0"), bottom=Side(style="thin", color="E2E8F0"),
        ),
        "border_kpi": Border(
            left=Side(style="thin", color="B0C4DE"), right=Side(style="thin", color="B0C4DE"),
            top=Side(style="thin",  color="B0C4DE"), bottom=Side(style="thin", color="B0C4DE"),
        ),
        # --- Alineaciones ---
        "align_center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "align_left":   Alignment(horizontal="left",   vertical="center", wrap_text=True),
        "align_right":  Alignment(horizontal="right",  vertical="center"),
    }


def write_kpi_card(ws, start_row: int, start_col: int,
                   label: str, value, bg_fill, val_font, s: dict) -> None:
    """
    Escribe una tarjeta KPI de 2 filas × 2 columnas en la hoja ws.

    Args:
        ws:         Hoja openpyxl activa.
        start_row:  Fila inicial (1-indexed).
        start_col:  Columna inicial (1-indexed).
        label:      Texto de la etiqueta superior.
        value:      Valor numérico o texto a mostrar.
        bg_fill:    PatternFill de fondo (ej. s["fill_kpi_gray"]).
        val_font:   Font para el valor (ej. s["font_kpi_blue"]).
        s:          Dict de estilos retornado por get_styles().
    """
    c1 = ws.cell(row=start_row,   column=start_col, value=label)
    c1.font = s["font_kpi_lbl"]
    c1.alignment = s["align_center"]

    c2 = ws.cell(row=start_row+1, column=start_col, value=value)
    c2.font = val_font
    c2.alignment = s["align_center"]

    for r in range(start_row, start_row + 2):
        for c in range(start_col, start_col + 2):
            cell = ws.cell(row=r, column=c)
            cell.fill = bg_fill
            cell.border = s["border_kpi"]

    ws.merge_cells(start_row=start_row,   start_column=start_col, end_row=start_row,   end_column=start_col + 1)
    ws.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+1, end_column=start_col + 1)
