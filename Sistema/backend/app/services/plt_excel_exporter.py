"""
app/services/plt_excel_exporter.py — Generador de Reporte Excel Profesional Multi-Hoja QA/QC para PLT.
Estructura y estilos 100% calcados del motor de auditoría de Mapeo Geomecánico (generar_excel_reporte_core).
"""

from collections import Counter, defaultdict
import os
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.excel_styles import get_styles, write_kpi_card
from app.core.audit_helpers import safe_int, safe_float, get_safe_sheet_name
from app.core.audit_plt_helpers import get_plt_incidence_category_name
from app.core.rules_plt import CATEGORIES_REGISTRY_PLT, RULES_REGISTRY_PLT, COMPACT_FIELD_CATEGORIES


def export_plt_audit_to_excel(diag: dict, compact: dict, filtered: list) -> openpyxl.Workbook:
    """
    Genera el libro Excel (.xlsx) completo de auditoría QA/QC para Ensayos PLT.
    Calcado 1:1 de la estructura, jerarquía, hipervínculos, estilos y gráficos de Mapeo Geomecánico.
    """
    s = get_styles()
    font_title = s["font_title"]
    font_subtitle = s["font_subtitle"]
    font_section = s["font_section"]
    font_header = s["font_header"]
    font_bold = s["font_bold"]
    font_regular = s["font_regular"]
    font_kpi_lbl = s["font_kpi_lbl"]

    font_kpi_val_blue = s["font_kpi_blue"]
    font_kpi_val_green = s["font_kpi_green"]
    font_kpi_val_red = s["font_kpi_red"]
    font_kpi_val_orange = s["font_kpi_orange"]

    fill_primary = s["fill_primary"]
    fill_accent_green = s["fill_green"]
    fill_accent_yellow = s["fill_yellow"]
    fill_accent_orange = s["fill_orange"]
    fill_accent_red = s["fill_red"]
    fill_zebra = s["fill_zebra"]
    fill_kpi_gray = s["fill_kpi_gray"]

    border_thin = s["border_thin"]
    border_kpi = s["border_kpi"]

    alignment_center = s["align_center"]
    alignment_left = s["align_left"]
    alignment_right = s["align_right"]

    formato_tipo = diag.get("formato_detectado", "FORMAT_STANDARD_34")

    wb = openpyxl.Workbook()

    def write_kpi_card_opt(ws, start_row, start_col, label, value, bg_fill, val_font):
        write_kpi_card(ws, start_row, start_col, label, value, bg_fill, val_font, s)

    # =========================================================================
    # --- HOJA 1: DASHBOARD EJECUTIVO ---
    # =========================================================================
    ws_dash = wb.active
    ws_dash.title = "📊 Dashboard Ejecutivo"
    ws_dash.views.sheetView[0].showGridLines = True

    ws_dash.cell(row=2, column=2, value="SISTEMA DE AUDITORÍA GEOTÉCNICA — ENSAYOS PLT").font = font_title
    subtitle_txt = (
        "Dashboard Ejecutivo de Control de Calidad, Consistencia Geomecánica y Secuencias ABCD (Formato de Campo)"
        if "COMPACT" in str(formato_tipo)
        else "Dashboard Ejecutivo de Control de Calidad, Consistencia de 34 Columnas y Secuencias ABCD"
    )
    ws_dash.cell(row=3, column=2, value=subtitle_txt).font = font_subtitle

    total_filas = compact.get("familia1", {}).get("total_registros", diag.get("total_filas_procesadas", 0))
    total_fields = compact.get("familia2", {}).get("total_fields", total_filas * (24 if "COMPACT" in str(formato_tipo) else 34))
    total_vacios = sum(1 for i in filtered if i.get("tipo_incidencia") == "VACIO")
    total_advertencias = sum(1 for i in filtered if i.get("tipo_incidencia") == "ADVERTENCIA")
    total_alertas = sum(1 for i in filtered if i.get("tipo_incidencia") == "ALERTA")
    total_correctos = max(0, total_fields - (total_vacios + total_advertencias + total_alertas))
    pct_integridad = (total_correctos / max(1, total_fields)) * 100

    total_celdas = len(compact.get("resumen_por_celda", diag.get("resumen_por_celda", {})))

    # Tarjetas KPI Superiores
    write_kpi_card_opt(ws_dash, 5, 2, "CELDAS EVALUADAS", total_celdas, fill_kpi_gray, font_kpi_val_blue)
    write_kpi_card_opt(ws_dash, 5, 4, "MUESTRAS REGISTRADAS", total_filas, fill_kpi_gray, font_kpi_val_blue)
    write_kpi_card_opt(ws_dash, 5, 6, "INTEGRIDAD GLOBAL", f"{pct_integridad:.2f}%", fill_accent_green, font_kpi_val_green)
    write_kpi_card_opt(ws_dash, 5, 8, "ALERTAS CRÍTICAS", total_alertas, fill_accent_red, font_kpi_val_red)
    write_kpi_card_opt(ws_dash, 5, 10, "ADVERTENCIAS", total_advertencias, fill_accent_orange, font_kpi_val_orange)

    # 1. Tabla: Desempeño por Campaña (solo si existen campañas registradas)
    dist_camp = compact.get("distribucion_campania", [])
    has_real_campaigns = any(str(r.get("campania", "")).strip() not in ["N/A", "Sin Campaña", ""] for r in dist_camp)

    r_camp = 11
    if has_real_campaigns:
        ws_dash.cell(row=9, column=2, value="DESEMPEÑO DE CONTROL POR CAMPAÑA").font = font_section
        headers_camp = ["Campaña", "Muestras", "Celdas Afectadas", "Muestras Afectadas", "Alertas (N)", "% Alertas", "Vacíos (N)", "% Vacíos"]
        for idx, col in enumerate(headers_camp, start=2):
            cell = ws_dash.cell(row=10, column=idx, value=col)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = alignment_center
            cell.border = border_thin

        for row in dist_camp:
            ws_dash.cell(row=r_camp, column=2, value=str(row.get("campania"))).font = font_bold
            ws_dash.cell(row=r_camp, column=2).alignment = alignment_center

            ws_dash.cell(row=r_camp, column=3, value=safe_int(row.get("registros"))).number_format = '#,##0'
            ws_dash.cell(row=r_camp, column=3).alignment = alignment_right

            ws_dash.cell(row=r_camp, column=4, value=safe_int(row.get("celdas_afectadas"))).number_format = '#,##0'
            ws_dash.cell(row=r_camp, column=4).alignment = alignment_right

            ws_dash.cell(row=r_camp, column=5, value=safe_int(row.get("registros_afectados", row.get("registros")))).number_format = '#,##0'
            ws_dash.cell(row=r_camp, column=5).alignment = alignment_right

            ws_dash.cell(row=r_camp, column=6, value=safe_int(row.get("alertas_cant"))).number_format = '#,##0'
            ws_dash.cell(row=r_camp, column=6).alignment = alignment_right

            ws_dash.cell(row=r_camp, column=7, value=safe_float(row.get("alertas_pct")) / 100.0).number_format = '0.00%'
            ws_dash.cell(row=r_camp, column=7).alignment = alignment_right

            ws_dash.cell(row=r_camp, column=8, value=safe_int(row.get("vacios_cant"))).number_format = '#,##0'
            ws_dash.cell(row=r_camp, column=8).alignment = alignment_right

            ws_dash.cell(row=r_camp, column=9, value=safe_float(row.get("vacios_pct")) / 100.0).number_format = '0.00%'
            ws_dash.cell(row=r_camp, column=9).alignment = alignment_right

            for col_idx in range(2, 10):
                ws_dash.cell(row=r_camp, column=col_idx).border = border_thin
                if r_camp % 2 == 0:
                    ws_dash.cell(row=r_camp, column=col_idx).fill = fill_zebra
            r_camp += 1
    else:
        r_camp = 9

    # 2. Tabla: Distribución por Tipo Litológico
    r_lito = r_camp + (2 if has_real_campaigns else 0)
    ws_dash.cell(row=r_lito, column=2, value="DISTRIBUCIÓN POR TIPO LITOLÓGICO").font = font_section
    headers_lito = ["Tipo Litológico", "Muestras", "Celdas Afectadas", "Alertas (N)", "% Alertas", "Vacíos (N)", "% Vacíos"]
    for idx, col in enumerate(headers_lito, start=2):
        cell = ws_dash.cell(row=r_lito + 1, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    curr_l_r = r_lito + 2
    for row in compact.get("distribucion_litologia", []):
        ws_dash.cell(row=curr_l_r, column=2, value=str(row.get("tipo_litologico"))).font = font_bold
        ws_dash.cell(row=curr_l_r, column=2).alignment = alignment_left

        ws_dash.cell(row=curr_l_r, column=3, value=safe_int(row.get("registros"))).number_format = '#,##0'
        ws_dash.cell(row=curr_l_r, column=3).alignment = alignment_right

        ws_dash.cell(row=curr_l_r, column=4, value=safe_int(row.get("celdas_afectadas", 0))).number_format = '#,##0'
        ws_dash.cell(row=curr_l_r, column=4).alignment = alignment_right

        ws_dash.cell(row=curr_l_r, column=5, value=safe_int(row.get("alertas_cant"))).number_format = '#,##0'
        ws_dash.cell(row=curr_l_r, column=5).alignment = alignment_right

        ws_dash.cell(row=curr_l_r, column=6, value=safe_float(row.get("alertas_pct")) / 100.0).number_format = '0.00%'
        ws_dash.cell(row=curr_l_r, column=6).alignment = alignment_right

        ws_dash.cell(row=curr_l_r, column=7, value=safe_int(row.get("vacios_cant"))).number_format = '#,##0'
        ws_dash.cell(row=curr_l_r, column=7).alignment = alignment_right

        ws_dash.cell(row=curr_l_r, column=8, value=safe_float(row.get("vacios_pct")) / 100.0).number_format = '0.00%'
        ws_dash.cell(row=curr_l_r, column=8).alignment = alignment_right

        for col_idx in range(2, 9):
            ws_dash.cell(row=curr_l_r, column=col_idx).border = border_thin
            if curr_l_r % 2 == 0:
                ws_dash.cell(row=curr_l_r, column=col_idx).fill = fill_zebra
        curr_l_r += 1

    # 3. Tabla: Top 5 Celdas con Mayor Incidencia
    r_worst = curr_l_r + 2
    ws_dash.cell(row=r_worst, column=2, value="TOP 5 CELDAS CON MAYOR INCIDENCIA").font = font_section
    headers_worst = ["Celda Mapeo", "Muestras", "Alertas (N)", "Advertencias (N)", "Vacíos (N)"]
    for idx, col in enumerate(headers_worst, start=2):
        cell = ws_dash.cell(row=r_worst + 1, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    curr_w_r = r_worst + 2
    for row in compact.get("worst_cells", [])[:5]:
        ws_dash.cell(row=curr_w_r, column=2, value=str(row.get("celda"))).font = font_bold
        ws_dash.cell(row=curr_w_r, column=2).alignment = alignment_left

        ws_dash.cell(row=curr_w_r, column=3, value=safe_int(row.get("total_muestras", 4))).number_format = '#,##0'
        ws_dash.cell(row=curr_w_r, column=3).alignment = alignment_right

        ws_dash.cell(row=curr_w_r, column=4, value=safe_int(row.get("alertas"))).number_format = '#,##0'
        ws_dash.cell(row=curr_w_r, column=4).alignment = alignment_right

        ws_dash.cell(row=curr_w_r, column=5, value=safe_int(row.get("advertencias"))).number_format = '#,##0'
        ws_dash.cell(row=curr_w_r, column=5).alignment = alignment_right

        ws_dash.cell(row=curr_w_r, column=6, value=safe_int(row.get("vacios"))).number_format = '#,##0'
        ws_dash.cell(row=curr_w_r, column=6).alignment = alignment_right

        for col_idx in range(2, 7):
            ws_dash.cell(row=curr_w_r, column=col_idx).border = border_thin
            if curr_w_r % 2 == 0:
                ws_dash.cell(row=curr_w_r, column=col_idx).fill = fill_zebra
        curr_w_r += 1

    # 4. Tabla: Integridad de Secuencias ABCD por Celda
    r_abcd = curr_w_r + 2
    ws_dash.cell(row=r_abcd, column=2, value="INTEGRIDAD DE SECUENCIAS ABCD POR CELDA").font = font_section
    headers_abcd = ["Estado de Secuencia", "Total Celdas (N)", "% Cumplimiento"]
    for idx, col in enumerate(headers_abcd, start=2):
        cell = ws_dash.cell(row=r_abcd + 1, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    integ = compact.get("integridad_celdas", {})
    tot_c = max(1, total_celdas)
    abcd_rows = [
        ("Correctas (4/4 ABCD)", integ.get("correctas_abcd", 0)),
        ("En Desorden", integ.get("desorden_abcd", 0)),
        ("Incompletas (< 4)", integ.get("incompletas_abcd", 0)),
        ("Excedentes (> 4)", integ.get("excedentes_abcd", 0)),
        ("Anómalas (#ERR)", integ.get("anomalas_abcd", 0)),
    ]

    curr_a_r = r_abcd + 2
    for label, count in abcd_rows:
        ws_dash.cell(row=curr_a_r, column=2, value=label).font = font_bold
        ws_dash.cell(row=curr_a_r, column=2).alignment = alignment_left

        ws_dash.cell(row=curr_a_r, column=3, value=safe_int(count)).number_format = '#,##0'
        ws_dash.cell(row=curr_a_r, column=3).alignment = alignment_right

        ws_dash.cell(row=curr_a_r, column=4, value=count / tot_c).number_format = '0.00%'
        ws_dash.cell(row=curr_a_r, column=4).alignment = alignment_right

        for col_idx in range(2, 5):
            ws_dash.cell(row=curr_a_r, column=col_idx).border = border_thin
            if curr_a_r % 2 == 0:
                ws_dash.cell(row=curr_a_r, column=col_idx).fill = fill_zebra
        curr_a_r += 1

    # 5. Tabla: Principales Desviaciones Críticas (Columna Derecha)
    ws_dash.cell(row=9, column=10, value="PRINCIPALES DESVIACIONES CRÍTICAS").font = font_section
    headers_top = ["Regla de Consistencia", "Cantidad (N)", "% Incidencia"]
    for idx, col in enumerate(headers_top, start=10):
        cell = ws_dash.cell(row=10, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    top_alertas = compact.get("top_5_alertas", [])
    if not top_alertas:
        top_alertas = compact.get("error_types_detailed", {}).get("alertas", [])[:5]

    for idx, a in enumerate(top_alertas[:5]):
        r_top = 11 + idx
        ws_dash.cell(row=r_top, column=10, value=a.get("mensaje", "Desviación Crítica")).font = font_regular
        ws_dash.cell(row=r_top, column=10).alignment = alignment_left

        ws_dash.cell(row=r_top, column=11, value=safe_int(a.get("cantidad"))).number_format = '#,##0'
        ws_dash.cell(row=r_top, column=11).alignment = alignment_right

        pct = safe_float(a.get("pct", a.get("cantidad", 0) / max(1, total_alertas) * 100)) / 100.0
        ws_dash.cell(row=r_top, column=12, value=pct).number_format = '0.00%'
        ws_dash.cell(row=r_top, column=12).alignment = alignment_right

        for col_idx in range(10, 13):
            ws_dash.cell(row=r_top, column=col_idx).border = border_thin
            if r_top % 2 == 0:
                ws_dash.cell(row=r_top, column=col_idx).fill = fill_zebra

    # Gráfica Nativa de Excel
    if len(top_alertas) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Frecuencia de Desviaciones Críticas Detectadas"
        chart.y_axis.title = "Cantidad de Ocurrencias"
        chart.x_axis.title = "Regla de Consistencia"
        max_r_chart = 10 + len(top_alertas[:5])
        chart_data = Reference(ws_dash, min_col=11, min_row=10, max_row=max_r_chart)
        chart_cats = Reference(ws_dash, min_col=10, min_row=11, max_row=max_r_chart)
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(chart_cats)
        chart.legend = None
        chart.width = 16
        chart.height = 10
        ws_dash.add_chart(chart, "J18")

    # =========================================================================
    # --- HOJA 2: REGISTRO MAESTRO DE ERRORES (CATÁLOGO / ÍNDICE) ---
    # =========================================================================
    ws_cat = wb.create_sheet(title="📋 Catálogo de Errores")
    ws_cat.views.sheetView[0].showGridLines = True

    ws_cat.cell(row=2, column=2, value="REGISTRO MAESTRO DE REGLAS DE CONSISTENCIA PLT").font = font_title
    ws_cat.cell(
        row=3,
        column=2,
        value="Catálogo completo de validación geomecánica ordenado por frecuencia. Use los hipervínculos para navegar.",
    ).font = font_subtitle

    # Campañas únicas
    all_campaigns = sorted(
        set(
            [str(row.get("campania")) for row in compact.get("distribucion_campania", []) if row.get("campania") and str(row.get("campania")) not in ["N/A", "None", "", "Sin Campaña"]]
            + [str(inc.get("campania")) for inc in filtered if inc.get("campania") and str(inc.get("campania")) not in ["N/A", "None", "", "Sin Campaña"]]
        ),
        key=lambda x: str(x),
    )

    headers_cat = ["ID", "Gravedad", "Regla de Consistencia Evaluada", "Total (N)"]
    if all_campaigns:
        headers_cat += [f"Año {c}" for c in all_campaigns]
    headers_cat.append("Enlace Directo")

    for idx, col in enumerate(headers_cat, start=2):
        cell = ws_cat.cell(row=5, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    incidencias_por_error = defaultdict(list)
    for inc in filtered:
        msg_simplificado = get_plt_incidence_category_name(inc)
        incidencias_por_error[msg_simplificado].append(inc)

    # Filtrar categorías aplicables según el formato detectado
    if "COMPACT" in str(formato_tipo):
        applicable_categories = [cat for code, cat in CATEGORIES_REGISTRY_PLT.items() if code in COMPACT_FIELD_CATEGORIES]
    else:
        applicable_categories = list(CATEGORIES_REGISTRY_PLT.values())

    catalog_frequencies = []
    for cat in applicable_categories:
        rule_msg = cat.name
        matches = incidencias_por_error[rule_msg]
        camp_counts = {}
        for c in all_campaigns:
            camp_counts[c] = sum(1 for m in matches if str(m.get("campania", "N/A")) == c)
        catalog_frequencies.append({
            "msg": rule_msg,
            "severity": cat.severity,
            "matches": matches,
            "count": len(matches),
            "camp_counts": camp_counts,
        })

    catalog_frequencies = sorted(catalog_frequencies, key=lambda x: x["count"], reverse=True)

    r_cat = 6
    active_sheets_mapping = {}

    for c_idx, rule in enumerate(catalog_frequencies, start=1):
        ws_cat.cell(row=r_cat, column=2, value=c_idx).font = font_regular
        ws_cat.cell(row=r_cat, column=2).alignment = alignment_center
        ws_cat.cell(row=r_cat, column=2).border = border_thin

        c_sev = ws_cat.cell(row=r_cat, column=3, value=rule["severity"])
        c_sev.font = font_bold
        c_sev.alignment = alignment_center
        c_sev.border = border_thin
        if rule["severity"] == "ALERTA":
            c_sev.fill = fill_accent_red
        elif rule["severity"] == "ADVERTENCIA":
            c_sev.fill = fill_accent_orange
        else:
            c_sev.fill = fill_accent_yellow

        ws_cat.cell(row=r_cat, column=4, value=rule["msg"]).font = font_bold if rule["count"] > 0 else font_regular
        ws_cat.cell(row=r_cat, column=4).border = border_thin

        c_count = ws_cat.cell(row=r_cat, column=5, value=rule["count"])
        c_count.font = font_bold
        c_count.alignment = alignment_right
        c_count.number_format = '#,##0'
        c_count.border = border_thin

        # Columnas por campaña
        for y_offset, camp_key in enumerate(all_campaigns):
            c_yr = ws_cat.cell(row=r_cat, column=6 + y_offset, value=rule["camp_counts"].get(camp_key, 0))
            c_yr.font = font_regular
            c_yr.alignment = alignment_right
            c_yr.number_format = '#,##0'
            c_yr.border = border_thin
            if rule["camp_counts"].get(camp_key, 0) == 0:
                c_yr.font = Font(name="Segoe UI", size=9, color="AAAAAA")

        link_col = 6 + len(all_campaigns)
        c_link = ws_cat.cell(row=r_cat, column=link_col)
        if rule["count"] > 0:
            tab_name = get_safe_sheet_name(rule["msg"], c_idx)
            active_sheets_mapping[rule["msg"]] = {"tab_name": tab_name, "records": rule["matches"]}

            c_link.value = f'=HYPERLINK("#{chr(39)}{tab_name}{chr(39)}!B2", "👉 Navegar a Registros")'
            c_link.font = Font(name="Segoe UI", size=10, bold=True, color="1B365D", underline="single")
            c_link.alignment = alignment_center
        else:
            c_link.value = "Limpio / 0 Incidencias"
            c_link.font = Font(name="Segoe UI", size=9, italic=True, color="7F8C8D")
            c_link.alignment = alignment_center
            c_link.fill = fill_accent_green

        c_link.border = border_thin
        r_cat += 1

    last_col_letter = get_column_letter(5 + len(all_campaigns) + 1)
    ws_cat.auto_filter.ref = f"B5:{last_col_letter}{r_cat - 1}"

    # =========================================================================
    # --- HOJA 3: DETALLE PLANO COMPLETO DE INCIDENCIAS ---
    # =========================================================================
    chunk_size = 1000000
    detail_chunks = [filtered[i:i + chunk_size] for i in range(0, len(filtered), chunk_size)]
    if not detail_chunks:
        detail_chunks = [[]]

    for chunk_idx, chunk_data in enumerate(detail_chunks):
        title = "📑 Detalle de Incidencias"
        if len(detail_chunks) > 1:
            title = f"📑 Detalle Incidencias ({chunk_idx + 1})"

        ws_detail = wb.create_sheet(title=title)
        ws_detail.views.sheetView[0].showGridLines = True

        ws_detail.cell(row=2, column=2, value="REGISTRO DETALLADO DE TODAS LAS INCIDENCIAS DETECTADAS").font = font_title
        ws_detail.cell(
            row=3,
            column=2,
            value="Base de datos plana de inconsistencias. Utilice los filtros en los encabezados para auditar registros específicos.",
        ).font = font_subtitle

        headers_detail = [
            "ID",
            "Fila Excel",
            "Gravedad",
            "Celda Mapeo",
            "Muestra",
            "Campaña",
            "Tipo Litológico",
            "Nivel",
            "Columna de Falla",
            "Valor Actual",
            "Código Regla",
            "Mensaje de Inconsistencia Geomecánica",
        ]

        grid_heading_row = 5
        for idx, col in enumerate(headers_detail, start=2):
            cell = ws_detail.cell(row=grid_heading_row, column=idx, value=col)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = alignment_center
            cell.border = border_thin

        r_detail_start = 6
        for d_idx, inc in enumerate(chunk_data, start=1):
            r_detail = r_detail_start + d_idx - 1
            ws_detail.cell(row=r_detail, column=2, value=d_idx).font = font_regular
            ws_detail.cell(row=r_detail, column=2).alignment = alignment_center

            ws_detail.cell(row=r_detail, column=3, value=safe_int(inc.get("fila_excel"))).alignment = alignment_center

            c_sev = ws_detail.cell(row=r_detail, column=4, value=inc.get("tipo_incidencia", "ALERTA"))
            c_sev.font = font_bold
            c_sev.alignment = alignment_center
            if inc.get("tipo_incidencia") == "ALERTA":
                c_sev.fill = fill_accent_red
            elif inc.get("tipo_incidencia") == "ADVERTENCIA":
                c_sev.fill = fill_accent_orange
            else:
                c_sev.fill = fill_accent_yellow

            ws_detail.cell(row=r_detail, column=5, value=str(inc.get("celda_mapeo", inc.get("celda_padre", "—")))).alignment = alignment_left
            ws_detail.cell(row=r_detail, column=6, value=str(inc.get("muestra", inc.get("celda_hija", "—")))).alignment = alignment_center
            ws_detail.cell(row=r_detail, column=7, value=str(inc.get("campania", "—"))).alignment = alignment_center
            ws_detail.cell(row=r_detail, column=8, value=str(inc.get("tipo_litologico", "—"))).alignment = alignment_left
            ws_detail.cell(row=r_detail, column=9, value=str(inc.get("nivel", "—"))).alignment = alignment_center
            ws_detail.cell(row=r_detail, column=10, value=str(inc.get("columna", "—"))).alignment = alignment_left

            raw_val = inc.get("valor_actual")
            val_str = "—" if raw_val is None or str(raw_val).strip() in ["", "-1"] else str(raw_val)
            ws_detail.cell(row=r_detail, column=11, value=val_str).alignment = alignment_center

            ws_detail.cell(row=r_detail, column=12, value=str(inc.get("rule_code", "—"))).alignment = alignment_center
            ws_detail.cell(row=r_detail, column=13, value=str(inc.get("mensaje", "—"))).alignment = alignment_left

            for col_idx in range(2, 14):
                cell_d = ws_detail.cell(row=r_detail, column=col_idx)
                cell_d.border = border_thin
                if r_detail % 2 == 0:
                    if col_idx != 4:
                        cell_d.fill = fill_zebra

        end_detail_row = max(grid_heading_row + 1, r_detail_start + len(chunk_data) - 1)
        ws_detail.auto_filter.ref = f"B{grid_heading_row}:M{end_detail_row}"

    # =========================================================================
    # --- HOJAS 4+: DETALLES INDIVIDUALES POR REGLA DE ERROR ---
    # =========================================================================
    for orig_msg, mapping_data in active_sheets_mapping.items():
        sh_name = mapping_data["tab_name"]
        err_records = mapping_data["records"]

        ws_err = wb.create_sheet(title=sh_name)
        ws_err.views.sheetView[0].showGridLines = True

        c_back = ws_err.cell(row=2, column=2)
        c_back.value = '=HYPERLINK("#\'📋 Catálogo de Errores\'!B2", "⬅️ Volver al Registro Maestro de Errores")'
        c_back.font = Font(name="Segoe UI", size=10, bold=True, color="1B365D", underline="single")
        c_back.alignment = alignment_left

        ws_err.cell(row=4, column=2, value="ANÁLISIS DE INCIDENCIA EN BASE DE DATOS").font = font_section
        cell_err_desc = ws_err.cell(row=5, column=2, value=f"Regla: {orig_msg.upper()}")
        cell_err_desc.font = Font(name="Segoe UI", size=11, bold=True, color="7F1D1D")
        cell_err_desc.fill = fill_accent_red
        cell_err_desc.border = border_thin
        ws_err.merge_cells(start_row=5, start_column=2, end_row=5, end_column=7)

        st_affected = len(set(x.get("celda_mapeo", x.get("celda_padre", "N/A")) for x in err_records))
        tot_affected = len(err_records)

        write_kpi_card_opt(ws_err, 7, 2, "CELDAS AFECTADAS", st_affected, fill_kpi_gray, font_kpi_val_blue)
        write_kpi_card_opt(ws_err, 7, 4, "MUESTRAS AFECTADAS (CANTIDAD)", tot_affected, fill_kpi_gray, font_kpi_val_blue)

        # Distribución por Campaña
        ws_err.cell(row=10, column=2, value="DISTRIBUCIÓN POR CAMPAÑA").font = font_section
        for idx, col in enumerate(["Campaña / Año", "Ocurrencias (N)", "% Contribución"], start=2):
            cell = ws_err.cell(row=11, column=idx, value=col)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = alignment_center
            cell.border = border_thin

        r_dist_yr = defaultdict(int)
        for r in err_records:
            r_dist_yr[str(r.get("campania", "N/A"))] += 1

        curr_y_r = 12
        for yr, y_qty in sorted(r_dist_yr.items()):
            ws_err.cell(row=curr_y_r, column=2, value=yr).font = font_bold
            ws_err.cell(row=curr_y_r, column=2).alignment = alignment_center
            ws_err.cell(row=curr_y_r, column=2).border = border_thin

            c_yq = ws_err.cell(row=curr_y_r, column=3, value=y_qty)
            c_yq.font = font_regular
            c_yq.alignment = alignment_right
            c_yq.number_format = '#,##0'
            c_yq.border = border_thin

            c_yp = ws_err.cell(row=curr_y_r, column=4, value=y_qty / max(1, tot_affected))
            c_yp.font = font_regular
            c_yp.alignment = alignment_right
            c_yp.number_format = '0.00%'
            c_yp.border = border_thin

            curr_y_r += 1

        unique_years = sorted(list(set(str(r.get("campania", "N/A")) for r in err_records)))

        # 1. Agrupar por Código de Regla
        rule_group = defaultdict(list)
        for r in err_records:
            rule_group[r.get("rule_code", "Desconocido")].append(r)

        rule_stats = []
        for rule_code, recs in rule_group.items():
            yr_counts = defaultdict(int)
            for r in recs:
                yr_counts[str(r.get("campania", "N/A"))] += 1
            rule_stats.append({
                "rule_code": rule_code,
                "total": len(recs),
                "yr_counts": yr_counts,
            })
        rule_stats.sort(key=lambda x: x["total"], reverse=True)

        # 2. Agrupar por Mensaje Único (Tabla B)
        msg_group = defaultdict(list)
        for r in err_records:
            msg_group[r.get("mensaje", "Desconocido")].append(r)

        msg_stats = []
        for msg_val, recs in msg_group.items():
            yr_counts = defaultdict(int)
            for r in recs:
                yr_counts[str(r.get("campania", "N/A"))] += 1
            msg_stats.append({
                "message": msg_val,
                "total": len(recs),
                "yr_counts": yr_counts,
            })
        msg_stats.sort(key=lambda x: x["total"], reverse=True)

        # Precalcular coordenadas de las secciones
        dist_table_height = len(r_dist_yr)
        jump_link_row = 12 + dist_table_height + 1

        table_a_start = jump_link_row + 2
        table_a_height = 2 + len(rule_stats)

        max_allowed_err = 1000000
        truncated_err = len(err_records) > max_allowed_err
        err_records_to_write = err_records[:max_allowed_err]

        indiv_start = table_a_start + table_a_height + 2
        indiv_height = 2 + len(err_records_to_write) + (1 if truncated_err else 0)

        table_b_start = indiv_start + indiv_height + 2

        # Enlace directo de salto rápido
        c_jump = ws_err.cell(row=jump_link_row, column=2)
        c_jump.value = f'=HYPERLINK("#\'{sh_name}\'!B{table_b_start}", "⬇️ Ir a Métricas de Mensajes de Inconsistencia Únicos (al final de la hoja)")'
        c_jump.font = Font(name="Segoe UI", size=10, bold=True, color="1B365D", underline="single")
        c_jump.alignment = alignment_left

        # --- TABLA A: RESUMEN POR REGLA ESPECÍFICA (CÓDIGO) ---
        ws_err.cell(row=table_a_start, column=2, value="MÉTRICAS POR REGLA ESPECÍFICA (CÓDIGO)").font = font_section

        ws_err.cell(row=table_a_start + 1, column=2, value="Código de Regla").font = font_header
        ws_err.cell(row=table_a_start + 1, column=2).fill = fill_primary
        ws_err.cell(row=table_a_start + 1, column=2).alignment = alignment_center
        ws_err.merge_cells(start_row=table_a_start + 1, start_column=2, end_row=table_a_start + 1, end_column=4)
        for c_idx in [3, 4]:
            ws_err.cell(row=table_a_start + 1, column=c_idx).fill = fill_primary

        headers_a = ["Ocurrencias Totales"] + [f"Año {y}" if y != "N/A" else "N/A" for y in unique_years]
        for col_offset, h_name in enumerate(headers_a, start=5):
            cell = ws_err.cell(row=table_a_start + 1, column=col_offset, value=h_name)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = alignment_center

        for col_idx in range(2, 5 + len(unique_years) + 1):
            ws_err.cell(row=table_a_start + 1, column=col_idx).border = border_thin

        for idx, stat in enumerate(rule_stats):
            r_row = table_a_start + 2 + idx
            ws_err.cell(row=r_row, column=2, value=stat["rule_code"]).font = font_bold
            ws_err.cell(row=r_row, column=2).alignment = alignment_center
            ws_err.merge_cells(start_row=r_row, start_column=2, end_row=r_row, end_column=4)

            c_tot = ws_err.cell(row=r_row, column=5, value=stat["total"])
            c_tot.font = font_bold
            c_tot.alignment = alignment_right
            c_tot.number_format = '#,##0'

            for y_idx, yr in enumerate(unique_years):
                val = stat["yr_counts"].get(yr, 0)
                c_val = ws_err.cell(row=r_row, column=6 + y_idx, value=val)
                c_val.font = font_regular
                c_val.alignment = alignment_right
                c_val.number_format = '#,##0'

            for col_idx in range(2, 5 + len(unique_years) + 1):
                cell = ws_err.cell(row=r_row, column=col_idx)
                cell.border = border_thin
                if r_row % 2 == 0:
                    cell.fill = fill_zebra

        # --- TABLA B: REGISTROS INDIVIDUALES AFECTADOS (LISTADO COMPLETO) ---
        ws_err.cell(row=indiv_start, column=2, value="REGISTROS INDIVIDUALES AFECTADOS (LISTADO COMPLETO)").font = font_section

        headers_inc = [
            "Fila Excel",
            "Celda Mapeo",
            "Muestra",
            "Campaña",
            "Tipo Litológico",
            "Nivel",
            "Columna de Falla",
            "Valor Actual",
            "Mensaje de Inconsistencia",
        ]
        header_row_idx = indiv_start + 1
        for col_idx, h_name in enumerate(headers_inc, start=2):
            cell = ws_err.cell(row=header_row_idx, column=col_idx, value=h_name)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = alignment_center
            cell.border = border_thin

        start_data_row = indiv_start + 2
        for idx, inc_item in enumerate(err_records_to_write):
            curr_row = start_data_row + idx
            ws_err.cell(row=curr_row, column=2, value=safe_int(inc_item.get("fila_excel")))
            ws_err.cell(row=curr_row, column=3, value=str(inc_item.get("celda_mapeo", inc_item.get("celda_padre", "—"))))
            ws_err.cell(row=curr_row, column=4, value=str(inc_item.get("muestra", inc_item.get("celda_hija", "—"))))
            ws_err.cell(row=curr_row, column=5, value=str(inc_item.get("campania", "—")))
            ws_err.cell(row=curr_row, column=6, value=str(inc_item.get("tipo_litologico", "—")))
            ws_err.cell(row=curr_row, column=7, value=str(inc_item.get("nivel", "—")))
            ws_err.cell(row=curr_row, column=8, value=str(inc_item.get("columna", "—")))

            raw_val = inc_item.get("valor_actual")
            val_str = "—" if raw_val is None or str(raw_val).strip() in ["", "-1"] else str(raw_val)
            ws_err.cell(row=curr_row, column=9, value=val_str)

            ws_err.cell(row=curr_row, column=10, value=str(inc_item.get("mensaje", "—")))

        if truncated_err:
            curr_row = start_data_row + len(err_records_to_write)
            ws_err.cell(row=curr_row, column=3, value="--- REPORTE TRUNCADO: SE SUPERÓ EL LÍMITE DE FILAS DE EXCEL (1,048,576) ---")

        end_data_row = start_data_row + len(err_records_to_write) - 1 + (1 if truncated_err else 0)

        for r_idx in range(start_data_row, end_data_row + 1):
            if r_idx <= start_data_row + 150:
                ws_err.cell(row=r_idx, column=2).alignment = alignment_center
                ws_err.cell(row=r_idx, column=3).alignment = alignment_center
                ws_err.cell(row=r_idx, column=4).alignment = alignment_center
                ws_err.cell(row=r_idx, column=5).alignment = alignment_center
                ws_err.cell(row=r_idx, column=6).alignment = alignment_left
                ws_err.cell(row=r_idx, column=7).alignment = alignment_center
                ws_err.cell(row=r_idx, column=8).alignment = alignment_left
                ws_err.cell(row=r_idx, column=9).alignment = alignment_center
                ws_err.cell(row=r_idx, column=10).alignment = alignment_left

                if r_idx % 2 == 0:
                    for col_idx in range(2, 11):
                        ws_err.cell(row=r_idx, column=col_idx).fill = fill_zebra
            else:
                ws_err.cell(row=r_idx, column=2).alignment = alignment_center
                ws_err.cell(row=r_idx, column=3).alignment = alignment_center

            for col_idx in range(2, 11):
                ws_err.cell(row=r_idx, column=col_idx).border = border_thin

        ws_err.auto_filter.ref = f"B{header_row_idx}:J{end_data_row}"

        # --- TABLA C: RESUMEN POR MENSAJE DE INCONSISTENCIA ÚNICO ---
        ws_err.cell(row=table_b_start, column=2, value="MÉTRICAS POR MENSAJE DE INCONSISTENCIA GEOMECÁNICA ÚNICO").font = font_section

        c_ret = ws_err.cell(row=table_b_start, column=5)
        c_ret.value = f'=HYPERLINK("#\'{sh_name}\'!B2", "⬆️ Volver al Inicio de la Hoja")'
        c_ret.font = Font(name="Segoe UI", size=10, bold=True, color="1B365D", underline="single")
        c_ret.alignment = alignment_left

        ws_err.cell(row=table_b_start + 1, column=2, value="Mensaje de Inconsistencia Geomecánica Único").font = font_header
        ws_err.cell(row=table_b_start + 1, column=2).fill = fill_primary
        ws_err.cell(row=table_b_start + 1, column=2).alignment = alignment_center

        ws_err.merge_cells(start_row=table_b_start + 1, start_column=2, end_row=table_b_start + 1, end_column=7)
        for c_idx in range(3, 8):
            ws_err.cell(row=table_b_start + 1, column=c_idx).fill = fill_primary

        headers_b = ["Ocurrencias Totales"] + [f"Año {y}" if y != "N/A" else "N/A" for y in unique_years]
        for col_offset, h_name in enumerate(headers_b, start=8):
            cell = ws_err.cell(row=table_b_start + 1, column=col_offset, value=h_name)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = alignment_center

        for col_idx in range(2, 8 + len(unique_years) + 1):
            ws_err.cell(row=table_b_start + 1, column=col_idx).border = border_thin

        for idx, stat in enumerate(msg_stats):
            r_row = table_b_start + 2 + idx
            ws_err.cell(row=r_row, column=2, value=stat["message"]).font = font_regular
            ws_err.cell(row=r_row, column=2).alignment = alignment_left
            ws_err.merge_cells(start_row=r_row, start_column=2, end_row=r_row, end_column=7)

            c_tot = ws_err.cell(row=r_row, column=8, value=stat["total"])
            c_tot.font = font_bold
            c_tot.alignment = alignment_right
            c_tot.number_format = '#,##0'

            for y_idx, yr in enumerate(unique_years):
                val = stat["yr_counts"].get(yr, 0)
                c_val = ws_err.cell(row=r_row, column=9 + y_idx, value=val)
                c_val.font = font_regular
                c_val.alignment = alignment_right
                c_val.number_format = '#,##0'

            for col_idx in range(2, 8 + len(unique_years) + 1):
                cell = ws_err.cell(row=r_row, column=col_idx)
                cell.border = border_thin
                if r_row % 2 == 0:
                    cell.fill = fill_zebra

    # =========================================================================
    # --- AUTO-AJUSTE DINÁMICO DE COLUMNAS ---
    # =========================================================================
    for ws in wb.worksheets:
        ws.column_dimensions['A'].width = 3
        if ws.title not in ["📋 Catálogo de Errores", "📑 Detalle de Incidencias", "📊 Dashboard Ejecutivo"]:
            ws.column_dimensions['B'].width = 11  # Fila Excel
            ws.column_dimensions['C'].width = 14  # Celda Mapeo
            ws.column_dimensions['D'].width = 14  # Muestra
            ws.column_dimensions['E'].width = 10  # Campaña
            ws.column_dimensions['F'].width = 18  # Tipo Litológico
            ws.column_dimensions['G'].width = 10  # Nivel
            ws.column_dimensions['H'].width = 24  # Columna de Falla
            ws.column_dimensions['I'].width = 14  # Valor Actual
            ws.column_dimensions['J'].width = 60  # Mensaje de Inconsistencia
        else:
            for col_idx in range(2, ws.max_column + 1):
                vals = []
                for row_idx in range(1, min(20, ws.max_row + 1)):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val is not None:
                        val_str = str(val)
                        if val_str.startswith("=HYPERLINK"):
                            vals.append("Ver Registros")
                        else:
                            vals.append(val_str)
                if not vals:
                    continue
                max_len = max(len(v) for v in vals)
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 11), 52)

    return wb


# Alias para retrocompatibilidad
generar_excel_reporte_plt = export_plt_audit_to_excel
generar_excel_reporte_core_plt = export_plt_audit_to_excel
