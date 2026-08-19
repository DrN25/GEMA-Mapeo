"""
services/excel_exporter.py — Motor de exportación a Excel basado en plantilla maestra.

Carga la plantilla maestra 'backend/templates/plantilla.xlsx' y rellena la información
geomecánica en:
  1. Hoja 'ventana':
     - Ficha técnica completa de mapeo geomecánico.
     - Celdas con fórmulas existentes protegidas (K5, N6:N8, AS11:AZ12, etc.).
     - Inyección de fórmulas W..AI solo en filas de discontinuidades activas (evita #¡VALOR!).
     - Adaptación de fórmulas de espaciamiento y condición ponderada (AW11/12, AY11/12).
     - Expansión dinámica para > 14 estructuras (copia fila 28 con constante ALLOW_EXPAND_STRUCTURES).
     - Expansión dinámica para > 3 familias (duplica filas PROM y actualiza fórmula de Jv).
     - Re-enrutamiento de referencias RMR por desplazamiento (AT11/12 apuntando a fila real de Jv, AV11 promedios).
     - Sanitización estricta de valores centinela (-1, '-1', 'None', etc. -> celda vacía).
     - Soporte multi-celda con 2 filas de separación (fila inicio = fin_anterior + 3).
  2. Hoja 'BD':
     - Tabla base de datos continua (sin filas vacías de relleno) donde cada fila es una discontinuidad.
     - Lógica Fila Padre (cabecera completa + discontinuidad 1) y Filas Hijas (herencia de celda + discontinuidad K).
     - Coordenadas 3D calculadas (x, y, z) mediante fórmulas trigonométricas de azimut e inclinación.
     - Punteros INDIRECT automáticos a las filas exactas de la hoja 'ventana'.
"""

import io
import os
import re
import copy
import base64
import logging
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Union
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage

logger = logging.getLogger(__name__)

# =============================================================================
# CONSTANTES DE CONFIGURACIÓN
# =============================================================================
# Si True, permite expandir la tabla de discontinuidades si una ventana tiene > 14 estructuras.
ALLOW_EXPAND_STRUCTURES: bool = True

# Si True, permite expandir las filas de familias si una ventana tiene > 3 familias.
ALLOW_EXPAND_FAMILIES: bool = True

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DEFAULT_TEMPLATE_PATH = os.path.join(BASE_DIR, "templates", "plantilla.xlsx")


def _sanitize_val(val: Any) -> Any:
    """
    Convierte valores centinela (-1, '-1', 'None', etc.) en None (celda vacía en Excel)
    para que las fórmulas matemáticas y lógicas de Excel no fallen con #¡VALOR!.
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return None if val == -1 else val
    if isinstance(val, str):
        v = val.strip()
        if v in ["-1", "-1.0", "-1.00", "None", "NONE", "null", "NULL", ""]:
            return None
        return v
    return val


def _shift_formula_rows(formula: str, delta: int, min_row: int = 1) -> str:
    """
    Desplaza las referencias a filas dentro de una fórmula de Excel sumando `delta`,
    solo si el número de fila es >= `min_row`.
    """
    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        return formula

    pattern = r"('(?:[^']|'')*'!|[A-Za-z0-9_]+!)?(\$?)([A-Z]{1,3})(\$?)(\d+)"

    def repl_token(match: re.Match) -> str:
        sheet_prefix = match.group(1) or ""
        col_abs = match.group(2) or ""
        col = match.group(3)
        row_abs = match.group(4) or ""
        row_num_str = match.group(5)
        row = int(row_num_str)

        if row < min_row:
            return match.group(0)

        if sheet_prefix:
            return match.group(0)

        if row_abs == "$":
            return match.group(0)

        return f"{sheet_prefix}{col_abs}{col}{row + delta}"

    return re.sub(pattern, repl_token, formula)


def _load_pil_image(item: Any, uploads_dir: str) -> Optional[PILImage.Image]:
    """Carga una imagen en objeto PIL.Image desde archivo, URL, Base64 o BytesIO."""
    if item is None:
        return None
    try:
        if isinstance(item, PILImage.Image):
            return item
        if isinstance(item, bytes):
            return PILImage.open(io.BytesIO(item))
        if isinstance(item, io.BytesIO):
            item.seek(0)
            return PILImage.open(item)
        if isinstance(item, str):
            item_clean = item.strip()
            if not item_clean:
                return None
            if item_clean.startswith("data:image"):
                b64_data = item_clean.split(",", 1)[-1]
                decoded = base64.b64decode(b64_data)
                return PILImage.open(io.BytesIO(decoded))
            if "api/uploads/" in item_clean:
                rel_path = item_clean.split("api/uploads/", 1)[-1].split("?")[0]
                local_path = os.path.join(uploads_dir, rel_path)
                if os.path.exists(local_path):
                    return PILImage.open(local_path)
            if os.path.exists(item_clean):
                return PILImage.open(item_clean)
    except Exception as e:
        logger.warning(f"Error cargando imagen para exportación Excel: {e}")
    return None


def _get_cell_photos(codigo: str, fotos_input: Optional[List[Any]] = None) -> List[PILImage.Image]:
    """Obtiene hasta 2 imágenes PIL para la celda dada (del payload o del almacenamiento en disco)."""
    results: List[PILImage.Image] = []
    uploads_dir = os.path.join(BASE_DIR, "uploads")

    # 1. Si se pasaron fotos explícitas en el payload
    if fotos_input:
        for f in fotos_input[:2]:
            pil_img = _load_pil_image(f, uploads_dir)
            if pil_img:
                results.append(pil_img)

    # 2. Si faltan fotos, buscar en disco en uploads/{codigo}
    if len(results) < 2 and codigo:
        code_up = str(codigo).strip().upper()
        cell_dir = os.path.join(uploads_dir, code_up)
        if os.path.exists(cell_dir):
            allowed_exts = ["jpg", "jpeg", "png", "webp", "bmp", "gif", "svg", "tiff"]
            for i in range(2):
                if len(results) > i:
                    continue
                found_img = None
                for ext in allowed_exts:
                    path_a = os.path.join(cell_dir, f"{code_up}-VENTANA-{i+1}.{ext}")
                    if os.path.exists(path_a):
                        found_img = _load_pil_image(path_a, uploads_dir)
                        if found_img:
                            break
                    path_b = os.path.join(cell_dir, f"foto_{i}.{ext}")
                    if os.path.exists(path_b):
                        found_img = _load_pil_image(path_b, uploads_dir)
                        if found_img:
                            break
                if found_img:
                    results.append(found_img)

    return results[:2]


def _insert_photos_into_box(
    ws: Worksheet,
    base_row: int,
    codigo: str,
    fotos_input: Optional[List[Any]] = None
):
    """
    Inserta hasta 2 imágenes dentro del cuadro BD5:BT19 (filas base_row + 1 a base_row + 15),
    manteniendo su relación de aspecto. Si hay 2 fotos, la primera va arriba y la segunda abajo.
    """
    try:
        pil_images = _get_cell_photos(codigo, fotos_input)
        if not pil_images:
            return

        box_top_row = base_row + 1      # ej. 5
        box_mid_row = base_row + 9      # ej. 13
        max_box_w = 1350

        if len(pil_images) == 1:
            img = pil_images[0]
            w, h = img.size
            if w <= 0 or h <= 0:
                return
            max_h = 310
            scale = min(max_box_w / w, max_h / h)
            final_w = max(1, int(w * scale))
            final_h = max(1, int(h * scale))

            img_buf = io.BytesIO()
            img.convert("RGB").save(img_buf, format="PNG")
            img_buf.seek(0)

            xl_img = OpenpyxlImage(img_buf)
            xl_img.width = final_w
            xl_img.height = final_h
            xl_img.anchor = f"BD{box_top_row}"
            ws.add_image(xl_img)

        elif len(pil_images) >= 2:
            # 1. Foto 1 (Superior)
            img1 = pil_images[0]
            w1, h1 = img1.size
            if w1 > 0 and h1 > 0:
                max_h1 = 145
                scale1 = min(max_box_w / w1, max_h1 / h1)
                final_w1 = max(1, int(w1 * scale1))
                final_h1 = max(1, int(h1 * scale1))

                img1_buf = io.BytesIO()
                img1.convert("RGB").save(img1_buf, format="PNG")
                img1_buf.seek(0)

                xl_img1 = OpenpyxlImage(img1_buf)
                xl_img1.width = final_w1
                xl_img1.height = final_h1
                xl_img1.anchor = f"BD{box_top_row}"
                ws.add_image(xl_img1)

            # 2. Foto 2 (Inferior)
            img2 = pil_images[1]
            w2, h2 = img2.size
            if w2 > 0 and h2 > 0:
                max_h2 = 155
                scale2 = min(max_box_w / w2, max_h2 / h2)
                final_w2 = max(1, int(w2 * scale2))
                final_h2 = max(1, int(h2 * scale2))

                img2_buf = io.BytesIO()
                img2.convert("RGB").save(img2_buf, format="PNG")
                img2_buf.seek(0)

                xl_img2 = OpenpyxlImage(img2_buf)
                xl_img2.width = final_w2
                xl_img2.height = final_h2
                xl_img2.anchor = f"BD{box_mid_row}"
                ws.add_image(xl_img2)
    except Exception as err:
        logger.warning(f"No se pudieron insertar imágenes en celda {codigo}: {err}")


def _normalize_ventana_input(item: Any) -> Dict[str, Any]:
    """
    Normaliza cualquier objeto de entrada (ORM models.Ventana, schemas.VentanaResponseSchema,
    diccionario con excel_data + estructuras, o WindowData con header + joints) a una estructura unificada.
    """
    if isinstance(item, dict):
        fotos = item.get("fotos") or item.get("photos") or []
        # 1. Si tiene 'header' o 'joints' (formato WindowData de la interfaz activa de mapeo)
        if "header" in item or "joints" in item:
            header = item.get("header") or {}
            joints = item.get("joints") or item.get("estructuras") or item.get("discontinuidades") or []
            if "excel_data" in item and isinstance(item["excel_data"], dict):
                merged_header = {**item["excel_data"], **header}
            else:
                merged_header = header

            codigo = item.get("codigo") or item.get("codigo_final") or merged_header.get("celda") or merged_header.get("codigo") or "VENTANA"
            comentarios = item.get("comentarios") or merged_header.get("comentario") or merged_header.get("comentarios") or ""
            if not fotos:
                fotos = merged_header.get("fotos") or merged_header.get("photos") or []
            return {
                "codigo": codigo,
                "header": merged_header,
                "estructuras": joints,
                "comentarios": comentarios,
                "fotos": fotos
            }

        # 2. Si tiene 'excel_data' y 'estructuras' (formato escáner IA / importador)
        if "excel_data" in item:
            excel_data = item.get("excel_data") or {}
            estructuras = item.get("estructuras") or item.get("discontinuidades") or item.get("joints") or []
            codigo = item.get("codigo_final") or item.get("codigo") or excel_data.get("codigo") or excel_data.get("celda") or "VENTANA"
            comentarios = excel_data.get("comentarios") or excel_data.get("comentario") or ""
            if not fotos:
                fotos = excel_data.get("fotos") or excel_data.get("photos") or []
            return {
                "codigo": codigo,
                "header": excel_data,
                "estructuras": estructuras,
                "comentarios": comentarios,
                "fotos": fotos
            }

        # 3. Formato plano (diccionario genérico)
        estructuras = item.get("discontinuidades") or item.get("estructuras") or item.get("joints") or []
        codigo = item.get("codigo") or item.get("codigo_celda") or item.get("celda") or "VENTANA"
        comentarios = item.get("comentarios") or item.get("comentario") or ""
        return {
            "codigo": codigo,
            "header": item,
            "estructuras": estructuras,
            "comentarios": comentarios,
            "fotos": fotos
        }

    if hasattr(item, "codigo"):
        discs = getattr(item, "discontinuidades", []) or []
        rmr_in = getattr(item, "rmr_input", None)
        coment = getattr(rmr_in, "comentario", "") if rmr_in else ""
        fotos = getattr(item, "fotos", []) or []
        return {
            "codigo": getattr(item, "codigo", "VENTANA"),
            "header": item,
            "estructuras": discs,
            "comentarios": coment or "",
            "fotos": fotos
        }

    return {"codigo": "VENTANA", "header": {}, "estructuras": [], "comentarios": "", "fotos": []}


def _get_val(source: Any, *keys: str, default: Any = None) -> Any:
    """Extrae el primer valor no nulo de un diccionario u objeto por varias claves posibles."""
    if source is None:
        return default
    if isinstance(source, dict):
        for k in keys:
            if k in source and source[k] is not None:
                return source[k]
        return default
    for k in keys:
        if hasattr(source, k):
            v = getattr(source, k)
            if v is not None:
                return v
    return default


def _fill_sheet_ventana_block(
    ws: Worksheet,
    base_row: int,
    ventana_dict: Dict[str, Any]
) -> Dict[str, int]:
    """
    Rellena un bloque de celda en la hoja 'ventana' adaptándose dinámicamente:
      - Soporta > 14 estructuras (inserta filas adicionales si ALLOW_EXPAND_STRUCTURES=True).
      - Soporta > 3 familias (expande filas de promedios si ALLOW_EXPAND_FAMILIES=True).
      - Actualiza fórmulas de RMR y JV con los números de fila reales.

    Retorna un diccionario con los metadatos de filas calculadas:
      {"base_row": base_row, "start_struct_row": ..., "end_struct_row": ..., "jv_row": ..., "end_cell_row": ...}
    """
    codigo = ventana_dict.get("codigo") or "VENTANA"
    header = ventana_dict.get("header") or {}
    estructuras = ventana_dict.get("estructuras") or []
    comentarios = ventana_dict.get("comentarios") or _get_val(header, "comentarios", "comentario") or ""

    # =========================================================================
    # 1. CABECERA (Filas base_row a base_row + 4)
    # =========================================================================
    ws[f"A{base_row}"] = _sanitize_val(codigo)

    # FROM (B5, D5, F5)
    ws[f"B{base_row + 1}"] = _sanitize_val(_get_val(header, "este_ini", "este_from", "este_inicio", "este"))
    ws[f"D{base_row + 1}"] = _sanitize_val(_get_val(header, "norte_ini", "norte_from", "norte_inicio", "norte"))
    ws[f"F{base_row + 1}"] = _sanitize_val(_get_val(header, "cota_ini", "cota_from", "cota_inicio", "cota"))

    # TO (B6, D6, F6)
    ws[f"B{base_row + 2}"] = _sanitize_val(_get_val(header, "este_fin", "este_to", "este_final"))
    ws[f"D{base_row + 2}"] = _sanitize_val(_get_val(header, "norte_fin", "norte_to", "norte_final"))
    ws[f"F{base_row + 2}"] = _sanitize_val(_get_val(header, "cota_fin", "cota_to", "cota_final"))

    # Altura (K6)
    ws[f"K{base_row + 2}"] = _sanitize_val(_get_val(header, "altura_m", "altura", "alto"))

    # Dip_talud (N5)
    ws[f"N{base_row + 1}"] = _sanitize_val(_get_val(header, "dip_talud", "inclinacion", "inclinacion_talud"))

    # Lito 3 (P4 y P7)
    lito3_val = _sanitize_val(_get_val(header, "lito_3", "litologia_3", "lito3", "litologia", "lito_1"))
    ws[f"P{base_row}"] = lito3_val
    ws[f"P{base_row + 3}"] = lito3_val

    # Alteracion (P5) e Intemperismo (P6)
    ws[f"P{base_row + 1}"] = _sanitize_val(_get_val(header, "alteracion", "alt_mapeo"))
    ws[f"P{base_row + 2}"] = _sanitize_val(_get_val(header, "intemperismo", "intemperia"))

    # Sector (U4 y U7)
    sector_val = _sanitize_val(_get_val(header, "sector", "sector_geotecnico", "sect_geot"))
    ws[f"U{base_row}"] = sector_val
    ws[f"U{base_row + 3}"] = sector_val

    # Fase (U5) | Nivel (U6)
    ws[f"U{base_row + 1}"] = _sanitize_val(_get_val(header, "fase"))
    ws[f"U{base_row + 2}"] = _sanitize_val(_get_val(header, "nivel"))

    # Mapeador (P8)
    ws[f"P{base_row + 4}"] = _sanitize_val(_get_val(header, "mapeador"))

    # Fecha (AK4)
    ws[f"AK{base_row}"] = _sanitize_val(_get_val(header, "fecha", "fecha_mapeo"))

    # Comentarios (BD21 en bloque 1, o base_row + 17)
    coment_row = base_row + 17
    ws[f"BD{coment_row}"] = _sanitize_val(comentarios)

    # =========================================================================
    # 2. RMR & GSI INPUTS (Filas base_row + 7 y base_row + 8 -> ej. 11 y 12)
    # =========================================================================
    r_76 = base_row + 7  # 11
    r_89 = base_row + 8  # 12

    # Agua
    agua_76 = _sanitize_val(_get_val(header, "condicion_agua_rmr76", "condicion_agua", "agua_codigo"))
    agua_89 = _sanitize_val(_get_val(header, "condicion_agua_rmr89", "condicion_agua", "agua_codigo")) or agua_76
    ws[f"AJ{r_76}"] = agua_76
    ws[f"AJ{r_89}"] = agua_89

    # Dureza
    dureza_76 = _sanitize_val(_get_val(header, "dureza_rmr76", "resistencia_ucs", "dureza", "resistencia_codigo"))
    dureza_89 = _sanitize_val(_get_val(header, "dureza_rmr89", "resistencia_ucs", "dureza", "resistencia_codigo")) or dureza_76
    ws[f"AL{r_76}"] = dureza_76
    ws[f"AL{r_89}"] = dureza_89

    # GSI Superficie y Estructura
    gsi_sup = _sanitize_val(_get_val(header, "gsi_superficie"))
    gsi_est = _sanitize_val(_get_val(header, "gsi_estructura"))
    ws[f"AN{r_76}"] = gsi_sup
    ws[f"AN{r_89}"] = gsi_sup
    ws[f"AO{r_76}"] = gsi_est
    ws[f"AO{r_89}"] = gsi_est

    # GSI Visual
    gsi_vis_76 = _sanitize_val(_get_val(header, "gsi_visual_rmr76", "gsi_visual"))
    gsi_vis_89 = _sanitize_val(_get_val(header, "gsi_visual_rmr89", "gsi_visual")) or gsi_vis_76
    ws[f"AP{r_76}"] = gsi_vis_76
    ws[f"AP{r_89}"] = gsi_vis_89

    # Control Estructural
    ctrl_76 = _sanitize_val(_get_val(header, "control_estructural_rmr76", "control_estructural"))
    ctrl_89 = _sanitize_val(_get_val(header, "control_estructural_rmr89", "control_estructural")) or ctrl_76
    ws[f"AQ{r_76}"] = ctrl_76
    ws[f"AQ{r_89}"] = ctrl_89

    # Efectos Voladura
    vol_76 = _sanitize_val(_get_val(header, "efectos_voladura_rmr76", "efectos_voladura"))
    vol_89 = _sanitize_val(_get_val(header, "efectos_voladura_rmr89", "efectos_voladura")) or vol_76
    ws[f"AR{r_76}"] = vol_76
    ws[f"AR{r_89}"] = vol_89

    # UCS e Is50
    ucs_val = _sanitize_val(_get_val(header, "ucs_mpa"))
    is50_val = _sanitize_val(_get_val(header, "is50_mpa"))
    ws[f"BA{r_76}"] = ucs_val
    ws[f"BA{r_89}"] = ucs_val
    ws[f"BB{r_76}"] = is50_val
    ws[f"BB{r_89}"] = is50_val

    # =========================================================================
    # 3. EXPANDIR FILAS SI NUM_STRUCTS > 14 (Notas Líneas 42-46)
    # =========================================================================
    start_struct_row = base_row + 11  # 15
    num_structs = len(estructuras)
    base_slots = 14  # filas 15 a 28

    extra_struct_rows = 0
    if num_structs > base_slots and ALLOW_EXPAND_STRUCTURES:
        extra_struct_rows = num_structs - base_slots
        insert_idx = start_struct_row + base_slots  # insertar después de la fila 28
        ws.insert_rows(idx=insert_idx, amount=extra_struct_rows)
        # Copiar estilos de la fila anterior a las nuevas filas
        ref_row = insert_idx - 1
        for extra_i in range(extra_struct_rows):
            new_r = insert_idx + extra_i
            ws.row_dimensions[new_r].height = ws.row_dimensions[ref_row].height
            for c in range(1, ws.max_column + 1):
                src = ws.cell(row=ref_row, column=c)
                tgt = ws.cell(row=new_r, column=c)
                if src.has_style:
                    tgt.font = copy.copy(src.font)
                    tgt.border = copy.copy(src.border)
                    tgt.fill = copy.copy(src.fill)
                    tgt.number_format = src.number_format
                    tgt.protection = copy.copy(src.protection)
                    tgt.alignment = copy.copy(src.alignment)

    total_struct_slots = base_slots + extra_struct_rows

    # Fórmulas de subratings base apuntando a la hoja RMR!
    subrating_formulas_template = {
        "W": "=VLOOKUP(V{r},RMR!$A$40:$C$45,3,FALSE)",
        "X": "=IF(O{r}=\"\",\"error\",MIN(O{r}:P{r}))",
        "Y": "=VLOOKUP(I{r},RMR!$A$31:$D$35,4,TRUE)",
        "Z": "=IF(G{r}=\"\",\"error\",VLOOKUP(G{r},RMR!$B$51:$E$55,4,TRUE))",
        "AA": "=VLOOKUP(T{r},RMR!$A$59:$C$67,3,TRUE)",
        "AB": "=SUM(X{r}+Y{r}+Z{r}+AA{r}+W{r})",
        "AC": "=VLOOKUP(V{r},RMR!$A$40:$C$45,2,FALSE)",
        "AD": "=IF(Q{r}=\"\",\"error\",MIN(Q{r}:R{r}))",
        "AE": "=VLOOKUP(I{r},RMR!$A$31:$D$35,3,TRUE)",
        "AF": "=IF(G{r}=\"\",\"error\",VLOOKUP(G{r},RMR!$B$51:$E$55,3,TRUE))",
        "AG": "=VLOOKUP(T{r},RMR!$A$59:$C$67,2,TRUE)",
        "AH": "=SUM(AC{r}+AD{r}+AE{r}+AF{r}+AG{r})",
        "AI": "=(5/J{r})/3"
    }

    # Fórmulas auxiliares de relleno y ratings O..R y BW..CA
    aux_fill_formulas_template = {
        "BW": "=IF(H{r}=\"\",\"error\",VLOOKUP(H{r},RMR!$N$73:$P$74,3,TRUE))",
        "BX": "=IF(OR(M{r}=-1,M{r}=\"\"),\"\",VLOOKUP(M{r},RMR!$A$72:$B$79,2,FALSE))",
        "BY": "=IF(BX{r}=3,5,IF(BX{r}=\"\",\"\",IF(AND(BX{r}=1,BW{r}=1),1,IF(AND(BX{r}=1,BW{r}=2),2,IF(AND(BX{r}=2,BW{r}=1),3,IF(AND(BX{r}=2,BW{r}=2),4,\"error\"))))))",
        "BZ": "=IF(OR(N{r}=-1,N{r}=\"\"),\"\",VLOOKUP(N{r},RMR!$A$73:$B$79,2,FALSE))",
        "CA": "=IF(BZ{r}=\"\",\"\",IF(AND(BZ{r}=1,BW{r}=1),1,IF(AND(BZ{r}=1,BW{r}=2),2,IF(AND(BZ{r}=2,BW{r}=1),3,IF(AND(BZ{r}=2,BW{r}=2),4,\"error\")))))",
        "O": "=IF(BY{r}=\"\",\"\",VLOOKUP(BY{r},RMR!$N$77:$Q$81,4,TRUE))",
        "P": "=IF(CA{r}=\"\",\"\",VLOOKUP(CA{r},RMR!$N$77:$Q$81,4,TRUE))",
        "Q": "=IF(BY{r}=\"\",\"\",VLOOKUP(BY{r},RMR!$N$77:$Q$81,3,TRUE))",
        "R": "=IF(CA{r}=\"\",\"\",VLOOKUP(CA{r},RMR!$N$76:$Q$79,3,TRUE))"
    }

    # =========================================================================
    # 4. INYECTAR DATOS EN TABLA DE DISCONTINUIDADES
    # =========================================================================
    active_rows_count = 0
    fam_to_rows: Dict[Any, List[int]] = {}

    for slot_idx in range(total_struct_slots):
        curr_row = start_struct_row + slot_idx
        if slot_idx < num_structs:
            s = estructuras[slot_idx]
            fam = _sanitize_val(_get_val(s, "familia_id", "familia", "numero_estructura"))
            dist = _sanitize_val(_get_val(s, "distancia_m", "distancia"))
            tipo = _sanitize_val(_get_val(s, "tipo_estructura", "tipo"))
            dip = _sanitize_val(_get_val(s, "dip"))
            dip_dir = _sanitize_val(_get_val(s, "dip_dir", "dipdir"))

            is_active = any(x is not None for x in [dist, tipo, dip, dip_dir])

            # Inyectar datos de entrada
            ws[f"A{curr_row}"] = fam
            ws[f"B{curr_row}"] = dist
            ws[f"C{curr_row}"] = tipo
            ws[f"D{curr_row}"] = dip
            ws[f"E{curr_row}"] = dip_dir
            ws[f"F{curr_row}"] = _sanitize_val(_get_val(s, "n_estructuras", "nstr", "numero_estructuras"))
            ws[f"G{curr_row}"] = _sanitize_val(_get_val(s, "abertura_mm", "abertura", "aber"))
            ws[f"H{curr_row}"] = _sanitize_val(_get_val(s, "espesor_mm", "espesor", "esp"))
            ws[f"I{curr_row}"] = _sanitize_val(_get_val(s, "continuidad_m", "continuidad", "cont"))
            ws[f"J{curr_row}"] = _sanitize_val(_get_val(s, "espaciamiento_m", "espaciamiento", "espac"))
            ws[f"K{curr_row}"] = _sanitize_val(_get_val(s, "n_extremos_visibles", "extremos_visibles", "next"))
            ws[f"L{curr_row}"] = _sanitize_val(_get_val(s, "terminacion", "term"))
            ws[f"M{curr_row}"] = _sanitize_val(_get_val(s, "relleno_1_codigo", "relleno1", "r1"))
            ws[f"N{curr_row}"] = _sanitize_val(_get_val(s, "relleno_2_codigo", "relleno2", "r2"))
            ws[f"S{curr_row}"] = _sanitize_val(_get_val(s, "jrc"))
            ws[f"T{curr_row}"] = _sanitize_val(_get_val(s, "rugosidad_codigo", "rugosidad", "rug"))
            ws[f"U{curr_row}"] = _sanitize_val(_get_val(s, "forma_estructura", "forma"))
            ws[f"V{curr_row}"] = _sanitize_val(_get_val(s, "alteracion_codigo", "alteracion", "alt"))

            if is_active:
                active_rows_count += 1
                if fam is not None:
                    fam_to_rows.setdefault(fam, []).append(curr_row)

                # Inyectar fórmulas adaptadas para esta fila en W..AI
                for col_letter, formula_tmpl in subrating_formulas_template.items():
                    ws[f"{col_letter}{curr_row}"] = formula_tmpl.format(r=curr_row)

                # Inyectar fórmulas de relleno y ratings en BW..CA y O..R
                for col_letter, formula_tmpl in aux_fill_formulas_template.items():
                    ws[f"{col_letter}{curr_row}"] = formula_tmpl.format(r=curr_row)
            else:
                # Fila sin registro: limpiar fórmulas W..AI y O..R / BW..CA
                for col_letter in subrating_formulas_template.keys():
                    ws[f"{col_letter}{curr_row}"] = None
                for col_letter in aux_fill_formulas_template.keys():
                    ws[f"{col_letter}{curr_row}"] = None
        else:
            # Fila vacía de la plantilla: limpiar entradas y fórmulas
            for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "S", "T", "U", "V"]:
                ws[f"{col_letter}{curr_row}"] = None
            for col_letter in subrating_formulas_template.keys():
                ws[f"{col_letter}{curr_row}"] = None
            for col_letter in aux_fill_formulas_template.keys():
                ws[f"{col_letter}{curr_row}"] = None

        # Asegurar color de texto blanco en columna AI para todas las filas de estructuras (activas, inactivas o vacías)
        ws[f"AI{curr_row}"].font = openpyxl.styles.Font(name="Arial", size=7.0, color="FFFFFFFF")

    # Normalizar contornos y bordes en cajas laterales AJ..BB y BD..BT respetando la plantilla
    # 1. Caja Foto / Esquema (BD..BT, filas base_row + 1 a base_row + 15 -> ej. 5 a 19): contorno NEGRO MEDIO
    photo_top = base_row + 1
    photo_bottom = base_row + 15
    for r in range(photo_top, photo_bottom + 1):
        is_top = (r == photo_top)
        is_bottom = (r == photo_bottom)
        for c_idx in range(openpyxl.utils.column_index_from_string("BD"), openpyxl.utils.column_index_from_string("BT") + 1):
            col_let = openpyxl.utils.get_column_letter(c_idx)
            is_left = (col_let == "BD")
            is_right = (col_let == "BT")
            ws[f"{col_let}{r}"].border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style="medium") if is_left else None,
                right=openpyxl.styles.Side(style="medium") if is_right else None,
                top=openpyxl.styles.Side(style="medium") if is_top else None,
                bottom=openpyxl.styles.Side(style="medium") if is_bottom else None
            )

    # 2. Caja Comentarios (BD..BT, filas base_row + 17 a base_row + 24 + extra_struct_rows -> ej. 21 a 28): contorno PUNTEADO
    coment_top = base_row + 17
    coment_bottom = base_row + 24 + extra_struct_rows
    for r in range(coment_top, coment_bottom + 1):
        is_top = (r == coment_top)
        is_bottom = (r == coment_bottom)
        for c_idx in range(openpyxl.utils.column_index_from_string("BD"), openpyxl.utils.column_index_from_string("BT") + 1):
            col_let = openpyxl.utils.get_column_letter(c_idx)
            is_left = (col_let == "BD")
            is_right = (col_let == "BT")
            if r == coment_top and is_left:
                ws[f"{col_let}{r}"].border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style="dotted"),
                    right=openpyxl.styles.Side(style="dotted"),
                    top=openpyxl.styles.Side(style="dotted"),
                    bottom=openpyxl.styles.Side(style="dotted")
                )
            else:
                ws[f"{col_let}{r}"].border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style="dotted") if is_left else None,
                    right=openpyxl.styles.Side(style="dotted") if is_right else None,
                    top=openpyxl.styles.Side(style="dotted") if is_top else None,
                    bottom=openpyxl.styles.Side(style="dotted") if is_bottom else None
                )

    # 3. Caja RMR (AJ..BB, filas base_row + 9 a base_row + 24 + extra_struct_rows -> ej. 13 a 28): contorno NEGRO MEDIO
    rmr_top = base_row + 9
    rmr_bottom = base_row + 24 + extra_struct_rows
    for r in range(rmr_top, rmr_bottom + 1):
        is_top = (r == rmr_top)
        is_bottom = (r == rmr_bottom)
        for c_idx in range(openpyxl.utils.column_index_from_string("AJ"), openpyxl.utils.column_index_from_string("BB") + 1):
            col_let = openpyxl.utils.get_column_letter(c_idx)
            is_left = (col_let == "AJ")
            is_right = (col_let == "BB")
            ws[f"{col_let}{r}"].border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style="medium") if is_left else None,
                right=openpyxl.styles.Side(style="medium") if is_right else None,
                top=openpyxl.styles.Side(style="medium") if is_top else None,
                bottom=openpyxl.styles.Side(style="medium") if is_bottom else None
            )

    # =========================================================================
    # 5. EXPANDIR FAMILIAS (PROM Y JV) SI FAMILIAS > 3 (Notas Líneas 155-175)
    # =========================================================================
    family_base_start_row = start_struct_row + total_struct_slots  # ej. 29
    unique_fams = [f for f in sorted(list(fam_to_rows.keys()), key=lambda x: str(x)) if f is not None]
    total_fams = max(3, len(unique_fams))

    extra_fam_rows = 0
    if total_fams > 3 and ALLOW_EXPAND_FAMILIES:
        extra_fam_rows = total_fams - 3
        # Insertar filas extras antes de la última fila de JV
        insert_fam_idx = family_base_start_row + 2  # antes de la fila 31
        ws.insert_rows(idx=insert_fam_idx, amount=extra_fam_rows)
        ref_fam_row = family_base_start_row + 1
        for extra_fi in range(extra_fam_rows):
            new_fr = insert_fam_idx + extra_fi
            ws.row_dimensions[new_fr].height = ws.row_dimensions[ref_fam_row].height
            for c in range(1, ws.max_column + 1):
                src = ws.cell(row=ref_fam_row, column=c)
                tgt = ws.cell(row=new_fr, column=c)
                if src.has_style:
                    tgt.font = copy.copy(src.font)
                    tgt.border = copy.copy(src.border)
                    tgt.fill = copy.copy(src.fill)
                    tgt.number_format = src.number_format
                    tgt.protection = copy.copy(src.protection)
                    tgt.alignment = copy.copy(src.alignment)

    # Actualizar nombres y fórmulas de cada familia PROM
    prom_j_rows = []
    for fam_idx in range(total_fams):
        curr_fam_row = family_base_start_row + fam_idx
        ws[f"H{curr_fam_row}"] = f"PROM {fam_idx + 1}"

        # Si tenemos filas registradas para esta familia, adaptamos el promedio exacto
        if fam_idx < len(unique_fams) and unique_fams[fam_idx] in fam_to_rows:
            rows_for_fam = fam_to_rows[unique_fams[fam_idx]]
            min_r, max_r = min(rows_for_fam), max(rows_for_fam)
            ws[f"J{curr_fam_row}"] = f"=AVERAGE(J{min_r}:J{max_r})"
        else:
            # Rango estándar de 3 en 3
            def_start = start_struct_row + (fam_idx * 3)
            def_end = min(def_start + 2, start_struct_row + total_struct_slots - 1)
            ws[f"J{curr_fam_row}"] = f"=AVERAGE(J{def_start}:J{def_end})"

        prom_j_rows.append(curr_fam_row)

    # Fila final de JV
    jv_row = family_base_start_row + total_fams - 1
    ws[f"M{jv_row}"] = "JV"
    jv_terms = "+".join([f"(1/J{r})" for r in prom_j_rows])
    ws[f"N{jv_row}"] = f"=({jv_terms})"

    # =========================================================================
    # 6. RE-ENRUTAR FÓRMULAS RMR CON DESPLAZAMIENTO (Notas Líneas 94-113)
    # =========================================================================
    # AT11 y AT12: RQD % = 115 - 3.3 * N{jv_row}
    ws[f"AT{r_76}"] = f"=115-3.3*N{jv_row}"
    ws[f"AT{r_89}"] = f"=115-3.3*N{jv_row}"

    # AV11 y AV12: Tamaño de bloque =(J29+J30+...+Jn)/total_fams
    prom_j_sum = "+".join([f"J{r}" for r in prom_j_rows])
    ws[f"AV{r_76}"] = f"=({prom_j_sum})/{total_fams}"
    ws[f"AV{r_89}"] = f"=AV{r_76}"

    # AW11 y AW12: Espaciamiento ponderado =(J15*F15+...+Jn*Fn)/SUMA(F15:Fn)
    end_calc_row = start_struct_row + max(1, active_rows_count) - 1
    if active_rows_count > 0:
        prod_j_f = "+".join([f"J{r}*F{r}" for r in range(start_struct_row, end_calc_row + 1)])
        ws[f"AW{r_76}"] = f"=({prod_j_f})/SUM(F{start_struct_row}:F{end_calc_row})"
        ws[f"AW{r_89}"] = f"=AW{r_76}"

        # AY11: Condición Discontinuidad 76 =(AB15*F15+...+ABn*Fn)/SUMA(F15:Fn)
        prod_ab_f = "+".join([f"AB{r}*F{r}" for r in range(start_struct_row, end_calc_row + 1)])
        ws[f"AY{r_76}"] = f"=({prod_ab_f})/SUM(F{start_struct_row}:F{end_calc_row})"

        # AY12: Condición Discontinuidad 89 =(AH15*F15+...+AHn*Fn)/SUMA(F15:Fn)
        prod_ah_f = "+".join([f"AH{r}*F{r}" for r in range(start_struct_row, end_calc_row + 1)])
        ws[f"AY{r_89}"] = f"=({prod_ah_f})/SUM(F{start_struct_row}:F{end_calc_row})"

    # Insertar fotografías de la celda en la caja BD..BT (filas 5 a 19)
    fotos = (
        ventana_dict.get("fotos")
        or ventana_dict.get("photos")
        or (header.get("fotos") if isinstance(header, dict) else None)
        or []
    )
    _insert_photos_into_box(ws, base_row, codigo, fotos)

    return {
        "base_row": base_row,
        "start_struct_row": start_struct_row,
        "end_struct_row": start_struct_row + total_struct_slots - 1,
        "active_rows_count": active_rows_count,
        "jv_row": jv_row,
        "end_cell_row": jv_row,
        "estructuras": estructuras
    }


def _populate_sheet_bd(ws_bd: Worksheet, ventanas_meta: List[Dict[str, Any]]):
    """
    Construye la hoja 'BD' como una tabla de datos continua y compacta:
      - 0 filas vacías de relleno (si una ventana tiene 3 estructuras, genera 3 filas exactas).
      - Fila Padre: Cabecera completa + discontinuidad 1 + fórmulas INDIRECT a 'ventana'.
      - Filas Hijas: Hereda código y FROM/TO (=C2, =F2:L2) + discontinuidades subsecuentes.
      - Coordenadas 3D (x, y, z) calculadas trigonométricamente.
    """
    curr_bd_row = 2

    for v_meta in ventanas_meta:
        base_v_row = v_meta["base_row"]
        start_struct_v = v_meta["start_struct_row"]
        estructuras = v_meta.get("estructuras") or []
        num_structs = max(1, len(estructuras))
        padre_bd_row = curr_bd_row

        for struct_idx in range(num_structs):
            r = curr_bd_row
            struct_v_row = start_struct_v + struct_idx

            # A: id autoincremental
            ws_bd[f"A{r}"] = r - 1

            if struct_idx == 0:
                # =============================================================
                # FILA PADRE: Metadatos completos de la Celda
                # =============================================================
                # Punteros numéricos a las filas exactas de 'ventana'
                ws_bd[f"B{r}"] = base_v_row          # Código A{B}
                ws_bd[f"E{r}"] = base_v_row + 1      # FROM B{E}, D{E}, F{E}
                ws_bd[f"I{r}"] = base_v_row + 2      # TO B{I}, D{I}, F{I}, Altura K{I}
                ws_bd[f"O{r}"] = base_v_row + 3      # Dip hole N{O}
                ws_bd[f"Q{r}"] = base_v_row + 4      # Az hole N{Q}
                ws_bd[f"S{r}"] = base_v_row + 1      # Dip talud N{S}
                ws_bd[f"U{r}"] = base_v_row + 2      # DipDir talud N{U}, Intemperismo P{U}
                ws_bd[f"X{r}"] = base_v_row + 7      # RMR 76 AJ{X}..BB{X}
                ws_bd[f"AP{r}"] = base_v_row + 8     # RMR 89 AJ{AP}..AZ{AP}
                ws_bd[f"BF{r}"] = base_v_row         # Fecha AK{BF}
                ws_bd[f"BH{r}"] = base_v_row + 17    # Comentarios BD{BH}
                ws_bd[f"CF{r}"] = base_v_row + 4     # Mapeador / Geotécnico P{CF}
                ws_bd[f"CH{r}"] = base_v_row + 7     # Is50 BB{CH}
                ws_bd[f"CJ{r}"] = base_v_row + 3     # Lito Modelo P{CJ}
                ws_bd[f"CL{r}"] = base_v_row         # Lito-3 P{CL}
                ws_bd[f"CN{r}"] = base_v_row + 3     # Sector U{CN}
                ws_bd[f"CP{r}"] = base_v_row + 2     # Nivel U{CP}

                # Fórmulas de extracción dinámica
                ws_bd[f"C{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!A"&B{r})'
                ws_bd[f"D{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!A"&B{r})'
                ws_bd[f"F{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!B"&E{r})'
                ws_bd[f"G{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!D"&E{r})'
                ws_bd[f"H{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!F"&E{r})'
                ws_bd[f"J{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!B"&I{r})'
                ws_bd[f"K{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!D"&I{r})'
                ws_bd[f"L{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!F"&I{r})'
                ws_bd[f"M{r}"] = f'=(((J{r}-F{r})^2)+((K{r}-G{r})^2)+((L{r}-H{r})^2))^(1/2)'
                ws_bd[f"N{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!K"&I{r})'
                ws_bd[f"P{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!N"&O{r})'
                ws_bd[f"R{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!N"&Q{r})'
                ws_bd[f"T{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!N"&S{r})'
                ws_bd[f"V{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!N"&U{r})'
                ws_bd[f"W{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!P"&U{r})'
                ws_bd[f"Y{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AJ"&X{r})'
                ws_bd[f"Z{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AK"&X{r})'
                ws_bd[f"AA{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AL"&X{r})'
                ws_bd[f"AB{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AM"&X{r})'
                ws_bd[f"AC{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AP"&X{r})'
                ws_bd[f"AD{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AQ"&X{r})'
                ws_bd[f"AE{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AR"&X{r})'
                ws_bd[f"AF{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AS"&X{r})'
                ws_bd[f"AG{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AT"&X{r})'
                ws_bd[f"AH{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AU"&X{r})'
                ws_bd[f"AI{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AV"&X{r})'
                ws_bd[f"AJ{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AW"&X{r})'
                ws_bd[f"AK{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AX"&X{r})'
                ws_bd[f"AL{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AY"&X{r})'
                ws_bd[f"AM{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AZ"&X{r})'
                ws_bd[f"AN{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!BA"&X{r})'
                ws_bd[f"AO{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!BB"&X{r})'
                ws_bd[f"AQ{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AJ"&AP{r})'
                ws_bd[f"AR{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AK"&AP{r})'
                ws_bd[f"AS{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AL"&AP{r})'
                ws_bd[f"AT{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AM"&AP{r})'
                ws_bd[f"AU{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AP"&AP{r})'
                ws_bd[f"AV{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AQ"&AP{r})'
                ws_bd[f"AW{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AR"&AP{r})'
                ws_bd[f"AX{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AS"&AP{r})'
                ws_bd[f"AY{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AT"&AP{r})'
                ws_bd[f"AZ{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AU"&AP{r})'
                ws_bd[f"BA{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AV"&AP{r})'
                ws_bd[f"BB{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AW"&AP{r})'
                ws_bd[f"BC{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AX"&AP{r})'
                ws_bd[f"BD{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AY"&AP{r})'
                ws_bd[f"BE{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AZ"&AP{r})'
                ws_bd[f"BG{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!AK"&BF{r})'
                ws_bd[f"BI{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!BD"&BH{r})'
                ws_bd[f"CG{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!P"&CF{r})'
                ws_bd[f"CI{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!BB"&CH{r})'
                ws_bd[f"CK{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!P"&CJ{r})'
                ws_bd[f"CM{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!P"&CL{r})'
                ws_bd[f"CO{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!U"&CN{r})'
                ws_bd[f"CQ{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!U"&CP{r})'
                ws_bd[f"CR{r}"] = f'=AJ{r}*1000'
            else:
                # =============================================================
                # FILAS HIJAS: Herencia de celda y FROM/TO
                # =============================================================
                ws_bd[f"C{r}"] = f'=C{padre_bd_row}'
                ws_bd[f"F{r}"] = f'=F{padre_bd_row}'
                ws_bd[f"G{r}"] = f'=G{padre_bd_row}'
                ws_bd[f"H{r}"] = f'=H{padre_bd_row}'
                ws_bd[f"J{r}"] = f'=J{padre_bd_row}'
                ws_bd[f"K{r}"] = f'=K{padre_bd_row}'
                ws_bd[f"L{r}"] = f'=L{padre_bd_row}'

            # =================================================================
            # DISCONTINUIDADES: En todas las filas activas
            # =================================================================
            ws_bd[f"BJ{r}"] = struct_v_row
            ws_bd[f"BK{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!B"&BJ{r})'
            ws_bd[f"BL{r}"] = f'=_xlfn.ACOT((K{r}-G{r})/(J{r}-F{r}))'
            ws_bd[f"BM{r}"] = f'=IF(L{r}=H{r},0,_xlfn.ACOT((J{r}-F{r})/(L{r}-H{r})))'
            ws_bd[f"BN{r}"] = f'=BK{r}*COS(BL{r})+F{r}'
            ws_bd[f"BO{r}"] = f'=BK{r}*SIN(BL{r})+G{r}'
            ws_bd[f"BP{r}"] = f'=BK{r}*COS(BL{r})*SIN(BM{r})+H{r}'
            ws_bd[f"BQ{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!C"&BJ{r})'
            ws_bd[f"BR{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!D"&BJ{r})'
            ws_bd[f"BS{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!E"&BJ{r})'
            ws_bd[f"BT{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!F"&BJ{r})'
            ws_bd[f"BU{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!G"&BJ{r})'
            ws_bd[f"BV{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!H"&BJ{r})'
            ws_bd[f"BW{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!I"&BJ{r})'
            ws_bd[f"BX{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!J"&BJ{r})'
            ws_bd[f"BY{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!K"&BJ{r})'
            ws_bd[f"BZ{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!M"&BJ{r})'
            ws_bd[f"CA{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!N"&BJ{r})'
            ws_bd[f"CB{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!S"&BJ{r})'
            ws_bd[f"CC{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!T"&BJ{r})'
            ws_bd[f"CD{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!U"&BJ{r})'
            ws_bd[f"CE{r}"] = f'=INDIRECT("\'" & "ventana" & "\'!V"&BJ{r})'

            curr_bd_row += 1

    # Limpiar cualquier fila remanente si existía en la plantilla
    if ws_bd.max_row >= curr_bd_row:
        for r_clean in range(curr_bd_row, ws_bd.max_row + 1):
            for c_clean in range(1, ws_bd.max_column + 1):
                ws_bd.cell(row=r_clean, column=c_clean).value = None


def export_ventanas_to_excel(
    ventanas: Union[List[Any], Any],
    template_path: Optional[str] = None
) -> io.BytesIO:
    """
    Genera un archivo Excel (.xlsx) a partir de la plantilla maestra, inyectando
    los datos en la hoja 'ventana' y construyendo la tabla consolidada continua 'BD'.

    Retorna un buffer BytesIO con el contenido del workbook listo para streaming.
    """
    path = template_path or DEFAULT_TEMPLATE_PATH
    if not os.path.exists(path):
        raise FileNotFoundError(f"No se encontró la plantilla de Excel en: {path}")

    # Lista normalizada de ventanas
    if not isinstance(ventanas, list):
        ventanas_list = [ventanas]
    else:
        ventanas_list = ventanas

    if not ventanas_list:
        raise ValueError("Se debe proporcionar al menos una ventana para exportar.")

    normalized_ventanas = [_normalize_ventana_input(v) for v in ventanas_list]

    # Cargar workbook preservando fórmulas y estilos
    wb = openpyxl.load_workbook(path, data_only=False)

    if "ventana" not in wb.sheetnames:
        raise ValueError(f"La plantilla no contiene la hoja requerida 'ventana'. Hojas disponibles: {wb.sheetnames}")

    ws_ventana = wb["ventana"]

    # =========================================================================
    # 1. SNAPSHOT PRÍSTINO DEL BLOQUE MAESTRO (FILAS 4..31)
    # =========================================================================
    template_row_heights: Dict[int, Optional[float]] = {}
    for r in range(4, 32):
        template_row_heights[r] = ws_ventana.row_dimensions[r].height

    template_cells: Dict[tuple, Dict[str, Any]] = {}
    for r in range(4, 32):
        for c in range(1, ws_ventana.max_column + 1):
            cell = ws_ventana.cell(row=r, column=c)
            template_cells[(r, c)] = {
                "value": cell.value,
                "font": copy.copy(cell.font) if cell.has_style else None,
                "border": copy.copy(cell.border) if cell.has_style else None,
                "fill": copy.copy(cell.fill) if cell.has_style else None,
                "number_format": cell.number_format,
                "protection": copy.copy(cell.protection) if cell.has_style else None,
                "alignment": copy.copy(cell.alignment) if cell.has_style else None,
            }

    # Snapshot de todas las celdas combinadas de la plantilla en filas 4..31
    template_merges = []
    for rng in list(ws_ventana.merged_cells.ranges):
        if rng.min_row >= 4 and rng.max_row <= 31:
            template_merges.append((rng.min_row - 4, rng.max_row - 4, rng.min_col, rng.max_col))

    # =========================================================================
    # 2. INYECTAR HOJA 1: 'ventana'
    # =========================================================================
    ventanas_meta: List[Dict[str, Any]] = []

    # Bloque 1: Celda 1 empieza siempre en base_row = 4
    meta_1 = _fill_sheet_ventana_block(ws_ventana, 4, normalized_ventanas[0])
    ventanas_meta.append(meta_1)

    current_end_row = meta_1["end_cell_row"]
    spacing = 2  # 2 filas en blanco de separación

    for v_idx in range(1, len(normalized_ventanas)):
        # Dejamos 2 filas vacías exactas (current_end_row + 1, current_end_row + 2)
        next_base_row = current_end_row + spacing + 1
        row_offset = next_base_row - 4

        # A) Clonar alturas exactas de fila de la plantilla
        for r_orig in range(4, 32):
            tgt_r = next_base_row + (r_orig - 4)
            ws_ventana.row_dimensions[tgt_r].height = template_row_heights[r_orig]

        # B) Clonar celdas, fórmulas con desplazamiento seguro y estilos
        for (r_orig, c), record in template_cells.items():
            tgt_r = next_base_row + (r_orig - 4)
            tgt_cell = ws_ventana.cell(row=tgt_r, column=c)

            val = record["value"]
            if val is not None:
                if str(val).startswith("="):
                    tgt_cell.value = _shift_formula_rows(val, row_offset)
                else:
                    tgt_cell.value = val

            if record["font"]: tgt_cell.font = copy.copy(record["font"])
            if record["border"]: tgt_cell.border = copy.copy(record["border"])
            if record["fill"]: tgt_cell.fill = copy.copy(record["fill"])
            if record["number_format"]: tgt_cell.number_format = record["number_format"]
            if record["protection"]: tgt_cell.protection = copy.copy(record["protection"])
            if record["alignment"]: tgt_cell.alignment = copy.copy(record["alignment"])

        # C) Replicar todas las celdas combinadas (merged_cells) de la plantilla
        for r_start_off, r_end_off, min_col, max_col in template_merges:
            ws_ventana.merge_cells(
                start_row=next_base_row + r_start_off,
                end_row=next_base_row + r_end_off,
                start_column=min_col,
                end_column=max_col
            )

        # D) Inyectar datos en el nuevo bloque limpio
        meta_next = _fill_sheet_ventana_block(ws_ventana, next_base_row, normalized_ventanas[v_idx])
        ventanas_meta.append(meta_next)
        current_end_row = meta_next["end_cell_row"]

    # =========================================================================
    # 3. INYECTAR HOJA 2: 'BD'
    # =========================================================================
    if "BD" in wb.sheetnames:
        _populate_sheet_bd(wb["BD"], ventanas_meta)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
