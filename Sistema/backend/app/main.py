from fastapi import FastAPI, HTTPException, Depends, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from typing import List, Dict, Any
from datetime import date, datetime
from collections import Counter, defaultdict
import io
import openpyxl
import math
import os
import shutil
import json
from fastapi.staticfiles import StaticFiles

import sys
from fastapi import BackgroundTasks

from app.database import get_db, Base, engine
from app import models, schemas, calculator
from sqlalchemy import text


# Agregar el directorio raíz al path para poder importar tu validador_geomecanico.py
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
try:
    from validador_geomecanico import validate_bulk_excel
except ImportError:
    # Fallback preventivo si se ejecuta desde otro contexto de directorios
    sys.path.append(os.path.dirname(os.path.abspath(__file__)))
    from validador_geomecanico import validate_bulk_excel

# Caches globales en memoria para respuestas instantáneas (< 10ms) al frontend
DIAGNOSTIC_CACHE = None
COMPACT_CACHE = None

# Auto-migrate database tables to add missing columns on startup
try:
    with engine.begin() as conn:
        try:
            conn.execute(text("ALTER TABLE ventana ADD turno VARCHAR(50) NULL"))
        except Exception:
            pass
        try:
            conn.execute(text("ALTER TABLE ventanas_final ADD turno VARCHAR(50) NULL"))
        except Exception:
            pass
        try:
            conn.execute(text("ALTER TABLE ventanas_final ADD campania INT NULL"))
        except Exception:
            pass
except Exception as e:
    print(f"Error checking/adding database columns: {e}")

app = FastAPI(title="Geomechanical Window Mapping API", version="1.0")

# Crear carpeta de uploads física si no existe
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
uploads_dir = os.path.join(BASE_DIR, "uploads")
os.makedirs(uploads_dir, exist_ok=True)

# Servir estáticos de forma pública desde /uploads
app.mount("/api/uploads", StaticFiles(directory=uploads_dir), name="uploads")

# Enable CORS for Vite frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# HELPER: Sync data to flat table ventanas_final
def sync_to_ventanas_final(db: Session, ventana_id: int):
    # Load normalized entities
    v = db.query(models.Ventana).filter_by(ventana_id=ventana_id).first()
    if not v:
        return
    
    # Run geomechanical calculations
    rows_data = []
    for d in v.discontinuidades:
        rows_data.append({
            "fam": d.familia_id,
            "dist": float(d.distancia_m) if d.distancia_m is not None else None,
            "tipo": d.tipo_estructura,
            "dip": float(d.dip),
            "dipdir": float(d.dip_dir),
            "aber": float(d.abertura_mm) if d.abertura_mm is not None else None,
            "esp": float(d.espesor_mm) if d.espesor_mm is not None else None,
            "cont": float(d.continuidad_m) if d.continuidad_m is not None else None,
            "espac": float(d.espaciamiento_m),
            "nstr": float(d.n_estructuras) if d.n_estructuras is not None else None,
            "next": d.n_extremos_visibles,
            "term": d.terminacion,
            "r1": d.relleno_1_codigo,
            "r2": d.relleno_2_codigo,
            "jrc": d.jrc,
            "rug": d.rugosidad_codigo,
            "forma": d.forma_estructura,
            "alt": d.alteracion_codigo
        })
        
    rmr_data = {
        "agua_codigo": v.rmr_input.agua_codigo if v.rmr_input else "C",
        "resistencia_codigo": v.rmr_input.resistencia_codigo if v.rmr_input else "R4",
        "gsi_estructura": v.rmr_input.gsi_estructura if v.rmr_input else "VB",
        "gsi_superficie": v.rmr_input.gsi_superficie if v.rmr_input else "G",
        "gsi_visual": v.rmr_input.gsi_visual if v.rmr_input else 50,
        "control_estructural": v.rmr_input.control_estructural if v.rmr_input else 4,
        "efectos_voladura": v.rmr_input.efectos_voladura if v.rmr_input else 3,
        "ucs_mpa": float(v.rmr_input.ucs_mpa) if (v.rmr_input and v.rmr_input.ucs_mpa is not None) else 74.0,
        "is50_mpa": float(v.rmr_input.is50_mpa) if (v.rmr_input and v.rmr_input.is50_mpa is not None) else 5.0,
        "comentario": v.rmr_input.comentario if v.rmr_input else ""
    }
    
    # Asegurar redondeo estricto a entero de largo_m
    largo_entero = int(round(float(v.largo_m))) if v.largo_m is not None else None

    header_data = {
        "este_ini": v.este_ini,
        "norte_ini": v.norte_ini,
        "cota_ini": v.cota_ini,
        "este_fin": v.este_fin,
        "norte_fin": v.norte_fin,
        "cota_fin": v.cota_fin,
        "largo_m": largo_entero
    }
    
    res = calculator.calculate_geomechanics(header_data, rows_data, rmr_data)
    
    # Delete existing rows in ventanas_final for this celda
    db.query(models.VentanasFinal).filter_by(celda=v.codigo).delete()
    db.flush()
    
    # Query current max id to manually assign
    max_id = db.execute(text("SELECT MAX(id) FROM ventanas_final")).scalar()
    next_id = (max_id or 0) + 1
    
    # Insert flat rows
    for r_idx, r_calc in enumerate(res["rows"]):
        row_norm = r_calc["row"]
        
        # nstr guardado como None si es -1
        final_nstr = int(row_norm["nstr"]) if (row_norm["nstr"] is not None and row_norm["nstr"] != -1) else None

        final_row = models.VentanasFinal(
            id=next_id,
            celda=v.codigo,
            este_from=float(v.este_ini),
            norte_from=float(v.norte_ini),
            cota_from=float(v.cota_ini),
            este_to=float(v.este_fin),
            norte_to=float(v.norte_fin),
            cota_to=float(v.cota_fin),
            dist_celda=largo_entero,
            altura=float(v.altura_m) if v.altura_m is not None else None,
            dip=float(v.dip_hw) if v.dip_hw is not None else (r_calc["alfa"] * 180 / math.pi if r_calc["alfa"] else None),  
            az_hole=float(v.az_hw) if v.az_hw is not None else (r_calc["teta"] * 180 / math.pi if r_calc["teta"] else None),
            dip_talud=float(v.dip_talud),
            dip_dir_talud=float(v.dipdir_talud) if v.dipdir_talud is not None else (float(v.dip_talud) + 90) % 360,
            intemperismo=v.intemperismo_codigo,
            cond_agua_76=v.rmr_input.agua_codigo if v.rmr_input else "C",
            cond_agua_valor_76=res["agua_r76"],
            dureza_76=v.rmr_input.resistencia_codigo if v.rmr_input else "R4",
            resistencia_est_valor_76=res["resist_r76"],
            gsi_visual_76=v.rmr_input.gsi_visual if v.rmr_input else 50,
            control_estructural_76=v.rmr_input.control_estructural if v.rmr_input else 4,
            efectos_voladura_76=v.rmr_input.efectos_voladura if v.rmr_input else 3,
            rqd_valor_76=res["rqd_r76"],
            rqd_76=res["rqd_pct"],
            freq_fractura_m_76=res["jv"],
            tam_bloques_m3_76=res["espac_prom"]**3 if res["espac_prom"] else None,
            espaciamiento_prom_76=res["espac_prom"],
            espaciamiento_valor_76=res["spacing_r76"],
            cond_discontinuidad_valor_76=res["condisc_r76"],
            rmr_76=res["rmr_76"],
            ucs_mpa=float(v.rmr_input.ucs_mpa) if v.rmr_input else None,
            is50_mpa=float(v.rmr_input.is50_mpa) if v.rmr_input else None,
            cond_agua_89=v.rmr_input.agua_codigo if v.rmr_input else "C",
            cond_agua_valor_89=res["agua_r89"],
            dureza_89=v.rmr_input.resistencia_codigo if v.rmr_input else "R4",
            resistencia_est_valor_89=res["resist_r89"],
            gsi_visual_89=v.rmr_input.gsi_visual if v.rmr_input else 50,
            control_estructural_89=v.rmr_input.control_estructural if v.rmr_input else 4,
            efecto_voladura_89=v.rmr_input.efectos_voladura if v.rmr_input else 3,
            rqd_valor_89=res["rqd_r89"],
            rqd_89=res["rqd_pct"],
            freq_fractura_m_89=res["jv"],
            tam_bloques_m3_89=res["espac_prom"]**3 if res["espac_prom"] else None,
            espaciamiento_prom_89=res["espac_prom"],
            espaciamiento_valor_89=res["spacing_r89"],
            cond_discontinuidad_valor_89=res["condisc_r89"],
            rmr_89=res["rmr_89"],
            fecha=datetime.combine(v.fecha_mapeo, datetime.min.time()) if v.fecha_mapeo else None,
            comentario=v.rmr_input.comentario if v.rmr_input else "",
            dist_estructura=row_norm["dist"],
            angulo_estruct_teta=r_calc["teta"],
            angulo_estruct_alfa=r_calc["alfa"],
            estruct_x=r_calc["wx"],
            struct_y=r_calc["wy"],
            struct_z=r_calc["wz"],
            tipo_estructura=row_norm["tipo"],
            dip_estructura=row_norm["dip"],
            dip_dir_estructura=row_norm["dipdir"],
            num_estructuras=final_nstr,
            abertura_mm=row_norm["aber"] if row_norm["aber"] is not None else 0.0,
            espesor_mm=row_norm["esp"] if row_norm["esp"] is not None else 0.0,
            continuidad_m=row_norm["cont"] if row_norm["cont"] is not None else 0.0,
            espaciamiento_m=row_norm["espac"],
            num_extremos_visibles=row_norm["next"],
            tipo_relleno_1=row_norm["r1"] if row_norm["r1"] else "cwf",
            tipo_relleno_2=row_norm["r2"] if row_norm["r2"] else "-1",
            jrc=row_norm["jrc"],
            rugosidad_estructuras=row_norm["rug"] if row_norm["rug"] is not None else 1,
            forma_estructura=row_norm["forma"] if row_norm["forma"] else "P",
            alteracion=row_norm["alt"] if row_norm["alt"] else "f",
            geotecnico=v.mapeador,
            nivel=v.nivel,
            lito_1=v.lito_1,
            lito_2=v.lito_2,
            lito_3=v.lito_3,
            unidad_litologica=v.unidad_litologica,
            sector_geotecnico=v.sector_geotecnico if v.sector_geotecnico else "E1",
            campania=v.campania if v.campania is not None else 2026,
            turno=v.turno
        )
        db.add(final_row)
        next_id += 1
    db.flush()

# API ENDPOINTS

@app.get("/api/ventanas", response_model=List[schemas.VentanaSummarySchema])
def get_ventanas(db: Session = Depends(get_db)):
    ventanas = db.query(models.Ventana).all()
    res = []
    for v in ventanas:
        res.append(schemas.VentanaSummarySchema(
            codigo=v.codigo,
            fecha_mapeo=v.fecha_mapeo,
            mapeador=v.mapeador,
            lito_1=v.lito_1,
            discontinuidades_count=len(v.discontinuidades),
            creado_en=v.creado_en
        ))
    return res

@app.get("/api/ventanas/{codigo}", response_model=schemas.VentanaSaveSchema, response_model_by_alias=False)
def get_ventana(codigo: str, db: Session = Depends(get_db)):
    v = db.query(models.Ventana).filter_by(codigo=codigo.strip().upper()).first()
    if not v:
        raise HTTPException(status_code=404, detail="Ventana no encontrada")
    
    # Construct schema
    discs = []
    for d in v.discontinuidades:
        discs.append(schemas.DiscontinuidadBase(
            familia_id=d.familia_id,
            distancia_m=float(d.distancia_m) if d.distancia_m is not None else None,
            tipo_estructura=d.tipo_estructura,
            dip=float(d.dip),
            dip_dir=float(d.dip_dir),
            abertura_mm=float(d.abertura_mm) if d.abertura_mm is not None else None,
            espesor_mm=float(d.espesor_mm) if d.espesor_mm is not None else None,
            continuidad_m=float(d.continuidad_m) if d.continuidad_m is not None else None,
            espaciamiento_m=float(d.espaciamiento_m),
            n_estructuras=float(d.n_estructuras) if d.n_estructuras is not None else -1.0, # Retornar -1.0 si es NULL
            n_extremos_visibles=d.n_extremos_visibles,
            terminacion=d.terminacion,
            relleno_1_codigo=d.relleno_1_codigo,
            relleno_2_codigo=d.relleno_2_codigo,
            jrc=d.jrc,
            rugosidad_codigo=d.rugosidad_codigo,
            forma_estructura=d.forma_estructura,
            alteracion_codigo=d.alteracion_codigo
        ))
        
    rmr = None
    if v.rmr_input:
        rmr = schemas.VentanaRmrInputBase(
            agua_codigo=v.rmr_input.agua_codigo,
            resistencia_codigo=v.rmr_input.resistencia_codigo,
            gsi_estructura=v.rmr_input.gsi_estructura,
            gsi_superficie=v.rmr_input.gsi_superficie,
            gsi_visual=v.rmr_input.gsi_visual,
            control_estructural=v.rmr_input.control_estructural,
            efectos_voladura=v.rmr_input.efectos_voladura,
            ucs_mpa=float(v.rmr_input.ucs_mpa) if v.rmr_input.ucs_mpa is not None else None,
            is50_mpa=float(v.rmr_input.is50_mpa) if v.rmr_input.is50_mpa is not None else None,
            comentario=v.rmr_input.comentario
        )
        
    return schemas.VentanaSaveSchema(
        codigo=v.codigo,
        fecha_mapeo=v.fecha_mapeo,
        mapeador=v.mapeador,
        campania=v.campania,
        este_ini=float(v.este_ini),
        norte_ini=float(v.norte_ini),
        cota_ini=float(v.cota_ini),
        este_fin=float(v.este_fin),
        norte_fin=float(v.norte_fin),
        cota_fin=float(v.cota_fin),
        largo_m=int(round(float(v.largo_m))) if v.largo_m is not None else None, # Retornar como entero
        altura_m=float(v.altura_m) if v.altura_m is not None else None,
        dip_talud=float(v.dip_talud),
        dipdir_talud=float(v.dipdir_talud) if v.dipdir_talud is not None else None,
        dip_hw=float(v.dip_hw) if v.dip_hw is not None else None,
        az_hw=float(v.az_hw) if v.az_hw is not None else None,
        alteracion_codigo=v.alteracion_codigo,
        intemperismo_codigo=v.intemperismo_codigo,
        lito_1=v.lito_1,
        lito_2=v.lito_2,
        lito_3=v.lito_3,
        unidad_litologica=v.unidad_litologica,
        sector=v.sector,
        fase=v.fase,
        nivel=v.nivel,
        sector_geotecnico=v.sector_geotecnico,
        turno=v.turno,
        discontinuidades=discs,
        rmr_input=rmr
    )

@app.post("/api/ventanas")
def save_ventana(data: schemas.VentanaSaveSchema, db: Session = Depends(get_db)):
    code_up = data.codigo.strip().upper()
    v = db.query(models.Ventana).filter_by(codigo=code_up).first()

    def clean_null_val(val):
        if val == -1 or val == -1.0 or val == "-1" or val == "":
            return None
        return val
    
    if v:
        # Update header details
        v.fecha_mapeo = data.fecha_mapeo
        v.mapeador = data.mapeador
        v.campania = data.campania
        v.este_ini = data.este_ini
        v.norte_ini = data.norte_ini
        v.cota_ini = data.cota_ini
        v.este_fin = data.este_fin
        v.norte_fin = data.norte_fin
        v.cota_fin = data.cota_fin
        if "sqlite" in str(db.bind.url).lower():
            v.largo_m = data.largo_m
        v.altura_m = data.altura_m
        v.dip_talud = data.dip_talud
        v.dipdir_talud = data.dipdir_talud
        v.dip_hw = data.dip_hw
        v.az_hw = data.az_hw
        v.alteracion_codigo = data.alteracion_codigo
        v.intemperismo_codigo = data.intemperismo_codigo
        v.lito_1 = data.lito_1
        v.lito_2 = data.lito_2
        v.lito_3 = data.lito_3
        v.unidad_litologica = data.unidad_litologica
        v.sector = data.sector
        v.fase = data.fase
        v.nivel = data.nivel
        v.sector_geotecnico = data.sector_geotecnico
        v.turno = data.turno
        
        # Delete old dependent children
        db.query(models.Discontinuidad).filter_by(ventana_id=v.ventana_id).delete()
        if v.rmr_input:
            db.delete(v.rmr_input)
    else:
        # Create new Ventana
        v = models.Ventana(
            codigo=code_up,
            fecha_mapeo=data.fecha_mapeo,
            mapeador=data.mapeador,
            campania=data.campania,
            este_ini=data.este_ini,
            norte_ini=data.norte_ini,
            cota_ini=data.cota_ini,
            este_fin=data.este_fin,
            norte_fin=data.norte_fin,
            cota_fin=data.cota_fin,
            altura_m=data.altura_m,
            dip_talud=data.dip_talud,
            dipdir_talud=data.dipdir_talud,
            dip_hw=data.dip_hw,
            az_hw=data.az_hw,
            alteracion_codigo=data.alteracion_codigo,
            intemperismo_codigo=data.intemperismo_codigo,
            lito_1=data.lito_1,
            lito_2=data.lito_2,
            lito_3=data.lito_3,
            unidad_litologica=data.unidad_litologica,
            sector=data.sector,
            fase=data.fase,
            nivel=data.nivel,
            sector_geotecnico=data.sector_geotecnico,
            turno=data.turno
        )
        if "sqlite" in str(db.bind.url).lower():
            v.largo_m = data.largo_m
        db.add(v)
        db.flush() # get ventana_id
        
    # Add discontinuidades
    for idx, d in enumerate(data.discontinuidades):
        disc = models.Discontinuidad(
            ventana_id=v.ventana_id,
            familia_id=d.fam,
            orden_en_familia=idx + 1,
            distancia_m=clean_null_val(d.dist),
            tipo_estructura=d.tipo,
            dip=clean_null_val(d.dip),
            dip_dir=clean_null_val(d.dipdir),
            abertura_mm=clean_null_val(d.aber),
            espesor_mm=clean_null_val(d.esp),
            continuidad_m=clean_null_val(d.cont),
            espaciamiento_m=clean_null_val(d.espac),
            n_estructuras=clean_null_val(d.nstr), # Guardar en SQL Server / SQLite de forma libre
            n_extremos_visibles=clean_null_val(d.next),
            terminacion=clean_null_val(d.term),
            relleno_1_codigo=d.r1 if d.r1 != "-1" else None,
            relleno_2_codigo=d.r2 if d.r2 != "-1" else None,
            jrc=clean_null_val(d.jrc),
            rugosidad_codigo=clean_null_val(d.rug),
            forma_estructura=d.forma if d.forma != "-1" else None,
            alteracion_codigo=d.alt if d.alt != "-1" else None
           )
        db.add(disc)

    db.flush()
        
    # Add rmr_input
    if data.rmr_input:
        ri = models.VentanaRmrInput(
            ventana_id=v.ventana_id,
            agua_codigo=data.rmr_input.agua_codigo,
            resistencia_codigo=data.rmr_input.resistencia_codigo,
            gsi_estructura=data.rmr_input.gsi_estructura,
            gsi_superficie=data.rmr_input.gsi_superficie,
            gsi_visual=data.rmr_input.gsi_visual,
            control_estructural=data.rmr_input.control_estructural,
            efectos_voladura=data.rmr_input.efectos_voladura,
            ucs_mpa=data.rmr_input.ucs_mpa,
            is50_mpa=data.rmr_input.is50_mpa,
            comentario=data.rmr_input.comentario
        )
        db.add(ri)
        
    db.flush()
    
    # Sync with ventanas_final flat table
    sync_to_ventanas_final(db, v.ventana_id)
    
    db.commit()
    return {"status": "success", "message": f"Ventana {code_up} guardada y sincronizada correctamente"}

@app.delete("/api/ventanas/{codigo}")
def delete_ventana(codigo: str, db: Session = Depends(get_db)):
    code_up = codigo.strip().upper()
    v = db.query(models.Ventana).filter_by(codigo=code_up).first()
    if not v:
        raise HTTPException(status_code=404, detail="Ventana no encontrada")
    
    # Cascades will automatically clean up discontinuidad and rmr_input
    db.delete(v)
    
    # Delete flat rows too
    db.query(models.VentanasFinal).filter_by(celda=code_up).delete()
    
    db.commit()
    return {"status": "success", "message": f"Ventana {code_up} eliminada correctamente de la base de datos"}

@app.post("/api/calculate")
def run_calculate(data: Dict[str, Any]):
    header = data.get("header", {})
    discs = data.get("discontinuidades", [])
    rmr_input = data.get("rmr_input", {})
    
    res = calculator.calculate_geomechanics(header, discs, rmr_input)
    return res

LITHOLOGY_CLASSIFICATION_DB = [
    {"grupo": "INTRUSIVOS", "unidad": "MZB", "litologia": "MZB", "codigo": "MZB_EQ"},
    {"grupo": "INTRUSIVOS", "unidad": "MZB", "litologia": "MZB", "codigo": "MZB_P"},
    {"grupo": "INTRUSIVOS", "unidad": "MBF1", "litologia": "MBF", "codigo": "MBF1"},
    {"grupo": "INTRUSIVOS", "unidad": "MBF2", "litologia": "MBF", "codigo": "MBF2"},
    {"grupo": "INTRUSIVOS", "unidad": "MBF2", "litologia": "MBF", "codigo": "MBF_P"},
    {"grupo": "INTRUSIVOS", "unidad": "MZM", "litologia": "MZM", "codigo": "MZM_F"},
    {"grupo": "INTRUSIVOS", "unidad": "MZM", "litologia": "MZM", "codigo": "MZM_M"},
    {"grupo": "INTRUSIVOS", "unidad": "MZH", "litologia": "MZH", "codigo": "MZH_1"},
    {"grupo": "INTRUSIVOS", "unidad": "MZH", "litologia": "MZH", "codigo": "MZH_2"},
    {"grupo": "INTRUSIVOS", "unidad": "MZD", "litologia": "MZD", "codigo": "MZD"},
    {"grupo": "INTRUSIVOS", "unidad": "MZQ", "litologia": "MZQ", "codigo": "MZQ"},
    {"grupo": "INTRUSIVOS", "unidad": "AN", "litologia": "LAM", "codigo": "LAM"},
    {"grupo": "SEDIMENTARIOS", "unidad": "LMT", "litologia": "LMT", "codigo": "LMT_M"},
    {"grupo": "SEDIMENTARIOS", "unidad": "LMT", "litologia": "LMT", "codigo": "LMT_MG"},
    {"grupo": "SEDIMENTARIOS", "unidad": "LMT", "litologia": "LMT", "codigo": "LMT_S"},
    {"grupo": "SEDIMENTARIOS", "unidad": "LMT", "litologia": "LMT", "codigo": "LMT_C"},
    {"grupo": "SEDIMENTARIOS", "unidad": "LMT", "litologia": "LMT", "codigo": "LMT_U"},
    {"grupo": "SEDIMENTARIOS", "unidad": "SHL", "litologia": "HFL", "codigo": "SHL_MA"},
    {"grupo": "METAMORFICAS", "unidad": "LMT", "litologia": "GSK", "codigo": "Varios"},
    {"grupo": "METAMORFICAS", "unidad": "LMT", "litologia": "PSK", "codigo": "Varios"},
    {"grupo": "METAMORFICAS", "unidad": "LMT", "litologia": "MSK", "codigo": "Varios"},
    {"grupo": "METAMORFICAS", "unidad": "LMT", "litologia": "ESK", "codigo": "Varios"},
    {"grupo": "METAMORFICAS", "unidad": "LMT", "litologia": "MBC", "codigo": "Varios"},
    {"grupo": "METAMORFICAS", "unidad": "LMT", "litologia": "MBL", "codigo": "Varios"},
    {"grupo": "METAMORFICAS", "unidad": "SHL", "litologia": "HFL", "codigo": "-"},
    {"grupo": "METAMORFICAS", "unidad": "SND", "litologia": "QZT", "codigo": "-"},
    {"grupo": "BRECHAS", "unidad": "TBX", "litologia": "TBX", "codigo": "TBX"},
    {"grupo": "BRECHAS", "unidad": "HBX", "litologia": "HBX", "codigo": "HBX"},
    {"grupo": "BRECHAS", "unidad": "MBX / varios", "litologia": "MBX", "codigo": "MBX"},
    {"grupo": "ENDOSKARN", "unidad": "MZM", "litologia": "EPG", "codigo": "-"},
    {"grupo": "ENDOSKARN", "unidad": "MZM", "litologia": "EGT", "codigo": "-"}
]

def resolve_lithology(lito3_code: str) -> dict:
    """
    Sanea y busca el código de Litología 3 en el catálogo para autocompletar 
    las 4 variables consistentes en la base de datos de forma autocurativa.
    """
    code_clean = str(lito3_code or "").strip().upper().replace(" ", "").replace("-", "")
    
    match = None
    for item in LITHOLOGY_CLASSIFICATION_DB:
        item_code = item["codigo"].upper().replace(" ", "").replace("-", "")
        if item_code and item_code == code_clean:
            match = item
            break
            
    if not match:
        for item in LITHOLOGY_CLASSIFICATION_DB:
            if code_clean in item["codigo"].upper() or code_clean in item["litologia"].upper():
                match = item
                break

    if match:
        return {
            "lito_1": match["unidad"],
            "lito_2": match["litologia"],
            "lito_3": match["codigo"],
            "unidad_litologica": match["grupo"]
        }
    
    # Fallback por defecto si no se encuentra en el catálogo
    return {
        "lito_1": lito3_code,
        "lito_2": "",
        "lito_3": lito3_code,
        "unidad_litologica": "INTRUSIVOS"
    }

@app.post("/api/importar-excel")
async def importar_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    
    imported_count = 0
    
    # CASO 1: Pestaña "ventana" estructurada cada 30 filas
    if "ventana" in wb.sheetnames:
        ws = wb["ventana"]
        
        for start in range(3, ws.max_row, 30):
            celda_val = ws.cell(row=start+1, column=1).value
            if not celda_val:
                celda_val = ws.cell(row=start, column=51).value
            
            if not celda_val or not str(celda_val).strip():
                continue
                
            codigo = str(celda_val).strip().upper()
            
            fecha_val = ws.cell(row=start+1, column=37).value
            if isinstance(fecha_val, str):
                try:
                    fecha_mapeo = datetime.strptime(fecha_val[:10], "%Y-%m-%d").date()
                except:
                    fecha_mapeo = date.today()
            elif isinstance(fecha_val, datetime):
                fecha_mapeo = fecha_val.date()
            else:
                fecha_mapeo = date.today()
                
            def get_num(r, c):
                val = ws.cell(row=r, column=c).value
                if val is None:
                    return 0.0
                try:
                    return float(val)
                except:
                    return 0.0
            
            def get_str(r, c):
                val = ws.cell(row=r, column=c).value
                return str(val).strip() if val is not None else ""

            este_ini = round(get_num(start+2, 2), 4)      # <- 4 decimales
            norte_ini = round(get_num(start+2, 4), 3)     # <- 3 decimales
            cota_ini = round(get_num(start+2, 6), 2)
            este_fin = round(get_num(start+3, 2), 4)      # <- 4 decimales
            norte_fin = round(get_num(start+3, 4), 3)     # <- 3 decimales
            cota_fin = round(get_num(start+3, 6), 2)
            
            largo = int(round(get_num(start+2, 11)))
            altura = round(get_num(start+3, 11), 1)
            dip_talud = round(get_num(start+2, 14), 2)
            
            # Recuperamos Lito 3 (codigo) de Row 7 Col P (Lito-Model)
            lito_model = get_str(start+4, 16)
            
            mapeador = get_str(start+5, 16)
            sector = get_str(start+1, 20)
            fase = int(get_num(start+2, 21))
            nivel = round(get_num(start+3, 21), 2)
            sect_geot = get_str(start+4, 21)
            intemp = get_str(start+3, 16)
            
            agua_code = get_str(start+8, 36)
            res_code = get_str(start+8, 38)
            gsi_cond = get_str(start+8, 40)
            gsi_est = get_str(start+8, 41)
            gsi_vis = int(get_num(start+8, 42))
            ctrl = int(get_num(start+8, 43))
            vol = int(get_num(start+8, 44))
            ucs = get_num(start+8, 53)
            is50 = get_num(start+8, 54)
            comentario = get_str(start+18, 56)
            
            # Resolución dinámica en cascada de los 4 campos litológicos
            lito_details = resolve_lithology(lito_model)
            
            discs = []
            for r_idx in range(start+12, start+25):
                fam_val = ws.cell(row=r_idx, column=1).value
                if fam_val is None or str(fam_val).strip() == "":
                    continue
                try:
                    fam_id = int(fam_val)
                except:
                    continue
                
                raw_nstr = int(round(get_num(r_idx, 6)))
                nstr = raw_nstr if raw_nstr > 0 else -1

                discs.append(schemas.DiscontinuidadBase(
                    familia_id=fam_id,
                    distancia_m=int(round(get_num(r_idx, 2))),
                    tipo_estructura=get_str(r_idx, 3) if get_str(r_idx, 3) else "JN",
                    dip=round(get_num(r_idx, 4), 2),
                    dip_dir=round(get_num(r_idx, 5), 2),
                    n_estructuras=nstr,
                    abertura_mm=round(get_num(r_idx, 7), 1),
                    espesor_mm=round(get_num(r_idx, 8), 1),
                    continuidad_m=round(get_num(r_idx, 9), 2),
                    espaciamiento_m=round(get_num(r_idx, 10), 2),
                    n_extremos_visibles=min(2, max(0, int(get_num(r_idx, 11)))) if ws.cell(row=r_idx, column=11).value is not None else None,
                    terminacion=min(3, max(0, int(get_num(r_idx, 12)))) if ws.cell(row=r_idx, column=12).value is not None else None,
                    relleno_1_codigo=get_str(r_idx, 13),
                    relleno_2_codigo=get_str(r_idx, 14),
                    jrc=min(20, max(0, int(get_num(r_idx, 19)))) if ws.cell(row=r_idx, column=19).value is not None else None,
                    rugosidad_codigo=min(9, max(0, int(get_num(r_idx, 20)))) if ws.cell(row=r_idx, column=20).value is not None else None,
                    forma_estructura=get_str(r_idx, 21),
                    alteracion_codigo=get_str(r_idx, 22)
                ))
                
            ri_schema = schemas.VentanaRmrInputBase(
                agua_codigo=agua_code if agua_code else "C",
                resistencia_codigo=res_code if res_code else "R4",
                gsi_estructura=gsi_est if gsi_est else "VB",
                gsi_superficie=gsi_cond if gsi_cond else "G",
                gsi_visual=gsi_vis if gsi_vis else 50,
                control_estructural=ctrl if ctrl else 4,
                efectos_voladura=vol if vol else 3,
                ucs_mpa=ucs if ucs else 74.0,
                is50_mpa=is50 if is50 else 5.0,
                comentario=comentario
            )
            
            ventana_schema = schemas.VentanaSaveSchema(
                codigo=codigo,
                fecha_mapeo=fecha_mapeo,
                mapeador=mapeador if mapeador else "RD/RB",
                campania=2026,
                este_ini=este_ini,
                norte_ini=norte_ini,
                cota_ini=cota_ini,
                este_fin=este_fin,
                norte_fin=norte_fin,
                cota_fin=cota_fin,
                largo_m=largo,
                altura_m=altura,
                dip_talud=dip_talud,
                alteracion_codigo=intemp,
                intemperismo_codigo=intemp,
                lito_1=lito_details["lito_1"],
                lito_2=lito_details["lito_2"],
                lito_3=lito_details["lito_3"],
                unidad_litologica=lito_details["unidad_litologica"],
                sector=sector,
                fase=fase,
                nivel=nivel,
                sector_geotecnico=sect_geot,
                discontinuidades=discs,
                rmr_input=ri_schema
            )
            
            save_ventana(ventana_schema, db)
            imported_count += 1
            
    # CASO 2: Pestaña plana desnormalizada "BD"
    elif "BD" in wb.sheetnames:
        ws = wb["BD"]
        celda_groups = {}
        for r_idx in range(2, ws.max_row + 1):
            celda_val = ws.cell(row=r_idx, column=3).value or ws.cell(row=r_idx, column=4).value
            if not celda_val or str(celda_val).strip() == "":
                continue
            celda_code = str(celda_val).strip().upper()
            if celda_code not in celda_groups:
                celda_groups[celda_code] = []
            celda_groups[celda_code].append(r_idx)
            
        for celda_code, rows_indices in celda_groups.items():
            f_row = rows_indices[0]
            
            def get_num(r, c):
                val = ws.cell(row=r, column=c).value
                if val is None:
                    return 0.0
                try:
                    return float(val)
                except:
                    return 0.0
            
            def get_str(r, c):
                val = ws.cell(row=r, column=c).value
                return str(val).strip() if val is not None else ""
                
            este_from = round(get_num(f_row, 6), 4)       # <- 4 decimales
            norte_from = round(get_num(f_row, 7), 3)      # <- 3 decimales
            cota_from = round(get_num(f_row, 8), 2)
            este_to = round(get_num(f_row, 10), 4)        # <- 4 decimales
            norte_to = round(get_num(f_row, 11), 3)       # <- 3 decimales
            cota_to = round(get_num(f_row, 12), 2)
            dist_celda = int(round(get_num(f_row, 13)))
            altura = round(get_num(f_row, 14), 1)
            dip_talud = round(get_num(f_row, 20), 2)
            intemp = get_str(f_row, 23)
            
            agua_code = get_str(f_row, 43)
            res_code = get_str(f_row, 45)
            gsi_vis = int(get_num(f_row, 47))
            ctrl = int(get_num(f_row, 48))
            vol = int(get_num(f_row, 49))
            ucs = get_num(f_row, 40)
            is50 = get_num(f_row, 41)
            fecha_val = ws.cell(row=f_row, column=59).value
            if isinstance(fecha_val, str):
                try:
                    fecha_mapeo = datetime.strptime(fecha_val[:10], "%Y-%m-%d").date()
                except:
                    fecha_mapeo = date.today()
            elif isinstance(fecha_val, datetime):
                fecha_mapeo = fecha_val.date()
            else:
                fecha_mapeo = date.today()
                
            comentario = get_str(f_row, 61)
            geot = get_str(f_row, 85)
            nivel = round(get_num(f_row, 95), 2)
            
            # Leemos Lito 3 (LITO3_MODELO) de Col CK (Index 89)
            l1 = get_str(f_row, 89)
            
            # Resolución dinámica en cascada de los 4 campos litológicos
            lito_details = resolve_lithology(l1)
            
            discs = []
            for r_idx in rows_indices:
                fam_val = ws.cell(row=r_idx, column=1).value or 1
                
                raw_nstr = int(round(get_num(r_idx, 72)))
                nstr = raw_nstr if raw_nstr > 0 else -1

                discs.append(schemas.DiscontinuidadBase(
                    familia_id=int(fam_val),
                    distancia_m=int(round(get_num(r_idx, 63))),
                    tipo_estructura=get_str(r_idx, 69) if get_str(r_idx, 69) else "JN",
                    dip=round(get_num(r_idx, 70), 2),
                    dip_dir=round(get_num(r_idx, 71), 2),
                    n_estructuras=nstr,
                    abertura_mm=round(get_num(r_idx, 73), 1),
                    espesor_mm=round(get_num(r_idx, 74), 1),
                    continuidad_m=round(get_num(r_idx, 75), 2),
                    espaciamiento_m=round(get_num(r_idx, 76), 2),
                    n_extremos_visibles=min(2, max(0, int(get_num(r_idx, 77)))) if ws.cell(row=r_idx, column=77).value is not None else None,
                    terminacion=3,
                    relleno_1_codigo=get_str(r_idx, 78),
                    relleno_2_codigo=get_str(r_idx, 79),
                    jrc=min(20, max(0, int(get_num(r_idx, 80)))) if ws.cell(row=r_idx, column=80).value is not None else None,
                    rugosidad_codigo=min(9, max(0, int(get_num(r_idx, 81)))) if ws.cell(row=r_idx, column=81).value is not None else None,
                    forma_estructura=get_str(r_idx, 82),
                    alteracion_codigo=get_str(r_idx, 83)
                ))
                
            ri_schema = schemas.VentanaRmrInputBase(
                agua_codigo=agua_code if agua_code else "C",
                resistencia_codigo=res_code if res_code else "R4",
                gsi_estructura="VB",
                gsi_superficie="G",
                gsi_visual=gsi_vis if gsi_vis else 50,
                control_estructural=ctrl if ctrl else 4,
                efectos_voladura=vol if vol else 3,
                ucs_mpa=ucs if ucs else 74.0,
                is50_mpa=is50 if is50 else 5.0,
                comentario=comentario
            )
            
            ventana_schema = schemas.VentanaSaveSchema(
                codigo=celda_code,
                fecha_mapeo=fecha_mapeo,
                mapeador=geot if geot else "RD/RB",
                campania=2026,
                este_ini=este_from,
                norte_ini=norte_from,
                cota_ini=cota_from,
                este_fin=este_to,
                norte_fin=norte_to,
                cota_fin=cota_to,
                largo_m=dist_celda,
                altura_m=altura,
                dip_talud=dip_talud,
                alteracion_codigo=intemp,
                intemperismo_codigo=intemp,
                lito_1=lito_details["lito_1"],
                lito_2=lito_details["lito_2"],
                lito_3=lito_details["lito_3"],
                unidad_litologica=lito_details["unidad_litologica"],
                sector="E1",
                fase=5,
                nivel=nivel,
                sector_geotecnico="E1",
                discontinuidades=discs,
                rmr_input=ri_schema
            )
            
            save_ventana(ventana_schema, db)
            imported_count += 1
            
    return {"status": "success", "message": f"Importación completada. {imported_count} ventanas importadas con éxito."}

@app.get("/api/ventanas/{codigo}/exportar")
def exportar_ventana_excel(codigo: str, db: Session = Depends(get_db)):
    code_up = codigo.strip().upper()
    
    # Consultar todas las filas calculadas para esta celda
    rows = db.query(models.VentanasFinal).filter_by(celda=code_up).order_by(models.VentanasFinal.id).all()
    if not rows:
        raise HTTPException(
            status_code=404, 
            detail=f"No se encontraron datos calculados para la celda {code_up}. Por favor, guarde la celda primero para sincronizar."
        )
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapeo Ventana"
    
    # Encabezados de columna idénticos a la maqueta de referencia
    headers = [
        "id", "CELDA", "CELDA", "ESTE_FROM", "NORTE_FROM", "COTA", "ESTE_TO", "NORTE_TO", "COTA",
        "Dist.Celda", "Altura", "DIP", "AZ_HOLE",
        "DIP_TALUD", "DIP_DIR_TALUD", "INTEMPERISMO",
        "CONDICION DE AGUA '76", "CONDICION DE AGUA VALOR '76", "DUREZA '76", "RESISTENCIA ESTIMADA VALOR '76",
        "GSI VISUAL '76", "CONTROL ESTRUCTURAL '76", "EFECTOS DE VOLADURA '76", "RQD - VALOR '76", "RQD '76",
        "FRECUENCIA DE FRACTURAMIENTO x m '76", "TAMAÑO DE BLOQUES x m3 '76", "ESPACIAMIENTO PROMEDIO '76",
        "ESPACIAMIENTO - VALOR '76", "CONDICION DE DISCONTINUIDAD - VALOR '76", "RMR '76",
        "( UCS ) (Mpa)", "is50 (Mpa)",
        "CONDICION DE AGUA '89", "CONDICION DE AGUA VALOR '89", "DUREZA '89", "RESISTENCIA ESTIMADA VALOR '89",
        "GSI VISUAL '89", "CONTROL ESTRUCTURAL '89", "EFECTOS DE VOLADURA '89", "RQD - VALOR '89", "RQD '89",
        "FRECUENCIA DE FRACTURAMIENTO x m '89", "TAMAÑO DE BLOQUES x m3 '89", "ESPACIAMIENTO PROMEDIO '89",
        "ESPACIAMIENTO - VALOR '89", "CONDICION DE DISCONTINUIDAD - VALOR '89", "RMR '89",
        "FECHA", "COMENTARIO",
        "Dist. de estr.", "teta", "alfa", "x", "y", "z",
        "TIPO DE ESTRUCT", "DIP", "DIP DIR", "NUMERO DE ESTRUCTURAS", "ABERTURA mm", "ESPESOR mm",
        "CONTINUIDAD m", "ESPACIAMIENTO m", "NUMERO DE EXTREMOS VISIBLES", "TIPO DE RELLENO 1", "TIPO DE RELLENO 2",
        "JRC", "RUGOSIDAD", "FORMA DE ESTRUCTURA", "ALTERACION", "GEOTECNICO", "Is50_Mpa", "LITO3_MODELO", "Sector", "Nivel"
    ]
    
    ws.append(headers)
    
    # Escribir los registros desnormalizados en el mismo orden
    for r in rows:
        row_data = [
            r.id,
            r.celda,
            r.celda,  # Duplicado según formato requerido
            float(r.este_from) if r.este_from is not None else 0.0,
            float(r.norte_from) if r.norte_from is not None else 0.0,
            float(r.cota_from) if r.cota_from is not None else 0.0,
            float(r.este_to) if r.este_to is not None else 0.0,
            float(r.norte_to) if r.norte_to is not None else 0.0,
            float(r.cota_to) if r.cota_to is not None else 0.0,
            r.dist_celda,
            float(r.altura) if r.altura is not None else None,
            float(r.dip) if r.dip is not None else None,
            float(r.az_hole) if r.az_hole is not None else None,
            float(r.dip_talud) if r.dip_talud is not None else None,
            float(r.dip_dir_talud) if r.dip_dir_talud is not None else None,
            r.intemperismo,
            r.cond_agua_76,
            r.cond_agua_valor_76,
            r.dureza_76,
            r.resistencia_est_valor_76,
            r.gsi_visual_76,
            r.control_estructural_76,
            r.efectos_voladura_76,
            r.rqd_valor_76,
            r.rqd_76,
            r.freq_fractura_m_76,
            r.tam_bloques_m3_76,
            r.espaciamiento_prom_76,
            r.espaciamiento_valor_76,
            r.cond_discontinuidad_valor_76,
            r.rmr_76,
            r.ucs_mpa,
            r.is50_mpa,
            r.cond_agua_89,
            r.cond_agua_valor_89,
            r.dureza_89,
            r.resistencia_est_valor_89,
            r.gsi_visual_89,
            r.control_estructural_89,
            r.efecto_voladura_89,
            r.rqd_valor_89,
            r.rqd_89,
            r.freq_fractura_m_89,
            r.tam_bloques_m3_89,
            r.espaciamiento_prom_89,
            r.espaciamiento_valor_89,
            r.cond_discontinuidad_valor_89,
            r.rmr_89,
            r.fecha.strftime("%Y-%m-%d") if r.fecha else "",
            r.comentario,
            r.dist_estructura,
            r.angulo_estruct_teta,
            r.angulo_estruct_alfa,
            r.estruct_x,
            r.struct_y,
            r.struct_z,
            r.tipo_estructura,
            r.dip_estructura,
            r.dip_dir_estructura,
            r.num_estructuras,
            r.abertura_mm,
            r.espesor_mm,
            r.continuidad_m,
            r.espaciamiento_m,
            r.num_extremos_visibles,
            r.tipo_relleno_1,
            r.tipo_relleno_2,
            r.jrc,
            r.rugosidad_estructuras,
            r.forma_estructura,
            r.alteracion,
            r.geotecnico,
            r.is50_mpa, # Duplicada en maqueta
            r.lito_3,
            r.sector_geotecnico,
            r.nivel
        ]
        ws.append(row_data)

    # --- ESTILIZACIÓN PREMIUM DE COLUMNAS SEGÚN MAQUETA ---
    # Fills en formato pastel elegante para no fatigar la lectura de datos
    fill_red = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")     # Rojo suave (A-I)
    fill_blue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")    # Azul suave (J-M)
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Amarillo suave (N-AE)
    fill_peach = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")   # Durazno/Salmón (AF-AG)
    fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")   # Verde suave (AH-AW)
    fill_brown = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")   # Gris/Marrón para proyección (BB-BG)

    font_header = Font(name="Arial", size=9, bold=True, color="333333")
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )

    # Aplicar diseño agrupado a la fila de cabeceras
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = border_thin
        
        # Agrupación por color según rangos de la maqueta
        if col_idx <= 9:                      # A - I
            cell.fill = fill_red
        elif 10 <= col_idx <= 13:             # J - M
            cell.fill = fill_blue
        elif 14 <= col_idx <= 31:             # N - AE
            cell.fill = fill_yellow
        elif 32 <= col_idx <= 33:             # AF - AG
            cell.fill = fill_peach
        elif 34 <= col_idx <= 48:             # AH - AW
            cell.fill = fill_green
        elif 51 <= col_idx <= 56:             # BB - BG
            cell.fill = fill_brown
        elif col_idx >= 57:                   # BH - Nivel
            cell.fill = fill_yellow

    # Dar formato limpio de visualización al cuerpo de la tabla
    font_body = Font(name="Arial", size=9, bold=False)
    for col_idx in range(1, len(headers) + 1):
        header_name = headers[col_idx - 1].upper()
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_body
            cell.border = border_thin
            if isinstance(cell.value, (int, float)):
                if "ESTE" in header_name or header_name == "X":
                    cell.number_format = '0.0000' # <- Visualizar 4 decimales en Excel
                elif "NORTE" in header_name or header_name == "Y":
                    cell.number_format = '0.000'  # <- Visualizar 3 decimales en Excel
                elif "COTA" in header_name or header_name == "Z" or "ELEVACION" in header_name:
                    cell.number_format = '0.00'   # <- Visualizar 2 decimales en Excel
                else:
                    cell.number_format = '0.00'   # fallback general

    # Autoajustar anchos de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    # Transmitir el archivo de regreso
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=mapeo_ventana_{code_up}.xlsx"}
    )

@app.get("/api/ensayos-plt", response_model=List[schemas.EnsayoPLTSaveSchema])
def get_ensayos_plt(db: Session = Depends(get_db)):
    res = db.query(models.EnsayoPLTIrregular).all()
    return res

@app.post("/api/ensayos-plt")
def save_ensayos_plt(data: List[schemas.EnsayoPLTSaveSchema], db: Session = Depends(get_db)):
    existing = {r.id: r for r in db.query(models.EnsayoPLTIrregular).all()}
    incoming_ids = {d.id for d in data if d.id is not None}
    
    # 1. Delete rows not in incoming data
    for rid, row in list(existing.items()):
        if rid not in incoming_ids:
            db.delete(row)
            
    # 2. Insert or update incoming rows
    for d in data:
        if d.id is not None and d.id in existing:
            # Update
            row = existing[d.id]
            row.campana = d.campana
            row.fecha_ensayo = d.fecha_ensayo
            row.sector_geotecnico = d.sector_geotecnico
            row.ejecutado_por = d.ejecutado_por
            row.zona_mapeo = d.zona_mapeo
            row.nivel = d.nivel
            row.celda_mapeo = d.celda_mapeo.strip().upper()
            row.muestra = d.muestra
            row.codigo_muestra = d.codigo_muestra
            row.litologia_1 = d.litologia_1
            row.litologia_2 = d.litologia_2
            row.litologia_3 = d.litologia_3
            row.tipo_litologico = d.tipo_litologico
            row.este = d.este
            row.norte = d.norte
            row.elevacion = d.elevacion
            row.espesor_d = d.espesor_d
            row.longitud_l = d.longitud_l
            row.ancho_w1 = d.ancho_w1
            row.ancho_w2 = d.ancho_w2
            row.fuerza_p = d.fuerza_p
            row.direccion_rotura = d.direccion_rotura
            row.tipo_fractura = d.tipo_fractura
            row.factor_conversion_k = d.factor_conversion_k
            row.observaciones = d.observaciones
        else:
            # Insert new
            row = models.EnsayoPLTIrregular(
                campana=d.campana,
                fecha_ensayo=d.fecha_ensayo,
                sector_geotecnico=d.sector_geotecnico,
                ejecutado_por=d.ejecutado_por,
                zona_mapeo=d.zona_mapeo,
                nivel=d.nivel,
                celda_mapeo=d.celda_mapeo.strip().upper(),
                muestra=d.muestra,
                codigo_muestra=d.codigo_muestra,
                litologia_1=d.litologia_1,
                litologia_2=d.litologia_2,
                litologia_3=d.litologia_3,
                tipo_litologico=d.tipo_litologico,
                este=d.este,
                norte=d.norte,
                elevacion=d.elevacion,
                espesor_d=d.espesor_d,
                longitud_l=d.longitud_l,
                ancho_w1=d.ancho_w1,
                ancho_w2=d.ancho_w2,
                fuerza_p=d.fuerza_p,
                direccion_rotura=d.direccion_rotura,
                tipo_fractura=d.tipo_fractura,
                factor_conversion_k=d.factor_conversion_k,
                observaciones=d.observaciones
            )
            db.add(row)
            
    db.commit()
    return {"status": "success", "message": "Ensayos PLT guardados con éxito"}

@app.post("/api/ventanas/{codigo}/fotos")
async def upload_foto(codigo: str, index: int, file: UploadFile = File(...)):
    code_up = codigo.strip().upper()
    dir_path = os.path.join(uploads_dir, code_up)
    os.makedirs(dir_path, exist_ok=True)
    
    # Validar tamaño máximo (5 MB)
    MAX_SIZE = 5 * 1024 * 1024
    contents = await file.read()
    if len(contents) > MAX_SIZE:
        raise HTTPException(status_code=400, detail="La fotografía excede el tamaño máximo permitido de 5MB")
        
    # Validar formato de imagen
    ext = file.filename.split(".")[-1].lower() if "." in file.filename else ""
    allowed_exts = ["jpg", "jpeg", "png", "webp", "bmp", "gif", "svg", "tiff"]
    if ext not in allowed_exts:
        raise HTTPException(
            status_code=400, 
            detail="Formato de imagen no soportado. Use JPG, JPEG, PNG, WEBP, BMP, GIF o SVG."
        )
    
    # Nueva convención de nombre solicitado: {codigo}-VENTANA-{index + 1}.{ext}
    new_filename = f"{code_up}-VENTANA-{index + 1}.{ext}"
    
    # Eliminar fotos previas en este índice (tanto del formato nuevo como del antiguo)
    # con otras extensiones para evitar archivos huérfanos duplicados en el disco
    for e in allowed_exts:
        # Limpiar formato antiguo si existía
        old_path = os.path.join(dir_path, f"foto_{index}.{e}")
        if os.path.exists(old_path):
            try:
                os.remove(old_path)
            except:
                pass
        
        # Limpiar formato nuevo con otra extensión
        new_path_diff_ext = os.path.join(dir_path, f"{code_up}-VENTANA-{index + 1}.{e}")
        if os.path.exists(new_path_diff_ext) and e != ext:
            try:
                os.remove(new_path_diff_ext)
            except:
                pass
                
    # Guardar la nueva fotografía con la extensión correspondiente
    file_path = os.path.join(dir_path, new_filename)
    with open(file_path, "wb") as f:
        f.write(contents)
        
    return {"status": "success", "url": f"/api/uploads/{code_up}/{new_filename}"}


@app.delete("/api/ventanas/{codigo}/fotos/{index}")
def delete_foto(codigo: str, index: int):
    code_up = codigo.strip().upper()
    dir_path = os.path.join(uploads_dir, code_up)
    allowed_exts = ["jpg", "jpeg", "png", "webp", "bmp", "gif", "svg", "tiff"]
    
    for e in allowed_exts:
        # Eliminar del formato nuevo
        new_path = os.path.join(dir_path, f"{code_up}-VENTANA-{index + 1}.{e}")
        if os.path.exists(new_path):
            try:
                os.remove(new_path)
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Error al eliminar archivo: {e}")
        
        # Eliminar del formato antiguo (compatibilidad)
        old_path = os.path.join(dir_path, f"foto_{index}.{e}")
        if os.path.exists(old_path):
            try:
                os.remove(old_path)
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Error al eliminar archivo: {e}")
                
    return {"status": "success"}

@app.post("/api/ventanas/{codigo}/fotos/meta")
def save_metadata(codigo: str, data: Dict[str, Any]):
    code_up = codigo.strip().upper()
    dir_path = os.path.join(uploads_dir, code_up)
    os.makedirs(dir_path, exist_ok=True)
    
    meta_path = os.path.join(dir_path, "metadata.json")
    import json
    try:
        with open(meta_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al guardar metadatos: {e}")
    return {"status": "success"}

@app.get("/api/ventanas/{codigo}/fotos")
def get_fotos(codigo: str):
    code_up = codigo.strip().upper()
    dir_path = os.path.join(uploads_dir, code_up)
    
    photos = ["", "", "", ""]
    captions = ["", "", "", ""]
    
    if os.path.exists(dir_path):
        # 1. Recuperar leyendas si existen en metadata.json
        meta_path = os.path.join(dir_path, "metadata.json")
        if os.path.exists(meta_path):
            try:
                import json
                with open(meta_path, 'r') as f: # o lectura estándar compatible
                    meta = json.load(f)
                    captions = meta.get("captions", ["", "", "", ""])
            except:
                # Si falla por bloqueo o formato, usamos fallback vacío
                pass
                
        # 2. Comprobar existencia física de las 4 fotos
        allowed_exts = ["jpg", "jpeg", "png", "webp", "bmp", "gif", "svg", "tiff"]
        for i in range(4):
            found = False
            # Intentar buscar la foto en el formato nuevo
            for e in allowed_exts:
                file_path = os.path.join(dir_path, f"{code_up}-VENTANA-{i+1}.{e}")
                if os.path.exists(file_path):
                    import time
                    photos[i] = f"/api/uploads/{code_up}/{code_up}-VENTANA-{i+1}.{e}?t={int(time.time())}"
                    found = True
                    break
            
            # Fallback de compatibilidad por si tenían el formato antiguo foto_{i}.{e}
            if not found:
                for e in allowed_exts:
                    file_path = os.path.join(dir_path, f"foto_{i}.{e}")
                    if os.path.exists(file_path):
                        import time
                        photos[i] = f"/api/uploads/{code_up}/foto_{i}.{e}?t={int(time.time())}"
                        break
                
    return {"photos": photos, "captions": captions}
    

def run_bulk_pipeline_with_id(file_path: str, audit_id: str):
    """
    Background Task: Ejecuta el validador, almacena el diagnóstico masivo 
    y pre-calcula los KPIs avanzados de las 8 familias geotécnicas.
    """
    history_dir = os.path.join(uploads_dir, "history")
    os.makedirs(history_dir, exist_ok=True)
    
    raw_json_out = os.path.join(history_dir, f"{audit_id}_diagnostico.json")
    compact_json_out = os.path.join(history_dir, f"{audit_id}_compact.json")
    
    # 1. Ejecutar el validador de consistencia
    validate_bulk_excel(file_path, raw_json_out)
    
    # Copiar también al path por defecto para mantener compatibilidad
    shutil.copyfile(raw_json_out, os.path.join(uploads_dir, "diagnostico_geomecanico.json"))
    
    # 2. Leer diagnóstico detallado para realizar agregaciones estadísticas
    with open(raw_json_out, "r", encoding="utf-8") as f:
        diag = json.load(f)
        
    compact = {k: v for k, v in diag.items() if k != "incidencias"}
    incidencias = diag.get("incidencias", [])
    total_filas = diag.get("total_filas_procesadas", 0)
    
    # --- METRICA 1: Mapeo de Celdas Padre ---
    resumen_celdas = diag.get("resumen_por_celda_padre", {})
    num_celdas_padre = len(resumen_celdas)
    promedio_hijas = sum(x["total_hijas"] for x in resumen_celdas.values()) / max(1, num_celdas_padre)
    
    # --- METRICA 2: Cantidad de campos totales (77 Columnas Obligatorias) ---
    total_fields = total_filas * 77
    total_vacios = sum(1 for i in incidencias if i.get("tipo_incidencia") == "VACIO")
    total_advertencias = sum(1 for i in incidencias if i.get("tipo_incidencia") == "ADVERTENCIA")
    total_alertas = sum(1 for i in incidencias if i.get("tipo_incidencia") == "ALERTA")
    total_correctos = total_fields - (total_vacios + total_advertencias + total_alertas)
    
    # --- METRICA 3: Cantidad de discontinuidades (Filas de Excel) ---
    row_errors = defaultdict(set)
    for i in incidencias:
        row_errors[i["fila_excel"]].add(i["tipo_incidencia"])
        
    discs_con_alerta = sum(1 for row, errs in row_errors.items() if "ALERTA" in errs)
    discs_con_advertencia = sum(1 for row, errs in row_errors.items() if "ADVERTENCIA" in errs and "ALERTA" not in errs)
    discs_con_vacio = sum(1 for row, errs in row_errors.items() if "VACIO" in errs)
    discs_correctas = total_filas - len(row_errors)
    
    # --- METRICAS 4, 5 y 6: Distribuciones por Campaña, Sector y Geólogo ---
    camp_stats = defaultdict(lambda: {"vacios": 0, "advertencias": 0, "alertas": 0, "filas": set()})
    geo_stats = defaultdict(lambda: {"vacios": 0, "advertencias": 0, "alertas": 0, "filas": set()})
    sector_stats = defaultdict(lambda: {"vacios": 0, "advertencias": 0, "alertas": 0, "filas": set()})
    
    for i in incidencias:
        c = i.get("campania", "N/A")
        g = i.get("geotecnico", "N/A")
        s = i.get("sector_geotecnico", "N/A")
        
        camp_stats[c]["filas"].add(i["fila_excel"])
        geo_stats[g]["filas"].add(i["fila_excel"])
        sector_stats[s]["filas"].add(i["fila_excel"])
        
        tipo = i.get("tipo_incidencia")
        if tipo == "VACIO":
            camp_stats[c]["vacios"] += 1
            geo_stats[g]["vacios"] += 1
            sector_stats[s]["vacios"] += 1
        elif tipo == "ADVERTENCIA":
            camp_stats[c]["advertencias"] += 1
            geo_stats[g]["advertencias"] += 1
            sector_stats[s]["advertencias"] += 1
        elif tipo == "ALERTA":
            camp_stats[c]["alertas"] += 1
            geo_stats[g]["alertas"] += 1
            sector_stats[s]["alertas"] += 1
            
    distribucion_campania = []
    for c, stats in camp_stats.items():
        rows_count = len(stats["filas"])
        total_fields_group = rows_count * 77
        distribucion_campania.append({
            "campania": c,
            "discontinuidades": rows_count,
            "vacios_cant": stats["vacios"],
            "vacios_pct": (stats["vacios"] / max(1, total_fields_group)) * 100,
            "advertencias_cant": stats["advertencias"],
            "advertencias_pct": (stats["advertencias"] / max(1, total_fields_group)) * 100,
            "alertas_cant": stats["alertas"],
            "alertas_pct": (stats["alertas"] / max(1, total_fields_group)) * 100,
        })
        
    distribucion_geotecnico = []
    for g, stats in geo_stats.items():
        rows_count = len(stats["filas"])
        total_fields_group = rows_count * 77
        distribucion_geotecnico.append({
            "geotecnico": g,
            "discontinuidades": rows_count,
            "vacios_cant": stats["vacios"],
            "vacios_pct": (stats["vacios"] / max(1, total_fields_group)) * 100,
            "advertencias_cant": stats["advertencias"],
            "advertencias_pct": (stats["advertencias"] / max(1, total_fields_group)) * 100,
            "alertas_cant": stats["alertas"],
            "alertas_pct": (stats["alertas"] / max(1, total_fields_group)) * 100,
        })
        
    distribucion_sector = []
    for s, stats in sector_stats.items():
        rows_count = len(stats["filas"])
        total_fields_group = rows_count * 77
        distribucion_sector.append({
            "sector": s,
            "discontinuidades": rows_count,
            "vacios_cant": stats["vacios"],
            "vacios_pct": (stats["vacios"] / max(1, total_fields_group)) * 100,
            "advertencias_cant": stats["advertencias"],
            "advertencias_pct": (stats["advertencias"] / max(1, total_fields_group)) * 100,
            "alertas_cant": stats["alertas"],
            "alertas_pct": (stats["alertas"] / max(1, total_fields_group)) * 100,
        })
        
    # --- METRICAS 7 y 8: Top Tipo Alertas e Inconsistencias Frecuentes ---
    msg_alertas = Counter(i.get("mensaje") for i in incidencias if i.get("tipo_incidencia") == "ALERTA")
    msg_advertencias = Counter(i.get("mensaje") for i in incidencias if i.get("tipo_incidencia") == "ADVERTENCIA")
    
    top_5_alertas = [{"mensaje": k, "cantidad": v, "pct": (v / max(1, total_alertas)) * 100} for k, v in msg_alertas.most_common(5)]
    lista_detallada_alertas = [{"mensaje": k, "cantidad": v, "pct": (v / max(1, total_alertas)) * 100} for k, v in msg_alertas.most_common(20)]
    lista_detallada_advertencias = [{"mensaje": k, "cantidad": v, "pct": (v / max(1, total_advertencias)) * 100} for k, v in msg_advertencias.most_common(20)]
    
    # 4. Consolidar el Metacompacto final
    compact["audit_id"] = audit_id
    compact["fecha_auditoria"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    compact["nombre_archivo"] = os.path.basename(file_path)
    
    compact["familia1"] = {
        "num_celdas_padre": num_celdas_padre,
        "promedio_hijas": round(promedio_hijas, 2),
        "total_discontinuidades": total_filas
    }
    compact["familia2"] = {
        "total_fields": total_fields,
        "total_vacios": total_vacios,
        "total_advertencias": total_advertencias,
        "total_alertas": total_alertas,
        "total_correctos": total_correctos
    }
    compact["familia3"] = {
        "total_discontinuidades": total_filas,
        "discontinuidades_alertas": discs_con_alerta,
        "discontinuidades_advertencias": discs_con_advertencia,
        "discontinuidades_vacios": discs_con_vacio,
        "discontinuidades_correctas": discs_correctas
    }
    compact["distribucion_campania"] = distribucion_campania
    compact["distribucion_sector"] = distribucion_sector
    compact["distribucion_geotecnico"] = distribucion_geotecnico
    compact["top_5_alertas"] = top_5_alertas
    compact["error_types_detailed"] = {
        "alertas": lista_detallada_alertas,
        "advertencias": lista_detallada_advertencias
    }
    
    # Worst Cells
    sorted_worst = sorted(
        resumen_celdas.items(),
        key=lambda x: (
            x[1].get("alertas", 0),
            x[1].get("vacios", 0),
            x[1].get("advertencias", 0)
        ),
        reverse=True
    )[:20]
    worst_cells = [{"celda": k, **v} for k, v in sorted_worst] # <-- Definida de manera correcta
    
    col_counter = Counter(i.get("columna", "Desconocido") for i in incidencias)
    compact["top_column_errors"] = [{"columna": k, "cantidad": v} for k, v in col_counter.most_common(15)]
    compact["worst_cells"] = worst_cells
    
    with open(compact_json_out, "w", encoding="utf-8") as f:
        json.dump(compact, f, ensure_ascii=False)
        
    # Copiar al default para compatibilidad
    shutil.copyfile(compact_json_out, os.path.join(uploads_dir, "resumen_geomecanico_ligero.json"))


@app.get("/api/geomecanica/incidencias-paginadas")
def obtener_incidencias_paginadas(
    page: int = 1,
    limit: int = 50,
    tipo: str = None,
    celda: str = None,
    columna: str = None,
    campania: str = None,
    geotecnico: str = None,
    sector_geotecnico: str = None,
    search: str = None,
    audit_id: str = None
):
    """Devuelve bloques de incidencias cruzando filtros avanzados con resiliencia en segundo plano."""
    if audit_id:
        raw_file = os.path.join(uploads_dir, "history", f"{audit_id}_diagnostico.json")
        excel_file = os.path.join(uploads_dir, "history", f"{audit_id}.xlsx")
    else:
        raw_file = os.path.join(uploads_dir, "diagnostico_geomecanico.json")
        excel_file = os.path.join(uploads_dir, "bulk_raw.xlsx")
        
    if not os.path.exists(raw_file):
        if os.path.exists(excel_file):
            # El archivo Excel se encuentra procesándose, se retorna vacío sin levantar 404
            return {
                "page": page,
                "limit": limit,
                "total_records": 0,
                "total_pages": 0,
                "data": [],
                "status": "procesando"
            }
        raise HTTPException(status_code=404, detail="El diagnóstico solicitado no existe.")

    with open(raw_file, "r", encoding="utf-8") as f:
        diag_data = json.load(f)

    incidencias = diag_data.get("incidencias", [])
    
    # Aplicar filtros cruzados avanzados en memoria
    filtered = incidencias
    if tipo:
        filtered = [i for i in filtered if i.get("tipo_incidencia") == tipo.upper()]
    if celda:
        celda_up = celda.upper()
        filtered = [i for i in filtered if i.get("celda_padre") == celda_up or i.get("celda_hija") == celda_up]
    if columna:
        filtered = [i for i in filtered if i.get("columna", "").upper() == columna.upper()]
    if campania:
        filtered = [i for i in filtered if i.get("campania") == campania]
    if geotecnico:
        geo_up = geotecnico.upper()
        filtered = [i for i in filtered if i.get("geotecnico", "").upper() == geo_up]
    if sector_geotecnico:
        sect_up = sector_geotecnico.upper()
        filtered = [i for i in filtered if i.get("sector_geotecnico", "").upper() == sect_up]
    if search:
        search_lower = search.lower()
        filtered = [
            i for i in filtered 
            if search_lower in i.get("mensaje", "").lower() 
            or search_lower in i.get("columna", "").lower()
            or search_lower in i.get("celda_padre", "").lower()
        ]
        
    total_records = len(filtered)
    start_idx = (page - 1) * limit
    end_idx = start_idx + limit
    
    return {
        "page": page,
        "limit": limit,
        "total_records": total_records,
        "total_pages": math.ceil(total_records / limit),
        "data": filtered[start_idx:end_idx],
        "status": "completado"
    }

@app.post("/api/geomecanica/importar-excel-bulk")
async def importar_excel_bulk(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo provisto no es una planilla Excel valida.")
        
    audit_id = f"audit_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    history_dir = os.path.join(uploads_dir, "history")
    os.makedirs(history_dir, exist_ok=True)
    
    file_path = os.path.join(history_dir, f"{audit_id}.xlsx")
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    background_tasks.add_task(run_bulk_pipeline_with_id, file_path, audit_id)
    return {"status": "procesando", "audit_id": audit_id, "filename": file.filename}


@app.get("/api/geomecanica/auditorias")
def listar_auditorias():
    history_dir = os.path.join(uploads_dir, "history")
    if not os.path.exists(history_dir):
        return []
    
    audits = []
    for f in os.listdir(history_dir):
        if f.endswith("_compact.json"):
            audit_id = f.replace("_compact.json", "")
            compact_file = os.path.join(history_dir, f)
            try:
                with open(compact_file, "r", encoding="utf-8") as file_content:
                    meta = json.load(file_content)
                    audits.append({
                        "audit_id": audit_id,
                        "fecha": meta.get("fecha_auditoria", "Desconocida"),
                        "archivo": meta.get("nombre_archivo", "Desconocido.xlsx"),
                        "total_filas": meta.get("familia1", {}).get("total_discontinuidades", 0),
                        "total_vacios": meta.get("familia2", {}).get("total_vacios", 0),
                        "total_advertencias": meta.get("familia2", {}).get("total_advertencias", 0),
                        "total_alertas": meta.get("familia2", {}).get("total_alertas", 0),
                    })
            except Exception:
                pass
                
    audits.sort(key=lambda x: x["fecha"], reverse=True)
    return audits


@app.get("/api/geomecanica/resumen-ligero")
def obtener_resumen_ligero(audit_id: str = None):
    """
    Retorna los KPIs compactos. Si el archivo final no existe pero el Excel
    de origen sí, indica al cliente que el proceso está en curso (202).
    """
    global COMPACT_CACHE
    if audit_id:
        compact_file = os.path.join(uploads_dir, "history", f"{audit_id}_compact.json")
        if os.path.exists(compact_file):
            with open(compact_file, "r", encoding="utf-8") as f:
                return json.load(f)
        
        # Si el JSON no existe, pero el Excel sí, es porque se está procesando
        excel_file = os.path.join(uploads_dir, "history", f"{audit_id}.xlsx")
        if os.path.exists(excel_file):
            return JSONResponse(
                status_code=202,
                content={"status": "procesando", "message": "Compilando estadísticas de las 8 familias geotécnicas..."}
            )
            
        raise HTTPException(status_code=404, detail="Auditoría no encontrada")
        
    if COMPACT_CACHE is None:
        compact_file = os.path.join(uploads_dir, "resumen_geomecanico_ligero.json")
        if os.path.exists(compact_file):
            with open(compact_file, "r", encoding="utf-8") as f:
                COMPACT_CACHE = json.load(f)
        else:
            return JSONResponse(status_code=202, content={"status": "procesando", "message": "Los datos generales se están procesando."})
    return COMPACT_CACHE


@app.get("/api/geomecanica/incidencias-paginadas")
def obtener_incidencias_paginadas(
    page: int = 1,
    limit: int = 50,
    tipo: str = None,
    celda: str = None,
    columna: str = None,
    campania: str = None,
    geotecnico: str = None,
    sector_geotecnico: str = None,
    search: str = None,
    audit_id: str = None
):
    if audit_id:
        raw_file = os.path.join(uploads_dir, "history", f"{audit_id}_diagnostico.json")
    else:
        raw_file = os.path.join(uploads_dir, "diagnostico_geomecanico.json")
        
    if not os.path.exists(raw_file):
        raise HTTPException(status_code=404, detail="El diagnóstico solicitado no existe.")

    with open(raw_file, "r", encoding="utf-8") as f:
        diag_data = json.load(f)

    incidencias = diag_data.get("incidencias", [])
    
    filtered = incidencias
    if tipo:
        filtered = [i for i in filtered if i.get("tipo_incidencia") == tipo.upper()]
    if celda:
        celda_up = celda.upper()
        filtered = [i for i in filtered if i.get("celda_padre") == celda_up or i.get("celda_hija") == celda_up]
    if columna:
        filtered = [i for i in filtered if i.get("columna", "").upper() == columna.upper()]
    if campania:
        filtered = [i for i in filtered if i.get("campania") == campania]
    if geotecnico:
        geo_up = geotecnico.upper()
        filtered = [i for i in filtered if i.get("geotecnico", "").upper() == geo_up]
    if sector_geotecnico:
        sect_up = sector_geotecnico.upper()
        filtered = [i for i in filtered if i.get("sector_geotecnico", "").upper() == sect_up]
    if search:
        search_lower = search.lower()
        filtered = [
            i for i in filtered 
            if search_lower in i.get("mensaje", "").lower() 
            or search_lower in i.get("columna", "").lower()
            or search_lower in i.get("celda_padre", "").lower()
        ]
        
    total_records = len(filtered)
    start_idx = (page - 1) * limit
    end_idx = start_idx + limit
    
    return {
        "page": page,
        "limit": limit,
        "total_records": total_records,
        "total_pages": math.ceil(total_records / limit),
        "data": filtered[start_idx:end_idx]
    }


@app.get("/api/geomecanica/reporte-markdown")
def descargar_reporte_markdown(
    tipo: str = None,
    celda: str = None,
    columna: str = None,
    campania: str = None,
    geotecnico: str = None,
    sector_geotecnico: str = None,
    search: str = None,
    audit_id: str = None
):
    if audit_id:
        raw_file = os.path.join(uploads_dir, "history", f"{audit_id}_diagnostico.json")
    else:
        raw_file = os.path.join(uploads_dir, "diagnostico_geomecanico.json")
        
    if not os.path.exists(raw_file):
        raise HTTPException(status_code=404, detail="El diagnóstico solicitado no ha sido generado.")
        
    with open(raw_file, "r", encoding="utf-8") as f:
        diag = json.load(f)
            
    incidencias = diag.get("incidencias", [])
    
    # Aplicar idénticos filtros cruzados
    filtered = incidencias
    if tipo:
        filtered = [i for i in filtered if i.get("tipo_incidencia") == tipo.upper()]
    if celda:
        celda_up = celda.upper()
        filtered = [i for i in filtered if i.get("celda_padre") == celda_up or i.get("celda_hija") == celda_up]
    if columna:
        filtered = [i for i in filtered if i.get("columna", "").upper() == columna.upper()]
    if campania:
        filtered = [i for i in filtered if i.get("campania") == campania]
    if geotecnico:
        geo_up = geotecnico.upper()
        filtered = [i for i in filtered if i.get("geotecnico", "").upper() == geo_up]
    if sector_geotecnico:
        sect_up = sector_geotecnico.upper()
        filtered = [i for i in filtered if i.get("sector_geotecnico", "").upper() == sect_up]
    if search:
        search_lower = search.lower()
        filtered = [
            i for i in filtered 
            if search_lower in i.get("mensaje", "").lower() 
            or search_lower in i.get("columna", "").lower()
        ]
        
    md_content = []
    md_content.append("# Reporte de Auditoria de Consistencia Geomecanica Detallado")
    md_content.append(f"**Generado el:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  ")
    md_content.append(f"**Total Incidencias Coincidentes:** {len(filtered):,}  \n")
    md_content.append("### Filtros de Auditoria Aplicados:")
    md_content.append(f"*   **Tipo de Incidencia:** `{tipo if tipo else 'TODOS'}`")
    md_content.append(f"*   **Columna del Excel:** `{columna if columna else 'TODAS'}`")
    md_content.append(f"*   **Estacion de Mapeo (Celda):** `{celda if celda else 'TODAS'}`")
    md_content.append(f"*   **Campaña (Año):** `{campania if campania else 'TODAS'}`")
    md_content.append(f"*   **Sector Geotécnico:** `{sector_geotecnico if sector_geotecnico else 'TODOS'}`")
    md_content.append(f"*   **Geotécnico Logueador:** `{geotecnico if geotecnico else 'TODOS'}`")
    if search:
        md_content.append(f"*   **Filtro por Buscador:** `{search}`")
    md_content.append("\n---")
    
    md_content.append("\n| Fila Excel | Celda Padre | Celda Hija | Columna | Valor Actual | Tipo | Mensaje de Retroalimentación |")
    md_content.append("| :-: | :--- | :--- | :--- | :---: | :---: | :--- |")
    
    limite_exportacion = 3000
    for inc in filtered[:limite_exportacion]:
        val = inc.get("valor_actual")
        val_display = val if val is not None else "—"
        md_content.append(
            f"| {inc.get('fila_excel', '—')} | "
            f"{inc.get('celda_padre', '—')} | "
            f"{inc.get('celda_hija', '—')} | "
            f"{inc.get('columna', '—')} | "
            f"{val_display} | "
            f"`{inc.get('tipo_incidencia')}` | "
            f"{inc.get('mensaje', '—')} |"
        )
        
    if len(filtered) > limite_exportacion:
        md_content.append(f"\n*Nota: Se han acotado los registros detallados a los primeros {limite_exportacion} elementos por estabilidad del archivo.*")
        
    md_string = "\n".join(md_content)
    
    return StreamingResponse(
        io.BytesIO(md_string.encode("utf-8")),
        media_type="text/markdown",
        headers={"Content-Disposition": "attachment; filename=reporte_auditoria_detallado.md"}
    )

# Funciones auxiliares de conversión segura para evitar TypeErrors en datos antiguos
def safe_int(val, default=0):
    if val is None:
        return default
    try:
        return int(val)
    except:
        return default

def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except:
        return default

@app.get("/api/geomecanica/reporte-excel")
def descargar_reporte_excel(
    tipo: str = None,
    celda: str = None,
    columna: str = None,
    campania: str = None,
    geotecnico: str = None,
    sector_geotecnico: str = None,
    search: str = None,
    audit_id: str = None
):
    # Ruta de archivos por defecto
    if audit_id:
        raw_file = os.path.join(uploads_dir, "history", f"{audit_id}_diagnostico.json")
        compact_file = os.path.join(uploads_dir, "history", f"{audit_id}_compact.json")
    else:
        raw_file = os.path.join(uploads_dir, "diagnostico_geomecanico.json")
        compact_file = os.path.join(uploads_dir, "resumen_geomecanico_ligero.json")
        
        # Búsqueda autocurativa en el historial si los archivos por defecto no están
        if not os.path.exists(raw_file) or not os.path.exists(compact_file):
            history_dir = os.path.join(uploads_dir, "history")
            if os.path.exists(history_dir):
                jsons = [f for f in os.listdir(history_dir) if f.endswith("_diagnostico.json")]
                if jsons:
                    jsons.sort(key=lambda x: os.path.getmtime(os.path.join(history_dir, x)), reverse=True)
                    latest_id = jsons[0].replace("_diagnostico.json", "")
                    raw_file = os.path.join(history_dir, f"{latest_id}_diagnostico.json")
                    compact_file = os.path.join(history_dir, f"{latest_id}_compact.json")
            
    if not os.path.exists(raw_file) or not os.path.exists(compact_file):
        raise HTTPException(status_code=404, detail="El diagnóstico solicitado no ha sido generado o está incompleto.")
        
    with open(raw_file, "r", encoding="utf-8") as f:
        diag = json.load(f)
    with open(compact_file, "r", encoding="utf-8") as f:
        compact = json.load(f)
        
    incidencias = diag.get("incidencias", [])
    
    # Aplicar filtros cruzados dinámicos
    filtered = incidencias
    if tipo:
        filtered = [i for i in filtered if i.get("tipo_incidencia") == tipo.upper()]
    if celda:
        celda_up = celda.upper()
        filtered = [i for i in filtered if i.get("celda_padre") == celda_up or i.get("celda_hija") == celda_up]
    if columna:
        filtered = [i for i in filtered if i.get("columna", "").upper() == columna.upper()]
    if campania:
        filtered = [i for i in filtered if i.get("campania") == campania]
    if geotecnico:
        geo_up = geotecnico.upper()
        filtered = [i for i in filtered if i.get("geotecnico", "").upper() == geo_up]
    if sector_geotecnico:
        sect_up = sector_geotecnico.upper()
        filtered = [i for i in filtered if i.get("sector_geotecnico", "").upper() == sect_up]
    if search:
        search_lower = search.lower()
        filtered = [
            i for i in filtered 
            if search_lower in i.get("mensaje", "").lower() 
            or search_lower in i.get("columna", "").lower()
            or search_lower in i.get("celda_padre", "").lower()
        ]

    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Estilos reutilizables
    font_title = Font(name="Calibri", size=16, bold=True, color="1B365D")
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="555555")
    font_section = Font(name="Calibri", size=11, bold=True, color="1B365D")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_regular = Font(name="Calibri", size=10, color="000000")
    
    fill_primary = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_accent_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_accent_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_accent_orange = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_accent_red = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")
    alignment_right = Alignment(horizontal="right", vertical="center")
    
    # ----------------------------------------------------
    # HOJA 1: RESUMEN EJECUTIVO (Poco volumen, estéticamente pulido)
    # ----------------------------------------------------
    ws1 = wb.create_sheet(title="Resumen Ejecutivo")
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.cell(row=2, column=2, value="AUDITORÍA DE INTEGRIDAD GEOTÉCNICA").font = font_title
    ws1.cell(row=3, column=2, value="Reporte consolidado del estado de consistencia física y lógica").font = font_subtitle
    
    meta_rows = [
        ("Archivo Auditado:", compact.get("nombre_archivo", "N/A")),
        ("Fecha de Auditoría:", compact.get("fecha_auditoria", "N/A")),
        ("Celdas Padre Evaluadas:", safe_int(compact.get("familia1", {}).get("num_celdas_padre", 0))),
        ("Total de Estructuras (Filas):", safe_int(compact.get("familia1", {}).get("total_discontinuidades", 0))),
        ("Total de Datos (Campos):", safe_int(compact.get("familia2", {}).get("total_fields", 0)))
    ]
    
    start_row = 5
    for label, val in meta_rows:
        c1 = ws1.cell(row=start_row, column=2, value=label)
        c1.font = font_bold
        c1.border = border_thin
        c2 = ws1.cell(row=start_row, column=3, value=val)
        c2.font = font_regular
        c2.border = border_thin
        if isinstance(val, (int, float)):
            c2.number_format = '#,##0'
            c2.alignment = alignment_right
        start_row += 1
        
    ws1.cell(row=11, column=2, value="MÉTRICAS DE ESTRUCTURAS Y CELDAS").font = font_section
    header_row = 12
    cols = ["Métrica", "Valor", "Descripción"]
    for idx, col in enumerate(cols, start=2):
        cell = ws1.cell(row=header_row, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin
        
    fam1 = compact.get("familia1", {})
    general_metrics = [
        ("Celdas Padre", safe_int(fam1.get("num_celdas_padre", 0)), "Estaciones totales auditadas"),
        ("Promedio de Capas Hijas", safe_float(fam1.get("promedio_hijas", 0)), "Estructuras promedio por estación"),
        ("Total de Estructuras Mapeadas", safe_int(fam1.get("total_discontinuidades", 0)), "Total de filas de discontinuidades")
    ]
    
    curr_row = 13
    for m, v, d in general_metrics:
        ws1.cell(row=curr_row, column=2, value=m).font = font_regular
        ws1.cell(row=curr_row, column=2).border = border_thin
        
        val_cell = ws1.cell(row=curr_row, column=3, value=v)
        val_cell.font = font_bold
        val_cell.border = border_thin
        val_cell.alignment = alignment_right
        if isinstance(v, int):
            val_cell.number_format = '#,##0'
        elif isinstance(v, float):
            val_cell.number_format = '#,##0.00'
            
        ws1.cell(row=curr_row, column=4, value=d).font = font_regular
        ws1.cell(row=curr_row, column=4).border = border_thin
        curr_row += 1
        
    ws1.cell(row=17, column=2, value="AUDITORÍA DE DATOS INDIVIDUALES (CAMPOS)").font = font_section
    header_row = 18
    for idx, col in enumerate(["Estado de Campo", "Cantidad de Campos", "Porcentaje", "Acción Recomendada"], start=2):
        cell = ws1.cell(row=header_row, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin
        
    fam2 = compact.get("familia2", {})
    total_fields = max(1, safe_int(fam2.get("total_fields", 1)))
    fields_metrics = [
        ("Campos OK", safe_int(fam2.get("total_correctos", 0)), fill_accent_green, "Datos validados, listos para análisis geomecánico"),
        ("Campos Vacíos", safe_int(fam2.get("total_vacios", 0)), fill_accent_yellow, "Completar celdas obligatorias según scanline"),
        ("Advertencias", safe_int(fam2.get("total_advertencias", 0)), fill_accent_orange, "Revisar posibles inconsistencias lógicas"),
        ("Alertas Críticas", safe_int(fam2.get("total_alertas", 0)), fill_accent_red, "Corregir inmediatamente para evitar distorsión RMR")
    ]
    
    curr_row = 19
    for name, qty, fill, rcmd in fields_metrics:
        ws1.cell(row=curr_row, column=2, value=name).font = font_regular
        ws1.cell(row=curr_row, column=2).border = border_thin
        
        qty_cell = ws1.cell(row=curr_row, column=3, value=qty)
        qty_cell.font = font_bold
        qty_cell.border = border_thin
        qty_cell.alignment = alignment_right
        qty_cell.number_format = '#,##0'
        
        pct_cell = ws1.cell(row=curr_row, column=4, value=qty / total_fields)
        pct_cell.font = font_bold
        pct_cell.border = border_thin
        pct_cell.alignment = alignment_right
        pct_cell.number_format = '0.00%'
        
        rc_cell = ws1.cell(row=curr_row, column=5, value=rcmd)
        rc_cell.font = font_regular
        rc_cell.border = border_thin
        rc_cell.fill = fill
        
        curr_row += 1
        
    ws1.cell(row=24, column=2, value="AUDITORÍA POR ESTRUCTURAS (FILAS)").font = font_section
    header_row = 25
    for idx, col in enumerate(["Estado de Estructura", "Cantidad de Filas", "Porcentaje", "Efecto en Calidad de Base de Datos"], start=2):
        cell = ws1.cell(row=header_row, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin
        
    fam3 = compact.get("familia3", {})
    total_discs = max(1, safe_int(fam3.get("total_discontinuidades", 1)))
    discs_metrics = [
        ("Filas 100% OK", safe_int(fam3.get("discontinuidades_correctas", 0)), fill_accent_green, "Integridad de registros geomecánicos perfecta"),
        ("Filas con Vacíos", safe_int(fam3.get("discontinuidades_vacios", 0)), fill_accent_yellow, "Registros incompletos con datos vacíos"),
        ("Filas con Advertencias", safe_int(fam3.get("discontinuidades_advertencias", 0)), fill_accent_orange, "Registros con desvíos leves de consistencia"),
        ("Filas con Alertas", safe_int(fam3.get("discontinuidades_alertas", 0)), fill_accent_red, "Incompatibilidad física o geométrica grave detectada")
    ]
    
    curr_row = 26
    for name, qty, fill, eff in discs_metrics:
        ws1.cell(row=curr_row, column=2, value=name).font = font_regular
        ws1.cell(row=curr_row, column=2).border = border_thin
        
        qty_cell = ws1.cell(row=curr_row, column=3, value=qty)
        qty_cell.font = font_bold
        qty_cell.border = border_thin
        qty_cell.alignment = alignment_right
        qty_cell.number_format = '#,##0'
        
        pct_cell = ws1.cell(row=curr_row, column=4, value=qty / total_discs)
        pct_cell.font = font_bold
        pct_cell.border = border_thin
        pct_cell.alignment = alignment_right
        pct_cell.number_format = '0.00%'
        
        eff_cell = ws1.cell(row=curr_row, column=5, value=eff)
        eff_cell.font = font_regular
        eff_cell.border = border_thin
        eff_cell.fill = fill
        
        curr_row += 1
        
    # --- HOJA 2: DISTRIBUCIONES & CELDAS ---
    ws2 = wb.create_sheet(title="Distribuciones & Celdas")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.cell(row=2, column=2, value="DISTRIBUCIÓN POR CAMPAÑA DE LOGUEO (AÑO)").font = font_section
    r_cam = 3
    for idx, col in enumerate(["Campaña", "Discontinuidades", "Alertas (cant)", "Alertas (%)", "Vacíos (cant)", "Vacíos (%)"], start=2):
        cell = ws2.cell(row=r_cam, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin
        
    for row in compact.get("distribucion_campania", []):
        r_cam += 1
        ws2.cell(row=r_cam, column=2, value=row.get("campania")).font = font_bold
        ws2.cell(row=r_cam, column=3, value=safe_int(row.get("discontinuidades"))).font = font_regular
        ws2.cell(row=r_cam, column=4, value=safe_int(row.get("alertas_cant"))).font = font_regular
        
        alertas_pct = safe_float(row.get("alertas_pct")) / 100.0
        ws2.cell(row=r_cam, column=5, value=alertas_pct).font = font_regular
        
        ws2.cell(row=r_cam, column=6, value=safe_int(row.get("vacios_cant"))).font = font_regular
        
        vacios_pct = safe_float(row.get("vacios_pct")) / 100.0
        ws2.cell(row=r_cam, column=7, value=vacios_pct).font = font_regular
        
        ws2.cell(row=r_cam, column=2).alignment = alignment_center
        ws2.cell(row=r_cam, column=3).number_format = '#,##0'
        ws2.cell(row=r_cam, column=3).alignment = alignment_right
        ws2.cell(row=r_cam, column=4).number_format = '#,##0'
        ws2.cell(row=r_cam, column=4).alignment = alignment_right
        ws2.cell(row=r_cam, column=5).number_format = '0.00%'
        ws2.cell(row=r_cam, column=5).alignment = alignment_right
        ws2.cell(row=r_cam, column=6).number_format = '#,##0'
        ws2.cell(row=r_cam, column=6).alignment = alignment_right
        ws2.cell(row=r_cam, column=7).number_format = '0.00%'
        ws2.cell(row=r_cam, column=7).alignment = alignment_right
        
        for col_idx in range(2, 8):
            ws2.cell(row=r_cam, column=col_idx).border = border_thin
            
    r_sec = r_cam + 3
    ws2.cell(row=r_sec-1, column=2, value="DISTRIBUCIÓN POR SECTOR GEOTÉCNICO").font = font_section
    for idx, col in enumerate(["Sector", "Discontinuidades", "Alertas (cant)", "Alertas (%)", "Vacíos (cant)", "Vacíos (%)"], start=2):
        cell = ws2.cell(row=r_sec, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin
        
    for row in compact.get("distribucion_sector", []):
        r_sec += 1
        ws2.cell(row=r_sec, column=2, value=row.get("sector")).font = font_bold
        ws2.cell(row=r_sec, column=3, value=safe_int(row.get("discontinuidades"))).font = font_regular
        ws2.cell(row=r_sec, column=4, value=safe_int(row.get("alertas_cant"))).font = font_regular
        
        alertas_pct = safe_float(row.get("alertas_pct")) / 100.0
        ws2.cell(row=r_sec, column=5, value=alertas_pct).font = font_regular
        
        ws2.cell(row=r_sec, column=6, value=safe_int(row.get("vacios_cant"))).font = font_regular
        
        vacios_pct = safe_float(row.get("vacios_pct")) / 100.0
        ws2.cell(row=r_sec, column=7, value=vacios_pct).font = font_regular
        
        ws2.cell(row=r_sec, column=2).alignment = alignment_center
        ws2.cell(row=r_sec, column=3).number_format = '#,##0'
        ws2.cell(row=r_sec, column=3).alignment = alignment_right
        ws2.cell(row=r_sec, column=4).number_format = '#,##0'
        ws2.cell(row=r_sec, column=4).alignment = alignment_right
        ws2.cell(row=r_sec, column=5).number_format = '0.00%'
        ws2.cell(row=r_sec, column=5).alignment = alignment_right
        ws2.cell(row=r_sec, column=6).number_format = '#,##0'
        ws2.cell(row=r_sec, column=6).alignment = alignment_right
        ws2.cell(row=r_sec, column=7).number_format = '0.00%'
        ws2.cell(row=r_sec, column=7).alignment = alignment_right
        
        for col_idx in range(2, 8):
            ws2.cell(row=r_sec, column=col_idx).border = border_thin
            
    r_geo = r_sec + 3
    ws2.cell(row=r_geo-1, column=2, value="CALIDAD DE REGISTRO POR GEOTÉCNICO / PERSONA").font = font_section
    for idx, col in enumerate(["Geotécnico", "Discontinuidades", "Alertas (cant)", "Alertas (%)", "Vacíos (cant)", "Vacíos (%)"], start=2):
        cell = ws2.cell(row=r_geo, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin
        
    for row in compact.get("distribucion_geotecnico", []):
        r_geo += 1
        ws2.cell(row=r_geo, column=2, value=row.get("geotecnico")).font = font_bold
        ws2.cell(row=r_geo, column=3, value=safe_int(row.get("discontinuidades"))).font = font_regular
        ws2.cell(row=r_geo, column=4, value=safe_int(row.get("alertas_cant"))).font = font_regular
        
        alertas_pct = safe_float(row.get("alertas_pct")) / 100.0
        ws2.cell(row=r_geo, column=5, value=alertas_pct).font = font_regular
        
        ws2.cell(row=r_geo, column=6, value=safe_int(row.get("vacios_cant"))).font = font_regular
        
        vacios_pct = safe_float(row.get("vacios_pct")) / 100.0
        ws2.cell(row=r_geo, column=7, value=vacios_pct).font = font_regular
        
        ws2.cell(row=r_geo, column=2).alignment = alignment_center
        ws2.cell(row=r_geo, column=3).number_format = '#,##0'
        ws2.cell(row=r_geo, column=3).alignment = alignment_right
        ws2.cell(row=r_geo, column=4).number_format = '#,##0'
        ws2.cell(row=r_geo, column=4).alignment = alignment_right
        ws2.cell(row=r_geo, column=5).number_format = '0.00%'
        ws2.cell(row=r_geo, column=5).alignment = alignment_right
        ws2.cell(row=r_geo, column=6).number_format = '#,##0'
        ws2.cell(row=r_geo, column=6).alignment = alignment_right
        ws2.cell(row=r_geo, column=7).number_format = '0.00%'
        ws2.cell(row=r_geo, column=7).alignment = alignment_right
        
        for col_idx in range(2, 8):
            ws2.cell(row=r_geo, column=col_idx).border = border_thin
            
    r_worst = r_geo + 3
    ws2.cell(row=r_worst-1, column=2, value="PEORES CELDAS DE ESTACIÓN (MAYOR ACUMULACIÓN DE OBSERVACIONES)").font = font_section
    for idx, col in enumerate(["Estación (Celda)", "Total Hijas", "Vacíos", "Advertencias", "Alertas", "Calificación"], start=2):
        cell = ws2.cell(row=r_worst, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin
        
    for row in compact.get("worst_cells", []):
        r_worst += 1
        ws2.cell(row=r_worst, column=2, value=row.get("celda")).font = font_bold
        ws2.cell(row=r_worst, column=3, value=safe_int(row.get("total_hijas"))).font = font_regular
        ws2.cell(row=r_worst, column=4, value=safe_int(row.get("vacios"))).font = font_regular
        ws2.cell(row=r_worst, column=5, value=safe_int(row.get("advertencias"))).font = font_regular
        ws2.cell(row=r_worst, column=6, value=safe_int(row.get("alertas"))).font = font_regular
        
        status = row.get("estado_celda", "OK")
        status_cell = ws2.cell(row=r_worst, column=7, value=status)
        status_cell.font = font_bold
        if status == "ALERTA":
            status_cell.fill = fill_accent_red
        elif status == "ADVERTENCIA":
            status_cell.fill = fill_accent_orange
        else:
            status_cell.fill = fill_accent_green
            
        ws2.cell(row=r_worst, column=2).alignment = alignment_center
        ws2.cell(row=r_worst, column=3).number_format = '#,##0'
        ws2.cell(row=r_worst, column=3).alignment = alignment_right
        ws2.cell(row=r_worst, column=4).number_format = '#,##0'
        ws2.cell(row=r_worst, column=4).alignment = alignment_right
        ws2.cell(row=r_worst, column=5).number_format = '#,##0'
        ws2.cell(row=r_worst, column=5).alignment = alignment_right
        ws2.cell(row=r_worst, column=6).number_format = '#,##0'
        ws2.cell(row=r_worst, column=6).alignment = alignment_right
        ws2.cell(row=r_worst, column=7).alignment = alignment_center
        
        for col_idx in range(2, 8):
            ws2.cell(row=r_worst, column=col_idx).border = border_thin
            
    # --- HOJA 3: TOP FRECUENCIA DE ERRORES ---
    ws3 = wb.create_sheet(title="Top Frecuencia de Errores")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.cell(row=2, column=2, value="ALERTAS CRÍTICAS CON MAYOR CANTIDAD DE OCURRENCIAS").font = font_section
    r_err = 3
    for idx, col in enumerate(["Ranking", "Mensaje de Alerta Crítica", "Ocurrencias", "Porcentaje (%)"], start=2):
        cell = ws3.cell(row=r_err, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin
        
    for i_idx, item in enumerate(compact.get("error_types_detailed", {}).get("alertas", [])):
        r_err += 1
        ws3.cell(row=r_err, column=2, value=i_idx+1).font = font_bold
        ws3.cell(row=r_err, column=3, value=item.get("mensaje")).font = font_regular
        ws3.cell(row=r_err, column=4, value=safe_int(item.get("cantidad"))).font = font_regular
        
        pct = safe_float(item.get("pct")) / 100.0
        ws3.cell(row=r_err, column=5, value=pct).font = font_regular
        
        ws3.cell(row=r_err, column=2).alignment = alignment_center
        ws3.cell(row=r_err, column=4).number_format = '#,##0'
        ws3.cell(row=r_err, column=4).alignment = alignment_right
        ws3.cell(row=r_err, column=5).number_format = '0.00%'
        ws3.cell(row=r_err, column=5).alignment = alignment_right
        
        for col_idx in range(2, 6):
            cell_border = ws3.cell(row=r_err, column=col_idx)
            cell_border.border = border_thin
            if i_idx < 3:
                cell_border.fill = fill_accent_red
                
    r_warn = r_err + 3
    ws3.cell(row=r_warn-1, column=2, value="ADVERTENCIAS DE CONSISTENCIA CON MAYOR CANTIDAD DE OCURRENCIAS").font = font_section
    for idx, col in enumerate(["Ranking", "Mensaje de Advertencia", "Ocurrencias", "Porcentaje (%)"], start=2):
        cell = ws3.cell(row=r_warn, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin
        
    for i_idx, item in enumerate(compact.get("error_types_detailed", {}).get("advertencias", [])):
        r_warn += 1
        ws3.cell(row=r_warn, column=2, value=i_idx+1).font = font_bold
        ws3.cell(row=r_warn, column=3, value=item.get("mensaje")).font = font_regular
        ws3.cell(row=r_warn, column=4, value=safe_int(item.get("cantidad"))).font = font_regular
        
        pct = safe_float(item.get("pct")) / 100.0
        ws3.cell(row=r_warn, column=5, value=pct).font = font_regular
        
        ws3.cell(row=r_warn, column=2).alignment = alignment_center
        ws3.cell(row=r_warn, column=4).number_format = '#,##0'
        ws3.cell(row=r_warn, column=4).alignment = alignment_right
        ws3.cell(row=r_warn, column=5).number_format = '0.00%'
        ws3.cell(row=r_warn, column=5).alignment = alignment_right
        
        for col_idx in range(2, 6):
            cell_border = ws3.cell(row=r_warn, column=col_idx)
            cell_border.border = border_thin
            if i_idx < 3:
                cell_border.fill = fill_accent_orange
                
    # ----------------------------------------------------
    # HOJA 4: DETALLE INDIVIDUAL FILTRADO (CORREGIDO)
    # ----------------------------------------------------
    ws4 = wb.create_sheet(title="Listado Incidencias Detalle")
    ws4.views.sheetView[0].showGridLines = True
    
    headers_inc = [
        "Fila Excel", "Celda Padre", "Celda Hija", "Campaña", 
        "Geotécnico", "Sector Geotécnico", "Columna Evaluada", 
        "Valor Actual", "Tipo Incidencia", "Mensaje"
    ]
    
    ws4.cell(row=2, column=2, value="DETALLE INDIVIDUAL DE REGISTROS CON OBSERVACIONES (FILTRADO)").font = font_section
    ws4.cell(row=3, column=2, value="Listado dinámico según filtros cruzados aplicados en el panel de auditoría").font = font_subtitle
    
    # Escribir cabecera (Fila 5)
    for idx, col in enumerate(headers_inc, start=2):
        cell = ws4.cell(row=5, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin
        
    # Acotar exportación a un límite seguro de filas (ej. 12,000) para asegurar tiempos óptimos de descarga
    limite_filas = 12000
    r_inc = 5
    for inc in filtered[:limite_filas]:
        r_inc += 1
        ws4.cell(row=r_inc, column=2, value=safe_int(inc.get("fila_excel"))).font = font_regular
        ws4.cell(row=r_inc, column=3, value=inc.get("celda_padre")).font = font_bold
        ws4.cell(row=r_inc, column=4, value=inc.get("celda_hija")).font = font_regular
        ws4.cell(row=r_inc, column=5, value=inc.get("campania")).font = font_regular
        ws4.cell(row=r_inc, column=6, value=inc.get("geotecnico")).font = font_regular
        ws4.cell(row=r_inc, column=7, value=inc.get("sector_geotecnico")).font = font_regular
        ws4.cell(row=r_inc, column=8, value=inc.get("columna")).font = font_regular
        
        val_act = inc.get("valor_actual")
        ws4.cell(row=r_inc, column=9, value=val_act if val_act is not None else "—").font = font_regular
        
        tipo_inc = inc.get("tipo_incidencia")
        tipo_cell = ws4.cell(row=r_inc, column=10, value=tipo_inc)
        tipo_cell.font = font_bold
        if tipo_inc == "ALERTA":
            tipo_cell.fill = fill_accent_red
        elif tipo_inc == "ADVERTENCIA":
            tipo_cell.fill = fill_accent_orange
        else:
            tipo_cell.fill = fill_accent_yellow
            
        ws4.cell(row=r_inc, column=11, value=inc.get("mensaje")).font = font_regular
        
        # Estilos rápidos por celda
        ws4.cell(row=r_inc, column=2).alignment = alignment_center
        ws4.cell(row=r_inc, column=3).alignment = alignment_center
        ws4.cell(row=r_inc, column=4).alignment = alignment_center
        ws4.cell(row=r_inc, column=5).alignment = alignment_center
        ws4.cell(row=r_inc, column=6).alignment = alignment_left
        ws4.cell(row=r_inc, column=7).alignment = alignment_center
        ws4.cell(row=r_inc, column=8).alignment = alignment_left
        ws4.cell(row=r_inc, column=9).alignment = alignment_center
        ws4.cell(row=r_inc, column=10).alignment = alignment_center
        ws4.cell(row=r_inc, column=11).alignment = alignment_left
        
        for col_idx in range(2, 12):
            ws4.cell(row=r_inc, column=col_idx).border = border_thin
            
    # Auto-ajustar anchos para evitar desbordes y textos truncados (###)
    for ws in [ws1, ws2, ws3, ws4]:
        ws.column_dimensions['A'].width = 3
        for col in ws.columns:
            if col[0].column == 1:
                continue
            vals = [str(cell.value or '') for cell in col if cell.value is not None]
            if not vals:
                continue
            max_len = max(len(v) for v in vals)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 48)
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"reporte_auditoria_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )