"""
parsers/excel_a.py — Parser de Excel formato A (Estaciones).

Lee una hoja de Excel con bloques de estación geomecánica (formato "Estaciones"),
detectando cada bloque por el ancla "UBICACIÓN" y extrayendo:

  - Cabecera de la estación (coordenadas INI/FIN, largo, altura, orientación,
    litología LITO-3, alteración, intemperismo, mapeador, fase, nivel, sector).
  - Tabla de discontinuidades (ID = familia, distancia, tipo, dip, dip_dir,
    rellenos, JRC, rugosidad, forma, alteración, etc.).
  - Tabla resumen RMR/G SI (JC76 y JC89) con fecha de mapeo.

La lógica proviene de `docs/import_excel_A.py` (script de referencia),
adaptada a un módulo reutilizable con salida de diccionarios (sin CSV/JSON),
pensada para integrarse al endpoint de previsualización de importación.

Funciones públicas:
  - detect_format(ws)            -> 'a' | 'b'
  - parse_excel_a(ws)            -> list[dict] (estaciones crudas)
  - normalize_station_to_celda(station, db) -> dict compatible con el
    shape de `celdas_preview` del importador Excel B.
"""

import math
import re
import unicodedata
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Utilidades de texto (tolerantes a tildes / mayúsculas / espacios extra)
# ---------------------------------------------------------------------------


def normalize(text: Any) -> str:
    if text is None:
        return ""
    text = str(text).strip().lower()
    text = "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )
    text = re.sub(r"\s+", " ", text)
    text = text.rstrip(":.")
    return text


def text_matches(cell_value: Any, *variants: str) -> bool:
    """True si cell_value (normalizado) es igual a, o contiene como PALABRA
    COMPLETA, alguna de las variantes dadas. Evita falsos positivos como
    que la etiqueta 'ALT' matchee dentro de 'ALTURA'."""
    if not isinstance(cell_value, str):
        return False
    norm = normalize(cell_value)
    for v in variants:
        v = normalize(v)
        if not v:
            continue
        if norm == v:
            return True
        if re.search(r"(?<![a-z0-9])" + re.escape(v) + r"(?![a-z0-9])", norm):
            return True
    return False


# ---------------------------------------------------------------------------
# Búsqueda genérica de celdas-ancla dentro de una ventana de filas/cols
# ---------------------------------------------------------------------------


def find_anchor(ws: Worksheet, row_range: Tuple[int, int], col_range: Tuple[int, int], *label_variants: str) -> Optional[Tuple[int, int]]:
    """Busca la primera celda cuyo texto matchee alguna variante dentro del
    rango dado. Devuelve (row, col) o None."""
    for r in range(row_range[0], row_range[1] + 1):
        for c in range(col_range[0], col_range[1] + 1):
            v = ws.cell(row=r, column=c).value
            if text_matches(v, *label_variants):
                return (r, c)
    return None


def value_after_label(ws: Worksheet, r: int, c: int, max_skip: int = 6) -> Any:
    """Primer valor no vacío a la derecha de la celda-etiqueta (r, c)."""
    for dc in range(1, max_skip + 1):
        v = ws.cell(row=r, column=c + dc).value
        if v is not None:
            return v
    return None


def values_after_label(ws: Worksheet, r: int, c: int, count: int, max_skip: int = 10) -> List[Any]:
    """Primeros `count` valores no vacíos hacia la derecha (filas INI/FIN)."""
    found: List[Any] = []
    for dc in range(1, max_skip + 1):
        v = ws.cell(row=r, column=c + dc).value
        if v is not None:
            found.append(v)
            if len(found) == count:
                break
    while len(found) < count:
        found.append(None)
    return found


# ---------------------------------------------------------------------------
# Detección de bloques de estación
# ---------------------------------------------------------------------------


def get_station_starts(ws: Worksheet, max_col: int = 60) -> List[Dict[str, Any]]:
    """Un bloque de estación empieza en la fila con el texto 'UBICACIÓN'.
    El código de estación (TD1, TD2...) se busca cerca de la esquina
    izquierda de esa fila (o 1 fila arriba/abajo)."""
    starts: List[Dict[str, Any]] = []
    for r in range(1, ws.max_row + 1):
        anchor = None
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if text_matches(v, "UBICACION"):
                anchor = (r, c)
                break
        if anchor is None:
            continue

        code: Optional[str] = None
        code_row = anchor[0]
        for dr in (0, -1, 1):
            for c in range(1, 6):
                v = ws.cell(row=anchor[0] + dr, column=c).value
                if isinstance(v, str) and v.strip() and not text_matches(v, "UBICACION"):
                    code = v.strip()
                    code_row = anchor[0] + dr
                    break
            if code:
                break
        if code:
            starts.append({"header_row": anchor[0], "code_row": code_row, "codigo": code})
    return starts


# ---------------------------------------------------------------------------
# Cabecera / ficha técnica de la estación
# ---------------------------------------------------------------------------


def parse_header_block(ws: Worksheet, block_top: int, block_bottom: int) -> Dict[str, Any]:
    row_range = (max(1, block_top - 2), block_bottom)
    col_range = (1, 40)

    def anchor_value(*variants: str, count: int = 1, max_skip: int = 6):
        a = find_anchor(ws, row_range, col_range, *variants)
        if a is None:
            return None if count == 1 else [None] * count
        r, c = a
        if count == 1:
            return value_after_label(ws, r, c, max_skip=max_skip)
        return values_after_label(ws, r, c, count, max_skip=max_skip)

    este_i, norte_i, cota_i = anchor_value("INI", count=3)
    este_f, norte_f, cota_f = anchor_value("FIN", count=3)

    return {
        "ubicacion_ini": {"este": este_i, "norte": norte_i, "cota": cota_i},
        "ubicacion_fin": {"este": este_f, "norte": norte_f, "cota": cota_f},
        "largo": anchor_value("LARGO"),
        "altura": anchor_value("ALTURA"),
        "dip_talud": anchor_value("Dip_talud", "Dip talud"),
        "dipdir_talud": anchor_value("DipDir_Talud", "Dip Dir Talud"),
        "dip_hole": anchor_value("Dip_hole", "Dip hole"),
        "az_hole": anchor_value("Az_hole", "Az hole"),
        "lito_3": anchor_value("LITO-3", "LITO 3"),
        "lito_modelo": anchor_value("Lito-Model", "Lito Model", "Lito-Modelo"),
        "alt": anchor_value("ALT"),
        "int": anchor_value("INT"),
        "mapeador": anchor_value("Mapeador"),
        "fase": anchor_value("FASE"),
        "nivel": anchor_value("NIVEL"),
        "sector_geotecnico": anchor_value("Sect. GEOT", "Sector Geotecnico", "Sect GEOT"),
    }


# ---------------------------------------------------------------------------
# Tabla de discontinuidades (desambiguación JC76 / JC89)
# ---------------------------------------------------------------------------


def build_column_names(ws: Worksheet, header_row: int, last_col: int) -> Dict[int, str]:
    """Mapeo columna -> nombre, desambiguando campos repetidos JC1976/JC1989."""
    group_spans: List[Tuple[int, int, str]] = []
    for rr in range(header_row - 1, max(header_row - 4, 0), -1):
        row_has_group = False
        for c in range(1, last_col + 1):
            v = ws.cell(row=rr, column=c).value
            if isinstance(v, str) and ("1976" in v or "1989" in v):
                row_has_group = True
                suf = "JC76" if "1976" in v else "JC89"
                col_ini = c
                col_fin = last_col
                for merge in ws.merged_cells.ranges:
                    if merge.min_row == rr and merge.min_col == c:
                        col_fin = merge.max_col
                        break
                group_spans.append((col_ini, col_fin, suf))
        if row_has_group:
            break
    group_spans.sort()
    for i in range(len(group_spans) - 1):
        col_ini, _, suf = group_spans[i]
        next_ini = group_spans[i + 1][0]
        group_spans[i] = (col_ini, next_ini - 1, suf)

    def group_suffix(c: int) -> Optional[str]:
        for ci, cf, suf in group_spans:
            if ci <= c <= cf:
                return suf
        return None

    # Caso especial: par "Valor de relleno 1/2" duplicado (JC76 luego JC89)
    relleno_pair_cols: Dict[int, str] = {}
    tipo_relleno_1_col = None
    for c in range(1, last_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if text_matches(v, "Tipo de Relleno 1"):
            tipo_relleno_1_col = c
            break
    if tipo_relleno_1_col:
        c2 = tipo_relleno_1_col + 1
        pair76 = (c2 + 1, c2 + 2)
        pair89 = (c2 + 3, c2 + 4)
        relleno_pair_cols = {
            pair76[0]: "JC76", pair76[1]: "JC76",
            pair89[0]: "JC89", pair89[1]: "JC89",
        }

    names: Dict[int, str] = {}
    seen: Dict[str, int] = {}
    for c in range(1, last_col + 1):
        raw = ws.cell(row=header_row, column=c).value
        if raw is None:
            continue
        base = str(raw).strip()

        suf = group_suffix(c)
        if suf is None and c in relleno_pair_cols:
            suf = relleno_pair_cols[c]

        name = f"{base} ({suf})" if suf else base

        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        names[c] = name

    return names


def parse_structures_table(ws: Worksheet, block_top: int, block_bottom: int) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Optional[int]]:
    search_top = max(1, block_top - 3)
    header_row: Optional[int] = None
    for r in range(search_top, block_bottom + 1):
        v = ws.cell(row=r, column=1).value
        if text_matches(v, "ID"):
            header_row = r
            break
    if header_row is None:
        return [], [], None

    last_col = 1
    for c in range(1, 60):
        if ws.cell(row=header_row, column=c).value is not None:
            last_col = c
    names = build_column_names(ws, header_row, last_col)

    first_data_row = header_row + 1
    for merge in ws.merged_cells.ranges:
        if merge.min_row == header_row and merge.min_col == 1:
            first_data_row = max(first_data_row, merge.max_row + 1)

    rows: List[Dict[str, Any]] = []
    r = first_data_row
    while r <= block_bottom:
        a = ws.cell(row=r, column=1).value
        h_label = ws.cell(row=r, column=8).value
        if isinstance(h_label, str) and normalize(h_label).startswith("prom"):
            break
        is_id = isinstance(a, (int, float)) and (not isinstance(a, float) or a == int(a))
        if is_id:
            record = {names[c]: ws.cell(row=r, column=c).value for c in names}
            rows.append(record)
            r += 1
            continue
        if a is None and ws.cell(row=r, column=2).value is None:
            break
        r += 1

    promedios: List[Dict[str, Any]] = []
    rp = first_data_row
    while rp <= block_bottom:
        h_label = ws.cell(row=rp, column=8).value
        if isinstance(h_label, str) and normalize(h_label).startswith("prom"):
            promedios.append({
                "familia": h_label.strip(),
                "espaciamiento_promedio": ws.cell(row=rp, column=10).value,
            })
        rp += 1

    return rows, promedios, header_row


# ---------------------------------------------------------------------------
# Tabla resumen GSI / RMR (JC76 y JC89)
# ---------------------------------------------------------------------------


def parse_rmr_table(ws: Worksheet, block_top: int, block_bottom: int) -> Tuple[Optional[str], List[Dict[str, Any]]]:
    row_range = (max(1, block_top - 3), block_bottom)
    col_range = (1, 160)

    fecha_anchor = find_anchor(ws, row_range, col_range, "Fecha")
    header_anchor = find_anchor(ws, row_range, col_range, "Condicion de Agua")
    if header_anchor is None:
        return None, []

    header_row, start_col = header_anchor

    fecha = None
    if fecha_anchor:
        fecha = value_after_label(ws, *fecha_anchor, max_skip=3)

    last_col = start_col
    c = start_col
    while c < start_col + 25:
        if ws.cell(row=header_row, column=c).value is not None:
            last_col = c
        c += 1

    headers: Dict[int, str] = {}
    for c in range(start_col, last_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None:
            headers[c] = str(v).strip()

    rows: List[Dict[str, Any]] = []
    started = False
    for r in range(header_row + 1, block_bottom + 1):
        metodo = ws.cell(row=r, column=start_col - 1).value
        if metodo in (76, 89):
            record: Dict[str, Any] = {"metodo_JC": metodo, "fecha": fecha}
            for c, name in headers.items():
                record[name] = ws.cell(row=r, column=c).value
            rows.append(record)
            started = True
        elif started:
            break

    return fecha, rows


# ---------------------------------------------------------------------------
# Detección de formato (A = estaciones, B = base de datos)
# ---------------------------------------------------------------------------


def detect_format(ws: Worksheet) -> str:
    """Detecta si la hoja es formato A (estaciones por bloque 'UBICACIÓN')
    o formato B (tabla plana con columna 'CELDA')."""
    for r in range(1, min(ws.max_row, 200) + 1):
        for c in range(1, 40):
            v = ws.cell(row=r, column=c).value
            if text_matches(v, "UBICACION"):
                return "a"
    for r in range(1, min(ws.max_row, 200) + 1):
        for c in range(1, 40):
            v = ws.cell(row=r, column=c).value
            if text_matches(v, "CELDA"):
                return "b"
    return "a"


# ---------------------------------------------------------------------------
# Orquestación
# ---------------------------------------------------------------------------


def parse_excel_a(ws: Worksheet) -> List[Dict[str, Any]]:
    """Devuelve la lista de estaciones crudas del formato A."""
    starts = get_station_starts(ws)
    stations: List[Dict[str, Any]] = []
    for i, st in enumerate(starts):
        block_top = st["header_row"]
        block_bottom = (starts[i + 1]["header_row"] - 1) if i + 1 < len(starts) else ws.max_row

        header = parse_header_block(ws, block_top, block_bottom)
        estructuras, promedios, _ = parse_structures_table(ws, block_top, block_bottom)
        fecha, rmr = parse_rmr_table(ws, block_top, block_bottom)

        stations.append({
            "estacion": st["codigo"],
            "fila_inicio_bloque": block_top,
            "cabecera": header,
            "estructuras": estructuras,
            "promedios_espaciamiento": promedios,
            "fecha_mapeo": fecha,
            "rmr": rmr,
        })
    return stations


# ---------------------------------------------------------------------------
# Normalización a la estructura común del preview del importador
# ---------------------------------------------------------------------------


def _to_float(val: Any) -> Optional[float]:
    if val is None:
        return None
    try:
        f = float(val)
        if f in (-1.0, -1):
            return None
        return f
    except (ValueError, TypeError):
        return None


def _to_int(val: Any) -> Optional[int]:
    """Trunca un valor decimal a entero (mismo comportamiento que
    `sanitize_value(val, int)` usado en el importador Excel B:
    int(float(val)) — NO redondea)."""
    f = _to_float(val)
    if f is None:
        return None
    return int(f)


def _to_str(val: Any) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "-1", "-1.0", "None", "nan", "NaN"):
        return None
    return s


def _norm_tipo_estructura(val: Any) -> str:
    """Normaliza el tipo de estructura: 'J'/'JS' -> 'JN' (igual que el
    resolver del backend), devolviendo '-' si no hay valor."""
    s = _to_str(val) or '-'
    up = s.strip().upper()
    if up in ("J", "JS"):
        return "JN"
    return s


def normalize_station_to_celda(station: Dict[str, Any], infer_lito: callable = None) -> Dict[str, Any]:
    """Convierte una estación cruda del formato A al shape común de
    `celdas_preview` del importador Excel B:
        { codigo, is_duplicate, excel_data, existing_data, estructuras }
    """
    cab = station.get("cabecera") or {}
    ini = cab.get("ubicacion_ini") or {}
    fin = cab.get("ubicacion_fin") or {}

    estacion = _to_str(station.get("estacion")) or ""

    # Fecha y campaña derivada del año
    fecha_raw = station.get("fecha_mapeo")
    fecha_str = None
    if fecha_raw:
        s = str(fecha_raw).strip()
        if len(s) >= 10:
            fecha_str = s[:10]
    campania_val = f"Campaña {fecha_str[:4]}" if fecha_str and fecha_str[:4].isdigit() else "Campaña 2026"

    # Litología: solo LITO-3 viene en el Excel A; se infiere una
    # combinación válida lito1/lito2/unidad a partir del catálogo.
    lito3 = _to_str(cab.get("lito_3")) or ""
    lito1, lito2, unidad = "", "", ""
    if lito3 and infer_lito:
        inferido = infer_lito(lito3)
        if inferido:
            lito1 = inferido.get("lito1") or ""
            lito2 = inferido.get("lito2") or ""
            unidad = inferido.get("grupo") or ""

    mapeador = _to_str(cab.get("mapeador")) or "SRK"
    sector = _to_str(cab.get("sector_geotecnico")) or "PENDIENTE"
    nivel_val = cab.get("nivel")
    fase_val = cab.get("fase")
    alt = _to_str(cab.get("alt")) or ""
    intem = _to_str(cab.get("int")) or ""

    # RMR 76 / 89 (el parser A devuelve 0, 1 o 2 filas con metodo_JC)
    rmr_rows = station.get("rmr") or []
    rmr76 = next((r for r in rmr_rows if r.get("metodo_JC") == 76), None)
    rmr89 = next((r for r in rmr_rows if r.get("metodo_JC") == 89), None)
    # GSI superficie/estructura son campos ÚNICOS de cabecera (no por método);
    # se toman del primer bloque RMR disponible.
    rmr_ref = rmr76 or rmr89

    excel_data: Dict[str, Any] = {
        "codigo": estacion,
        "campania": campania_val,
        "sector": sector,
        "este_ini": _to_float(ini.get("este")) or 0.0,
        "norte_ini": _to_float(ini.get("norte")) or 0.0,
        "cota_ini": _to_float(ini.get("cota")) or 0.0,
        "este_fin": _to_float(fin.get("este")) or 0.0,
        "norte_fin": _to_float(fin.get("norte")) or 0.0,
        "cota_fin": _to_float(fin.get("cota")) or 0.0,
        "largo_m": _to_float(cab.get("largo")),
        "altura_m": _to_float(cab.get("altura")),
        "dip": _to_float(cab.get("dip_hole")),
        "azimut_hole": _to_float(cab.get("az_hole")),
        "dip_talud": _to_float(cab.get("dip_talud")),
        "dipdir_talud": _to_float(cab.get("dipdir_talud")),
        "intemperismo": intem or None,
        "alteracion": alt or None,
        "fase": str(fase_val).strip() if fase_val is not None else None,
        "nivel": str(nivel_val).strip() if nivel_val is not None else None,
        "gsi_superficie": _to_str(rmr_ref.get("Condición de la SUP(GSI)")) if rmr_ref else None,
        "gsi_estructura": _to_str(rmr_ref.get("Estructura(GSI)")) if rmr_ref else None,
        "lito_1": lito1,
        "lito_2": lito2,
        "lito_3": lito3,
        "unidad_litologica": unidad,
        "mapeador": mapeador,
        "fecha": fecha_str or datetime.now().strftime("%Y-%m-%d"),
        "n_discontinuidades": len(station.get("estructuras") or []),
        # RMR 76
        "condicion_agua_rmr76": _to_str(rmr76.get("Condicion de Agua")) if rmr76 else None,
        "dureza_rmr76": _to_str(rmr76.get("Resistencia Estimada")) if rmr76 else None,
        "gsi_visual_rmr76": _to_float(rmr76.get("GSI (Visual)")) if rmr76 else None,
        "control_estructural_rmr76": _to_str(rmr76.get("Control Estructural")) if rmr76 else None,
        "efectos_voladura_rmr76": _to_str(rmr76.get("Efectos de la Voladura")) if rmr76 else None,
        "ucs_mpa": _to_float(rmr76.get("( UCS )  (Mpa)")) if rmr76 else None,
        "is50_mpa": _to_float(rmr76.get("is50 (Mpa)")) if rmr76 else None,
        "rmr_76": _to_float(rmr76.get("RMR")) if rmr76 else None,
        # RMR 89
        "condicion_agua_rmr89": _to_str(rmr89.get("Condicion de Agua")) if rmr89 else None,
        "dureza_rmr89": _to_str(rmr89.get("Resistencia Estimada")) if rmr89 else None,
        "gsi_visual_rmr89": _to_float(rmr89.get("GSI (Visual)")) if rmr89 else None,
        "control_estructural_rmr89": _to_str(rmr89.get("Control Estructural")) if rmr89 else None,
        "efectos_voladura_rmr89": _to_str(rmr89.get("Efectos de la Voladura")) if rmr89 else None,
        "rmr_89": _to_float(rmr89.get("RMR")) if rmr89 else None,
    }

    estructuras: List[Dict[str, Any]] = []
    for idx, e in enumerate(station.get("estructuras") or []):
        familia = _to_int(e.get("ID")) or (math.ceil((idx + 1) / 3.0))
        estructuras.append({
            "numero_estructura": idx + 1,
            "familia_id": familia,
            "tipo_estructura": _norm_tipo_estructura(e.get("Tipo de Estructura")),
            "dip": _to_float(e.get("Dip")) or 0.0,
            "dip_dir": _to_float(e.get("Dip Dir")) or 0.0,
            "distancia_m": _to_float(e.get("Distancia (m)")),
            "abertura_mm": _to_float(e.get("Abertura(mm)")),
            "espesor_mm": _to_float(e.get("Espesor (mm)")),
            "continuidad_m": _to_float(e.get("Continuidad (m)")),
            "espaciamiento_m": _to_float(e.get("Espaciamiento (m)")),
            "n_estructuras": _to_int(e.get("N de Estructuras")),
            "n_extremos_visibles": _to_int(e.get("N de Extremos Visibles")),
            "terminacion": _to_int(e.get("TERMINACION")),
            "relleno_1_codigo": _to_str(e.get("Tipo de Relleno 1")),
            "relleno_2_codigo": _to_str(e.get("Tipo de Relleno 2")),
            "jrc": _to_float(e.get("JRC")),
            "rugosidad_codigo": _to_str(e.get("Rugosidad de Estructura")),
            "forma_estructura": _to_str(e.get("Forma de Estructura")),
            "alteracion_codigo": _to_str(e.get("Alteracion")),
        })

    return {
        "codigo": estacion,
        "is_duplicate": False,  # se completa en el endpoint con la BD
        "excel_data": excel_data,
        "existing_data": None,  # se completa en el endpoint con la BD
        "estructuras": estructuras,
    }
