import os
import io
import math
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse

from app.utils.validator import clean_and_rename_columns, get_row_val, sanitize_value
from app.utils.interpolation import rating_continuo_rqd, rating_continuo_r1
from app.core.catalogs import RESISTENCIA_RATING_CATALOG

router = APIRouter()

# Estilos de Excel Premium
FONT_TITLE = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FILL_HEADER = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Azul oscuro / Slate 800
FILL_GREEN = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")  # Verde suave
FONT_GREEN = Font(name="Calibri", size=10, bold=True, color="137333")
FILL_RED = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")    # Rojo suave
FONT_RED = Font(name="Calibri", size=10, bold=True, color="C5221F")
FONT_DATA = Font(name="Calibri", size=10)
FONT_MONO = Font(name="Consolas", size=10)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

BORDER_THIN = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

def round_half_up(n):
    if n is None:
        return None
    return math.floor(n + 0.5)

def style_row_cell(cell, fill=None, font=FONT_DATA, alignment=ALIGN_LEFT, border=BORDER_THIN):
    if fill:
        cell.fill = fill
    cell.font = font
    cell.alignment = alignment
    cell.border = border

def apply_auto_filter_and_styles(ws):
    # Ajustar ancho de columnas y añadir filtros automáticos
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Cabecera siempre negrita
            if cell.row == 1:
                cell.font = FONT_TITLE
                cell.fill = FILL_HEADER
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_THIN
            else:
                if cell.border is None or cell.border == Border():
                    cell.border = BORDER_THIN
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
    
    if ws.max_row > 1:
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

def get_resolved_campaign(row_dict):
    camp = sanitize_value(get_row_val(row_dict, 'Campaña'), int)
    if camp is not None:
        return camp
    
    # Intenta resolver a partir de nombres alternativos de columna
    for alt in ['Año', 'Ano', 'Campanha', 'CAMPANIA', 'AÑO', 'CAMPAÑA']:
        val = sanitize_value(get_row_val(row_dict, alt), int)
        if val is not None:
            return val
    return None

def evaluate_resistencia_76(row_dict):
    dureza = sanitize_value(get_row_val(row_dict, "DUREZA  '76"), str)
    dureza_val = sanitize_value(get_row_val(row_dict, "RESISTENCIA ESTIMADA VALOR  '76"), float)
    
    if dureza is None and dureza_val is None:
        return None
    
    if dureza is not None:
        code_upper = dureza.upper()
        if code_upper in RESISTENCIA_RATING_CATALOG:
            expected = RESISTENCIA_RATING_CATALOG[code_upper]["r76"]
            if dureza_val is not None:
                dureza_val = round(dureza_val, 2)
                correct = abs(dureza_val - expected) <= 0.2
                msg = "Correcto" if correct else f"Valor ingresado {dureza_val:.1f} no coincide con el discreto de {code_upper} ({expected})"
                return correct, expected, msg
            else:
                return False, expected, "Valor numérico faltante"
        else:
            return False, None, f"Dureza '{dureza}' no es válida en catálogo"
    
    return False, None, "Código de dureza faltante"

def evaluate_rqd_76(row_dict):
    rqd_pct = sanitize_value(get_row_val(row_dict, "RQD  '76"), float)
    rqd_val = sanitize_value(get_row_val(row_dict, "RQD - VALOR  '76"), float)
    
    if rqd_pct is None and rqd_val is None:
        return None
        
    if rqd_pct is not None:
        rqd_int = int(round_half_up(rqd_pct))
        expected = 3 if rqd_int < 25 else (8 if rqd_int < 50 else (13 if rqd_int < 75 else (17 if rqd_int < 90 else 20)))
        if rqd_val is not None:
            rqd_val = round(rqd_val, 2)
            correct = abs(rqd_val - expected) <= 0.2
            msg = "Correcto" if correct else f"Valor ingresado {rqd_val:.1f} no coincide con Bieniawski '76 ({expected}) para {rqd_pct:.1f}%"
            return correct, expected, msg
        else:
            return False, expected, "Rating de RQD faltante"
            
    return False, None, "Porcentaje RQD faltante"

def evaluate_resistencia_89(row_dict, camp):
    dureza = sanitize_value(get_row_val(row_dict, "DUREZA '89"), str)
    dureza_val = sanitize_value(get_row_val(row_dict, "RESISTENCIA ESTIMADA VALOR '89"), float)
    ucs = sanitize_value(get_row_val(row_dict, "( UCS )  (Mpa)"), float)
    
    if dureza is None and dureza_val is None:
        return None
        
    if camp in [2021, 2022, 2023]:
        if ucs is not None:
            expected = rating_continuo_r1(ucs)
            if dureza_val is not None:
                dureza_val = round(dureza_val, 2)
                correct = abs(dureza_val - expected) <= 0.2
                msg = "Correcto" if correct else f"Valor ingresado {dureza_val:.2f} difiere del ábaco continuo ({expected:.2f}) para UCS {ucs:.2f} MPa"
                return correct, round(expected, 2), msg
            else:
                return False, round(expected, 2), "Valor numérico faltante (Ábaco)"
        else:
            # Fallback a discreto si no hay UCS en años de ábaco
            if dureza is not None:
                code_upper = dureza.upper()
                if code_upper in RESISTENCIA_RATING_CATALOG:
                    expected = RESISTENCIA_RATING_CATALOG[code_upper]["r89"]
                    if dureza_val is not None:
                        dureza_val = round(dureza_val, 2)
                        correct = abs(dureza_val - expected) <= 0.2
                        msg = "Correcto (Fallback)" if correct else f"Valor ingresado {dureza_val:.1f} difiere de {code_upper} ({expected})"
                        return correct, expected, msg
            return False, None, "Falta valor de UCS para cálculo de ábaco continuo (Campaña 2021-2023)"
    else:
        # 2024+ o N/A: Discreto
        if dureza is not None:
            code_upper = dureza.upper()
            if code_upper in RESISTENCIA_RATING_CATALOG:
                expected = RESISTENCIA_RATING_CATALOG[code_upper]["r89"]
                if dureza_val is not None:
                    dureza_val = round(dureza_val, 2)
                    correct = abs(dureza_val - expected) <= 0.2
                    msg = "Correcto" if correct else f"Valor ingresado {dureza_val:.1f} difiere de Bieniawski '89 ({expected}) para {code_upper}"
                    return correct, expected, msg
                else:
                    return False, expected, "Valor numérico faltante"
            else:
                return False, None, f"Dureza '{dureza}' no es válida en catálogo"
        return False, None, "Código de dureza faltante"

def evaluate_rqd_89(row_dict, camp):
    rqd_pct = sanitize_value(get_row_val(row_dict, "RQD '89"), float)
    rqd_val = sanitize_value(get_row_val(row_dict, "RQD - VALOR '89"), float)
    
    if rqd_pct is None and rqd_val is None:
        return None
        
    if camp == 2021:
        # Discreto en 2021 con redondeo a entero
        if rqd_pct is not None:
            rqd_int = int(round_half_up(rqd_pct))
            expected = 3 if rqd_int < 25 else (8 if rqd_int < 50 else (13 if rqd_int < 75 else (17 if rqd_int < 90 else 20)))
            if rqd_val is not None:
                rqd_val = round(rqd_val, 2)
                correct = abs(rqd_val - expected) <= 0.2
                msg = "Correcto" if correct else f"Valor ingresado {rqd_val:.1f} difiere de Bieniawski '89 ({expected}) para {rqd_pct:.1f}% (2021)"
                return correct, expected, msg
            else:
                return False, expected, "Rating de RQD faltante"
        return False, None, "Porcentaje RQD faltante"
    else:
        # 2022+ o N/A: Continuo
        if rqd_pct is not None:
            expected = rating_continuo_rqd(rqd_pct)
            if rqd_val is not None:
                rqd_val = round(rqd_val, 2)
                correct = abs(rqd_val - expected) <= 0.2
                msg = "Correcto" if correct else f"Valor ingresado {rqd_val:.2f} difiere del ábaco continuo ({expected:.2f}) para {rqd_pct:.2f}%"
                return correct, round(expected, 2), msg
            else:
                return False, round(expected, 2), "Rating de RQD faltante (Ábaco)"
        return False, None, "Porcentaje RQD faltante"

@router.post("/congruencia/auditar")
async def auditar_congruencia(file: UploadFile = File(...)):
    # 1. Leer el archivo Excel
    contents = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(contents), engine='openpyxl')
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo Excel: {e}")

    # Identificar celdas padre antes del renombrado y limpieza
    # Si la columna CELDA contiene un valor válido, la marcamos como padre
    raw_cols = [str(c).strip().upper() for c in df.columns]
    celda_col_name = None
    for col in df.columns:
        if str(col).strip().upper() in ['CELDA', 'CELDA_PADRE']:
            celda_col_name = col
            break
            
    if not celda_col_name:
        raise HTTPException(status_code=400, detail="No se encontró la columna 'CELDA' o 'CELDA_PADRE' en el archivo Excel.")

    # Guardamos el indicador de fila padre original (donde no es nulo/vacío)
    is_parent_series = df[celda_col_name].notna() & (df[celda_col_name].astype(str).str.strip() != '') & (df[celda_col_name].astype(str).str.strip() != '-1') & (df[celda_col_name].astype(str).str.strip() != '-1.0')

    # Limpiar columnas
    df = clean_and_rename_columns(df)
    
    # Llenar hacia adelante de manera controlada para heredar campaña/taladro
    df['CELDA_PADRE'] = df['CELDA_PADRE'].ffill()
    
    # Si hay columnas de campaña/taladro, hacemos ffill también
    for c in df.columns:
        if c in ['Campaña', 'Año', 'Ano', 'Campanha', 'GEOTECNICO', 'Sector Geotecnico', 'Sector']:
            df[c] = df[c].ffill()

    records = df.to_dict(orient="records")
    
    # Crear Workbook de Excel nuevo para reporte de salida
    wb = openpyxl.Workbook()
    
    # Eliminar hoja por defecto
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Inicializar las 4 hojas requeridas
    ws_res_76 = wb.create_sheet(title="Resistencia 76")
    ws_rqd_76 = wb.create_sheet(title="RQD 76")
    ws_res_89 = wb.create_sheet(title="Resistencia 89")
    ws_rqd_89 = wb.create_sheet(title="RQD 89")
    
    # Escribir Cabeceras
    ws_res_76.append(["Número de Fila", "ID", "Celda", "Campaña", "Dureza 76", "Resistencia Estimada Valor 76", "Correcto (Si/No)", "Valor Esperado"])
    ws_rqd_76.append(["Número de Fila", "ID", "Celda", "Campaña", "RQD 76", "RQD Valor 76", "Correcto (Si/No)", "Valor Esperado"])
    ws_res_89.append(["Número de Fila", "ID", "Celda", "Campaña", "Dureza 89", "Resistencia Estimada Valor 89", "Correcto (Si/No)", "Valor Esperado"])
    ws_rqd_89.append(["Número de Fila", "ID", "Celda", "Campaña", "RQD 89", "UCS (MPa)", "RQD Valor 89", "Correcto (Si/No)", "Valor Esperado"])

    # Iterar sobre las filas padre únicamente
    for idx, row_dict in enumerate(records):
        fila_excel = idx + 2
        if not is_parent_series.iloc[idx]:
            continue
            
        celda = sanitize_value(get_row_val(row_dict, 'CELDA_PADRE'), str)
        id_val = sanitize_value(get_row_val(row_dict, 'ID'), str) or f"F{fila_excel}"
        camp = get_resolved_campaign(row_dict)
        camp_str = str(camp) if camp else "N/A"
        
        # 1. Evaluar Resistencia 76
        res_76_res = evaluate_resistencia_76(row_dict)
        if res_76_res is not None:
            correct, expected, msg = res_76_res
            val_esp = "" if correct else expected
            corr_str = "Si" if correct else "No"
            dureza_76 = sanitize_value(get_row_val(row_dict, "DUREZA  '76"), str) or ""
            dureza_val_76 = sanitize_value(get_row_val(row_dict, "RESISTENCIA ESTIMADA VALOR  '76"), float)
            dureza_val_76_str = f"{dureza_val_76:.1f}" if dureza_val_76 is not None else ""
            
            ws_res_76.append([fila_excel, id_val, celda, camp_str, dureza_76, dureza_val_76_str, corr_str, val_esp])
            # Aplicar estilos a la última fila agregada
            r_idx = ws_res_76.max_row
            fill_cell = FILL_GREEN if correct else FILL_RED
            font_cell = FONT_GREEN if correct else FONT_RED
            
            style_row_cell(ws_res_76.cell(row=r_idx, column=1), font=FONT_MONO, alignment=ALIGN_CENTER)
            style_row_cell(ws_res_76.cell(row=r_idx, column=2), font=FONT_MONO, alignment=ALIGN_CENTER)
            style_row_cell(ws_res_76.cell(row=r_idx, column=3), alignment=ALIGN_LEFT)
            style_row_cell(ws_res_76.cell(row=r_idx, column=4), font=FONT_MONO, alignment=ALIGN_CENTER)
            style_row_cell(ws_res_76.cell(row=r_idx, column=5), alignment=ALIGN_CENTER)
            style_row_cell(ws_res_76.cell(row=r_idx, column=6), font=FONT_MONO, alignment=ALIGN_RIGHT)
            style_row_cell(ws_res_76.cell(row=r_idx, column=7), fill=fill_cell, font=font_cell, alignment=ALIGN_CENTER)
            style_row_cell(ws_res_76.cell(row=r_idx, column=8), font=FONT_MONO, alignment=ALIGN_RIGHT if isinstance(val_esp, (int, float)) else ALIGN_LEFT)
            
        # 2. Evaluar RQD 76
        rqd_76_res = evaluate_rqd_76(row_dict)
        if rqd_76_res is not None:
            correct, expected, msg = rqd_76_res
            val_esp = "" if correct else expected
            corr_str = "Si" if correct else "No"
            rqd_76 = sanitize_value(get_row_val(row_dict, "RQD  '76"), float)
            rqd_76_str = f"{rqd_76:.2f}" if rqd_76 is not None else ""
            rqd_val_76 = sanitize_value(get_row_val(row_dict, "RQD - VALOR  '76"), float)
            rqd_val_76_str = f"{rqd_val_76:.1f}" if rqd_val_76 is not None else ""
            
            ws_rqd_76.append([fila_excel, id_val, celda, camp_str, rqd_76_str, rqd_val_76_str, corr_str, val_esp])
            r_idx = ws_rqd_76.max_row
            fill_cell = FILL_GREEN if correct else FILL_RED
            font_cell = FONT_GREEN if correct else FONT_RED
            
            style_row_cell(ws_rqd_76.cell(row=r_idx, column=1), font=FONT_MONO, alignment=ALIGN_CENTER)
            style_row_cell(ws_rqd_76.cell(row=r_idx, column=2), font=FONT_MONO, alignment=ALIGN_CENTER)
            style_row_cell(ws_rqd_76.cell(row=r_idx, column=3), alignment=ALIGN_LEFT)
            style_row_cell(ws_rqd_76.cell(row=r_idx, column=4), font=FONT_MONO, alignment=ALIGN_CENTER)
            style_row_cell(ws_rqd_76.cell(row=r_idx, column=5), font=FONT_MONO, alignment=ALIGN_RIGHT)
            style_row_cell(ws_rqd_76.cell(row=r_idx, column=6), font=FONT_MONO, alignment=ALIGN_RIGHT)
            style_row_cell(ws_rqd_76.cell(row=r_idx, column=7), fill=fill_cell, font=font_cell, alignment=ALIGN_CENTER)
            style_row_cell(ws_rqd_76.cell(row=r_idx, column=8), font=FONT_MONO, alignment=ALIGN_RIGHT if isinstance(val_esp, (int, float)) else ALIGN_LEFT)

        # 3. Evaluar Resistencia 89
        res_89_res = evaluate_resistencia_89(row_dict, camp)
        if res_89_res is not None:
            correct, expected, msg = res_89_res
            val_esp = "" if correct else expected
            corr_str = "Si" if correct else "No"
            dureza_89 = sanitize_value(get_row_val(row_dict, "DUREZA '89"), str) or ""
            dureza_val_89 = sanitize_value(get_row_val(row_dict, "RESISTENCIA ESTIMADA VALOR '89"), float)
            dureza_val_89_str = f"{dureza_val_89:.2f}" if dureza_val_89 is not None else ""
            
            ws_res_89.append([fila_excel, id_val, celda, camp_str, dureza_89, dureza_val_89_str, corr_str, val_esp])
            r_idx = ws_res_89.max_row
            fill_cell = FILL_GREEN if correct else FILL_RED
            font_cell = FONT_GREEN if correct else FONT_RED
            
            style_row_cell(ws_res_89.cell(row=r_idx, column=1), font=FONT_MONO, alignment=ALIGN_CENTER)
            style_row_cell(ws_res_89.cell(row=r_idx, column=2), font=FONT_MONO, alignment=ALIGN_CENTER)
            style_row_cell(ws_res_89.cell(row=r_idx, column=3), alignment=ALIGN_LEFT)
            style_row_cell(ws_res_89.cell(row=r_idx, column=4), font=FONT_MONO, alignment=ALIGN_CENTER)
            style_row_cell(ws_res_89.cell(row=r_idx, column=5), alignment=ALIGN_CENTER)
            style_row_cell(ws_res_89.cell(row=r_idx, column=6), font=FONT_MONO, alignment=ALIGN_RIGHT)
            style_row_cell(ws_res_89.cell(row=r_idx, column=7), fill=fill_cell, font=font_cell, alignment=ALIGN_CENTER)
            style_row_cell(ws_res_89.cell(row=r_idx, column=8), font=FONT_MONO, alignment=ALIGN_RIGHT if isinstance(val_esp, (int, float)) else ALIGN_LEFT)

        # 4. Evaluar RQD 89
        rqd_89_res = evaluate_rqd_89(row_dict, camp)
        if rqd_89_res is not None:
            correct, expected, msg = rqd_89_res
            val_esp = "" if correct else expected
            corr_str = "Si" if correct else "No"
            rqd_89 = sanitize_value(get_row_val(row_dict, "RQD '89"), float)
            rqd_89_str = f"{rqd_89:.2f}" if rqd_89 is not None else ""
            ucs = sanitize_value(get_row_val(row_dict, "( UCS )  (Mpa)"), float)
            ucs_str = f"{ucs:.2f}" if ucs is not None else ""
            rqd_val_89 = sanitize_value(get_row_val(row_dict, "RQD - VALOR '89"), float)
            rqd_val_89_str = f"{rqd_val_89:.2f}" if rqd_val_89 is not None else ""
            
            ws_rqd_89.append([fila_excel, id_val, celda, camp_str, rqd_89_str, ucs_str, rqd_val_89_str, corr_str, val_esp])
            r_idx = ws_rqd_89.max_row
            fill_cell = FILL_GREEN if correct else FILL_RED
            font_cell = FONT_GREEN if correct else FONT_RED
            
            style_row_cell(ws_rqd_89.cell(row=r_idx, column=1), font=FONT_MONO, alignment=ALIGN_CENTER)
            style_row_cell(ws_rqd_89.cell(row=r_idx, column=2), font=FONT_MONO, alignment=ALIGN_CENTER)
            style_row_cell(ws_rqd_89.cell(row=r_idx, column=3), alignment=ALIGN_LEFT)
            style_row_cell(ws_rqd_89.cell(row=r_idx, column=4), font=FONT_MONO, alignment=ALIGN_CENTER)
            style_row_cell(ws_rqd_89.cell(row=r_idx, column=5), font=FONT_MONO, alignment=ALIGN_RIGHT)
            style_row_cell(ws_rqd_89.cell(row=r_idx, column=6), font=FONT_MONO, alignment=ALIGN_RIGHT)
            style_row_cell(ws_rqd_89.cell(row=r_idx, column=7), font=FONT_MONO, alignment=ALIGN_RIGHT)
            style_row_cell(ws_rqd_89.cell(row=r_idx, column=8), fill=fill_cell, font=font_cell, alignment=ALIGN_CENTER)
            style_row_cell(ws_rqd_89.cell(row=r_idx, column=9), font=FONT_MONO, alignment=ALIGN_RIGHT if isinstance(val_esp, (int, float)) else ALIGN_LEFT)

    # Aplicar filtros y autoajuste de columnas
    apply_auto_filter_and_styles(ws_res_76)
    apply_auto_filter_and_styles(ws_rqd_76)
    apply_auto_filter_and_styles(ws_res_89)
    apply_auto_filter_and_styles(ws_rqd_89)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="Auditoria_Congruencia_Consolidada.xlsx"'
    }
    return StreamingResponse(output, headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def _get_parent_cells_map(contents) -> dict:
    df = pd.read_excel(io.BytesIO(contents), engine='openpyxl')
    
    # Encontrar columna celda
    celda_col_name = None
    for col in df.columns:
        if str(col).strip().upper() in ['CELDA', 'CELDA_PADRE']:
            celda_col_name = col
            break
    if not celda_col_name:
        raise ValueError("No se encontró la columna 'CELDA' o 'CELDA_PADRE' en el archivo Excel.")
        
    is_parent_series = df[celda_col_name].notna() & (df[celda_col_name].astype(str).str.strip() != '') & (df[celda_col_name].astype(str).str.strip() != '-1')
    
    # Normalizar columnas
    df = clean_and_rename_columns(df)
    
    df['CELDA_PADRE'] = df['CELDA_PADRE'].ffill()
    for c in df.columns:
        if c in ['Campaña', 'Año', 'Ano', 'Campanha', 'GEOTECNICO', 'Sector Geotecnico', 'Sector']:
            df[c] = df[c].ffill()
            
    records = df.to_dict(orient="records")
    
    cells_map = {}
    for idx, row_dict in enumerate(records):
        if not is_parent_series.iloc[idx]:
            continue
            
        celda = sanitize_value(get_row_val(row_dict, 'CELDA_PADRE'), str)
        if not celda:
            continue
            
        camp = get_resolved_campaign(row_dict)
        key = (celda.strip().upper(), camp)
        
        cells_map[key] = {
            "fila": idx + 2,
            "id": sanitize_value(get_row_val(row_dict, 'ID'), str) or f"F{idx+2}",
            "celda": celda,
            "campania": camp,
            "dureza_76": sanitize_value(get_row_val(row_dict, "DUREZA  '76"), str) or "",
            "resistencia_76": sanitize_value(get_row_val(row_dict, "RESISTENCIA ESTIMADA VALOR  '76"), float),
            "rqd_76": sanitize_value(get_row_val(row_dict, "RQD  '76"), float),
            "rqd_valor_76": sanitize_value(get_row_val(row_dict, "RQD - VALOR  '76"), float),
            "dureza_89": sanitize_value(get_row_val(row_dict, "DUREZA '89"), str) or "",
            "resistencia_89": sanitize_value(get_row_val(row_dict, "RESISTENCIA ESTIMADA VALOR '89"), float),
            "ucs": sanitize_value(get_row_val(row_dict, "( UCS )  (Mpa)"), float),
            "rqd_89": sanitize_value(get_row_val(row_dict, "RQD '89"), float),
            "rqd_valor_89": sanitize_value(get_row_val(row_dict, "RQD - VALOR '89"), float)
        }
    return cells_map

@router.post("/congruencia/comparar")
async def comparar_congruencia(antes: UploadFile = File(...), despues: UploadFile = File(...)):
    contents_antes = await antes.read()
    contents_despues = await despues.read()
    
    try:
        map_antes = _get_parent_cells_map(contents_antes)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error leyendo archivo ANTES: {e}")
        
    try:
        map_despues = _get_parent_cells_map(contents_despues)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error leyendo archivo DESPUES: {e}")
        
    # Crear Excel de Salida
    wb = openpyxl.Workbook()
    ws_comp = wb.active
    ws_comp.title = "Comparaciones"
    ws_miss = wb.create_sheet(title="Celdas Faltantes")
    
    # Escribir Cabecera Comparaciones
    comp_header = [
        "Celda", "Campaña",
        "Fila Antes", "Fila Después", "Cambio Fila",
        "ID Antes", "ID Después", "Cambio ID",
        "Dureza 76 Antes", "Dureza 76 Después", "Cambio Dureza 76",
        "Resistencia 76 Antes", "Resistencia 76 Después", "Cambio Resistencia 76",
        "RQD 76 Antes", "RQD 76 Después", "Cambio RQD 76",
        "RQD Valor 76 Antes", "RQD Valor 76 Después", "Cambio RQD Valor 76",
        "Dureza 89 Antes", "Dureza 89 Después", "Cambio Dureza 89",
        "Resistencia 89 Antes", "Resistencia 89 Después", "Cambio Resistencia 89",
        "UCS Antes", "UCS Después", "Cambio UCS",
        "RQD 89 Antes", "RQD 89 Después", "Cambio RQD 89",
        "RQD Valor 89 Antes", "RQD Valor 89 Después", "Cambio RQD Valor 89"
    ]
    ws_comp.append(comp_header)
    
    # Cabecera Celdas Faltantes
    ws_miss.append(["Celda", "Campaña", "Fila", "ID", "Archivo Origen"])
    
    all_keys = set(map_antes.keys()) | set(map_despues.keys())
    
    # Ordenar claves por celda y campaña
    sorted_keys = sorted(list(all_keys), key=lambda x: (x[0], x[1] or 0))
    
    def format_val(val):
        if val is None:
            return ""
        if isinstance(val, float):
            return round(val, 2)
        return val

    for key in sorted_keys:
        a_data = map_antes.get(key)
        d_data = map_despues.get(key)
        
        if a_data and d_data:
            # Emparejada
            row_data = [key[0], str(key[1]) if key[1] else "N/A"]
            
            # Fila
            fila_a, fila_d = a_data["fila"], d_data["fila"]
            ch_fila = "Si" if fila_a != fila_d else "No"
            row_data.extend([fila_a, fila_d, ch_fila])
            
            # ID
            id_a, id_d = a_data["id"], d_data["id"]
            ch_id = "Si" if id_a != id_d else "No"
            row_data.extend([id_a, id_d, ch_id])
            
            # Dureza 76
            d76_a, d76_d = a_data["dureza_76"], d_data["dureza_76"]
            ch_d76 = "Si" if d76_a != d76_d else "No"
            row_data.extend([d76_a, d76_d, ch_d76])
            
            # Resistencia 76
            r76_a, r76_d = a_data["resistencia_76"], d_data["resistencia_76"]
            ch_r76 = "Si" if r76_a != r76_d else "No"
            row_data.extend([format_val(r76_a), format_val(r76_d), ch_r76])
            
            # RQD 76
            rq76_a, rq76_d = a_data["rqd_76"], d_data["rqd_76"]
            ch_rq76 = "Si" if rq76_a != rq76_d else "No"
            row_data.extend([format_val(rq76_a), format_val(rq76_d), ch_rq76])
            
            # RQD Valor 76
            rqv76_a, rqv76_d = a_data["rqd_valor_76"], d_data["rqd_valor_76"]
            ch_rqv76 = "Si" if rqv76_a != rqv76_d else "No"
            row_data.extend([format_val(rqv76_a), format_val(rqv76_d), ch_rqv76])
            
            # Dureza 89
            d89_a, d89_d = a_data["dureza_89"], d_data["dureza_89"]
            ch_d89 = "Si" if d89_a != d89_d else "No"
            row_data.extend([d89_a, d89_d, ch_d89])
            
            # Resistencia 89
            r89_a, r89_d = a_data["resistencia_89"], d_data["resistencia_89"]
            ch_r89 = "Si" if r89_a != r89_d else "No"
            row_data.extend([format_val(r89_a), format_val(r89_d), ch_r89])
            
            # UCS
            ucs_a, ucs_d = a_data["ucs"], d_data["ucs"]
            ch_ucs = "Si" if ucs_a != ucs_d else "No"
            row_data.extend([format_val(ucs_a), format_val(ucs_d), ch_ucs])
            
            # RQD 89
            rq89_a, rq89_d = a_data["rqd_89"], d_data["rqd_89"]
            ch_rq89 = "Si" if rq89_a != rq89_d else "No"
            row_data.extend([format_val(rq89_a), format_val(rq89_d), ch_rq89])
            
            # RQD Valor 89
            rqv89_a, rqv89_d = a_data["rqd_valor_89"], d_data["rqd_valor_89"]
            ch_rqv89 = "Si" if rqv89_a != rqv89_d else "No"
            row_data.extend([format_val(rqv89_a), format_val(rqv89_d), ch_rqv89])
            
            ws_comp.append(row_data)
            
            # Estilizar las celdas de cambio
            r_idx = ws_comp.max_row
            
            # Formatos de datos base
            style_row_cell(ws_comp.cell(row=r_idx, column=1), alignment=ALIGN_LEFT)
            style_row_cell(ws_comp.cell(row=r_idx, column=2), font=FONT_MONO, alignment=ALIGN_CENTER)
            
            # Aplicar estilo Si/No a cada columna de cambio
            change_cols = [5, 8, 11, 14, 17, 20, 23, 26, 29, 32, 35]
            for col_i in range(3, ws_comp.max_column + 1):
                cell = ws_comp.cell(row=r_idx, column=col_i)
                
                # Alinear según tipo de datos
                if col_i in change_cols:
                    is_change = cell.value == "Si"
                    cell.fill = FILL_RED if is_change else FILL_GREEN
                    cell.font = FONT_RED if is_change else FONT_GREEN
                    cell.alignment = ALIGN_CENTER
                else:
                    if isinstance(cell.value, (int, float)):
                        cell.font = FONT_MONO
                        cell.alignment = ALIGN_RIGHT
                    elif col_i in [3, 4, 6, 7]:
                        cell.font = FONT_MONO
                        cell.alignment = ALIGN_CENTER
                    else:
                        cell.alignment = ALIGN_LEFT
                cell.border = BORDER_THIN
                
        else:
            # Celdas faltantes (desparejadas)
            orig = "Antes" if a_data else "Después"
            dat = a_data if a_data else d_data
            ws_miss.append([
                dat["celda"],
                str(dat["campania"]) if dat["campania"] else "N/A",
                dat["fila"],
                dat["id"],
                orig
            ])
            r_idx = ws_miss.max_row
            fill_cell = FILL_RED if orig == "Antes" else FILL_GREEN
            font_cell = FONT_RED if orig == "Antes" else FONT_GREEN
            
            style_row_cell(ws_miss.cell(row=r_idx, column=1), alignment=ALIGN_LEFT)
            style_row_cell(ws_miss.cell(row=r_idx, column=2), font=FONT_MONO, alignment=ALIGN_CENTER)
            style_row_cell(ws_miss.cell(row=r_idx, column=3), font=FONT_MONO, alignment=ALIGN_CENTER)
            style_row_cell(ws_miss.cell(row=r_idx, column=4), font=FONT_MONO, alignment=ALIGN_CENTER)
            style_row_cell(ws_miss.cell(row=r_idx, column=5), fill=fill_cell, font=font_cell, alignment=ALIGN_CENTER)

    apply_auto_filter_and_styles(ws_comp)
    apply_auto_filter_and_styles(ws_miss)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="Comparacion_Celdas_Padre.xlsx"'
    }
    return StreamingResponse(output, headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
