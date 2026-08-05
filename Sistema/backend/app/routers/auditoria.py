import os
import io
import json
import shutil
import math
import openpyxl
import time
import traceback
from datetime import datetime
from collections import Counter, defaultdict
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks
from fastapi.responses import StreamingResponse, JSONResponse, FileResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.utils.validator import validate_bulk_excel, AuditCancelledError
from app.core.catalogs import MANDATORY_COLS_COUNT

router = APIRouter()
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
uploads_dir = os.path.join(BASE_DIR, "uploads")

# --- SISTEMA CENTRALIZADO (SSOT) DE REGLAS DE VALIDACIÓN GEOMECÁNICA ---
from app.core.rules import RULES_REGISTRY, CATEGORIES_REGISTRY
from app.core.audit_helpers import (
    safe_int,
    safe_float,
    get_safe_sheet_name,
    safe_replace,
    get_incidence_category_name,
    aggregate_audit_metrics,
)
from app.core.excel_styles import get_styles, write_kpi_card

def generar_excel_reporte_core(diag: dict, compact: dict, filtered: list):
    s = get_styles()
    font_title = s["font_title"]
    font_subtitle = s["font_subtitle"]
    font_section = s["font_section"]
    font_header = s["font_header"]
    font_bold = s["font_bold"]
    font_regular = s["font_regular"]
    font_kpi_lbl = s["font_kpi_lbl"]
    
    font_kpi_val_blue = s["font_kpi_blue"]
    font_kpi_val_green = s["font_kpi_green"]
    font_kpi_val_red = s["font_kpi_red"]
    font_kpi_val_orange = s["font_kpi_orange"]

    fill_primary = s["fill_primary"]
    fill_accent_green = s["fill_green"]
    fill_accent_yellow = s["fill_yellow"]
    fill_accent_orange = s["fill_orange"]
    fill_accent_red = s["fill_red"]
    fill_zebra = s["fill_zebra"]
    fill_kpi_gray = s["fill_kpi_gray"]

    border_thin = s["border_thin"]
    border_kpi = s["border_kpi"]

    alignment_center = s["align_center"]
    alignment_left = s["align_left"]
    alignment_right = s["align_right"]

    wb = openpyxl.Workbook()
    
    def write_kpi_card_opt(ws, start_row, start_col, label, value, bg_fill, val_font):
        write_kpi_card(ws, start_row, start_col, label, value, bg_fill, val_font, s)

    # --- HOJA 1: DASHBOARD EJECUTIVO ---
    ws_dash = wb.active
    ws_dash.title = "📊 Dashboard Ejecutivo"
    ws_dash.views.sheetView[0].showGridLines = True
    
    ws_dash.cell(row=2, column=2, value="SISTEMA DE AUDITORÍA GEOTÉCNICA").font = font_title
    ws_dash.cell(row=3, column=2, value="Dashboard de Control de Calidad y Consistencia de Mapeo Geomecánico").font = font_subtitle
    
    total_filas = compact.get("familia1", {}).get("total_discontinuidades", 0)
    total_fields = compact.get("familia2", {}).get("total_fields", 0)
    total_vacios = sum(1 for i in filtered if i.get("tipo_incidencia") == "VACIO")
    total_advertencias = sum(1 for i in filtered if i.get("tipo_incidencia") == "ADVERTENCIA")
    total_alertas = sum(1 for i in filtered if i.get("tipo_incidencia") == "ALERTA")
    total_correctos = total_fields - (total_vacios + total_advertencias + total_alertas)
    pct_integridad = (total_correctos / max(1, total_fields)) * 100

    write_kpi_card_opt(ws_dash, 5, 2, "ESTACIONES EVALUADAS", len(compact.get("resumen_por_celda_padre", {})), fill_kpi_gray, font_kpi_val_blue)
    write_kpi_card_opt(ws_dash, 5, 4, "ESTRUCTURAS REGISTRADAS", total_filas, fill_kpi_gray, font_kpi_val_blue)
    write_kpi_card_opt(ws_dash, 5, 6, "INTEGRIDAD GLOBAL", f"{pct_integridad:.2f}%", fill_accent_green, font_kpi_val_green)
    write_kpi_card_opt(ws_dash, 5, 8, "ALERTAS CRÍTICAS", total_alertas, fill_accent_red, font_kpi_val_red)
    write_kpi_card_opt(ws_dash, 5, 10, "ADVERTENCIAS", total_advertencias, fill_accent_orange, font_kpi_val_orange)

    # Tabla: Distribución por Campaña
    ws_dash.cell(row=9, column=2, value="DESEMPEÑO DE CONTROL POR CAMPAÑA").font = font_section
    headers_camp = ["Campaña", "Estructuras", "Celdas Afectadas", "Estructuras Afectadas", "Alertas (N)", "% Alertas", "Vacíos (N)", "% Vacíos"]
    for idx, col in enumerate(headers_camp, start=2):
        cell = ws_dash.cell(row=10, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    r_camp = 11
    for row in compact.get("distribucion_campania", []):
        ws_dash.cell(row=r_camp, column=2, value=row.get("campania")).font = font_bold
        ws_dash.cell(row=r_camp, column=2).alignment = alignment_center
        
        ws_dash.cell(row=r_camp, column=3, value=safe_int(row.get("discontinuidades"))).number_format = '#,##0'
        ws_dash.cell(row=r_camp, column=3).alignment = alignment_right
        
        ws_dash.cell(row=r_camp, column=4, value=safe_int(row.get("celdas_afectadas"))).number_format = '#,##0'
        ws_dash.cell(row=r_camp, column=4).alignment = alignment_right
        
        ws_dash.cell(row=r_camp, column=5, value=safe_int(row.get("estructuras_afectadas"))).number_format = '#,##0'
        ws_dash.cell(row=r_camp, column=5).alignment = alignment_right
        
        ws_dash.cell(row=r_camp, column=6, value=safe_int(row.get("alertas_cant"))).number_format = '#,##0'
        ws_dash.cell(row=r_camp, column=6).alignment = alignment_right
        
        ws_dash.cell(row=r_camp, column=7, value=safe_float(row.get("alertas_pct")) / 100.0).number_format = '0.00%'
        ws_dash.cell(row=r_camp, column=7).alignment = alignment_right
        
        ws_dash.cell(row=r_camp, column=8, value=safe_int(row.get("vacios_cant"))).number_format = '#,##0'
        ws_dash.cell(row=r_camp, column=8).alignment = alignment_right
        
        ws_dash.cell(row=r_camp, column=9, value=safe_float(row.get("vacios_pct")) / 100.0).number_format = '0.00%'
        ws_dash.cell(row=r_camp, column=9).alignment = alignment_right
        
        for col_idx in range(2, 10):
            ws_dash.cell(row=r_camp, column=col_idx).border = border_thin
            if r_camp % 2 == 0:
                ws_dash.cell(row=r_camp, column=col_idx).fill = fill_zebra
        r_camp += 1

    # Tabla: Distribución por Sectores
    r_sect = r_camp + 2
    ws_dash.cell(row=r_sect, column=2, value="DISTRIBUCIÓN POR SECTOR GEOTÉCNICO").font = font_section
    
    r_sect += 1
    headers_sect = ["Sector", "Estructuras", "Alertas (N)", "% Alertas", "Vacíos (N)", "% Vacíos"]
    for idx, col in enumerate(headers_sect, start=2):
        cell = ws_dash.cell(row=r_sect, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    for row in compact.get("distribucion_sector", []):
        r_sect += 1
        ws_dash.cell(row=r_sect, column=2, value=row.get("sector")).font = font_bold
        ws_dash.cell(row=r_sect, column=2).alignment = alignment_center
        
        ws_dash.cell(row=r_sect, column=3, value=safe_int(row.get("discontinuidades"))).number_format = '#,##0'
        ws_dash.cell(row=r_sect, column=3).alignment = alignment_right
        
        ws_dash.cell(row=r_sect, column=4, value=safe_int(row.get("alertas_cant"))).number_format = '#,##0'
        ws_dash.cell(row=r_sect, column=4).alignment = alignment_right
        
        ws_dash.cell(row=r_sect, column=5, value=safe_float(row.get("alertas_pct")) / 100.0).number_format = '0.00%'
        ws_dash.cell(row=r_sect, column=5).alignment = alignment_right
        
        ws_dash.cell(row=r_sect, column=6, value=safe_int(row.get("vacios_cant"))).number_format = '#,##0'
        ws_dash.cell(row=r_sect, column=6).alignment = alignment_right
        
        ws_dash.cell(row=r_sect, column=7, value=safe_float(row.get("vacios_pct")) / 100.0).number_format = '0.00%'
        ws_dash.cell(row=r_sect, column=7).alignment = alignment_right
        
        for col_idx in range(2, 8):
            ws_dash.cell(row=r_sect, column=col_idx).border = border_thin
            if r_sect % 2 == 0:
                ws_dash.cell(row=r_sect, column=col_idx).fill = fill_zebra
        r_sect += 1

    # Tabla: Distribución de Celdas más Afectadas
    r_worst = r_sect + 2
    ws_dash.cell(row=r_worst-1, column=2, value="TOP 5 ESTACIONES CON MAYOR INCIDENCIA").font = font_section
    headers_worst = ["Estación (Celda)", "Total Estructuras", "Vacíos", "Advertencias", "Alertas", "Calificación"]
    for idx, col in enumerate(headers_worst, start=2):
        cell = ws_dash.cell(row=r_worst, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    for row in compact.get("worst_cells", [])[:5]:
        r_worst += 1
        ws_dash.cell(row=r_worst, column=2, value=row.get("celda")).font = font_bold
        ws_dash.cell(row=r_worst, column=2).alignment = alignment_center
        
        ws_dash.cell(row=r_worst, column=3, value=safe_int(row.get("total_hijas"))).number_format = '#,##0'
        ws_dash.cell(row=r_worst, column=3).alignment = alignment_right
        
        ws_dash.cell(row=r_worst, column=4, value=safe_int(row.get("vacios"))).number_format = '#,##0'
        ws_dash.cell(row=r_worst, column=4).alignment = alignment_right
        
        ws_dash.cell(row=r_worst, column=5, value=safe_int(row.get("advertencias"))).number_format = '#,##0'
        ws_dash.cell(row=r_worst, column=5).alignment = alignment_right
        
        ws_dash.cell(row=r_worst, column=6, value=safe_int(row.get("alertas"))).number_format = '#,##0'
        ws_dash.cell(row=r_worst, column=6).alignment = alignment_right
        
        status = row.get("estado_celda", "OK")
        status_cell = ws_dash.cell(row=r_worst, column=7, value=status)
        status_cell.font = font_bold
        status_cell.alignment = alignment_center
        if status == "ALERTA": status_cell.fill = fill_accent_red
        elif status == "ADVERTENCIA": status_cell.fill = fill_accent_orange
        else: status_cell.fill = fill_accent_green
        
        for col_idx in range(2, 8):
            ws_dash.cell(row=r_worst, column=col_idx).border = border_thin
            if r_worst % 2 == 0:
                ws_dash.cell(row=r_worst, column=col_idx).fill = fill_zebra

    # Tabla para Gráfica Directa: Top 5 Alertas Críticas
    ws_dash.cell(row=9, column=9, value="PRINCIPALES DESVIACIONES CRÍTICAS").font = font_section
    headers_graph = ["Error Geotécnico", "Frecuencia"]
    for idx, col in enumerate(headers_graph, start=9):
        cell = ws_dash.cell(row=10, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    top_errs_list = Counter(get_incidence_category_name(i) for i in filtered if i.get("tipo_incidencia") == "ALERTA").most_common(5)
    r_graph = 11
    for msg, qty in top_errs_list:
        ws_dash.cell(row=r_graph, column=9, value=msg).font = font_regular
        ws_dash.cell(row=r_graph, column=9).border = border_thin
        
        c_qty = ws_dash.cell(row=r_graph, column=10, value=qty)
        c_qty.font = font_bold
        c_qty.alignment = alignment_right
        c_qty.number_format = '#,##0'
        c_qty.border = border_thin
        c_qty.fill = fill_accent_red
        
        r_graph += 1
        
    for dummy in range(r_graph, 16):
        ws_dash.cell(row=dummy, column=9, value="—").font = font_regular
        ws_dash.cell(row=dummy, column=9).border = border_thin
        ws_dash.cell(row=dummy, column=10, value=0).font = font_regular
        ws_dash.cell(row=dummy, column=10).border = border_thin
        ws_dash.cell(row=dummy, column=10).number_format = '#,##0'

    # Gráfica Nativa de Excel
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Frecuencia de Desviaciones Críticas Detectadas"
    chart.y_axis.title = "Cantidad de Ocurrencias"
    chart.x_axis.title = "Regla de Consistencia"
    
    chart_data = Reference(ws_dash, min_col=10, min_row=10, max_row=15)
    chart_cats = Reference(ws_dash, min_col=9, min_row=11, max_row=15)
    chart.add_data(chart_data, titles_from_data=True)
    chart.set_categories(chart_cats)
    chart.legend = None
    chart.width = 15
    chart.height = 11
    ws_dash.add_chart(chart, "I17")

    # --- HOJA 2: REGISTRO MAESTRO DE ERRORES (CATÁLOGO / ÍNDICE) ---
    ws_cat = wb.create_sheet(title="❌ Catálogo de Errores")
    ws_cat.views.sheetView[0].showGridLines = True
    
    ws_cat.cell(row=2, column=2, value="REGISTRO MAESTRO DE REGLAS DE CONSISTENCIA").font = font_title
    ws_cat.cell(row=3, column=2, value="Catálogo completo de validación geomecánica ordenado por frecuencia. Use los hipervínculos para navegar.").font = font_subtitle
    
    # Recopilar los años únicos de campaña presentes en todas las incidencias
    all_campaigns = sorted(set(
        str(inc.get("campania", "N/A")) for inc in filtered
        if inc.get("campania") and str(inc.get("campania")) not in ["N/A", "None", ""]
    ))

    headers_cat = ["ID", "Gravedad", "Regla de Consistencia Evaluada", "Total (N)"] + [f"Año {c}" for c in all_campaigns] + ["Enlace Directo"]
    for idx, col in enumerate(headers_cat, start=2):
        cell = ws_cat.cell(row=5, column=idx, value=col)
        cell.font = font_header
        cell.fill = fill_primary
        cell.alignment = alignment_center
        cell.border = border_thin

    incidencias_por_error = defaultdict(list)
    for inc in filtered:
        msg_simplificado = get_incidence_category_name(inc)
        incidencias_por_error[msg_simplificado].append(inc)

    catalog_frequencies = []
    for cat in CATEGORIES_REGISTRY.values():
        rule_msg = cat.name
        matches = incidencias_por_error[rule_msg]
        # Calcular conteos por campaña
        camp_counts = {}
        for c in all_campaigns:
            camp_counts[c] = sum(1 for m in matches if str(m.get("campania", "N/A")) == c)
        catalog_frequencies.append({
            "msg": rule_msg,
            "severity": cat.severity,
            "matches": matches,
            "count": len(matches),
            "camp_counts": camp_counts
        })
        
    catalog_frequencies = sorted(catalog_frequencies, key=lambda x: x["count"], reverse=True)

    r_cat = 6
    active_sheets_mapping = {}
    
    for c_idx, rule in enumerate(catalog_frequencies, start=1):
        ws_cat.cell(row=r_cat, column=2, value=c_idx).font = font_regular
        ws_cat.cell(row=r_cat, column=2).alignment = alignment_center
        ws_cat.cell(row=r_cat, column=2).border = border_thin
        
        c_sev = ws_cat.cell(row=r_cat, column=3, value=rule["severity"])
        c_sev.font = font_bold
        c_sev.alignment = alignment_center
        c_sev.border = border_thin
        if rule["severity"] == "ALERTA": c_sev.fill = fill_accent_red
        elif rule["severity"] == "ADVERTENCIA": c_sev.fill = fill_accent_orange
        else: c_sev.fill = fill_accent_yellow
        
        ws_cat.cell(row=r_cat, column=4, value=rule["msg"]).font = font_bold if rule["count"] > 0 else font_regular
        ws_cat.cell(row=r_cat, column=4).border = border_thin
        
        c_count = ws_cat.cell(row=r_cat, column=5, value=rule["count"])
        c_count.font = font_bold
        c_count.alignment = alignment_right
        c_count.number_format = '#,##0'
        c_count.border = border_thin

        # Columnas por año de campaña
        for y_offset, camp_key in enumerate(all_campaigns):
            c_yr = ws_cat.cell(row=r_cat, column=6 + y_offset, value=rule["camp_counts"].get(camp_key, 0))
            c_yr.font = font_regular
            c_yr.alignment = alignment_right
            c_yr.number_format = '#,##0'
            c_yr.border = border_thin
            if rule["camp_counts"].get(camp_key, 0) == 0:
                c_yr.font = Font(name="Segoe UI", size=9, color="AAAAAA")

        link_col = 6 + len(all_campaigns)
        c_link = ws_cat.cell(row=r_cat, column=link_col)
        if rule["count"] > 0:
            tab_name = get_safe_sheet_name(rule["msg"], c_idx)
            active_sheets_mapping[rule["msg"]] = {"tab_name": tab_name, "records": rule["matches"]}
            
            c_link.value = f'=HYPERLINK("#{chr(39)}{tab_name}{chr(39)}!B2", "🔍 Navegar a Registros")'
            c_link.font = Font(name="Segoe UI", size=10, bold=True, color="1B365D", underline="single")
            c_link.alignment = alignment_center
        else:
            c_link.value = "Limpio / 0 Incidencias"
            c_link.font = Font(name="Segoe UI", size=9, italic=True, color="7F8C8D")
            c_link.alignment = alignment_center
            c_link.fill = fill_accent_green
            
        c_link.border = border_thin
        r_cat += 1

    # AutoFilter para permitir filtrar por año directamente en Excel
    last_col_letter = chr(ord('A') + (5 + len(all_campaigns)))  # B=col2, C=col3, ...
    ws_cat.auto_filter.ref = f"B5:{last_col_letter}{r_cat - 1}"

    # --- HOJA 3: DETALLE PLANO COMPLETO DE INCIDENCIAS ---
    chunk_size = 1000000
    detail_chunks = [filtered[i:i + chunk_size] for i in range(0, len(filtered), chunk_size)]
    if not detail_chunks:
        detail_chunks = [[]]

    for chunk_idx, chunk_data in enumerate(detail_chunks):
        title = "📋 Detalle de Incidencias"
        if len(detail_chunks) > 1:
            title = f"📋 Detalle Incidencias ({chunk_idx + 1})"

        ws_detail = wb.create_sheet(title=title)
        ws_detail.views.sheetView[0].showGridLines = True

        ws_detail.cell(row=2, column=2, value="DETALLE COMPLETO DE INCIDENCIAS").font = font_title
        ws_detail.cell(row=3, column=2, value="Listado plano consolidado de todas las desviaciones y vacíos detectados.").font = font_subtitle

        headers_detail = [
            "Fila Excel", "Gravedad", "Estación Padre", "Estructura Hija", "Campaña", 
            "Logger Geotécnico", "Sector Geotécnico", "Tipo de Mapeo", "Columna de Falla", 
            "Valor Actual", "Mensaje de Inconsistencia Geomecánica"
        ]

        ws_detail.append([]) 
        ws_detail.append([None] + headers_detail) 
        grid_heading_row = ws_detail.max_row

        for idx in range(2, 13):
            cell = ws_detail.cell(row=grid_heading_row, column=idx)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = alignment_center
            cell.border = border_thin

        start_detail_row = ws_detail.max_row + 1
        for inc_item in chunk_data:
            row_data = [
                None,
                safe_int(inc_item.get("fila_excel")),
                inc_item.get("tipo_incidencia", "ALERTA"),
                inc_item.get("celda_padre"),
                inc_item.get("celda_hija"),
                inc_item.get("campania"),
                inc_item.get("geotecnico"),
                inc_item.get("sector_geotecnico"),
                inc_item.get("tipo_mapeo", "Mapeo de Celdas"),
                inc_item.get("columna"),
                inc_item.get("valor_actual") if inc_item.get("valor_actual") is not None else "—",
                inc_item.get("mensaje")
            ]
            ws_detail.append(row_data)

        end_detail_row = ws_detail.max_row

        for r_idx in range(start_detail_row, end_detail_row + 1):
            if r_idx <= start_detail_row + 150:
                ws_detail.cell(row=r_idx, column=2).alignment = alignment_center
                ws_detail.cell(row=r_idx, column=3).alignment = alignment_center
                ws_detail.cell(row=r_idx, column=4).alignment = alignment_center
                ws_detail.cell(row=r_idx, column=5).alignment = alignment_center
                ws_detail.cell(row=r_idx, column=6).alignment = alignment_center
                ws_detail.cell(row=r_idx, column=7).alignment = alignment_left
                ws_detail.cell(row=r_idx, column=8).alignment = alignment_center
                ws_detail.cell(row=r_idx, column=9).alignment = alignment_center
                ws_detail.cell(row=r_idx, column=10).alignment = alignment_left
                ws_detail.cell(row=r_idx, column=11).alignment = alignment_center
                ws_detail.cell(row=r_idx, column=12).alignment = alignment_left

                if r_idx % 2 == 0:
                    for col_idx in range(2, 13):
                        if col_idx != 3:
                            ws_detail.cell(row=r_idx, column=col_idx).fill = fill_zebra
            else:
                ws_detail.cell(row=r_idx, column=2).alignment = alignment_center
                ws_detail.cell(row=r_idx, column=3).alignment = alignment_center

            cell_sev = ws_detail.cell(row=r_idx, column=3)
            sev = cell_sev.value
            if sev == "ALERTA": cell_sev.fill = fill_accent_red
            elif sev == "ADVERTENCIA": cell_sev.fill = fill_accent_orange
            else: cell_sev.fill = fill_accent_yellow
            cell_sev.font = font_bold

            for col_idx in range(2, 13):
                ws_detail.cell(row=r_idx, column=col_idx).border = border_thin

        ws_detail.auto_filter.ref = f"B{grid_heading_row}:L{end_detail_row}"

    # --- HOJAS 4+: DETALLES INDIVIDUALES POR REGLA DE ERROR ---
    for orig_msg, mapping_data in active_sheets_mapping.items():
        sh_name = mapping_data["tab_name"]
        err_records = mapping_data["records"]
        
        ws_err = wb.create_sheet(title=sh_name)
        ws_err.views.sheetView[0].showGridLines = True
        
        c_back = ws_err.cell(row=2, column=2)
        c_back.value = '=HYPERLINK("#\'❌ Catálogo de Errores\'!B2", "⬅ Volver al Registro Maestro de Errores")'
        c_back.font = Font(name="Segoe UI", size=10, bold=True, color="1B365D", underline="single")
        c_back.alignment = alignment_left
        
        ws_err.cell(row=4, column=2, value="ANÁLISIS DE INCIDENCIA EN BASE DE DATOS").font = font_section
        cell_err_desc = ws_err.cell(row=5, column=2, value=f"Regla: {orig_msg.upper()}")
        cell_err_desc.font = Font(name="Segoe UI", size=11, bold=True, color="7F1D1D")
        cell_err_desc.fill = fill_accent_red
        cell_err_desc.border = border_thin
        ws_err.merge_cells(start_row=5, start_column=2, end_row=5, end_column=7)
        
        st_affected = len(set(x.get("celda_padre", "N/A") for x in err_records))
        tot_affected = len(err_records)
        
        write_kpi_card_opt(ws_err, 7, 2, "ESTACIONES AFECTADAS (CELDA PADRE)", st_affected, fill_kpi_gray, font_kpi_val_blue)
        write_kpi_card_opt(ws_err, 7, 4, "ESTRUCTURAS AFECTADAS (CANTIDAD)", tot_affected, fill_kpi_gray, font_kpi_val_blue)
        
        # Distribución por Campaña
        ws_err.cell(row=10, column=2, value="DISTRIBUCIÓN POR CAMPAÑA").font = font_section
        for idx, col in enumerate(["Campaña / Año", "Ocurrencias (N)", "% Contribución"], start=2):
            cell = ws_err.cell(row=11, column=idx, value=col)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = alignment_center
            cell.border = border_thin
            
        r_dist_yr = defaultdict(int)
        for r in err_records:
            r_dist_yr[str(r.get("campania", "N/A"))] += 1
            
        curr_y_r = 12
        for yr, y_qty in sorted(r_dist_yr.items()):
            ws_err.cell(row=curr_y_r, column=2, value=yr).font = font_bold
            ws_err.cell(row=curr_y_r, column=2).alignment = alignment_center
            ws_err.cell(row=curr_y_r, column=2).border = border_thin
            
            c_yq = ws_err.cell(row=curr_y_r, column=3, value=y_qty)
            c_yq.font = font_regular
            c_yq.alignment = alignment_right
            c_yq.number_format = '#,##0'
            c_yq.border = border_thin
            
            c_yp = ws_err.cell(row=curr_y_r, column=4, value=y_qty / max(1, tot_affected))
            c_yp.font = font_regular
            c_yp.alignment = alignment_right
            c_yp.number_format = '0.00%'
            c_yp.border = border_thin
            
            curr_y_r += 1

        unique_years = sorted(list(set(str(r.get("campania", "N/A")) for r in err_records)))
        rule_group = defaultdict(list)
        for r in err_records:
            rule_group[r.get("rule_code", "Desconocido")].append(r)
            
        rule_stats = []
        for rule_code, recs in rule_group.items():
            yr_counts = defaultdict(int)
            for r in recs:
                yr_counts[str(r.get("campania", "N/A"))] += 1
            rule_stats.append({
                "rule_code": rule_code,
                "total": len(recs),
                "yr_counts": yr_counts
            })
        rule_stats.sort(key=lambda x: x["total"], reverse=True)
        
        # 2. Agrupar por Mensaje Único (Tabla B)
        msg_group = defaultdict(list)
        for r in err_records:
            msg_group[r.get("mensaje", "Desconocido")].append(r)
            
        msg_stats = []
        for msg_val, recs in msg_group.items():
            yr_counts = defaultdict(int)
            for r in recs:
                yr_counts[str(r.get("campania", "N/A"))] += 1
            msg_stats.append({
                "message": msg_val,
                "total": len(recs),
                "yr_counts": yr_counts
            })
        msg_stats.sort(key=lambda x: x["total"], reverse=True)

        # Precalcular coordenadas de las secciones para los hipervínculos de navegación
        dist_table_height = len(r_dist_yr)
        jump_link_row = 12 + dist_table_height + 1
        
        table_a_start = jump_link_row + 2
        table_a_height = 2 + len(rule_stats)
        
        max_allowed_err = 1000000
        truncated_err = len(err_records) > max_allowed_err
        err_records_to_write = err_records[:max_allowed_err]

        indiv_start = table_a_start + table_a_height + 2
        indiv_height = 2 + len(err_records_to_write) + (1 if truncated_err else 0)
        
        table_b_start = indiv_start + indiv_height + 2

        # --- ENLACE DIRECTO DE NAVEGACIÓN RÁPIDA ---
        c_jump = ws_err.cell(row=jump_link_row, column=2)
        c_jump.value = f'=HYPERLINK("#\'{sh_name}\'!B{table_b_start}", "🔍 Ir a Métricas de Mensajes de Inconsistencia Únicos (al final de la hoja)")'
        c_jump.font = Font(name="Segoe UI", size=10, bold=True, color="1B365D", underline="single")
        c_jump.alignment = alignment_left

        # --- TABLA A: RESUMEN POR REGLA ESPECÍFICA ---
        ws_err.cell(row=table_a_start, column=2, value="MÉTRICAS POR REGLA ESPECÍFICA (CÓDIGO)").font = font_section
        
        ws_err.cell(row=table_a_start + 1, column=2, value="Código de Regla").font = font_header
        ws_err.cell(row=table_a_start + 1, column=2).fill = fill_primary
        ws_err.cell(row=table_a_start + 1, column=2).alignment = alignment_center
        
        ws_err.merge_cells(start_row=table_a_start + 1, start_column=2, end_row=table_a_start + 1, end_column=4)
        for c_idx in [3, 4]:
            ws_err.cell(row=table_a_start + 1, column=c_idx).fill = fill_primary
            
        headers_a = ["Ocurrencias Totales"] + [f"Año {y}" if y != "N/A" else "N/A" for y in unique_years]
        for col_offset, h_name in enumerate(headers_a, start=5):
            cell = ws_err.cell(row=table_a_start + 1, column=col_offset, value=h_name)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = alignment_center
            
        for col_idx in range(2, 5 + len(unique_years) + 1):
            ws_err.cell(row=table_a_start + 1, column=col_idx).border = border_thin
            
        for idx, stat in enumerate(rule_stats):
            r_row = table_a_start + 2 + idx
            ws_err.cell(row=r_row, column=2, value=stat["rule_code"]).font = font_bold
            ws_err.cell(row=r_row, column=2).alignment = alignment_center
            ws_err.merge_cells(start_row=r_row, start_column=2, end_row=r_row, end_column=4)
            
            c_tot = ws_err.cell(row=r_row, column=5, value=stat["total"])
            c_tot.font = font_bold
            c_tot.alignment = alignment_right
            c_tot.number_format = '#,##0'
            
            for y_idx, yr in enumerate(unique_years):
                val = stat["yr_counts"].get(yr, 0)
                c_val = ws_err.cell(row=r_row, column=6 + y_idx, value=val)
                c_val.font = font_regular
                c_val.alignment = alignment_right
                c_val.number_format = '#,##0'
                
            for col_idx in range(2, 5 + len(unique_years) + 1):
                cell = ws_err.cell(row=r_row, column=col_idx)
                cell.border = border_thin
                if r_row % 2 == 0:
                    cell.fill = fill_zebra

        # --- LISTADO DETALLADO INDIVIDUAL DE INCIDENCIAS (CON COLUMNA 4 SWAP) ---
        ws_err.cell(row=indiv_start, column=2, value="REGISTROS INDIVIDUALES AFECTADOS (LISTADO COMPLETO)").font = font_section
        
        headers_inc = [
            "Fila Excel", "Estación Padre", "Estructura Hija", "Tipo de Mapeo", "Campaña", 
            "Logger Geotécnico", "Sector Geotécnico", "Columna de Falla", 
            "Valor Actual", "Mensaje de Inconsistencia Geomecánica"
        ]
        header_row_idx = indiv_start + 1
        for col_idx, h_name in enumerate(headers_inc, start=2):
            cell = ws_err.cell(row=header_row_idx, column=col_idx, value=h_name)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = alignment_center
            cell.border = border_thin
            
        start_data_row = indiv_start + 2
        for idx, inc_item in enumerate(err_records_to_write):
            curr_row = start_data_row + idx
            ws_err.cell(row=curr_row, column=2, value=safe_int(inc_item.get("fila_excel")))
            ws_err.cell(row=curr_row, column=3, value=inc_item.get("celda_padre"))
            ws_err.cell(row=curr_row, column=4, value=inc_item.get("celda_hija"))
            ws_err.cell(row=curr_row, column=5, value=inc_item.get("tipo_mapeo", "Mapeo de Celdas"))
            ws_err.cell(row=curr_row, column=6, value=inc_item.get("campania"))
            ws_err.cell(row=curr_row, column=7, value=inc_item.get("geotecnico"))
            ws_err.cell(row=curr_row, column=8, value=inc_item.get("sector_geotecnico"))
            ws_err.cell(row=curr_row, column=9, value=inc_item.get("columna"))
            ws_err.cell(row=curr_row, column=10, value=inc_item.get("valor_actual") if inc_item.get("valor_actual") is not None else "—")
            ws_err.cell(row=curr_row, column=11, value=inc_item.get("mensaje"))
            
        if truncated_err:
            curr_row = start_data_row + len(err_records_to_write)
            ws_err.cell(row=curr_row, column=3, value="--- REPORTE TRUNCADO: SE SUPERÓ EL LÍMITE DE FILAS DE EXCEL (1,048,576) ---")

        end_data_row = start_data_row + len(err_records_to_write) - 1 + (1 if truncated_err else 0)
        
        for r_idx in range(start_data_row, end_data_row + 1):
            if r_idx <= start_data_row + 150:
                ws_err.cell(row=r_idx, column=2).alignment = alignment_center
                ws_err.cell(row=r_idx, column=3).alignment = alignment_center
                ws_err.cell(row=r_idx, column=4).alignment = alignment_center
                ws_err.cell(row=r_idx, column=5).alignment = alignment_center
                ws_err.cell(row=r_idx, column=6).alignment = alignment_center
                ws_err.cell(row=r_idx, column=7).alignment = alignment_left
                ws_err.cell(row=r_idx, column=8).alignment = alignment_center
                ws_err.cell(row=r_idx, column=9).alignment = alignment_left
                ws_err.cell(row=r_idx, column=10).alignment = alignment_center
                ws_err.cell(row=r_idx, column=11).alignment = alignment_left
                
                if r_idx % 2 == 0:
                    for col_idx in range(2, 12):
                        ws_err.cell(row=r_idx, column=col_idx).fill = fill_zebra
            else:
                ws_err.cell(row=r_idx, column=2).alignment = alignment_center
                ws_err.cell(row=r_idx, column=3).alignment = alignment_center
                
            for col_idx in range(2, 12):
                ws_err.cell(row=r_idx, column=col_idx).border = border_thin
                
        ws_err.auto_filter.ref = f"B{header_row_idx}:K{end_data_row}"

        # --- TABLA B: RESUMEN POR MENSAJE DE INCONSISTENCIA ÚNICO ---
        ws_err.cell(row=table_b_start, column=2, value="MÉTRICAS POR MENSAJE DE INCONSISTENCIA GEOMECÁNICA ÚNICO").font = font_section
        
        c_ret = ws_err.cell(row=table_b_start, column=5)
        c_ret.value = f'=HYPERLINK("#\'{sh_name}\'!B2", "⬅ Volver al Inicio de la Hoja")'
        c_ret.font = Font(name="Segoe UI", size=10, bold=True, color="1B365D", underline="single")
        c_ret.alignment = alignment_left

        ws_err.cell(row=table_b_start + 1, column=2, value="Mensaje de Inconsistencia Geomecánica Único").font = font_header
        ws_err.cell(row=table_b_start + 1, column=2).fill = fill_primary
        ws_err.cell(row=table_b_start + 1, column=2).alignment = alignment_center
        
        ws_err.merge_cells(start_row=table_b_start + 1, start_column=2, end_row=table_b_start + 1, end_column=7)
        for c_idx in range(3, 8):
            ws_err.cell(row=table_b_start + 1, column=c_idx).fill = fill_primary
            
        headers_b = ["Ocurrencias Totales"] + [f"Año {y}" if y != "N/A" else "N/A" for y in unique_years]
        for col_offset, h_name in enumerate(headers_b, start=8):
            cell = ws_err.cell(row=table_b_start + 1, column=col_offset, value=h_name)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = alignment_center
            
        for col_idx in range(2, 8 + len(unique_years) + 1):
            ws_err.cell(row=table_b_start + 1, column=col_idx).border = border_thin
            
        for idx, stat in enumerate(msg_stats):
            r_row = table_b_start + 2 + idx
            ws_err.cell(row=r_row, column=2, value=stat["message"]).font = font_regular
            ws_err.cell(row=r_row, column=2).alignment = alignment_left
            ws_err.merge_cells(start_row=r_row, start_column=2, end_row=r_row, end_column=7)
            
            c_tot = ws_err.cell(row=r_row, column=8, value=stat["total"])
            c_tot.font = font_bold
            c_tot.alignment = alignment_right
            c_tot.number_format = '#,##0'
            
            for y_idx, yr in enumerate(unique_years):
                val = stat["yr_counts"].get(yr, 0)
                c_val = ws_err.cell(row=r_row, column=9 + y_idx, value=val)
                c_val.font = font_regular
                c_val.alignment = alignment_right
                c_val.number_format = '#,##0'
                
            for col_idx in range(2, 8 + len(unique_years) + 1):
                cell = ws_err.cell(row=r_row, column=col_idx)
                cell.border = border_thin
                if r_row % 2 == 0:
                    cell.fill = fill_zebra

    # --- AUTO-AJUSTE DINÁMICO DE COLUMNAS ---
    for ws in wb.worksheets:
        ws.column_dimensions['A'].width = 3
        if ws.title not in ["❌ Catálogo de Errores", "📋 Detalle de Incidencias", "📊 Panel de Control"]:
            # Hojas de error con anchos específicos optimizados
            ws.column_dimensions['B'].width = 11  # Fila Excel
            ws.column_dimensions['C'].width = 14  # Estación Padre
            ws.column_dimensions['D'].width = 14  # Estructura Hija
            ws.column_dimensions['E'].width = 20  # Tipo de Mapeo
            ws.column_dimensions['F'].width = 9   # Campaña
            ws.column_dimensions['G'].width = 16  # Logger Geotécnico
            ws.column_dimensions['H'].width = 16  # Sector Geotécnico
            ws.column_dimensions['I'].width = 24  # Columna de Falla
            ws.column_dimensions['J'].width = 12  # Valor Actual
            ws.column_dimensions['K'].width = 60  # Mensaje de Inconsistencia
        else:
            for col_idx in range(2, ws.max_column + 1):
                vals = []
                for row_idx in range(1, min(15, ws.max_row + 1)):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val is not None:
                        val_str = str(val)
                        if val_str.startswith("=HYPERLINK"):
                            vals.append("Ver Registros")
                        else:
                            vals.append(val_str)
                if not vals: continue
                max_len = max(len(v) for v in vals)
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 11), 52)

    return wb

# --- Helpers de cancelación cooperativa y lock de concurrencia ---
# El pipeline valida en loop con pandas; para poder abortarlo desde la UI se usa un
# flag de cancelación por audit_id que el validator revisa periódicamente, y un lock
# global que impide lanzar dos auditorías pesadas simultáneas en la misma instancia.

def _cancel_flag_path(audit_id: str) -> str:
    return os.path.join(uploads_dir, "history", f"{audit_id}.cancel")

def _lock_file_path() -> str:
    return os.path.join(uploads_dir, "history", "processing.lock")

def _is_cancelled(audit_id: str) -> bool:
    return os.path.exists(_cancel_flag_path(audit_id))

def _cleanup_audit_files(audit_id: str) -> None:
    history_dir = os.path.join(uploads_dir, "history")
    for suffix in [".xlsx", "_diagnostico.json", "_compact.json", "_reporte_completo.xlsx", ".cancel"]:
        try:
            os.remove(os.path.join(history_dir, f"{audit_id}{suffix}"))
        except OSError:
            pass

def _acquire_lock(audit_id: str) -> bool:
    lock_path = _lock_file_path()
    if os.path.exists(lock_path):
        try:
            with open(lock_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            age = time.time() - data.get("started_at", 0)
            if age < 3 * 3600:  # 3 horas: cubre validaciones de archivos muy grandes
                return False
        except Exception:
            pass
        try:
            os.remove(lock_path)
        except OSError:
            pass
    with open(lock_path, "w", encoding="utf-8") as f:
        json.dump({"audit_id": audit_id, "started_at": time.time()}, f)
    return True

def _release_lock(audit_id: str) -> None:
    lock_path = _lock_file_path()
    try:
        if os.path.exists(lock_path):
            with open(lock_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if data.get("audit_id") == audit_id:
                os.remove(lock_path)
    except Exception:
        pass

def run_bulk_pipeline_with_id(file_path: str, audit_id: str, original_filename: str = None):
    t_start = time.time()
    print(f"\n======================================================================")
    print(f"[*] [AUDITORÍA {audit_id}] INICIANDO PIPELINE DE PROCESAMIENTO ASÍNCRONO")
    print(f"[*] Archivo cargado: {os.path.basename(file_path)}")
    print(f"======================================================================")
    
    history_dir = os.path.join(uploads_dir, "history")
    os.makedirs(history_dir, exist_ok=True)
    raw_json_out = os.path.join(history_dir, f"{audit_id}_diagnostico.json")
    compact_json_out = os.path.join(history_dir, f"{audit_id}_compact.json")
    excel_pregenerated_out = os.path.join(history_dir, f"{audit_id}_reporte_completo.xlsx")
    
    try:
        # Paso 1: Ejecutar la validación masiva en validator.py
        print(f"[*] [AUDITORÍA {audit_id}] Paso 1/5: Ejecutando motor de validación QA/QC geomecánica...")
        validate_bulk_excel(file_path, raw_json_out, cancel_flag_path=_cancel_flag_path(audit_id))
        if _is_cancelled(audit_id):
            _cleanup_audit_files(audit_id)
            print(f"[x] [AUDITORÍA {audit_id}] Cancelada por el usuario tras la validación.")
            return

        # Paso 2: Copiar al diagnóstico público estático
        print(f"[*] [AUDITORÍA {audit_id}] Paso 2/5: Publicando diagnóstico en caché pública...")
        if _is_cancelled(audit_id):
            _cleanup_audit_files(audit_id)
            print(f"[x] [AUDITORÍA {audit_id}] Cancelada por el usuario antes de publicar el diagnóstico.")
            return
        shutil.copyfile(raw_json_out, os.path.join(uploads_dir, "diagnostico_geomecanico.json"))

        # Paso 3: Cargar resultados y compactar métricas para la UI mediante la función unificada
        print(f"[*] [AUDITORÍA {audit_id}] Paso 3/5: Compilando métricas compactas para Dashboard...")
        if _is_cancelled(audit_id):
            _cleanup_audit_files(audit_id)
            print(f"[x] [AUDITORÍA {audit_id}] Cancelada por el usuario antes de compilar métricas.")
            return
        with open(raw_json_out, "r", encoding="utf-8") as f:
            diag = json.load(f)

        diag["nombre_archivo"] = original_filename or os.path.basename(file_path)
        compact = aggregate_audit_metrics(diag)
        compact["audit_id"] = audit_id

        print(f"[*] [AUDITORÍA {audit_id}] Paso 4/5: Escribiendo JSON de resumen ligero...")
        if _is_cancelled(audit_id):
            _cleanup_audit_files(audit_id)
            print(f"[x] [AUDITORÍA {audit_id}] Cancelada por el usuario antes de publicar el resumen.")
            return
        compact_json_tmp = compact_json_out + ".tmp"
        with open(compact_json_tmp, "w", encoding="utf-8") as f:
            json.dump(compact, f, ensure_ascii=False)
        safe_replace(compact_json_tmp, compact_json_out)
        
        public_compact = os.path.join(uploads_dir, "resumen_geomecanico_ligero.json")
        public_compact_tmp = public_compact + ".tmp"
        shutil.copyfile(compact_json_out, public_compact_tmp)
        safe_replace(public_compact_tmp, public_compact)

        # Paso 5: Generar archivo Excel cacheado
        # IMPORTANTE: Extraemos las incidencias del dict principal y vaciamos la referencia en diag
        # para que el GC pueda liberar el dict grande mientras openpyxl construye el workbook,
        # evitando el MemoryError por tener dos copias gigantes en RAM simultáneamente.
        print(f"[*] [AUDITORÍA {audit_id}] Paso 5/5: Pre-generando reporte de Excel (.xlsx) en segundo plano...")
        if _is_cancelled(audit_id):
            _cleanup_audit_files(audit_id)
            print(f"[x] [AUDITORÍA {audit_id}] Cancelada por el usuario antes de generar el reporte Excel.")
            return
        incidencias_list = diag.pop("incidencias", [])
        wb = generar_excel_reporte_core(diag, compact, incidencias_list)
        del incidencias_list  # liberar RAM antes de guardar el .xlsx
        
        excel_tmp = excel_pregenerated_out + ".tmp"
        wb.save(excel_tmp)
        safe_replace(excel_tmp, excel_pregenerated_out)
        
        public_excel = os.path.join(uploads_dir, "reporte_completo_ultimo.xlsx")
        public_excel_tmp = public_excel + ".tmp"
        shutil.copyfile(excel_pregenerated_out, public_excel_tmp)
        safe_replace(public_excel_tmp, public_excel)
        
        elapsed = time.time() - t_start
        print(f"======================================================================")
        print(f"[+] [AUDITORÍA {audit_id}] PIPELINE COMPLETADO EXITOSAMENTE")
        print(f"[*] Tiempo total: {elapsed:.2f} segundos")
        print(f"======================================================================\n")

    except AuditCancelledError as e:
        print(f"\n[x] [AUDITORÍA {audit_id}] AUDITORÍA CANCELADA POR EL USUARIO: {e}")
        _cleanup_audit_files(audit_id)
        print(f"[x] [AUDITORÍA {audit_id}] Archivos parciales eliminados.")

    except Exception as e:
        elapsed = time.time() - t_start
        print(f"\n======================================================================")
        print(f"[-] [AUDITORÍA {audit_id}] PIPELINE ABORTADO POR ERROR CRÍTICO")
        print(f"[-] Detalle del error: {str(e)}")
        print(f"[-] Tiempo ejecutado antes del fallo: {elapsed:.2f} segundos")
        print(f"======================================================================")
        traceback.print_exc()
        
        error_data = {
            "audit_id": audit_id,
            "status": "error",
            "fecha_auditoria": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "nombre_archivo": os.path.basename(file_path),
            "error_message": str(e),
            "familia1": {"total_discontinuidades": 0},
            "familia2": {"total_vacios": 0, "total_advertencias": 0, "total_alertas": 0}
        }
        try:
            with open(compact_json_out, "w", encoding="utf-8") as f:
                json.dump(error_data, f, ensure_ascii=False)
            shutil.copyfile(compact_json_out, os.path.join(uploads_dir, "resumen_geomecanico_ligero.json"))
        except Exception as write_err:
            print(f"[-] Error al guardar JSON de contingencia para error de auditoría: {write_err}")
    finally:
        _release_lock(audit_id)

@router.post("/geomecanica/cancelar")
def cancelar_auditoria(audit_id: str = None):
    """
    Cancela cooperativamente una auditoría en curso: crea un flag que el pipeline
    revisa periódicamente (en el loop de validación y entre pasos) para abortar
    y limpiar los archivos parciales.
    """
    if not audit_id:
        raise HTTPException(status_code=400, detail="Parámetro 'audit_id' requerido.")
    with open(_cancel_flag_path(audit_id), "w", encoding="utf-8") as f:
        f.write("1")
    _release_lock(audit_id)
    return {"status": "cancelado", "audit_id": audit_id}

@router.post("/geomecanica/importar-excel-bulk")
async def importar_excel_bulk(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Formato no soportado.")
    audit_id = f"audit_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    history_dir = os.path.join(uploads_dir, "history")
    os.makedirs(history_dir, exist_ok=True)

    # Guard de concurrencia: una sola auditoría pesada a la vez (evita OOM por
    # dos pipelines simultáneos de archivos grandes en instancias pequeñas).
    if not _acquire_lock(audit_id):
        raise HTTPException(
            status_code=409,
            detail="Ya hay una auditoría en proceso en el servidor. Espera a que termine o cancélala antes de subir otro archivo."
        )

    file_path = os.path.join(history_dir, f"{audit_id}.xlsx")
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    # Validar de forma síncrona si es un reporte generado por el sistema
    try:
        import pandas as pd
        xls = pd.ExcelFile(file_path, engine='openpyxl')
        sheet_names = xls.sheet_names
        is_report = any("DASHBOARD" in s.upper() or "ERRORES" in s.upper() or "INCIDENCIAS" in s.upper() for s in sheet_names)
        if is_report:
            try:
                os.remove(file_path)
            except:
                pass
            raise HTTPException(
                status_code=400,
                detail="El archivo cargado es un reporte de auditoría generado por el sistema. Por favor, cargue la base de datos geomecánica original con sus celdas de mapeo."
            )
    except HTTPException:
        _release_lock(audit_id)
        raise
    except Exception as e:
        try:
            os.remove(file_path)
        except:
            pass
        _release_lock(audit_id)
        raise HTTPException(
            status_code=400,
            detail=f"No se pudo leer el archivo Excel. Verifique que no esté corrupto o posea un formato inválido. Detalle: {str(e)}"
        )

    background_tasks.add_task(run_bulk_pipeline_with_id, file_path, audit_id, file.filename)
    return {"status": "procesando", "audit_id": audit_id, "filename": file.filename}

@router.get("/geomecanica/auditorias")
def listar_auditorias():
    history_dir = os.path.join(uploads_dir, "history")
    if not os.path.exists(history_dir): return []
    audits = []
    for f in os.listdir(history_dir):
        if f.endswith("_compact.json"):
            audit_id = f.replace("_compact.json", "")
            compact_file = os.path.join(history_dir, f)
            try:
                with open(compact_file, "r", encoding="utf-8") as file_content:
                    meta = json.load(file_content)
                    audits.append({
                        "audit_id": audit_id, "fecha": meta.get("fecha_auditoria", "Desconocida"),
                        "archivo": meta.get("nombre_archivo", "Desconocido.xlsx"),
                        "total_filas": meta.get("familia1", {}).get("total_discontinuidades", 0),
                        "total_vacios": meta.get("familia2", {}).get("total_vacios", 0),
                        "total_advertencias": meta.get("familia2", {}).get("total_advertencias", 0),
                        "total_alertas": meta.get("familia2", {}).get("total_alertas", 0)
                    })
            except: pass
    return sorted(audits, key=lambda x: x["fecha"], reverse=True)

@router.get("/geomecanica/status")
def obtener_status_auditoria(audit_id: str = None):
    """
    Estado canónico y liviano (~200 bytes) de una auditoría para el polling del frontend.
    Reemplaza el uso de /resumen-ligero como chequeo de estado: no carga JSONs grandes.

    Respuestas:
      {"status": "procesando"}   -> el pipeline aún corre (existe el .xlsx subido o el diagnóstico)
      {"status": "listo", "reporte_listo": bool, "nombre_archivo": ..., "fecha_auditoria": ...}
      {"status": "error", "detail": ...}  -> fallo persistido por el pipeline en el compact
      404                             -> el audit_id ya no existe (archivos perdidos por reinicio/redeploy)
    """
    if not audit_id:
        raise HTTPException(status_code=400, detail="Parámetro 'audit_id' requerido.")

    history_dir = os.path.join(uploads_dir, "history")
    excel_file = os.path.join(history_dir, f"{audit_id}.xlsx")
    raw_file = os.path.join(history_dir, f"{audit_id}_diagnostico.json")
    compact_file = os.path.join(history_dir, f"{audit_id}_compact.json")
    reporte_file = os.path.join(history_dir, f"{audit_id}_reporte_completo.xlsx")

    if os.path.exists(compact_file):
        try:
            with open(compact_file, "r", encoding="utf-8") as f:
                meta = json.load(f)
        except Exception:
            meta = {}
        if meta.get("status") == "error":
            return {"status": "error", "detail": meta.get("error_message", "El procesamiento falló en el servidor.")}
        return {
            "status": "listo",
            "reporte_listo": os.path.exists(reporte_file),
            "nombre_archivo": meta.get("nombre_archivo"),
            "fecha_auditoria": meta.get("fecha_auditoria"),
        }

    if os.path.exists(excel_file) or os.path.exists(raw_file):
        return {"status": "procesando"}

    raise HTTPException(
        status_code=404,
        detail="La auditoría solicitada ya no existe en el servidor (sus archivos se perdieron, probablemente por un reinicio o redeploy).",
    )

@router.get("/geomecanica/resumen-ligero")
def obtener_resumen_ligero(audit_id: str = None, years: str = None):
    # 1. Resolver rutas de archivos correspondientes
    if audit_id:
        raw_file = os.path.join(uploads_dir, "history", f"{audit_id}_diagnostico.json")
        compact_file = os.path.join(uploads_dir, "history", f"{audit_id}_compact.json")
        excel_file = os.path.join(uploads_dir, "history", f"{audit_id}.xlsx")
        
        # Check first if the compact_file has error status
        if os.path.exists(compact_file):
            try:
                with open(compact_file, "r", encoding="utf-8") as f:
                    meta = json.load(f)
                if meta.get("status") == "error":
                    raise HTTPException(status_code=400, detail=meta.get("error_message", "Error de procesamiento en segundo plano."))
            except HTTPException:
                raise
            except Exception:
                pass

        if not os.path.exists(compact_file) or not os.path.exists(raw_file):
            if os.path.exists(excel_file):
                return JSONResponse(
                    status_code=202, 
                    content={"status": "procesando", "message": "Procesando base de datos masiva y cruzando variables..."}
                )
            raise HTTPException(status_code=404, detail="La auditoría solicitada no existe en el servidor.")
    else:
        raw_file = os.path.join(uploads_dir, "diagnostico_geomecanico.json")
        compact_file = os.path.join(uploads_dir, "resumen_geomecanico_ligero.json")
        if not os.path.exists(raw_file) or not os.path.exists(compact_file):
            history_dir = os.path.join(uploads_dir, "history")
            if os.path.exists(history_dir):
                jsons = [f for f in os.listdir(history_dir) if f.endswith("_diagnostico.json") and not f.endswith(".tmp")]
                if jsons:
                    jsons.sort(key=lambda x: os.path.getmtime(os.path.join(history_dir, x)), reverse=True)
                    latest_id = jsons[0].replace("_diagnostico.json", "")
                    raw_file = os.path.join(history_dir, f"{latest_id}_diagnostico.json")
                    compact_file = os.path.join(history_dir, f"{latest_id}_compact.json")
            
            # Check default compact_file for error
            if os.path.exists(compact_file):
                try:
                    with open(compact_file, "r", encoding="utf-8") as f:
                        meta = json.load(f)
                    if meta.get("status") == "error":
                        raise HTTPException(status_code=400, detail=meta.get("error_message", "Error en segundo plano."))
                except HTTPException:
                    raise
                except Exception:
                    pass

            if not os.path.exists(raw_file) or not os.path.exists(compact_file):
                return JSONResponse(
                    status_code=202, 
                    content={"status": "procesando", "message": "Esperando inicialización de datos de auditoría..."}
                )

    # 2. FAST PATH: Cargar el reporte compacto directamente si no hay filtros aplicados
    if (not years or years == "TODOS" or years == "") and os.path.exists(compact_file):
        try:
            with open(compact_file, "r", encoding="utf-8") as f:
                meta = json.load(f)
            if meta.get("status") == "error":
                raise HTTPException(status_code=400, detail=meta.get("error_message", "Error en segundo plano."))
            return meta
        except HTTPException:
            raise
        except Exception:
            pass 

    # 3. BAJO DEMANDA / FILTRADO: Procesar dinámicamente si se especificaron años o falla la caché
    with open(raw_file, "r", encoding="utf-8") as f:
        diag = json.load(f)
        
    diag["nombre_archivo"] = os.path.basename(raw_file)
    compact = aggregate_audit_metrics(diag, years_filter=years)
    compact["audit_id"] = audit_id or "default"
    
    return compact

@router.get("/geomecanica/incidencias-paginadas")
def obtener_incidencias_paginadas(
    page: int = 1, limit: int = 50, tipo: str = None, celda: str = None, columna: str = None,
    campania: str = None, geotecnico: str = None, sector_geotecnico: str = None, search: str = None, audit_id: str = None
):
    if audit_id: 
        raw_file = os.path.join(uploads_dir, "history", f"{audit_id}_diagnostico.json")
        if not os.path.exists(raw_file):
            return {"data": [], "page": 1, "total_pages": 1, "total_records": 0}
    else: 
        raw_file = os.path.join(uploads_dir, "diagnostico_geomecanico.json")
        
    if not os.path.exists(raw_file): 
        return {"data": [], "page": 1, "total_pages": 1, "total_records": 0}
        
    with open(raw_file, "r", encoding="utf-8") as f:
        diag_data = json.load(f)
        
    incidencias = diag_data.get("incidencias", [])
    filtered = incidencias
    if tipo: filtered = [i for i in filtered if i.get("tipo_incidencia") == tipo.upper()]
    if celda:
        celda_up = celda.upper()
        filtered = [i for i in filtered if i.get("celda_padre") == celda_up or i.get("celda_hija") == celda_up]
    if columna: filtered = [i for i in filtered if i.get("columna", "").upper() == columna.upper()]
    if campania:
        years_list = [y.strip() for y in campania.split(",") if y.strip()]
        if len(years_list) > 0 and "TODOS" not in years_list:
            filtered = [i for i in filtered if str(i.get("campania")) in years_list]
    if geotecnico:
        geo_up = geotecnico.upper()
        filtered = [i for i in filtered if i.get("geotecnico", "").upper() == geo_up]
    if sector_geotecnico:
        sect_up = sector_geotecnico.upper()
        filtered = [i for i in filtered if i.get("sector_geotecnico", "").upper() == sect_up]
    
    if search:
        search_lower = search.lower()
        filtered = [
            i for i in filtered 
            if search_lower in i.get("mensaje", "").lower() 
            or search_lower in get_incidence_category_name(i).lower()
            or search_lower in i.get("columna", "").lower()
            or search_lower in i.get("celda_padre", "").lower()
        ]
        
    total_records = len(filtered)
    start_idx = (page - 1) * limit
    return {"page": page, "limit": limit, "total_records": total_records, "total_pages": math.ceil(total_records / limit), "data": filtered[start_idx:start_idx+limit]}

@router.get("/geomecanica/reporte-excel")
def descargar_reporte_excel(
    tipo: str = None, celda: str = None, columna: str = None, campania: str = None,
    geotecnico: str = None, sector_geotecnico: str = None, search: str = None, audit_id: str = None
):
    filtered_vals = []
    for val in [tipo, celda, columna, campania, geotecnico, sector_geotecnico, search]:
        if val is not None:
            v_str = str(val).strip().upper()
            if v_str not in ["", "NONE", "NULL", "UNDEFINED", "TODOS"]:
                filtered_vals.append(val)
                
    is_filtered = len(filtered_vals) > 0
    filename = f"reporte_auditoria_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    if audit_id:
        pregenerated_file = os.path.join(uploads_dir, "history", f"{audit_id}_reporte_completo.xlsx")
    else:
        pregenerated_file = os.path.join(uploads_dir, "reporte_completo_ultimo.xlsx")

    if not is_filtered and os.path.exists(pregenerated_file):
        return FileResponse(
            pregenerated_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename
        )

    if audit_id:
        raw_file = os.path.join(uploads_dir, "history", f"{audit_id}_diagnostico.json")
        compact_file = os.path.join(uploads_dir, "history", f"{audit_id}_compact.json")
    else:
        raw_file = os.path.join(uploads_dir, "diagnostico_geomecanico.json")
        compact_file = os.path.join(uploads_dir, "resumen_geomecanico_ligero.json")
        if not os.path.exists(raw_file) or not os.path.exists(compact_file):
            history_dir = os.path.join(uploads_dir, "history")
            if os.path.exists(history_dir):
                jsons = [f for f in os.listdir(history_dir) if f.endswith("_diagnostico.json") and not f.endswith(".tmp")]
                if jsons:
                    jsons.sort(key=lambda x: os.path.getmtime(os.path.join(history_dir, x)), reverse=True)
                    latest_id = jsons[0].replace("_diagnostico.json", "")
                    raw_file = os.path.join(history_dir, f"{latest_id}_diagnostico.json")
                    compact_file = os.path.join(history_dir, f"{latest_id}_compact.json")

    if not os.path.exists(raw_file) or not os.path.exists(compact_file):
        raise HTTPException(status_code=404, detail="El diagnóstico solicitado no ha sido generado o está incompleto.")

    with open(raw_file, "r", encoding="utf-8") as f:
        diag = json.load(f)
    with open(compact_file, "r", encoding="utf-8") as f:
        compact = json.load(f)
        
    incidencias = diag.get("incidencias", [])
    
    filtered = incidencias
    if tipo: filtered = [i for i in filtered if i.get("tipo_incidencia") == tipo.upper()]
    if celda:
        celda_up = celda.upper()
        filtered = [i for i in filtered if i.get("celda_padre") == celda_up or i.get("celda_hija") == celda_up]
    if columna: filtered = [i for i in filtered if i.get("columna", "").upper() == columna.upper()]
    if campania:
        years_list = [y.strip() for y in campania.split(",") if y.strip()]
        if len(years_list) > 0 and "TODOS" not in years_list:
            filtered = [i for i in filtered if str(i.get("campania")) in years_list]
    if geotecnico:
        geo_up = geotecnico.upper()
        filtered = [i for i in filtered if i.get("geotecnico", "").upper() == geo_up]
    if sector_geotecnico:
        sect_up = sector_geotecnico.upper()
        filtered = [i for i in filtered if i.get("sector_geotecnico", "").upper() == sect_up]
    
    if search:
        search_lower = search.lower()
        filtered = [
            i for i in filtered 
            if search_lower in i.get("mensaje", "").lower() 
            or search_lower in get_incidence_category_name(i).lower()
            or search_lower in i.get("columna", "").lower()
            or search_lower in i.get("celda_padre", "").lower()
        ]

    wb = generar_excel_reporte_core(diag, compact, filtered)
    
    if not is_filtered:
        os.makedirs(os.path.dirname(pregenerated_file), exist_ok=True)
        excel_tmp = pregenerated_file + ".tmp"
        wb.save(excel_tmp)
        safe_replace(excel_tmp, pregenerated_file)
        
        if not audit_id:
            public_excel = os.path.join(uploads_dir, "reporte_completo_ultimo.xlsx")
            public_excel_tmp = public_excel + ".tmp"
            shutil.copyfile(pregenerated_file, public_excel_tmp)
            safe_replace(public_excel_tmp, public_excel)
            
        return FileResponse(
            pregenerated_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )