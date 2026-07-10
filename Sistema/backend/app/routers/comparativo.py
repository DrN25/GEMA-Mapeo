"""
comparativo.py — Router independiente para generar reportes Excel comparativos
entre dos auditorias geomecanicas. Completamente desacoplado de auditoria.py.
"""
import os
import io
import re
import json
import shutil
import time
import tempfile
from datetime import datetime
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from fastapi import APIRouter, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse

from app.utils.validator import validate_bulk_excel
from app.core.catalogs import MANDATORY_COLS_COUNT
from app.core.rules import RULES_REGISTRY, CATEGORIES_REGISTRY
from app.routers.auditoria import get_incidence_category_name

router = APIRouter()
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
uploads_dir = os.path.join(BASE_DIR, "uploads")

# ─── Paleta de estilos ───────────────────────────────────────────────────────

def _make_styles():
    return {
        "font_title":      Font(name="Segoe UI", size=16, bold=True,  color="1B365D"),
        "font_subtitle":   Font(name="Segoe UI", size=10, italic=True, color="555555"),
        "font_section":    Font(name="Segoe UI", size=11, bold=True,  color="1B365D"),
        "font_header":     Font(name="Segoe UI", size=10, bold=True,  color="FFFFFF"),
        "font_bold":       Font(name="Segoe UI", size=10, bold=True,  color="000000"),
        "font_regular":    Font(name="Segoe UI", size=10,             color="000000"),
        "font_kpi_lbl":    Font(name="Segoe UI", size=9,  bold=True,  color="555555"),
        "font_kpi_blue":   Font(name="Segoe UI", size=18, bold=True,  color="1B365D"),
        "font_kpi_green":  Font(name="Segoe UI", size=18, bold=True,  color="375623"),
        "font_kpi_red":    Font(name="Segoe UI", size=18, bold=True,  color="C00000"),
        "font_delta_green":Font(name="Segoe UI", size=10, bold=True,  color="375623"),
        "font_delta_red":  Font(name="Segoe UI", size=10, bold=True,  color="C00000"),
        "font_delta_gray": Font(name="Segoe UI", size=10,             color="555555"),
        "font_link":       Font(name="Segoe UI", size=10, bold=True,  color="1B365D", underline="single"),
        "fill_primary":    PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid"),
        "fill_a":          PatternFill(start_color="2B4F8C", end_color="2B4F8C", fill_type="solid"),
        "fill_b":          PatternFill(start_color="14532D", end_color="14532D", fill_type="solid"),
        "fill_green":      PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "fill_orange":     PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "fill_red":        PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),
        "fill_yellow":     PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "fill_zebra":      PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid"),
        "fill_kpi_gray":   PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid"),
        "fill_new":        PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid"),
        "fill_resolved":   PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "border_thin": Border(
            left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),  bottom=Side(style="thin", color="E2E8F0")
        ),
        "border_kpi": Border(
            left=Side(style="thin", color="B0C4DE"),  right=Side(style="thin", color="B0C4DE"),
            top=Side(style="thin", color="B0C4DE"),   bottom=Side(style="thin", color="B0C4DE")
        ),
        "align_center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "align_left":   Alignment(horizontal="left",   vertical="center", wrap_text=True),
        "align_right":  Alignment(horizontal="right",  vertical="center"),
    }

def _safe_int(v, d=0):
    try: return int(v)
    except: return d

def _safe_float(v, d=0.0):
    try: return float(v)
    except: return d

def _get_safe_sheet_name(title, index):
    clean_title = re.sub(r'[:\/?*\[\]\'"]', '', title).strip()
    suffix = f" ({index})"
    max_title_len = 31 - len(suffix)
    return f"{clean_title[:max_title_len].strip()}{suffix}"

# ─── Lógica de comparación con intersección de campañas ───────────────────────

def _build_key_set(incs: list) -> set:
    return {(i.get("celda_padre",""), i.get("columna",""), i.get("rule_code","")) for i in incs}

def _build_key_map(incs: list) -> dict:
    m = {}
    for i in incs:
        k = (i.get("celda_padre",""), i.get("columna",""), i.get("rule_code",""))
        if k not in m:
            m[k] = i
    return m

def _load_values_from_excel(excel_path, keys_to_query):
    values_map = {}
    celda_counts = {}
    if not keys_to_query or not excel_path or not os.path.exists(excel_path):
        return values_map, celda_counts
    try:
        import pandas as pd
        import numpy as np
        from app.utils.validator import clean_and_rename_columns, sanitize_value
        
        try:
            import python_calamine
            df = pd.read_excel(excel_path, engine='calamine')
        except ImportError:
            df = pd.read_excel(excel_path, engine='openpyxl')
            
        df = clean_and_rename_columns(df)
        if 'CELDA_PADRE' not in df.columns:
            return values_map, celda_counts
            
        df['CELDA_PADRE'] = df['CELDA_PADRE'].replace([-1, -1.0, '-1', '-1.0'], np.nan)
        df['CELDA_PADRE'] = df['CELDA_PADRE'].ffill()
        
        c_padres = df['CELDA_PADRE'].tolist()
        
        from collections import defaultdict
        celda_to_indices = defaultdict(list)
        for idx, val in enumerate(c_padres):
            cp = sanitize_value(val, str)
            if cp:
                celda_to_indices[cp].append(idx)
                
        celda_counts = {cp: len(indices) for cp, indices in celda_to_indices.items()}
                
        for celda_p, n_th, columna in keys_to_query:
            indices = celda_to_indices.get(celda_p)
            if indices and len(indices) >= n_th:
                row_idx = indices[n_th - 1]
                if columna in df.columns:
                    val = df.at[row_idx, columna]
                    if pd.notna(val) and val is not None:
                        values_map[(celda_p, n_th, columna)] = val
    except Exception as e:
        print(f"Error loading Excel for comparison values ({excel_path}): {e}")
    return values_map, celda_counts

def _compare_audits(diag_a, compact_a, diag_b, compact_b, excel_path_a=None, excel_path_b=None) -> dict:
    inc_a = diag_a.get("incidencias", [])
    inc_b = diag_b.get("incidencias", [])

    # 1. Analizar e intersectar campañas para comparación justa
    campanias_a = {str(i.get("campania")) for i in inc_a if i.get("campania")}
    campanias_b = {str(i.get("campania")) for i in inc_b if i.get("campania")}
    campanias_a.discard("N/A"); campanias_a.discard("")
    campanias_b.discard("N/A"); campanias_b.discard("")
    
    common_campaigns = campanias_a & campanias_b
    has_camp_mismatch = (campanias_a != campanias_b)
    is_filtered_by_common = False
    
    # Filtrar incidencias si hay desajuste y existen campañas comunes
    if has_camp_mismatch and len(common_campaigns) > 0:
        is_filtered_by_common = True
        inc_a_filtered = [i for i in inc_a if str(i.get("campania")) in common_campaigns]
        inc_b_filtered = [i for i in inc_b if str(i.get("campania")) in common_campaigns]
    else:
        inc_a_filtered = inc_a
        inc_b_filtered = inc_b

    keys_a = _build_key_set(inc_a_filtered)
    keys_b = _build_key_set(inc_b_filtered)
    map_a  = _build_key_map(inc_a_filtered)
    map_b  = _build_key_map(inc_b_filtered)

    new_keys      = keys_b - keys_a
    resolved_keys = keys_a - keys_b
    persistent    = keys_a & keys_b

    # Construir conjuntos de consulta para la carga optimizada de Excel
    keys_to_query_b = set()
    for k in resolved_keys:
        rec_a = map_a.get(k)
        if rec_a:
            celda_p = rec_a.get("celda_padre", "")
            celda_h = rec_a.get("celda_hija", "")
            columna = rec_a.get("columna", "")
            n_th = 1
            if celda_h and "-" in celda_h:
                parts = celda_h.split("-")
                if len(parts) >= 2 and parts[-1].isdigit():
                    n_th = int(parts[-1])
            keys_to_query_b.add((celda_p, n_th, columna))
            
    keys_to_query_a = set()
    for k in new_keys:
        rec_b = map_b.get(k)
        if rec_b:
            celda_p = rec_b.get("celda_padre", "")
            celda_h = rec_b.get("celda_hija", "")
            columna = rec_b.get("columna", "")
            n_th = 1
            if celda_h and "-" in celda_h:
                parts = celda_h.split("-")
                if len(parts) >= 2 and parts[-1].isdigit():
                    n_th = int(parts[-1])
            keys_to_query_a.add((celda_p, n_th, columna))

    values_map_a, celda_counts_a = _load_values_from_excel(excel_path_a, keys_to_query_a)
    values_map_b, celda_counts_b = _load_values_from_excel(excel_path_b, keys_to_query_b)

    fam1_a = compact_a.get("familia1", {})
    fam2_a = compact_a.get("familia2", {})
    fam1_b = compact_b.get("familia1", {})
    fam2_b = compact_b.get("familia2", {})

    total_filas_a = _safe_int(fam1_a.get("total_discontinuidades", 0))
    total_filas_b = _safe_int(fam1_b.get("total_discontinuidades", 0))
    celdas_a      = len(compact_a.get("resumen_por_celda_padre", {}))
    celdas_b      = len(compact_b.get("resumen_por_celda_padre", {}))
    
    # Recalcular KPIs basados en el filtro si aplica
    if is_filtered_by_common:
        alertas_a = sum(1 for i in inc_a_filtered if i.get("tipo_incidencia") == "ALERTA")
        alertas_b = sum(1 for i in inc_b_filtered if i.get("tipo_incidencia") == "ALERTA")
        vacios_a  = sum(1 for i in inc_a_filtered if i.get("tipo_incidencia") == "VACIO")
        vacios_b  = sum(1 for i in inc_b_filtered if i.get("tipo_incidencia") == "VACIO")
        adv_a     = sum(1 for i in inc_a_filtered if i.get("tipo_incidencia") == "ADVERTENCIA")
        adv_b     = sum(1 for i in inc_b_filtered if i.get("tipo_incidencia") == "ADVERTENCIA")
        
        celdas_a  = len({i.get("celda_padre") for i in inc_a_filtered if i.get("celda_padre")})
        celdas_b  = len({i.get("celda_padre") for i in inc_b_filtered if i.get("celda_padre")})
        
        # Estimar cantidad de discontinuidades/filas de campañas comunes
        total_filas_a = len({i.get("fila_excel") for i in inc_a_filtered if i.get("fila_excel")})
        total_filas_b = len({i.get("fila_excel") for i in inc_b_filtered if i.get("fila_excel")})
    else:
        alertas_a = _safe_int(fam2_a.get("total_alertas", 0))
        alertas_b = _safe_int(fam2_b.get("total_alertas", 0))
        vacios_a  = _safe_int(fam2_a.get("total_vacios", 0))
        vacios_b  = _safe_int(fam2_b.get("total_vacios", 0))
        adv_a     = _safe_int(fam2_a.get("total_advertencias", 0))
        adv_b     = _safe_int(fam2_b.get("total_advertencias", 0))

    fields_a = total_filas_a * MANDATORY_COLS_COUNT
    fields_b = total_filas_b * MANDATORY_COLS_COUNT
    correctos_a = fields_a - alertas_a - vacios_a - adv_a
    correctos_b = fields_b - alertas_b - vacios_b - adv_b
    integrid_a = (correctos_a / max(1, fields_a)) * 100.0
    integrid_b = (correctos_b / max(1, fields_b)) * 100.0

    # Top errores A y B
    top_a = Counter(i.get("rule_code", i.get("mensaje","?")) for i in inc_a_filtered if i.get("tipo_incidencia")=="ALERTA")
    top_b = Counter(i.get("rule_code", i.get("mensaje","?")) for i in inc_b_filtered if i.get("tipo_incidencia")=="ALERTA")
    all_err_keys = set(top_a.keys()) | set(top_b.keys())
    top_errors = sorted(
        [{"codigo": k, "cant_a": top_a.get(k,0), "cant_b": top_b.get(k,0)} for k in all_err_keys],
        key=lambda x: max(x["cant_a"], x["cant_b"]), reverse=True
    )[:15]

    # Distribución por campaña
    camp_map = defaultdict(lambda: {"alertas_a":0,"vacios_a":0,"adv_a":0,"celdas_a":set(),
                                     "alertas_b":0,"vacios_b":0,"adv_b":0,"celdas_b":set()})
    for i in inc_a_filtered:
        c = str(i.get("campania","N/A")); t = i.get("tipo_incidencia","")
        camp_map[c]["celdas_a"].add(i.get("celda_padre",""))
        if t=="ALERTA": camp_map[c]["alertas_a"]+=1
        elif t=="VACIO": camp_map[c]["vacios_a"]+=1
        elif t=="ADVERTENCIA": camp_map[c]["adv_a"]+=1
    for i in inc_b_filtered:
        c = str(i.get("campania","N/A")); t = i.get("tipo_incidencia","")
        camp_map[c]["celdas_b"].add(i.get("celda_padre",""))
        if t=="ALERTA": camp_map[c]["alertas_b"]+=1
        elif t=="VACIO": camp_map[c]["vacios_b"]+=1
        elif t=="ADVERTENCIA": camp_map[c]["adv_b"]+=1

    dist_campana = sorted([{
        "campania": c,
        "celdas_a": len(v["celdas_a"]), "celdas_b": len(v["celdas_b"]),
        "alertas_a": v["alertas_a"],    "alertas_b": v["alertas_b"],
        "vacios_a":  v["vacios_a"],     "vacios_b":  v["vacios_b"],
        "adv_a":     v["adv_a"],        "adv_b":     v["adv_b"],
    } for c,v in camp_map.items()], key=lambda x: x["campania"])

    # Distribución por sector
    sect_map = defaultdict(lambda: {"alertas_a":0,"vacios_a":0,"celdas_a":set(),
                                     "alertas_b":0,"vacios_b":0,"celdas_b":set()})
    for i in inc_a_filtered:
        s2 = str(i.get("sector_geotecnico","N/A")); t = i.get("tipo_incidencia","")
        sect_map[s2]["celdas_a"].add(i.get("celda_padre",""))
        if t=="ALERTA": sect_map[s2]["alertas_a"]+=1
        elif t=="VACIO": sect_map[s2]["vacios_a"]+=1
    for i in inc_b_filtered:
        s2 = str(i.get("sector_geotecnico","N/A")); t = i.get("tipo_incidencia","")
        sect_map[s2]["celdas_b"].add(i.get("celda_padre",""))
        if t=="ALERTA": sect_map[s2]["alertas_b"]+=1
        elif t=="VACIO": sect_map[s2]["vacios_b"]+=1
    dist_sector = sorted([{
        "sector": s2,
        "celdas_a": len(v["celdas_a"]), "celdas_b": len(v["celdas_b"]),
        "alertas_a": v["alertas_a"],    "alertas_b": v["alertas_b"],
        "vacios_a":  v["vacios_a"],     "vacios_b":  v["vacios_b"],
    } for s2,v in sect_map.items()], key=lambda x: x["sector"])

    # Agrupar celdas
    res_a = compact_a.get("resumen_por_celda_padre", {})
    res_b = compact_b.get("resumen_por_celda_padre", {})
    all_celdas = set(res_a.keys()) | set(res_b.keys())
    celdas_peor, celdas_mejor = [], []
    for celda in all_celdas:
        da = res_a.get(celda, {"alertas":0,"vacios":0})
        db = res_b.get(celda, {"alertas":0,"vacios":0})
        d_alert = _safe_int(db.get("alertas")) - _safe_int(da.get("alertas"))
        d_vacio = _safe_int(db.get("vacios"))  - _safe_int(da.get("vacios"))
        entry = {"celda": celda, "alertas_a": da.get("alertas",0), "alertas_b": db.get("alertas",0),
                 "vacios_a": da.get("vacios",0), "vacios_b": db.get("vacios",0), "delta": d_alert}
        if d_alert > 0: celdas_peor.append(entry)
        elif d_alert < 0 or d_vacio < 0: celdas_mejor.append(entry)
    celdas_peor.sort(key=lambda x: x["delta"], reverse=True)
    celdas_mejor.sort(key=lambda x: x["delta"])

    return {
        "meta_a": {"nombre": diag_a.get("nombre_archivo","Auditoria A"), "fecha": compact_a.get("fecha_auditoria","N/A"), "campanias": sorted(list(campanias_a))},
        "meta_b": {"nombre": diag_b.get("nombre_archivo","Auditoria B"), "fecha": compact_b.get("fecha_auditoria","N/A"), "campanias": sorted(list(campanias_b))},
        "campanias_comunes": sorted(list(common_campaigns)),
        "is_filtered_by_common": is_filtered_by_common,
        "has_camp_mismatch": has_camp_mismatch,
        "kpis": {
            "total_filas":   (total_filas_a,  total_filas_b),
            "celdas":        (celdas_a,        celdas_b),
            "alertas":       (alertas_a,       alertas_b),
            "vacios":        (vacios_a,        vacios_b),
            "advertencias":  (adv_a,           adv_b),
            "integridad":    (integrid_a,      integrid_b),
        },
        "top_errors":   top_errors,
        "dist_campana": dist_campana,
        "dist_sector":  dist_sector,
        "celdas_peor":  celdas_peor[:20],
        "celdas_mejor": celdas_mejor[:20],
        "new_incidencias":      [map_b[k] for k in new_keys],
        "resolved_incidencias": [map_a[k] for k in resolved_keys],
        "totales": {"nuevas": len(new_keys), "resueltas": len(resolved_keys), "persistentes": len(persistent)},
        "inc_a_filtered": inc_a_filtered,
        "inc_b_filtered": inc_b_filtered,
        "values_map_a": values_map_a,
        "values_map_b": values_map_b,
        "celda_counts_a": celda_counts_a,
        "celda_counts_b": celda_counts_b,
    }

# ─── Helpers de escritura Excel ──────────────────────────────────────────────

def _write_header_row(ws, row, cols, s, fill=None):
    f = fill or s["fill_primary"]
    for ci, text in enumerate(cols, start=2):
        c = ws.cell(row=row, column=ci, value=text)
        c.font = s["font_header"]; c.fill = f
        c.alignment = s["align_center"]; c.border = s["border_thin"]

def _delta_cell(ws, row, col, val_a, val_b, s, fmt_pct=False, invert=False):
    if val_a is None or val_b is None:
        ws.cell(row=row, column=col, value="—").alignment = s["align_center"]; return
    delta = _safe_float(val_b) - _safe_float(val_a)
    disp = f"{delta:+.2f}%" if fmt_pct else (f"{int(delta):+d}" if delta == int(delta) else f"{delta:+.2f}")
    improved = (delta < 0) if not invert else (delta > 0)
    c = ws.cell(row=row, column=col, value=disp)
    c.alignment = s["align_center"]; c.border = s["border_thin"]
    if delta == 0:      c.font = s["font_delta_gray"]
    elif improved:      c.font = s["font_delta_green"]; c.fill = s["fill_green"]
    else:               c.font = s["font_delta_red"];   c.fill = s["fill_red"]

def _kpi_card_comparativo(ws, row, col, label, va, vb, s, fmt="n", invert=False):
    """
    Crea una tarjeta KPI limpia de 3 celdas horizontales.
    Formato:
      Fila 5: ESTRUCTURAS (combinado)
      Fila 6: Base (A) | Actual (B) | Δ (Diferencia)
      Fila 7:   76,005 |    8,782  |  -67,223 (verde/rojo)
    """
    # 1. Título de la tarjeta
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)
    lbl = ws.cell(row=row, column=col, value=label)
    lbl.font = s["font_kpi_lbl"]
    lbl.alignment = s["align_center"]
    lbl.fill = s["fill_kpi_gray"]
    lbl.border = s["border_kpi"]

    # 2. Sub-encabezados de columnas de la tarjeta
    headers = ["Base (A)", "Actual (B)", "Δ"]
    for i, h in enumerate(headers):
        c_sub = ws.cell(row=row+1, column=col+i, value=h)
        c_sub.font = Font(name="Segoe UI", size=8, bold=True, color="555555")
        c_sub.alignment = s["align_center"]
        c_sub.fill = s["fill_kpi_gray"]
        c_sub.border = s["border_kpi"]

    # 3. Valores y Delta
    fmt_val = (lambda v: f"{_safe_float(v):.2f}%") if fmt=="pct" else (lambda v: f"{_safe_int(v):,}")
    da, db = _safe_float(va), _safe_float(vb)
    delta = db - da
    improved = (delta < 0) if not invert else (delta > 0)
    
    if fmt == "pct":
        dstr = f"{delta:+.2f}%"
    else:
        dstr = f"{int(delta):+d}" if delta == int(delta) else f"{delta:+.2f}"

    c_a = ws.cell(row=row+2, column=col, value=fmt_val(va))
    c_b = ws.cell(row=row+2, column=col+1, value=fmt_val(vb))
    c_d = ws.cell(row=row+2, column=col+2, value=dstr if delta != 0 else "=")

    c_a.font = s["font_kpi_blue"]; c_a.fill = s["fill_zebra"]
    c_b.font = s["font_kpi_blue"]; c_b.fill = s["fill_zebra"]
    
    if delta == 0:
        c_d.font = s["font_delta_gray"]
        c_d.fill = s["fill_zebra"]
    elif improved:
        c_d.font = s["font_delta_green"]
        c_d.fill = s["fill_green"]
    else:
        c_d.font = s["font_delta_red"]
        c_d.fill = s["fill_red"]

    # Aplicar bordes
    for r in range(row, row+3):
        for c_idx in range(col, col+3):
            ws.cell(row=r, column=c_idx).border = s["border_kpi"]

# ─── Generador de Excel ───────────────────────────────────────────────────────

def generar_excel_comparativo_core(datos: dict) -> openpyxl.Workbook:
    s      = _make_styles()
    meta_a = datos["meta_a"]
    meta_b = datos["meta_b"]
    kpis   = datos["kpis"]
    totales = datos["totales"]
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    inc_a_filtered = datos.get("inc_a_filtered", [])
    inc_b_filtered = datos.get("inc_b_filtered", [])
    values_map_a = datos.get("values_map_a", {})
    values_map_b = datos.get("values_map_b", {})
    celda_counts_a = datos.get("celda_counts_a", {})
    celda_counts_b = datos.get("celda_counts_b", {})

    wb = openpyxl.Workbook()

    # ── HOJA 1: DASHBOARD ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "📊 Dashboard"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for ci in range(2, 25):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.row_dimensions[2].height = 38

    ws.merge_cells("B2:P2")
    t = ws.cell(row=2, column=2, value="COMPARATIVA DE AUDITORÍA Y CONTROL DE CALIDAD")
    t.font = s["font_title"]; t.alignment = s["align_center"]; t.fill = s["fill_primary"]

    # Alerta o Banner si hay desajuste de campañas
    row_start_kpi = 5
    if datos["has_camp_mismatch"]:
        ws.merge_cells("B4:P4")
        if datos["is_filtered_by_common"]:
            warning_text = f"⚠️ ADVERTENCIA: Mismatch de campañas. Se filtraron solo las campañas comunes: {', '.join(datos['campanias_comunes'])}."
            fill_warn = s["fill_orange"]
            font_warn = Font(name="Segoe UI", size=9, bold=True, color="C65911")
        else:
            warning_text = f"❌ ALERTA: Desajuste total de campañas. A: {datos['meta_a']['campanias']} vs B: {datos['meta_b']['campanias']}."
            fill_warn = s["fill_red"]
            font_warn = Font(name="Segoe UI", size=9, bold=True, color="C00000")
            
        c_warn = ws.cell(row=4, column=2, value=warning_text)
        c_warn.font = font_warn; c_warn.fill = fill_warn; c_warn.alignment = s["align_center"]
        c_warn.border = s["border_thin"]
        row_start_kpi = 6

    # KPI Cards con layout corregido
    kpi_definitions = [
        ("ESTRUCTURAS",   kpis["total_filas"][0],  kpis["total_filas"][1],  "n", False),
        ("ESTACIONES",    kpis["celdas"][0],        kpis["celdas"][1],        "n", False),
        ("ALERTAS",       kpis["alertas"][0],       kpis["alertas"][1],       "n", False),
        ("VACIOS",        kpis["vacios"][0],        kpis["vacios"][1],        "n", False),
        ("ADVERTENCIAS",  kpis["advertencias"][0],  kpis["advertencias"][1],  "n", False),
    ]
    for ki, (lbl, va, vb, fmt, inv) in enumerate(kpi_definitions):
        _kpi_card_comparativo(ws, row_start_kpi, 2 + ki*3, lbl, va, vb, s, fmt, inv)

    # Fila de Integridad global
    row_int = row_start_kpi + 4
    ws.merge_cells(start_row=row_int, start_column=2, end_row=row_int, end_column=4)
    ws.cell(row=row_int, column=2, value="INTEGRIDAD GLOBAL").font = s["font_kpi_lbl"]
    ws.cell(row=row_int, column=2).alignment = s["align_center"]; ws.cell(row=row_int, column=2).fill = s["fill_kpi_gray"]
    ws.cell(row=row_int, column=2).border = s["border_kpi"]
    
    ws.cell(row=row_int+1, column=2, value="Base (A)").font = Font(name="Segoe UI", size=8, bold=True, color="555555")
    ws.cell(row=row_int+1, column=2).alignment = s["align_center"]
    ws.cell(row=row_int+1, column=2).fill = s["fill_kpi_gray"]
    ws.cell(row=row_int+1, column=2).border = s["border_kpi"]
    
    ws.cell(row=row_int+1, column=3, value="Actual (B)").font = Font(name="Segoe UI", size=8, bold=True, color="555555")
    ws.cell(row=row_int+1, column=3).alignment = s["align_center"]
    ws.cell(row=row_int+1, column=3).fill = s["fill_kpi_gray"]
    ws.cell(row=row_int+1, column=3).border = s["border_kpi"]
    
    ws.cell(row=row_int+1, column=4, value="Δ").font = Font(name="Segoe UI", size=8, bold=True, color="555555")
    ws.cell(row=row_int+1, column=4).alignment = s["align_center"]
    ws.cell(row=row_int+1, column=4).fill = s["fill_kpi_gray"]
    ws.cell(row=row_int+1, column=4).border = s["border_kpi"]
    
    for ci2, val in [(2, kpis["integridad"][0]), (3, kpis["integridad"][1])]:
        c = ws.cell(row=row_int+2, column=ci2, value=f"{_safe_float(val):.2f}%")
        c.font = s["font_kpi_green"]; c.alignment = s["align_center"]; c.fill = s["fill_zebra"]; c.border = s["border_kpi"]
    _delta_cell(ws, row_int+2, 4, kpis["integridad"][0], kpis["integridad"][1], s, fmt_pct=True, invert=True)

    # Tabla resumen de cambios
    row_res = row_int + 5
    ws.cell(row=row_res, column=2, value="CONSOLIDADO DE CAMBIOS DE INCIDENCIAS").font = s["font_section"]
    for ri2, (lbl2, qty, fill2) in enumerate([
        ("NUEVAS (Regresiones)",    totales["nuevas"],      s["fill_new"]),
        ("RESUELTAS (Mejoras)",  totales["resueltas"],   s["fill_resolved"]),
        ("PERSISTENTES",         totales["persistentes"],s["fill_kpi_gray"]),
    ], start=row_res+1):
        c1 = ws.cell(row=ri2, column=2, value=lbl2)
        c1.font = s["font_bold"]; c1.fill = fill2; c1.border = s["border_thin"]
        c2 = ws.cell(row=ri2, column=3, value=qty)
        c2.font = s["font_bold"]; c2.alignment = s["align_right"]; c2.fill = fill2; c2.border = s["border_thin"]
        c2.number_format = '#,##0'

    # Tabla oculta / auxiliar para Gráficos
    row_chart_tbl = row_res + 6
    ws.cell(row=row_chart_tbl, column=2, value="CATEGORIA").font = s["font_bold"]
    ws.cell(row=row_chart_tbl, column=3, value="BASE (A)").font = s["font_bold"]
    ws.cell(row=row_chart_tbl, column=4, value="ACTUAL (B)").font = s["font_bold"]
    
    chart_metrics = [
        ("Alertas", kpis["alertas"][0], kpis["alertas"][1]),
        ("Vacios", kpis["vacios"][0], kpis["vacios"][1]),
        ("Advertencias", kpis["advertencias"][0], kpis["advertencias"][1])
    ]
    for idx, (cat_name, val_a, val_b) in enumerate(chart_metrics, start=1):
        ws.cell(row=row_chart_tbl+idx, column=2, value=cat_name).font = s["font_regular"]
        ws.cell(row=row_chart_tbl+idx, column=3, value=val_a).font = s["font_regular"]
        ws.cell(row=row_chart_tbl+idx, column=4, value=val_b).font = s["font_regular"]
        ws.cell(row=row_chart_tbl+idx, column=3).number_format = '#,##0'
        ws.cell(row=row_chart_tbl+idx, column=4).number_format = '#,##0'

    # Gráfico de barras en 3D/Clustered de Alertas, Vacíos y Advertencias
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Frecuencia de Desviaciones (Base vs Actual)"
    chart.y_axis.title = "Cantidad de Ocurrencias"
    chart.x_axis.title = "Categoría"
    
    data = Reference(ws, min_col=3, min_row=row_chart_tbl, max_col=4, max_row=row_chart_tbl+3)
    cats = Reference(ws, min_col=2, min_row=row_chart_tbl+1, max_row=row_chart_tbl+3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 18
    chart.height = 10
    ws.add_chart(chart, "G14")

    # ── HOJA 2: RESUMEN EJECUTIVO ─────────────────────────────────────────
    ws2 = wb.create_sheet("📋 Resumen Ejecutivo")
    ws2.column_dimensions["A"].width = 2; ws2.column_dimensions["B"].width = 30
    for ci2 in range(3,9): ws2.column_dimensions[get_column_letter(ci2)].width = 18
    ws2.merge_cells("B2:H2")
    ws2.cell(row=2, column=2, value="RESUMEN EJECUTIVO COMPARATIVO").font = s["font_title"]
    ws2.cell(row=2, column=2).alignment = s["align_center"]; ws2.cell(row=2, column=2).fill = s["fill_primary"]
    _write_header_row(ws2, 4, ["Indicador", f"A ({meta_a['nombre'][:18]})", f"B ({meta_b['nombre'][:18]})", "Delta Abs.", "Delta %", "Tendencia"], s)
    exec_rows = [
        ("Estructuras",    kpis["total_filas"][0],    kpis["total_filas"][1],    False, "n"),
        ("Estaciones",     kpis["celdas"][0],          kpis["celdas"][1],          False, "n"),
        ("Alertas criticas",kpis["alertas"][0],        kpis["alertas"][1],         False, "n"),
        ("Vacios",         kpis["vacios"][0],          kpis["vacios"][1],          False, "n"),
        ("Advertencias",   kpis["advertencias"][0],    kpis["advertencias"][1],    False, "n"),
        ("Integridad (%)", kpis["integridad"][0],      kpis["integridad"][1],      True,  "pct"),
        ("Incidencias nuevas",    0,                   totales["nuevas"],           False, "n"),
        ("Incidencias resueltas", totales["resueltas"],0,                           True,  "n"),
    ]
    for ri2, (lbl, va, vb, inv, fmt) in enumerate(exec_rows, start=5):
        va_f, vb_f = _safe_float(va), _safe_float(vb)
        delta_abs  = vb_f - va_f
        delta_pct2 = ((vb_f - va_f) / max(1.0, abs(va_f))) * 100.0
        improved   = (delta_abs < 0) if not inv else (delta_abs > 0)
        trend      = "Mejora" if improved else ("Deterioro" if delta_abs != 0 else "Igual")
        row_fill2  = s["fill_green"] if improved else (s["fill_red"] if delta_abs != 0 else None)
        disp_a     = f"{va_f:.2f}%" if fmt=="pct" else _safe_int(va)
        disp_b     = f"{vb_f:.2f}%" if fmt=="pct" else _safe_int(vb)
        disp_da    = f"{delta_abs:+.2f}%" if fmt=="pct" else f"{int(delta_abs):+d}"
        disp_dp    = f"{delta_pct2:+.1f}%"
        for ci2, val2 in [(2,lbl),(3,disp_a),(4,disp_b),(5,disp_da),(6,disp_dp),(7,trend)]:
            c = ws2.cell(row=ri2, column=ci2, value=val2)
            c.border = s["border_thin"]
            c.alignment = s["align_right"] if ci2>3 else s["align_left"]
            c.font = s["font_bold"] if ci2==2 else (s["font_delta_green"] if (ci2==7 and improved) else (s["font_delta_red"] if (ci2==7 and delta_abs!=0) else s["font_regular"]))
            if row_fill2 and ci2>=5: c.fill = row_fill2
            elif ri2%2==0: c.fill = s["fill_zebra"]
            
    ws2.auto_filter.ref = f"B4:H{len(exec_rows)+4}"

    # ── HOJA 3: POR CAMPAÑA ───────────────────────────────────────────────
    ws3 = wb.create_sheet("📅 Por Campaña")
    ws3.column_dimensions["A"].width=2; ws3.column_dimensions["B"].width=14
    for ci2 in range(3,15): ws3.column_dimensions[get_column_letter(ci2)].width=13
    ws3.merge_cells("B2:M2")
    ws3.cell(row=2, column=2, value="DISTRIBUCION COMPARATIVA POR CAMPANA").font = s["font_title"]
    ws3.cell(row=2, column=2).alignment=s["align_center"]; ws3.cell(row=2, column=2).fill=s["fill_primary"]
    _write_header_row(ws3, 4, ["Campana","Celdas A","Celdas B","Alertas A","Alertas B","D.Alertas","Vacios A","Vacios B","D.Vacios","Adv A","Adv B","D.Adv"], s)
    for ri2, row in enumerate(datos["dist_campana"], start=5):
        vals = [row["campania"],row["celdas_a"],row["celdas_b"],
                row["alertas_a"],row["alertas_b"],row["alertas_b"]-row["alertas_a"],
                row["vacios_a"], row["vacios_b"], row["vacios_b"]-row["vacios_a"],
                row["adv_a"],    row["adv_b"],    row["adv_b"]-row["adv_a"]]
        for ci2, val2 in enumerate(vals, start=2):
            c = ws3.cell(row=ri2, column=ci2, value=val2)
            c.border = s["border_thin"]
            c.alignment = s["align_right"] if ci2>2 else s["align_left"]
            if ci2 in (7,10,13):
                d = _safe_int(val2)
                if d<0: c.font=s["font_delta_green"]; c.fill=s["fill_green"]
                elif d>0: c.font=s["font_delta_red"];  c.fill=s["fill_red"]
                else: c.font=s["font_delta_gray"]
            elif ri2%2==0: c.fill=s["fill_zebra"]
            
    ws3.auto_filter.ref = f"B4:M{len(datos['dist_campana'])+4}"

    # ── HOJA 4: POR SECTOR ────────────────────────────────────────────────
    ws4 = wb.create_sheet("📍 Por Sector")
    ws4.column_dimensions["A"].width=2; ws4.column_dimensions["B"].width=22
    for ci2 in range(3,12): ws4.column_dimensions[get_column_letter(ci2)].width=14
    ws4.merge_cells("B2:J2")
    ws4.cell(row=2, column=2, value="DISTRIBUCION COMPARATIVA POR SECTOR GEOTECNICO").font=s["font_title"]
    ws4.cell(row=2, column=2).alignment=s["align_center"]; ws4.cell(row=2, column=2).fill=s["fill_primary"]
    _write_header_row(ws4, 4, ["Sector","Celdas A","Celdas B","Alertas A","Alertas B","D.Alertas","Vacios A","Vacios B","D.Vacios"], s)
    for ri2, row in enumerate(datos["dist_sector"], start=5):
        vals=[row["sector"],row["celdas_a"],row["celdas_b"],
              row["alertas_a"],row["alertas_b"],row["alertas_b"]-row["alertas_a"],
              row["vacios_a"], row["vacios_b"], row["vacios_b"]-row["vacios_a"]]
        for ci2, val2 in enumerate(vals, start=2):
            c=ws4.cell(row=ri2, column=ci2, value=val2); c.border=s["border_thin"]
            c.alignment=s["align_right"] if ci2>2 else s["align_left"]
            if ci2 in (7,10):
                d=_safe_int(val2)
                if d<0: c.font=s["font_delta_green"]; c.fill=s["fill_green"]
                elif d>0: c.font=s["font_delta_red"];  c.fill=s["fill_red"]
                else: c.font=s["font_delta_gray"]
            elif ri2%2==0: c.fill=s["fill_zebra"]
            
    ws4.auto_filter.ref = f"B4:J{len(datos['dist_sector'])+4}"

    # ── HOJA 5: MASTER REGISTRO DE ERRORES (CATÁLOGO / ÍNDICE) ───────────
    ws_cat = wb.create_sheet(title="❌ Catálogo de Errores")
    ws_cat.views.sheetView[0].showGridLines = True
    
    ws_cat.cell(row=2, column=2, value="REGISTRO MAESTRO DE ERRORES COMPARATIVO").font = s["font_title"]
    ws_cat.cell(row=3, column=2, value="Lista consolidada de desviaciones geomecánicas. Use los hipervínculos para navegar.").font = s["font_subtitle"]
    
    headers_cat = ["ID", "Gravedad", "Regla de Consistencia Evaluada", "Casos A", "Casos B", "Delta (B-A)", "Enlace Detalle"]
    _write_header_row(ws_cat, 5, headers_cat, s)

    # Agrupar incidencias por su nombre simplificado / categoría
    incidencias_por_error_a = defaultdict(list)
    for inc in (inc_a_filtered):
        msg_simplificado = get_incidence_category_name(inc)
        incidencias_por_error_a[msg_simplificado].append(inc)
        
    incidencias_por_error_b = defaultdict(list)
    for inc in (inc_b_filtered):
        msg_simplificado = get_incidence_category_name(inc)
        incidencias_por_error_b[msg_simplificado].append(inc)

    catalog_frequencies = []
    for c_idx, cat in enumerate(CATEGORIES_REGISTRY.values(), start=1):
        rule_msg = cat.name
        matches_a = incidencias_por_error_a[rule_msg]
        matches_b = incidencias_por_error_b[rule_msg]
        
        if len(matches_a) > 0 or len(matches_b) > 0:
            catalog_frequencies.append({
                "msg": rule_msg,
                "severity": cat.severity,
                "count_a": len(matches_a),
                "count_b": len(matches_b),
                "delta": len(matches_b) - len(matches_a),
                "matches_a": matches_a,
                "matches_b": matches_b
            })
            
    catalog_frequencies = sorted(catalog_frequencies, key=lambda x: max(x["count_a"], x["count_b"]), reverse=True)

    r_cat = 6
    active_sheets_mapping = {}
    
    for c_idx, rule in enumerate(catalog_frequencies, start=1):
        ws_cat.cell(row=r_cat, column=2, value=c_idx).font = s["font_regular"]
        ws_cat.cell(row=r_cat, column=2).alignment = s["align_center"]
        ws_cat.cell(row=r_cat, column=2).border = s["border_thin"]
        
        c_sev = ws_cat.cell(row=r_cat, column=3, value=rule["severity"])
        c_sev.font = s["font_bold"]; c_sev.alignment = s["align_center"]; c_sev.border = s["border_thin"]
        if rule["severity"] == "ALERTA": c_sev.fill = s["fill_red"]
        elif rule["severity"] == "ADVERTENCIA": c_sev.fill = s["fill_orange"]
        else: c_sev.fill = s["fill_yellow"]
        
        ws_cat.cell(row=r_cat, column=4, value=rule["msg"]).font = s["font_bold"]
        ws_cat.cell(row=r_cat, column=4).border = s["border_thin"]
        
        c_ca = ws_cat.cell(row=r_cat, column=5, value=rule["count_a"])
        c_ca.font = s["font_bold"]; c_ca.alignment = s["align_right"]; c_ca.border = s["border_thin"]
        c_ca.number_format = '#,##0'
        
        c_cb = ws_cat.cell(row=r_cat, column=6, value=rule["count_b"])
        c_cb.font = s["font_bold"]; c_cb.alignment = s["align_right"]; c_cb.border = s["border_thin"]
        c_cb.number_format = '#,##0'
        
        c_delta = ws_cat.cell(row=r_cat, column=7, value=rule["delta"])
        c_delta.font = s["font_bold"]; c_delta.alignment = s["align_right"]; c_delta.border = s["border_thin"]
        c_delta.number_format = '#,##0'
        
        # Semaforización del delta en el catálogo
        if rule["delta"] < 0:
            c_delta.font = s["font_delta_green"]; c_delta.fill = s["fill_green"]
        elif rule["delta"] > 0:
            c_delta.font = s["font_delta_red"]; c_delta.fill = s["fill_red"]

        c_link = ws_cat.cell(row=r_cat, column=8)
        tab_name = _get_safe_sheet_name(rule["msg"], c_idx)
        active_sheets_mapping[rule["msg"]] = {
            "tab_name": tab_name, 
            "matches_a": rule["matches_a"], 
            "matches_b": rule["matches_b"],
            "severity": rule["severity"]
        }
        
        c_link.value = f'=HYPERLINK("#\'Ref {c_idx}\'!B2", "🔍 Ver Cambios")'
        c_link.font = s["font_link"]; c_link.alignment = s["align_center"]; c_link.border = s["border_thin"]
        
        if r_cat % 2 == 0 and c_delta.fill == s["fill_zebra"]:
            ws_cat.cell(row=r_cat, column=4).fill = s["fill_zebra"]
            c_ca.fill = s["fill_zebra"]
            c_cb.fill = s["fill_zebra"]
            
        r_cat += 1
        
    ws_cat.auto_filter.ref = f"B5:H{r_cat-1}"

    # ── HOJA 6: CELDAS CRÍTICAS ───────────────────────────────────────────
    ws6 = wb.create_sheet("⚠️ Celdas Críticas")
    ws6.column_dimensions["A"].width=2; ws6.column_dimensions["B"].width=22
    for ci2 in range(3,9): ws6.column_dimensions[get_column_letter(ci2)].width=14
    ws6.merge_cells("B2:H2")
    h_peor = ws6.cell(row=2, column=2, value="CELDAS QUE EMPEORARON (mas alertas en B)")
    h_peor.font=s["font_section"]; h_peor.alignment=s["align_center"]
    h_peor.fill=PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    _write_header_row(ws6, 3, ["Celda","Alertas A","Alertas B","Vacios A","Vacios B","Delta Alert."],
                      s, fill=PatternFill(start_color="9B1C1C", end_color="9B1C1C", fill_type="solid"))
    for ri2, row in enumerate(datos["celdas_peor"], start=4):
        for ci2, val2 in [(2,row["celda"]),(3,row["alertas_a"]),(4,row["alertas_b"]),
                          (5,row["vacios_a"]),(6,row["vacios_b"]),(7,f'+{row["delta"]}')]:
            c=ws6.cell(row=ri2, column=ci2, value=val2); c.border=s["border_thin"]
            c.fill=s["fill_new"]; c.font=s["font_bold"] if ci2 in (2,7) else s["font_regular"]
            c.alignment=s["align_right"] if ci2>2 else s["align_left"]

    r_mejor = len(datos["celdas_peor"]) + 6
    ws6.merge_cells(start_row=r_mejor, start_column=2, end_row=r_mejor, end_column=7)
    h_mejor = ws6.cell(row=r_mejor, column=2, value="CELDAS QUE MEJORARON (menos alertas en B)")
    h_mejor.font=s["font_section"]; h_mejor.alignment=s["align_center"]
    h_mejor.fill=PatternFill(start_color="166534", end_color="166534", fill_type="solid")
    _write_header_row(ws6, r_mejor+1, ["Celda","Alertas A","Alertas B","Vacios A","Vacios B","Delta Alert."],
                      s, fill=PatternFill(start_color="14532D", end_color="14532D", fill_type="solid"))
    for ri2, row in enumerate(datos["celdas_mejor"], start=r_mejor+2):
        for ci2, val2 in [(2,row["celda"]),(3,row["alertas_a"]),(4,row["alertas_b"]),
                          (5,row["vacios_a"]),(6,row["vacios_b"]),(7,str(row["delta"]))]:
            c=ws6.cell(row=ri2, column=ci2, value=val2); c.border=s["border_thin"]
            c.fill=s["fill_resolved"]; c.font=s["font_bold"] if ci2 in (2,7) else s["font_regular"]
            c.alignment=s["align_right"] if ci2>2 else s["align_left"]

    # ── HOJAS 7+: DETALLES INDIVIDUALES POR REGLA DE ERROR COMPARATIVO ────
    # Mapea clave (celda, col) para rastrear de forma exacta
    for idx_rule, (rule_msg, mapping_data) in enumerate(active_sheets_mapping.items(), start=1):
        sh_name = f"Ref {idx_rule}" # Nombre acotado y seguro para Excel
        matches_a = mapping_data["matches_a"]
        matches_b = mapping_data["matches_b"]
        severity = mapping_data["severity"]
        
        ws_err = wb.create_sheet(title=sh_name)
        ws_err.views.sheetView[0].showGridLines = True
        
        # Link para regresar al catálogo maestro
        c_back = ws_err.cell(row=2, column=2)
        c_back.value = '=HYPERLINK("#\'❌ Catálogo de Errores\'!B2", "⬅ Volver al Catálogo de Errores")'
        c_back.font = s["font_link"]
        c_back.alignment = s["align_left"]
        
        ws_err.cell(row=4, column=2, value="ANÁLISIS COMPARATIVO DE INCIDENCIA").font = s["font_section"]
        cell_err_desc = ws_err.cell(row=5, column=2, value=f"Regla: {rule_msg.upper()}")
        cell_err_desc.font = Font(name="Segoe UI", size=10, bold=True, color="7F1D1D" if severity=="ALERTA" else "7F5F1D")
        cell_err_desc.fill = s["fill_red"] if severity=="ALERTA" else s["fill_orange"]
        cell_err_desc.border = s["border_thin"]
        ws_err.merge_cells(start_row=5, start_column=2, end_row=5, end_column=11)
        
        # Subheaders
        headers_detail = ["Fila A", "Fila B", "Celda Padre", "Celda Hija", "Campaña", "Sector", "Columna de Falla", "Valor Antiguo (A)", "Valor Nuevo (B)", "Estado de Cambio"]
        _write_header_row(ws_err, 7, headers_detail, s)
        
        # Combinar incidencias de A y B por su clave única (celda_padre, columna)
        combined_records = {}
        for r_a in matches_a:
            key_rec = (r_a.get("celda_padre"), r_a.get("columna"))
            combined_records[key_rec] = {"a": r_a, "b": None}
            
        for r_b in matches_b:
            key_rec = (r_b.get("celda_padre"), r_b.get("columna"))
            if key_rec in combined_records:
                combined_records[key_rec]["b"] = r_b
            else:
                combined_records[key_rec] = {"a": None, "b": r_b}
                
        r_idx = 8
        for (celda_p, col_falla), cmp_data in combined_records.items():
            rec_a = cmp_data["a"]
            rec_b = cmp_data["b"]
            
            # Resolver estado del cambio
            if rec_a and rec_b:
                estado_str = "PERSISTENTE"
                fill_row = s["fill_new"]
                fila_a = rec_a.get("fila_excel", "")
                fila_b = rec_b.get("fila_excel", "")
                val_antiguo = rec_a.get("valor_actual", "—")
                val_nuevo = rec_b.get("valor_actual", "—")
                resolved_camp = rec_b.get("campania", "")
                resolved_sector = rec_b.get("sector_geotecnico", "")
                celda_h = rec_b.get("celda_hija", "")
            elif rec_a and not rec_b:
                estado_str = "RESUELTA (Mejorado)"
                fill_row = s["fill_resolved"]
                fila_a = rec_a.get("fila_excel", "")
                fila_b = "—"
                val_antiguo = rec_a.get("valor_actual", "—")
                
                # Intentar obtener el valor corregido real de Excel B
                val_nuevo = "CORREGIDO / OK"
                
                celda_p = rec_a.get("celda_padre", "")
                celda_h = rec_a.get("celda_hija", "")
                columna = rec_a.get("columna", "")
                
                n_th = 1
                if celda_h and "-" in celda_h:
                    parts = celda_h.split("-")
                    if len(parts) >= 2 and parts[-1].isdigit():
                        n_th = int(parts[-1])
                        
                # Si el número de estructura es mayor que las existentes en B, es que la fila fue eliminada
                if celda_counts_b and (celda_p not in celda_counts_b or n_th > celda_counts_b[celda_p]):
                    val_nuevo = "FILA ELIMINADA / OK"
                elif values_map_b:
                    raw_val_b = values_map_b.get((celda_p, n_th, columna))
                    if raw_val_b is not None:
                        from app.utils.validator import format_raw_value_for_report
                        formatted_b = format_raw_value_for_report(raw_val_b)
                        if formatted_b is not None:
                            val_nuevo = formatted_b
                            
                resolved_camp = rec_a.get("campania", "")
                resolved_sector = rec_a.get("sector_geotecnico", "")
                celda_h = rec_a.get("celda_hija", "")
            else:
                estado_str = "NUEVA (Regresión)"
                fill_row = s["fill_new"]
                fila_a = "—"
                fila_b = rec_b.get("fila_excel", "")
                
                # Intentar obtener el valor antiguo correcto real de Excel A
                val_antiguo = "—"
                
                celda_p = rec_b.get("celda_padre", "")
                celda_h = rec_b.get("celda_hija", "")
                columna = rec_b.get("columna", "")
                
                n_th = 1
                if celda_h and "-" in celda_h:
                    parts = celda_h.split("-")
                    if len(parts) >= 2 and parts[-1].isdigit():
                        n_th = int(parts[-1])
                        
                # Si el número de estructura es mayor que las existentes en A, es que la fila es nueva/creada
                if celda_counts_a and (celda_p not in celda_counts_a or n_th > celda_counts_a[celda_p]):
                    val_antiguo = "FILA CREADA (Nueva)"
                elif values_map_a:
                    raw_val_a = values_map_a.get((celda_p, n_th, columna))
                    if raw_val_a is not None:
                        from app.utils.validator import format_raw_value_for_report
                        formatted_a = format_raw_value_for_report(raw_val_a)
                        if formatted_a is not None:
                            val_antiguo = formatted_a
                            
                val_nuevo = rec_b.get("valor_actual", "—")
                resolved_camp = rec_b.get("campania", "")
                resolved_sector = rec_b.get("sector_geotecnico", "")
                celda_h = rec_b.get("celda_hija", "")

            row_values = [
                fila_a,
                fila_b,
                celda_p,
                celda_h,
                resolved_camp,
                resolved_sector,
                col_falla,
                val_antiguo,
                val_nuevo,
                estado_str
            ]
            
            for col_idx, val in enumerate(row_values, start=2):
                cell = ws_err.cell(row=r_idx, column=col_idx, value=val)
                cell.font = s["font_regular"]
                cell.border = s["border_thin"]
                cell.alignment = s["align_left"] if col_idx in (4,5,7,8,9,10) else s["align_center"]
                
                # Resaltado condicional por estado de cambio
                if col_idx == 11:
                    cell.font = s["font_bold"]
                    if estado_str.startswith("RESUELTA"):
                        cell.fill = s["fill_resolved"]
                        cell.font = s["font_delta_green"]
                    elif estado_str.startswith("NUEVA"):
                        cell.fill = s["fill_new"]
                        cell.font = s["font_delta_red"]
                    else:
                        cell.fill = s["fill_yellow"]
                        cell.font = s["font_delta_gray"]
                elif r_idx % 2 == 0 and col_idx < 11:
                    cell.fill = s["fill_zebra"]
                    
            r_idx += 1
            
        # Configurar filtros automáticos en las hojas de detalles
        ws_err.auto_filter.ref = f"B7:K{r_idx-1}"

    # --- AUTO-AJUSTE DINÁMICO DE COLUMNAS AL FINAL ---
    for ws in wb.worksheets:
        ws.column_dimensions['A'].width = 3
        if ws.title.startswith("Ref "):
            ws.column_dimensions['B'].width = 11  # Fila A
            ws.column_dimensions['C'].width = 11  # Fila B
            ws.column_dimensions['D'].width = 16  # Celda Padre
            ws.column_dimensions['E'].width = 16  # Celda Hija
            ws.column_dimensions['F'].width = 11  # Campaña
            ws.column_dimensions['G'].width = 24  # Sector
            ws.column_dimensions['H'].width = 24  # Columna de Falla
            ws.column_dimensions['I'].width = 20  # Valor Antiguo
            ws.column_dimensions['J'].width = 20  # Valor Nuevo
            ws.column_dimensions['K'].width = 24  # Estado de Cambio
        elif ws.title == "❌ Catálogo de Errores":
            ws.column_dimensions['B'].width = 6   # ID
            ws.column_dimensions['C'].width = 16  # Gravedad
            ws.column_dimensions['D'].width = 52  # Regla
            ws.column_dimensions['E'].width = 12  # Casos A
            ws.column_dimensions['F'].width = 12  # Casos B
            ws.column_dimensions['G'].width = 12  # Delta
            ws.column_dimensions['H'].width = 22  # Enlace
        elif ws.title == "📊 Dashboard":
            # Mantener los anchos definidos para el dashboard
            pass
        else:
            # Auto-ajuste dinámico de las otras hojas (Ej. Resumen Ejecutivo, Campaña, Sector)
            for col_idx in range(2, ws.max_column + 1):
                vals = []
                for row_idx in range(1, min(15, ws.max_row + 1)):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val is not None:
                        val_str = str(val)
                        if val_str.startswith("=HYPERLINK"):
                            vals.append("Ver Registros")
                        else:
                            vals.append(val_str)
                if not vals: continue
                max_len = max(len(v) for v in vals)
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 11), 52)

    return wb

# ─── Helpers para cargar archivos de historial ───────────────────────────────

def _load_audit_files(audit_id: str):
    history_dir = os.path.join(uploads_dir, "history")
    raw_f    = os.path.join(history_dir, f"{audit_id}_diagnostico.json")
    comp_f   = os.path.join(history_dir, f"{audit_id}_compact.json")
    if not os.path.exists(raw_f) or not os.path.exists(comp_f):
        raise HTTPException(status_code=404,
            detail=f"Auditoria '{audit_id}' no encontrada o incompleta en el historial.")
    with open(raw_f,  "r", encoding="utf-8") as f: diag    = json.load(f)
    with open(comp_f, "r", encoding="utf-8") as f: compact = json.load(f)
    return diag, compact

def _stream_wb(wb, fname):
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})

# ─── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/geomecanica/comparativo/reporte")
def reporte_comparativo_historial(audit_id_a: str, audit_id_b: str):
    """Genera Excel comparativo entre dos auditorias ya procesadas en el historial."""
    if not audit_id_a or not audit_id_b:
        raise HTTPException(status_code=400, detail="Se requieren audit_id_a y audit_id_b.")
    if audit_id_a == audit_id_b:
        raise HTTPException(status_code=400, detail="Los dos audit_id no pueden ser iguales.")
    diag_a, compact_a = _load_audit_files(audit_id_a)
    diag_b, compact_b = _load_audit_files(audit_id_b)
    excel_path_a = os.path.join(uploads_dir, "history", f"{audit_id_a}.xlsx")
    excel_path_b = os.path.join(uploads_dir, "history", f"{audit_id_b}.xlsx")
    datos = _compare_audits(diag_a, compact_a, diag_b, compact_b, excel_path_a=excel_path_a, excel_path_b=excel_path_b)
    wb    = generar_excel_comparativo_core(datos)
    fname = f"comparativo_{audit_id_a}_vs_{audit_id_b}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _stream_wb(wb, fname)

@router.post("/geomecanica/comparativo/importar-y-comparar")
async def reporte_comparativo_importar(file_a: UploadFile = File(...), file_b: UploadFile = File(...)):
    """
    Sube dos Excel crudos, los audita secuencialmente y genera el comparativo.
    Puede tardar varios minutos dependiendo del tamano de los archivos.
    """
    for f in [file_a, file_b]:
        if not f.filename.endswith((".xlsx", ".xls")):
            raise HTTPException(status_code=400, detail=f"Formato no soportado: {f.filename}")
    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    tmp_dir = tempfile.mkdtemp(dir=uploads_dir)
    try:
        path_a = os.path.join(tmp_dir, f"cmp_a_{ts}.xlsx")
        path_b = os.path.join(tmp_dir, f"cmp_b_{ts}.xlsx")
        raw_a  = os.path.join(tmp_dir, "diag_a.json")
        raw_b  = os.path.join(tmp_dir, "diag_b.json")

        with open(path_a, "wb") as buf: shutil.copyfileobj(file_a.file, buf)
        with open(path_b, "wb") as buf: shutil.copyfileobj(file_b.file, buf)

        validate_bulk_excel(path_a, raw_a)
        with open(raw_a, "r", encoding="utf-8") as f2: diag_a = json.load(f2)
        diag_a["nombre_archivo"] = file_a.filename

        validate_bulk_excel(path_b, raw_b)
        with open(raw_b, "r", encoding="utf-8") as f2: diag_b = json.load(f2)
        diag_b["nombre_archivo"] = file_b.filename

        from app.routers.auditoria import aggregate_audit_metrics
        compact_a = aggregate_audit_metrics(diag_a)
        compact_a["fecha_auditoria"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        compact_b = aggregate_audit_metrics(diag_b)
        compact_b["fecha_auditoria"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        datos = _compare_audits(diag_a, compact_a, diag_b, compact_b, excel_path_a=path_a, excel_path_b=path_b)
        wb    = generar_excel_comparativo_core(datos)
        return _stream_wb(wb, f"comparativo_{ts}.xlsx")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)
