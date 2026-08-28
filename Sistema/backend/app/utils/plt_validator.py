"""
app/utils/plt_validator.py — Motor de Validación QA/QC para Ensayos PLT Irregulares.
Procesa archivos Excel de Ensayos PLT, detecta inconsistencias contra las reglas SSOT (rules_plt.py),
analiza la integridad de celdas A-B-C-D y exporta el diagnóstico estructurado.

Soporta múltiples formatos:
  1. FORMATO ESTÁNDAR 34 COLUMNAS (ej. agos_01_22.xlsx, agosto_plt_cargado.xlsx, junio_julio_26.xlsx)
  2. FORMATO DE CAMPO COMPACTO (ej. 03 feb1.xlsx - Encabezados fila 3, datos fila 6+)
"""

from collections import defaultdict
from datetime import datetime
import json
import math
import os
import re
import time
import unicodedata
from typing import Any, Callable, Dict, List, Optional, Tuple, Union

import openpyxl
import pandas as pd

from app.core.catalogs import (
    LITHOLOGY_FULL_CATALOG,
    LITHOLOGY_GROUP_SYNONYMS,
)
from app.core.rules_plt import RULES_REGISTRY_PLT, CATEGORIES_REGISTRY_PLT


PLT_MANDATORY_COLS_COUNT = 34
CAT_DIRECCION_ROTURA = ["Pa", "Pe", "NA"]
CAT_TIPO_FRACTURA = ["M", "E", "C"]

# ---------------------------------------------------------------------------
# Catálogo oficial ISRM (SSOT)
# ---------------------------------------------------------------------------
ISRM_TABLE_PLT = [
    {"indice": "R0", "min_ucs": 0.25, "max_ucs": 1.0, "denominacion": "Extremadamente débil"},
    {"indice": "R1", "min_ucs": 1.0, "max_ucs": 5.0, "denominacion": "Muy débil"},
    {"indice": "R2", "min_ucs": 5.0, "max_ucs": 25.0, "denominacion": "Débil"},
    {"indice": "R3", "min_ucs": 25.0, "max_ucs": 50.0, "denominacion": "Moderadamente resistente"},
    {"indice": "R4", "min_ucs": 50.0, "max_ucs": 100.0, "denominacion": "Resistente"},
    {"indice": "R5", "min_ucs": 100.0, "max_ucs": 250.0, "denominacion": "Muy resistente"},
    {"indice": "R6", "min_ucs": 250.0, "max_ucs": float("inf"), "denominacion": "Extremadamente resistente"},
]

# Sinónimos canónicos para mapear columnas de ambos formatos de Excel
PLT_CANONICAL_COLUMNS = {
    "campania": ["CAMPANA SRK", "CAMPANA", "CAMPANIA"],
    "fecha_ensayo": ["FECHA DE ENSAYO2", "FECHA DE ENSAYO", "FECHAENSAYO", "FECHA_ENSAYO", "FECHA"],
    "tipo_ensayo": ["TIPO DE ENSAYO", "TIPO_ENSAYO", "TIPO ENSAYO"],
    "ejecutado_por": ["EJECUTADO POR", "EJECUCION DE ENSAYO", "EJECUTADOPOR"],
    "zona_muestreo": ["ZONA DE MUESTREO", "ZONADEMUESTREO", "ZONA DE MAPEO", "ZONA_MAPEO", "ZONA"],
    "nivel": ["NIVEL"],
    "celda_mapeo": ["CELDA DE MAPEO", "CELDAMAPEO", "CELDA_MAPEO", "CELDA"],
    "muestra": ["MUESTRA"],
    "codigo_muestra": ["CODIGO DE MUESTRA", "CODIGO MUESTRA", "CODIGO_MUESTRA"],
    "litologia_1": ["LITOLOGIA 1", "LITOLOGIA_1", "LITO 1", "LITO1"],
    "litologia_2": ["LITOLOGIA 2", "LITOLOGIA_2", "LITO 2", "LITO2"],
    "litologia_3": ["LITOLOGIA 3", "LITOLOGIA_3", "LITO 3", "LITO3"],
    "tipo_litologico": ["TIPO LITOLOGICO", "TIPO_LITOLOGICO", "TIPO LITOLICO", "TIPOLITOLOGICO"],
    "este": ["ESTE (M)", "ESTE(M)", "ESTE", "EAST"],
    "norte": ["NORTE (M)", "NORTE(M)", "NORTE", "NORTH"],
    "elevacion": ["ELEVACION (MSNM)", "ELEVACION(MSNM)", "ELEVACION", "COTA", "Z"],
    "espesor_d": ["ESPESOR D (CM)", "ESPESOR D", "ESPESORD(CM)", "ESPESORD", "ESPESOR"],
    "longitud_l": ["LONGITUD L (CM)", "LONGITUD L", "LONGITUDL(CM)", "LONGITUDL", "LONGITUD"],
    "ancho_w1": ["ANCHO W1 (CM)", "ANCHO W1", "ANCHOW1(CM)", "ANCHOW1"],
    "ancho_w2": ["ANCHO W2 (CM)", "ANCHO W2", "ANCHOW2(CM)", "ANCHOW2"],
    "ancho_w": ["ANCHO W (CM)", "ANCHO W", "ANCHOW(CM)", "ANCHOW"],
    "muestra_valida_long": ["MUESTRA VALIDA - LONGITUD", "MUESTRA VALIDA LONGITUD", "VALIDA LONGITUD"],
    "muestra_valida_ancho": ["MUESTRA VALIDA - ANCHO", "MUESTRA VALIDA ANCHO", "VALIDA ANCHO"],
    "fuerza_p": ["FUERZA P (KN)", "FUERZA P", "FUERZAP(KN)", "FUERZAP", "FUERZA"],
    "direccion_rotura": ["DIRECCION DE ROTURA", "DIRECCION DE RUPTURA", "DIRECCION_ROTURA", "DIRECCION ROTURA"],
    "tipo_fractura": ["TIPO DE FRACTURA", "TIPO FRACTURA", "TIPO_FRACTURA"],
    "diametro_equiv": ["DIAMETRO EQUIVALENTE (CM)", "DIAMETRO EQUIVALENTE", "DIAMETROEQUIVALENTE"],
    "f": ["F", "FACTOR F", "FACTOR_F"],
    "is_mpa": ["IS (MPA)", "IS(MPA)", "IS"],
    "is50_mpa": ["IS(50) (MPA)", "IS50 (MPA)", "IS(50)", "IS50"],
    "factor_k": ["FACTOR DE CONVERSION K", "FACTOR DE CONVERSION", "FACTOR K", "FACTOR_K", "K"],
    "ucs_mpa": ["RCS/UCS (MPA)", "RCS (MPA)", "UCS (MPA)", "RCS", "UCS"],
    "resistencia_isrm": ["RESISTENCIA ISRM", "RESISTENCIA_ISRM", "ISRM"],
    "observaciones": ["OBSERVACIONES", "OBSERVACION", "COMENTARIOS", "COMENTARIO"],
}


def _strip_accents(text: str) -> str:
    """Remueve acentos y caracteres diacríticos."""
    if not text:
        return ""
    text = unicodedata.normalize('NFD', str(text))
    return "".join(c for c in text if unicodedata.category(c) != 'Mn')


def _norm_str(val: Any) -> str:
    """Normaliza texto eliminando acentos, saltos de línea y espacios redundantes."""
    if val is None or pd.isna(val):
        return ""
    s = _strip_accents(str(val)).strip().replace("\r", " ").replace("\n", " ")
    s = re.sub(r"\s+", " ", s).upper()
    return s


def sanitize_number(val: Any) -> Optional[float]:
    """Convierte cualquier valor a float seguro, o None si es nulo/inválido/error de fórmula."""
    if val is None or pd.isna(val):
        return None
    s = str(val).strip().replace(",", "")
    if s == "" or s.upper() in ("NONE", "NAN", "NULL", "N/A", "-", "#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#N/A"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def sanitize_date(val: Any) -> Tuple[Optional[datetime], Optional[str]]:
    """Parsea una fecha desde datetime, pandas Timestamp, string o serial numérico de Excel."""
    if val is None or pd.isna(val):
        return None, None

    if isinstance(val, (datetime, pd.Timestamp)):
        dt_val = val.to_pydatetime() if hasattr(val, "to_pydatetime") else val
        return dt_val, dt_val.strftime("%Y-%m-%d")

    if isinstance(val, (int, float)):
        try:
            dt_val = pd.to_datetime(val, unit='D', origin='1899-12-30').to_pydatetime()
            return dt_val, dt_val.strftime("%Y-%m-%d")
        except Exception:
            return None, None

    s = str(val).strip()
    if not s or s.upper() in ("NONE", "NAN", "NULL", "N/A", "-"):
        return None, None

    formats = [
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y",
        "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y", "%m/%d/%Y"
    ]
    for fmt in formats:
        try:
            dt_val = datetime.strptime(s, fmt)
            return dt_val, dt_val.strftime("%Y-%m-%d")
        except ValueError:
            pass

    return None, s


def is_formula_error(val: Any) -> bool:
    """Detecta si un valor contiene un error típico de fórmulas Excel (#VALUE!, #REF!, #DIV/0!, etc.)."""
    if val is None:
        return False
    s = str(val).strip().upper()
    return any(err in s for err in ("#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#N/A", "#NUM!", "#NULL!", "#ERR"))


def normalize_lithology_group(group_val: Any) -> str:
    """Normaliza el tipo/grupo litológico según los sinónimos geológicos permitidos."""
    norm = _norm_str(group_val)
    if not norm:
        return ""
    for canonical, synonyms in LITHOLOGY_GROUP_SYNONYMS.items():
        if norm == _norm_str(canonical) or any(norm == _norm_str(syn) for syn in synonyms):
            return canonical
    return str(group_val).strip().upper()


def resolve_expected_lithology(l1: str, l2: str, l3: str) -> Tuple[Optional[str], Optional[float]]:
    """
    Resuelve la combinación litológica según la cascada oficial de Ensayos PLT.
    Retorna (tipo_litologico_esperado, factor_k_esperado).
    """
    for item in LITHOLOGY_FULL_CATALOG:
        m1 = _norm_str(item.get("lito1")) == _norm_str(l1) if item.get("lito1") else True
        m2 = _norm_str(item.get("lito2")) == _norm_str(l2) if item.get("lito2") else True
        m3 = _norm_str(item.get("lito3")) == _norm_str(l3) if item.get("lito3") else True

        if m1 and m2 and m3:
            tipo_lito = item.get("grupo") or item.get("tipo_litologico")
            k_val = item.get("k") or item.get("factor_k")
            return tipo_lito, k_val
    return None, None


def get_expected_isrm(ucs_val: Optional[float]) -> Optional[str]:
    """Calcula el rango de resistencia ISRM esperado a partir del valor UCS (MPa)."""
    if ucs_val is None:
        return None
    for entry in ISRM_TABLE_PLT:
        if entry["min_ucs"] <= ucs_val < entry["max_ucs"]:
            return entry["indice"]
    if ucs_val >= 250.0:
        return "R6"
    return None


def normalize_sample_code(celda: str, muestra: str) -> str:
    """Formatea el código de muestra oficial (ej. 'WB1-A')."""
    return f"{celda.strip().upper()}-{muestra.strip().upper()}"


def are_sample_codes_compatible(actual: str, celda: str, muestra: str) -> bool:
    """Verifica si el código de muestra coincide con la celda y letra de muestra."""
    if not actual:
        return False
    act_norm = re.sub(r"[\s\-_]", "", str(actual).strip().upper())
    exp_norm = re.sub(r"[\s\-_]", "", f"{celda}{muestra}".strip().upper())
    return act_norm == exp_norm


def extract_cell_and_sample(codigo_muestra_raw: Any, celda_raw: Any) -> Tuple[str, str]:
    """
    Extrae (celda, muestra) de 'Código de muestra' (ej. 'QV1-A' -> 'QV1', 'A').
    Soporta separadores '-', '_', espacios o concatenación directa.
    """
    cod_str = str(codigo_muestra_raw or "").strip().upper()
    celda_str = str(celda_raw or "").strip().upper()

    if not cod_str and not celda_str:
        return "", ""

    m = re.match(r"^([A-Z0-9_-]+?)[\s\-_]*([A-F])$", cod_str)
    if m:
        c_part = m.group(1).replace("-", "").replace("_", "").strip()
        sample_letter = m.group(2)
        final_cell = celda_str if celda_str else c_part
        return final_cell, sample_letter

    return celda_str, ""


def detect_plt_file_format(file_path: str) -> Tuple[str, int]:
    """
    Autodetecta el formato del archivo Excel de Ensayos PLT.
    Retorna (FORMAT_TYPE, header_row_idx).
    - 'COMPACT_FIELD': Encabezados en fila 3 (index 2), datos desde fila 6.
    - 'STANDARD_34': Encabezados en fila 0 o 1, 34 columnas estándar.
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        sample_rows = []
        for r in range(1, 9):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, 30)]
            sample_rows.append(row_vals)
        wb.close()

        # Chequear si fila 3 (index 2) tiene 'Código de muestra' o 'Espesor'
        if len(sample_rows) >= 3:
            r3_text = " ".join([_norm_str(v) for v in sample_rows[2] if v is not None])
            if ("CODIGO DE MUESTRA" in r3_text or "MUESTRA VALIDA" in r3_text) and ("ESPESOR" in r3_text or "LITO" in r3_text):
                return "COMPACT_FIELD", 2

        # Chequear formato estándar
        for idx in range(min(4, len(sample_rows))):
            r_text = " ".join([_norm_str(v) for v in sample_rows[idx] if v is not None])
            if "CAMPANA" in r_text or "FECHA DE ENSAYO" in r_text or "LABORATORIO" in r_text:
                return "STANDARD_34", idx

        return "STANDARD_34", 1
    except Exception:
        return "STANDARD_34", 1


def detect_header_row(file_path: str, max_check: int = 10) -> int:
    """Detecta el índice de fila (0-indexed) donde residen los nombres de columna reales."""
    _, header_idx = detect_plt_file_format(file_path)
    return header_idx


# =============================================================================
# VALIDADOR 1: FORMATO ESTÁNDAR 34 COLUMNAS
# =============================================================================
def validate_plt_standard_34(
    file_path: str,
    header_row_idx: int = 1,
    tolerance: float = 0.1,
) -> dict:
    """
    Audita un archivo Excel de Ensayos PLT en formato estándar de 34 columnas.
    """
    df_raw = pd.read_excel(file_path, sheet_name=0, header=header_row_idx)

    # Mapeo dinámico de columnas del Excel a las claves canónicas
    col_map: Dict[str, str] = {}
    for canon_key, synonyms in PLT_CANONICAL_COLUMNS.items():
        found = False
        for raw_col in df_raw.columns:
            raw_norm = _norm_str(raw_col)
            if raw_norm in synonyms:
                col_map[canon_key] = raw_col
                found = True
                break
        if not found:
            for raw_col in df_raw.columns:
                raw_norm = _norm_str(raw_col)
                if any(s in raw_norm for s in synonyms if len(s) > 3):
                    col_map[canon_key] = raw_col
                    break

    incidencias: List[dict] = []
    total_filas = len(df_raw)

    # 1. Agrupamiento cronológico por (Fecha, Celda) para evaluar Integridad ABCD
    cell_groups: Dict[Tuple[str, str], List[dict]] = defaultdict(list)

    for idx, row in df_raw.iterrows():
        fila_excel = idx + header_row_idx + 2  # 1-indexed en Excel
        fecha_raw = row.get(col_map.get("fecha_ensayo"))
        _, fecha_str = sanitize_date(fecha_raw)
        celda_raw = str(row.get(col_map.get("celda_mapeo"), "") or "").strip().upper()
        muestra_raw = str(row.get(col_map.get("muestra"), "") or "").strip().upper()

        group_key = (fecha_str or "SIN_FECHA", celda_raw or f"FILA_{fila_excel}")
        cell_groups[group_key].append({
            "fila_excel": fila_excel,
            "muestra": muestra_raw,
            "row_data": row
        })

    # Resumen de celdas y evaluación de secuencia ABCD
    resumen_celdas: Dict[str, dict] = {}

    for (fecha_k, celda_k), group_rows in cell_groups.items():
        count_muestras = len(group_rows)
        muestras_secuencia = [r["muestra"] for r in group_rows]
        secuencia_str = "-".join(muestras_secuencia) if muestras_secuencia else "VACÍA"
        filas_nums = [r["fila_excel"] for r in group_rows]

        first_row = group_rows[0]["row_data"]
        camp_val = first_row.get(col_map.get("campania"))
        lito_val = normalize_lithology_group(first_row.get(col_map.get("tipo_litologico")))
        nivel_val = sanitize_number(first_row.get(col_map.get("nivel")))

        cell_summary_key = f"{celda_k} ({fecha_k})" if fecha_k != "SIN_FECHA" else celda_k

        has_formula_err = any(
            is_formula_error(r["row_data"].get(col_map.get(k)))
            for r in group_rows
            for k in ("is50_mpa", "ucs_mpa", "fuerza_p", "ancho_w")
            if col_map.get(k)
        )

        if has_formula_err:
            estado_secuencia = "ANÓMALA (#ERR FÓRMULA)"
            for r in group_rows:
                rule_obj = RULES_REGISTRY_PLT["ERR_PLT_CELDA_ANOMALA"]
                cat_obj = CATEGORIES_REGISTRY_PLT.get(rule_obj.category_code)
                incidencias.append({
                    "fila_excel": r["fila_excel"],
                    "tipo_incidencia": cat_obj.severity if cat_obj else "ALERTA",
                    "rule_code": "ERR_PLT_CELDA_ANOMALA",
                    "celda_mapeo": celda_k,
                    "campania": camp_val,
                    "columna": "Celda de mapeo",
                    "valor_actual": "#ERR",
                    "mensaje": f"Celda '{celda_k}': Contiene errores críticos de fórmula (#VALUE!, #REF!, etc.).",
                })
        elif count_muestras == 4:
            if muestras_secuencia == ["A", "B", "C", "D"]:
                estado_secuencia = "CORRECTO (A-B-C-D)"
            else:
                estado_secuencia = f"ORDEN INCORRECTO ({secuencia_str})"
                rule_obj = RULES_REGISTRY_PLT["WRN_PLT_SECUENCIA_DESORDEN"]
                cat_obj = CATEGORIES_REGISTRY_PLT.get(rule_obj.category_code)
                incidencias.append({
                    "fila_excel": filas_nums[0],
                    "tipo_incidencia": cat_obj.severity if cat_obj else "ADVERTENCIA",
                    "rule_code": "WRN_PLT_SECUENCIA_DESORDEN",
                    "celda_mapeo": celda_k,
                    "campania": camp_val,
                    "columna": "Muestra",
                    "valor_actual": secuencia_str,
                    "mensaje": rule_obj.format_message(celda=celda_k, secuencia=secuencia_str),
                })
        elif count_muestras < 4:
            estado_secuencia = f"INCOMPLETA ({count_muestras}/4: {secuencia_str})"
            rule_obj = RULES_REGISTRY_PLT["WRN_PLT_CELDA_INCOMPLETA"]
            cat_obj = CATEGORIES_REGISTRY_PLT.get(rule_obj.category_code)
            incidencias.append({
                "fila_excel": filas_nums[0],
                "tipo_incidencia": cat_obj.severity if cat_obj else "ADVERTENCIA",
                "rule_code": "WRN_PLT_CELDA_INCOMPLETA",
                "celda_mapeo": celda_k,
                "campania": camp_val,
                "columna": "Celda de mapeo",
                "valor_actual": secuencia_str,
                "mensaje": rule_obj.format_message(celda=celda_k, count=count_muestras, secuencia=secuencia_str),
            })
        else:
            estado_secuencia = f"EXCEDENTE ({count_muestras}/4: {secuencia_str})"
            rule_obj = RULES_REGISTRY_PLT["WRN_PLT_CELDA_EXCEDENTE"]
            cat_obj = CATEGORIES_REGISTRY_PLT.get(rule_obj.category_code)
            incidencias.append({
                "fila_excel": filas_nums[0],
                "tipo_incidencia": cat_obj.severity if cat_obj else "ADVERTENCIA",
                "rule_code": "WRN_PLT_CELDA_EXCEDENTE",
                "celda_mapeo": celda_k,
                "campania": camp_val,
                "columna": "Celda de mapeo",
                "valor_actual": secuencia_str,
                "mensaje": rule_obj.format_message(celda=celda_k, count=count_muestras, secuencia=secuencia_str),
            })

        resumen_celdas[cell_summary_key] = {
            "celda": celda_k,
            "fecha": fecha_k,
            "campania": camp_val,
            "tipo_litologico": lito_val,
            "nivel": nivel_val,
            "total_muestras": count_muestras,
            "secuencia": secuencia_str,
            "estado_secuencia": estado_secuencia,
            "filas": filas_nums,
            "alertas": 0,
            "advertencias": 0,
            "vacios": 0,
        }

    # 2. Validación fila a fila
    valid_litos_set = set(
        [_norm_str(x.get("lito1")) for x in LITHOLOGY_FULL_CATALOG if x.get("lito1")]
        + [_norm_str(x.get("lito2")) for x in LITHOLOGY_FULL_CATALOG if x.get("lito2")]
        + [_norm_str(x.get("lito3")) for x in LITHOLOGY_FULL_CATALOG if x.get("lito3")]
    )

    for idx, row in df_raw.iterrows():
        fila_excel = idx + header_row_idx + 2

        def get_val(key: str) -> Any:
            col_name = col_map.get(key)
            if col_name and col_name in row:
                v = row[col_name]
                return None if pd.isna(v) else v
            return None

        celda_raw = str(get_val("celda_mapeo") or "").strip().upper()
        camp_raw = get_val("campania")
        fecha_dt, fecha_str = sanitize_date(get_val("fecha_ensayo"))
        cell_key = f"{celda_raw} ({fecha_str})" if fecha_str else celda_raw

        def reg_err(col_key_name: str, val_actual: Any, rule_code: str, **kwargs):
            rule_obj = RULES_REGISTRY_PLT.get(rule_code)
            if not rule_obj:
                return
            cat_obj = CATEGORIES_REGISTRY_PLT.get(rule_obj.category_code)
            sev = cat_obj.severity if cat_obj else "ALERTA"
            msg = rule_obj.format_message(**kwargs)

            display_col = col_map.get(col_key_name, col_key_name)

            incidencias.append({
                "fila_excel": fila_excel,
                "tipo_incidencia": sev,
                "rule_code": rule_code,
                "celda_mapeo": celda_raw,
                "campania": camp_raw,
                "columna": display_col,
                "valor_actual": str(val_actual) if val_actual is not None else "—",
                "mensaje": msg,
            })

            if cell_key in resumen_celdas:
                if sev == "ALERTA":
                    resumen_celdas[cell_key]["alertas"] += 1
                elif sev == "ADVERTENCIA":
                    resumen_celdas[cell_key]["advertencias"] += 1
                elif sev == "VACIO":
                    resumen_celdas[cell_key]["vacios"] += 1

        # --- GRUPO 1: Metadata Administrativa y Proyecto ---
        if camp_raw is None or str(camp_raw).strip() == "":
            reg_err("campania", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Campaña")
        else:
            try:
                c_int = int(float(str(camp_raw).strip()))
                if not (2000 <= c_int <= 2050):
                    reg_err("campania", camp_raw, "ERR_PLT_CAMPANIA_RANGO", value=camp_raw)
            except (ValueError, TypeError):
                reg_err("campania", camp_raw, "ERR_PLT_CAMPANIA_RANGO", value=camp_raw)

        if not fecha_dt:
            reg_err("fecha_ensayo", get_val("fecha_ensayo"), "ERR_PLT_FORMATO_FECHA_INVALIDO", value=get_val("fecha_ensayo"))

        for col_req in ["tipo_ensayo", "ejecutado_por", "zona_muestreo"]:
            v_req = get_val(col_req)
            if v_req is None or str(v_req).strip() == "":
                reg_err(col_req, None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name=col_map.get(col_req, col_req))

        nivel_num = sanitize_number(get_val("nivel"))
        if nivel_num is None:
            reg_err("nivel", get_val("nivel"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Nivel")
        elif nivel_num < 0:
            reg_err("nivel", nivel_num, "ERR_PLT_NIVEL_RANGO", value=nivel_num)

        # --- GRUPO 2: Identificación Geomecánica y Muestras ---
        if not celda_raw:
            reg_err("celda_mapeo", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Celda de mapeo")

        muestra_raw = _norm_str(get_val("muestra"))
        if not muestra_raw:
            reg_err("muestra", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Muestra")
        elif muestra_raw not in ["A", "B", "C", "D"]:
            reg_err("muestra", muestra_raw, "ERR_PLT_MUESTRA_LETRA_INVALIDA", value=muestra_raw)

        cod_muestra_raw = get_val("codigo_muestra")
        if cod_muestra_raw is None or str(cod_muestra_raw).strip() == "":
            reg_err("codigo_muestra", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Código de muestra")
        elif celda_raw and muestra_raw in ["A", "B", "C", "D"]:
            if not are_sample_codes_compatible(str(cod_muestra_raw), celda_raw, muestra_raw):
                reg_err("codigo_muestra", cod_muestra_raw, "ERR_PLT_CODIGO_MUESTRA_INCONGRUENTE",
                        actual=cod_muestra_raw, celda=celda_raw, muestra=muestra_raw)

        # --- GRUPO 3: Clasificación Litológica y Factor K ---
        l1 = _norm_str(get_val("litologia_1"))
        l2 = _norm_str(get_val("litologia_2"))
        l3 = _norm_str(get_val("litologia_3"))
        tipo_lito_raw = _norm_str(get_val("tipo_litologico"))
        tipo_lito_norm = normalize_lithology_group(tipo_lito_raw)

        if not l1:
            reg_err("litologia_1", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Litología 1")
        elif l1 not in valid_litos_set:
            reg_err("litologia_1", l1, "ERR_PLT_LITO1_NO_RECONOCIDO", value=l1)

        if not l2:
            reg_err("litologia_2", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Litología 2")
        elif l2 not in valid_litos_set:
            reg_err("litologia_2", l2, "ERR_PLT_LITO2_NO_RECONOCIDO", value=l2)

        if not l3:
            reg_err("litologia_3", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Litología 3")
        elif l3 not in valid_litos_set:
            reg_err("litologia_3", l3, "ERR_PLT_LITO3_NO_RECONOCIDO", value=l3)

        exp_tipo, exp_k = resolve_expected_lithology(l1, l2, l3)
        if l1 and l2 and l3:
            if not exp_tipo:
                reg_err("litologia_1", f"{l1}/{l2}/{l3}", "ERR_PLT_COMBINACION_LITOLOGICA_NO_VALIDA",
                        lito1=l1, lito2=l2, lito3=l3)
            else:
                if not tipo_lito_raw:
                    reg_err("tipo_litologico", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Tipo litológico")
                elif tipo_lito_norm != normalize_lithology_group(exp_tipo):
                    reg_err("tipo_litologico", tipo_lito_raw, "ERR_PLT_TIPO_LITOLOGICO_INCONGRUENTE",
                            actual=tipo_lito_raw, expected=exp_tipo)

        # --- GRUPO 4: Coordenadas Espaciales WGS84 ---
        este_num = sanitize_number(get_val("este"))
        norte_num = sanitize_number(get_val("norte"))
        elev_num = sanitize_number(get_val("elevacion"))

        if este_num is None:
            reg_err("este", get_val("este"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Este (m)")
        elif not (100000.0 <= este_num <= 9999999.0):
            reg_err("este", este_num, "ERR_PLT_ESTE_RANGO", value=este_num)

        if norte_num is None:
            reg_err("norte", get_val("norte"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Norte (m)")
        elif not (1000000.0 <= norte_num <= 99999999.0):
            reg_err("norte", norte_num, "ERR_PLT_NORTE_RANGO", value=norte_num)

        if elev_num is None:
            reg_err("elevacion", get_val("elevacion"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Elevación (msnm)")
        elif not (0.0 <= elev_num <= 6000.0):
            reg_err("elevacion", elev_num, "ERR_PLT_ELEVACION_RANGO", value=elev_num)

        # --- GRUPO 5: Geometría de Probeta y Criterios de Validez ---
        d_num = sanitize_number(get_val("espesor_d"))
        l_num = sanitize_number(get_val("longitud_l"))
        w1_num = sanitize_number(get_val("ancho_w1"))
        w2_num = sanitize_number(get_val("ancho_w2"))
        w_num = sanitize_number(get_val("ancho_w"))

        if d_num is None:
            reg_err("espesor_d", get_val("espesor_d"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Espesor D (cm)")
        elif not (1.0 <= d_num <= 20.0):
            reg_err("espesor_d", d_num, "ERR_PLT_ESPESOR_D_RANGO", value=d_num)

        if l_num is None:
            reg_err("longitud_l", get_val("longitud_l"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Longitud L (cm)")
        elif not (1.0 <= l_num <= 50.0):
            reg_err("longitud_l", l_num, "ERR_PLT_LONGITUD_L_RANGO", value=l_num)

        if w1_num is None:
            reg_err("ancho_w1", get_val("ancho_w1"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Ancho W1 (cm)")
        elif not (1.0 <= w1_num <= 30.0):
            reg_err("ancho_w1", w1_num, "ERR_PLT_ANCHO_W1_RANGO", value=w1_num)

        if w2_num is None:
            reg_err("ancho_w2", get_val("ancho_w2"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Ancho W2 (cm)")
        elif not (1.0 <= w2_num <= 30.0):
            reg_err("ancho_w2", w2_num, "ERR_PLT_ANCHO_W2_RANGO", value=w2_num)

        calc_w = None
        if w1_num is not None and w2_num is not None:
            calc_w = round((w1_num + w2_num) / 2.0, 3)
            if w_num is None:
                reg_err("ancho_w", get_val("ancho_w"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Ancho W (cm)")
            elif abs(w_num - calc_w) > tolerance:
                reg_err("ancho_w", w_num, "ERR_PLT_ANCHO_W_INCONGRUENTE", actual=w_num, expected=calc_w)
            elif not (1.0 <= w_num <= 30.0):
                reg_err("ancho_w", w_num, "ERR_PLT_ANCHO_W_RANGO", value=w_num)

        effective_w = w_num if w_num is not None else calc_w

        valida_l_raw = get_val("muestra_valida_long")
        valida_l_norm = _norm_str(valida_l_raw)
        if d_num is not None and l_num is not None:
            exp_val_l = "SI" if l_num >= d_num else "NO"
            if not valida_l_norm:
                reg_err("muestra_valida_long", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Muestra válida - Longitud")
            elif valida_l_norm != exp_val_l:
                reg_err("muestra_valida_long", valida_l_raw, "ERR_PLT_MUESTRA_VALIDA_LONG_INCONGRUENTE",
                        actual=valida_l_raw, expected=exp_val_l)

        valida_w_raw = get_val("muestra_valida_ancho")
        valida_w_norm = _norm_str(valida_w_raw)
        if d_num is not None and effective_w is not None:
            exp_val_w = "SI" if (0.3 * effective_w < d_num < effective_w) else "NO"
            if not valida_w_norm:
                reg_err("muestra_valida_ancho", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Muestra válida - Ancho")
            elif valida_w_norm != exp_val_w:
                reg_err("muestra_valida_ancho", valida_w_raw, "ERR_PLT_MUESTRA_VALIDA_ANCHO_INCONGRUENTE",
                        actual=valida_w_raw, expected=exp_val_w)

        # --- GRUPO 6: Ensayo Físico, Fractura y Parámetros PLT ---
        fuerza_p = sanitize_number(get_val("fuerza_p"))
        dir_rot = _norm_str(get_val("direccion_rotura"))
        tipo_frac = _norm_str(get_val("tipo_fractura"))
        de_num = sanitize_number(get_val("diametro_equiv"))
        f_num = sanitize_number(get_val("f"))
        is_num = sanitize_number(get_val("is_mpa"))
        is50_num = sanitize_number(get_val("is50_mpa"))

        if fuerza_p is None:
            reg_err("fuerza_p", get_val("fuerza_p"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Fuerza P (kN)")
        elif not (0.01 <= fuerza_p <= 200.0):
            reg_err("fuerza_p", fuerza_p, "ERR_PLT_FUERZA_P_RANGO", value=fuerza_p)

        if not dir_rot:
            reg_err("direccion_rotura", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Dirección de rotura")
        elif dir_rot not in [x.upper() for x in CAT_DIRECCION_ROTURA]:
            reg_err("direccion_rotura", dir_rot, "ERR_PLT_DIRECCION_ROTURA_CATALOGO", value=dir_rot)

        if not tipo_frac:
            reg_err("tipo_fractura", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Tipo de fractura")
        elif tipo_frac not in [x.upper() for x in CAT_TIPO_FRACTURA]:
            reg_err("tipo_fractura", tipo_frac, "ERR_PLT_TIPO_FRACTURA_CATALOGO", value=tipo_frac)

        calc_de = None
        if effective_w is not None and d_num is not None and (effective_w * d_num) > 0:
            calc_de = round(math.sqrt((4.0 * effective_w * d_num) / math.pi), 2)
            if de_num is None:
                reg_err("diametro_equiv", get_val("diametro_equiv"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Diámetro equivalente (cm)")
            elif abs(de_num - calc_de) > tolerance:
                reg_err("diametro_equiv", de_num, "ERR_PLT_DIAMETRO_EQUIV_INCONGRUENTE", actual=de_num, expected=calc_de)

        effective_de = de_num if de_num is not None else calc_de

        calc_f = None
        if effective_de is not None and effective_de > 0:
            if 4.0 <= effective_de <= 6.0:
                calc_f = round(math.sqrt(effective_de / 5.0), 4)
            else:
                calc_f = round(math.pow(effective_de / 5.0, 0.45), 4)

            if f_num is None:
                reg_err("f", get_val("f"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Factor F")
            elif abs(f_num - calc_f) > tolerance:
                reg_err("f", f_num, "ERR_PLT_FACTOR_F_INCONGRUENTE", actual=f_num, expected=calc_f)

        calc_is = None
        if fuerza_p is not None and effective_de is not None and effective_de > 0:
            calc_is = round((fuerza_p * 1000.0) / math.pow(effective_de * 10.0, 2), 4)
            if is_num is None:
                reg_err("is_mpa", get_val("is_mpa"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Is (MPa)")
            elif abs(is_num - calc_is) > tolerance:
                reg_err("is_mpa", is_num, "ERR_PLT_IS_INCONGRUENTE", actual=is_num, expected=calc_is)

        effective_is = is_num if is_num is not None else calc_is
        effective_f = f_num if f_num is not None else calc_f
        if valida_w_norm == "SI" and effective_is is not None and effective_f is not None:
            calc_is50 = round(effective_is * effective_f, 4)
            if is50_num is None:
                reg_err("is50_mpa", get_val("is50_mpa"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Is(50) (MPa)")
            elif abs(is50_num - calc_is50) > tolerance:
                reg_err("is50_mpa", is50_num, "ERR_PLT_IS50_INCONGRUENTE", actual=is50_num, expected=calc_is50)
        elif valida_w_norm == "NO" and is50_num is not None:
            reg_err("is50_mpa", is50_num, "ERR_PLT_IS50_INCONGRUENTE", actual=is50_num, expected="NULL (Muestra Inválida)")

        # --- GRUPO 7: Resistencia de Roca (K, UCS, ISRM) ---
        factor_k_num = sanitize_number(get_val("factor_k"))
        ucs_num = sanitize_number(get_val("ucs_mpa"))
        isrm_raw = _norm_str(get_val("resistencia_isrm"))

        if factor_k_num is None:
            reg_err("factor_k", get_val("factor_k"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Factor de conversión K")
        else:
            if not (5.0 <= factor_k_num <= 30.0):
                reg_err("factor_k", factor_k_num, "ERR_PLT_FACTOR_K_RANGO", value=factor_k_num)
            elif exp_k is not None and abs(factor_k_num - exp_k) > tolerance:
                reg_err("factor_k", factor_k_num, "ERR_PLT_FACTOR_K_INCONGRUENTE", actual=factor_k_num, expected=exp_k)

        effective_k = factor_k_num if factor_k_num is not None else exp_k
        effective_is50 = is50_num if is50_num is not None else (calc_is50 if valida_w_norm == "SI" and 'calc_is50' in locals() else None)

        calc_ucs = None
        if effective_is50 is not None and effective_k is not None:
            calc_ucs = round(effective_is50 * effective_k, 3)
            if ucs_num is None:
                reg_err("ucs_mpa", get_val("ucs_mpa"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="RCS/UCS (MPa)")
            elif abs(ucs_num - calc_ucs) > tolerance:
                reg_err("ucs_mpa", ucs_num, "ERR_PLT_UCS_INCONGRUENTE", actual=ucs_num, expected=calc_ucs)

        effective_ucs = ucs_num if ucs_num is not None else calc_ucs
        if effective_ucs is not None:
            exp_isrm = get_expected_isrm(effective_ucs)
            if not isrm_raw:
                reg_err("resistencia_isrm", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Resistencia ISRM")
            elif exp_isrm and isrm_raw != exp_isrm:
                reg_err("resistencia_isrm", isrm_raw, "ERR_PLT_RESISTENCIA_ISRM_INCONGRUENTE",
                        actual=isrm_raw, ucs_val=round(effective_ucs, 2), expected=exp_isrm)

    return {
        "formato_detectado": "FORMAT_STANDARD_34",
        "total_filas_procesadas": total_filas,
        "incidencias": incidencias,
        "resumen_por_celda": resumen_celdas,
    }


# =============================================================================
# VALIDADOR 2: FORMATO DE CAMPO COMPACTO (03 feb1.xlsx)
# =============================================================================
def validate_plt_compact_field(
    file_path: str,
    header_row_idx: int = 2,
    tolerance: float = 0.1,
) -> dict:
    """
    Valida planillas de campo compactas de Ensayos PLT (ej. 03 feb1.xlsx).
    Encabezados en fila 3 (index 2), datos desde fila 6 (index 5).
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    rows_data: List[dict] = []
    current_cell = ""

    for r in range(6, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 30)]
        if not any(v is not None for v in row_vals):
            continue

        celda_raw = row_vals[0]
        cod_muestra_raw = row_vals[1]

        if celda_raw and str(celda_raw).strip():
            current_cell = str(celda_raw).strip().upper()

        extracted_cell, extracted_sample = extract_cell_and_sample(cod_muestra_raw, current_cell)
        if extracted_cell:
            current_cell = extracted_cell

        rows_data.append({
            "fila_excel": r,
            "celda": current_cell,
            "muestra": extracted_sample,
            "cod_muestra": cod_muestra_raw,
            "lito1": row_vals[2],
            "lito2": row_vals[3],
            "lito3": row_vals[4],
            "este": row_vals[5],
            "norte": row_vals[6],
            "cota": row_vals[7],
            "d": row_vals[8],
            "l": row_vals[9],
            "w1": row_vals[10],
            "w2": row_vals[11],
            "w": row_vals[12],
            "val_l": row_vals[13],
            "val_w": row_vals[14],
            "p": row_vals[15],
            "dir": row_vals[16],
            "frac": row_vals[17],
            "de": row_vals[18],
            "f": row_vals[19],
            "is": row_vals[20],
            "is50": row_vals[21],
            "ucs": row_vals[22],
            "isrm": row_vals[23],
            "obs": row_vals[24] if len(row_vals) > 24 else None,
        })

    incidencias: List[dict] = []
    total_filas = len(rows_data)

    # 1. Agrupamiento por Celda para evaluar Secuencia ABCD
    cell_groups: Dict[str, List[dict]] = defaultdict(list)
    for rd in rows_data:
        cell_groups[rd["celda"] or f"FILA_{rd['fila_excel']}"].append(rd)

    resumen_celdas: Dict[str, dict] = {}
    for celda_k, group_rows in cell_groups.items():
        count_muestras = len(group_rows)
        muestras_secuencia = [r["muestra"] for r in group_rows]
        secuencia_str = "-".join(muestras_secuencia) if muestras_secuencia else "VACÍA"
        filas_nums = [r["fila_excel"] for r in group_rows]

        first_row = group_rows[0]
        lito_val = str(first_row.get("lito1") or "").strip().upper()

        has_formula_err = any(
            is_formula_error(r.get(k))
            for r in group_rows
            for k in ("is50", "ucs", "p", "w")
        )

        if has_formula_err:
            estado_secuencia = "ANÓMALA (#ERR FÓRMULA)"
            for r in group_rows:
                rule_obj = RULES_REGISTRY_PLT["ERR_PLT_CELDA_ANOMALA"]
                cat_obj = CATEGORIES_REGISTRY_PLT.get(rule_obj.category_code)
                incidencias.append({
                    "fila_excel": r["fila_excel"],
                    "tipo_incidencia": cat_obj.severity if cat_obj else "ALERTA",
                    "rule_code": "ERR_PLT_CELDA_ANOMALA",
                    "celda_mapeo": celda_k,
                    "campania": "N/A",
                    "columna": "Celda de mapeo",
                    "valor_actual": "#ERR",
                    "mensaje": f"Celda '{celda_k}': Contiene errores críticos de fórmula (#VALUE!, #REF!, etc.).",
                })
        elif count_muestras == 4:
            if muestras_secuencia == ["A", "B", "C", "D"]:
                estado_secuencia = "CORRECTO (A-B-C-D)"
            else:
                estado_secuencia = f"ORDEN INCORRECTO ({secuencia_str})"
                rule_obj = RULES_REGISTRY_PLT["WRN_PLT_SECUENCIA_DESORDEN"]
                cat_obj = CATEGORIES_REGISTRY_PLT.get(rule_obj.category_code)
                incidencias.append({
                    "fila_excel": filas_nums[0],
                    "tipo_incidencia": cat_obj.severity if cat_obj else "ADVERTENCIA",
                    "rule_code": "WRN_PLT_SECUENCIA_DESORDEN",
                    "celda_mapeo": celda_k,
                    "campania": "N/A",
                    "columna": "Código de muestra",
                    "valor_actual": secuencia_str,
                    "mensaje": rule_obj.format_message(celda=celda_k, secuencia=secuencia_str),
                })
        elif count_muestras < 4:
            estado_secuencia = f"INCOMPLETA ({count_muestras}/4: {secuencia_str})"
            rule_obj = RULES_REGISTRY_PLT["WRN_PLT_CELDA_INCOMPLETA"]
            cat_obj = CATEGORIES_REGISTRY_PLT.get(rule_obj.category_code)
            incidencias.append({
                "fila_excel": filas_nums[0],
                "tipo_incidencia": cat_obj.severity if cat_obj else "ADVERTENCIA",
                "rule_code": "WRN_PLT_CELDA_INCOMPLETA",
                "celda_mapeo": celda_k,
                "campania": "N/A",
                "columna": "Código de muestra",
                "valor_actual": secuencia_str,
                "mensaje": rule_obj.format_message(celda=celda_k, count=count_muestras, secuencia=secuencia_str),
            })
        else:
            estado_secuencia = f"EXCEDENTE ({count_muestras}/4: {secuencia_str})"
            rule_obj = RULES_REGISTRY_PLT["WRN_PLT_CELDA_EXCEDENTE"]
            cat_obj = CATEGORIES_REGISTRY_PLT.get(rule_obj.category_code)
            incidencias.append({
                "fila_excel": filas_nums[0],
                "tipo_incidencia": cat_obj.severity if cat_obj else "ADVERTENCIA",
                "rule_code": "WRN_PLT_CELDA_EXCEDENTE",
                "celda_mapeo": celda_k,
                "campania": "N/A",
                "columna": "Código de muestra",
                "valor_actual": secuencia_str,
                "mensaje": rule_obj.format_message(celda=celda_k, count=count_muestras, secuencia=secuencia_str),
            })

        resumen_celdas[celda_k] = {
            "celda": celda_k,
            "fecha": "N/A",
            "campania": "N/A",
            "tipo_litologico": lito_val,
            "nivel": "N/A",
            "total_muestras": count_muestras,
            "secuencia": secuencia_str,
            "estado_secuencia": estado_secuencia,
            "filas": filas_nums,
            "alertas": 0,
            "advertencias": 0,
            "vacios": 0,
        }

    # 2. Validación Fila a Fila sobre columnas presentes en Formato de Campo
    valid_litos_set = set(
        [_norm_str(x.get("lito1")) for x in LITHOLOGY_FULL_CATALOG if x.get("lito1")]
        + [_norm_str(x.get("lito2")) for x in LITHOLOGY_FULL_CATALOG if x.get("lito2")]
        + [_norm_str(x.get("lito3")) for x in LITHOLOGY_FULL_CATALOG if x.get("lito3")]
    )

    for rd in rows_data:
        fila_excel = rd["fila_excel"]
        celda_act = rd["celda"]
        muestra_act = rd["muestra"]
        cod_muestra_act = rd["cod_muestra"]

        def reg_err_comp(col_name_display: str, val_actual: Any, rule_code: str, **kwargs):
            rule_obj = RULES_REGISTRY_PLT.get(rule_code)
            if not rule_obj:
                return
            cat_obj = CATEGORIES_REGISTRY_PLT.get(rule_obj.category_code)
            sev = cat_obj.severity if cat_obj else "ALERTA"
            msg = rule_obj.format_message(**kwargs)

            incidencias.append({
                "fila_excel": fila_excel,
                "tipo_incidencia": sev,
                "rule_code": rule_code,
                "celda_mapeo": celda_act,
                "campania": "N/A",
                "columna": col_name_display,
                "valor_actual": str(val_actual) if val_actual is not None else "—",
                "mensaje": msg,
            })

            if celda_act in resumen_celdas:
                if sev == "ALERTA":
                    resumen_celdas[celda_act]["alertas"] += 1
                elif sev == "ADVERTENCIA":
                    resumen_celdas[celda_act]["advertencias"] += 1
                elif sev == "VACIO":
                    resumen_celdas[celda_act]["vacios"] += 1

        # Celda y Muestra
        if not celda_act:
            reg_err_comp("Celda de mapeo", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Celda de mapeo")

        if not muestra_act:
            reg_err_comp("Código de muestra", cod_muestra_act, "ERR_PLT_MUESTRA_LETRA_INVALIDA", value=str(cod_muestra_act))
        elif muestra_act not in ["A", "B", "C", "D"]:
            reg_err_comp("Código de muestra", muestra_act, "ERR_PLT_MUESTRA_LETRA_INVALIDA", value=muestra_act)

        if not cod_muestra_act:
            reg_err_comp("Código de muestra", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Código de muestra")
        elif celda_act and muestra_act:
            if not are_sample_codes_compatible(str(cod_muestra_act), celda_act, muestra_act):
                reg_err_comp("Código de muestra", cod_muestra_act, "ERR_PLT_CODIGO_MUESTRA_INCONGRUENTE",
                             actual=cod_muestra_act, celda=celda_act, muestra=muestra_act)

        # Litologías 1, 2, 3
        lito1_val = _norm_str(rd.get("lito1"))
        lito2_val = _norm_str(rd.get("lito2"))
        lito3_val = _norm_str(rd.get("lito3"))

        if not lito1_val:
            reg_err_comp("Lito 1", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Lito 1")
        elif lito1_val not in valid_litos_set:
            reg_err_comp("Lito 1", lito1_val, "ERR_PLT_LITO1_NO_RECONOCIDO", value=lito1_val)

        if not lito2_val:
            reg_err_comp("Lito 2", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Lito 2")
        elif lito2_val not in valid_litos_set:
            reg_err_comp("Lito 2", lito2_val, "ERR_PLT_LITO2_NO_RECONOCIDO", value=lito2_val)

        if not lito3_val:
            reg_err_comp("Lito 3", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Lito 3")
        elif lito3_val not in valid_litos_set:
            reg_err_comp("Lito 3", lito3_val, "ERR_PLT_LITO3_NO_RECONOCIDO", value=lito3_val)

        exp_tipo, exp_k = resolve_expected_lithology(lito1_val, lito2_val, lito3_val)
        if lito1_val and lito2_val and lito3_val and not exp_tipo:
            reg_err_comp("Lito 1", f"{lito1_val}/{lito2_val}/{lito3_val}", "ERR_PLT_COMBINACION_LITOLOGICA_NO_VALIDA",
                         lito1=lito1_val, lito2=lito2_val, lito3=lito3_val)

        # Coordenadas
        este_num = sanitize_number(rd.get("este"))
        if este_num is None:
            reg_err_comp("Este", rd.get("este"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Este")
        elif not (100000.0 <= este_num <= 9999999.0):
            reg_err_comp("Este", este_num, "ERR_PLT_ESTE_RANGO", value=este_num)

        norte_num = sanitize_number(rd.get("norte"))
        if norte_num is None:
            reg_err_comp("Norte", rd.get("norte"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Norte")
        elif not (1000000.0 <= norte_num <= 99999999.0):
            reg_err_comp("Norte", norte_num, "ERR_PLT_NORTE_RANGO", value=norte_num)

        cota_num = sanitize_number(rd.get("cota"))
        if cota_num is None:
            reg_err_comp("Elevación", rd.get("cota"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Elevación")
        elif not (0.0 <= cota_num <= 6000.0):
            reg_err_comp("Elevación", cota_num, "ERR_PLT_ELEVACION_RANGO", value=cota_num)

        # Dimensiones D, L, W1, W2, W
        d_num = sanitize_number(rd.get("d"))
        if d_num is None:
            reg_err_comp("Espesor D (cm)", rd.get("d"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Espesor D (cm)")
        elif not (1.0 <= d_num <= 20.0):
            reg_err_comp("Espesor D (cm)", d_num, "ERR_PLT_ESPESOR_D_RANGO", value=d_num)

        l_num = sanitize_number(rd.get("l"))
        if l_num is None:
            reg_err_comp("Longitud L (cm)", rd.get("l"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Longitud L (cm)")
        elif not (1.0 <= l_num <= 50.0):
            reg_err_comp("Longitud L (cm)", l_num, "ERR_PLT_LONGITUD_L_RANGO", value=l_num)

        w1_num = sanitize_number(rd.get("w1"))
        if w1_num is None:
            reg_err_comp("Ancho W1 (cm)", rd.get("w1"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Ancho W1 (cm)")
        elif not (1.0 <= w1_num <= 30.0):
            reg_err_comp("Ancho W1 (cm)", w1_num, "ERR_PLT_ANCHO_W1_RANGO", value=w1_num)

        w2_num = sanitize_number(rd.get("w2"))
        if w2_num is None:
            reg_err_comp("Ancho W2 (cm)", rd.get("w2"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Ancho W2 (cm)")
        elif not (1.0 <= w2_num <= 30.0):
            reg_err_comp("Ancho W2 (cm)", w2_num, "ERR_PLT_ANCHO_W2_RANGO", value=w2_num)

        w_num = sanitize_number(rd.get("w"))
        calc_w = None
        if w1_num is not None and w2_num is not None:
            calc_w = round((w1_num + w2_num) / 2.0, 3)
            if w_num is None:
                reg_err_comp("Ancho W (cm)", rd.get("w"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Ancho W (cm)")
            elif abs(w_num - calc_w) > tolerance:
                reg_err_comp("Ancho W (cm)", w_num, "ERR_PLT_ANCHO_W_INCONGRUENTE", actual=w_num, expected=calc_w)
            elif not (1.0 <= w_num <= 30.0):
                reg_err_comp("Ancho W (cm)", w_num, "ERR_PLT_ANCHO_W_RANGO", value=w_num)

        effective_w = w_num if w_num is not None else calc_w

        valida_l_raw = rd.get("val_l")
        valida_l_norm = _norm_str(valida_l_raw)
        if d_num is not None and l_num is not None:
            exp_val_l = "SI" if l_num >= d_num else "NO"
            if not valida_l_norm:
                reg_err_comp("Muestra válida - longitud", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Muestra válida - longitud")
            elif valida_l_norm != exp_val_l:
                reg_err_comp("Muestra válida - longitud", valida_l_raw, "ERR_PLT_MUESTRA_VALIDA_LONG_INCONGRUENTE",
                             actual=valida_l_raw, expected=exp_val_l)

        valida_w_raw = rd.get("val_w")
        valida_w_norm = _norm_str(valida_w_raw)
        if d_num is not None and effective_w is not None:
            exp_val_w = "SI" if (0.3 * effective_w < d_num < effective_w) else "NO"
            if not valida_w_norm:
                reg_err_comp("Muestra válida - ancHP", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Muestra válida - ancHP")
            elif valida_w_norm != exp_val_w:
                reg_err_comp("Muestra válida - ancHP", valida_w_raw, "ERR_PLT_MUESTRA_VALIDA_ANCHO_INCONGRUENTE",
                             actual=valida_w_raw, expected=exp_val_w)

        # Fuerza P, Dirección de Rotura y Fractura
        p_num = sanitize_number(rd.get("p"))
        if p_num is None:
            reg_err_comp("Fuerza P (kN)", rd.get("p"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Fuerza P (kN)")
        elif not (0.01 <= p_num <= 200.0):
            reg_err_comp("Fuerza P (kN)", p_num, "ERR_PLT_FUERZA_P_RANGO", value=p_num)

        dir_rot = _norm_str(rd.get("dir"))
        if not dir_rot:
            reg_err_comp("Dirección de rotura", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Dirección de rotura")
        elif dir_rot not in [x.upper() for x in CAT_DIRECCION_ROTURA]:
            reg_err_comp("Dirección de rotura", dir_rot, "ERR_PLT_DIRECCION_ROTURA_CATALOGO", value=dir_rot)

        tipo_frac = _norm_str(rd.get("frac"))
        if not tipo_frac:
            reg_err_comp("Tipo de fractura", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Tipo de fractura")
        elif tipo_frac not in [x.upper() for x in CAT_TIPO_FRACTURA]:
            reg_err_comp("Tipo de fractura", tipo_frac, "ERR_PLT_TIPO_FRACTURA_CATALOGO", value=tipo_frac)

        # Diámetro Equivalente De
        de_num = sanitize_number(rd.get("de"))
        calc_de = None
        if effective_w is not None and d_num is not None and (effective_w * d_num) > 0:
            calc_de = round(math.sqrt((4.0 * effective_w * d_num) / math.pi), 2)
            if de_num is None:
                reg_err_comp("Diámetro equiv. (cm)", rd.get("de"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Diámetro equiv. (cm)")
            elif abs(de_num - calc_de) > tolerance:
                reg_err_comp("Diámetro equiv. (cm)", de_num, "ERR_PLT_DIAMETRO_EQUIV_INCONGRUENTE", actual=de_num, expected=calc_de)

        effective_de = de_num if de_num is not None else calc_de

        # Factor F
        f_num = sanitize_number(rd.get("f"))
        calc_f = None
        if effective_de is not None and effective_de > 0:
            if 4.0 <= effective_de <= 6.0:
                calc_f = round(math.sqrt(effective_de / 5.0), 4)
            else:
                calc_f = round(math.pow(effective_de / 5.0, 0.45), 4)

            if f_num is None:
                reg_err_comp("F", rd.get("f"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="F")
            elif abs(f_num - calc_f) > tolerance:
                reg_err_comp("F", f_num, "ERR_PLT_FACTOR_F_INCONGRUENTE", actual=f_num, expected=calc_f)

        effective_f = f_num if f_num is not None else calc_f

        # Is e Is(50)
        is_num = sanitize_number(rd.get("is"))
        calc_is = None
        if p_num is not None and effective_de is not None and effective_de > 0:
            calc_is = round((p_num * 1000.0) / math.pow(effective_de * 10.0, 2), 4)
            if is_num is None:
                reg_err_comp("Is", rd.get("is"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Is")
            elif abs(is_num - calc_is) > tolerance:
                reg_err_comp("Is", is_num, "ERR_PLT_IS_INCONGRUENTE", actual=is_num, expected=calc_is)

        effective_is = is_num if is_num is not None else calc_is

        is50_num = sanitize_number(rd.get("is50"))
        calc_is50 = None
        if valida_w_norm == "SI" and effective_is is not None and effective_f is not None:
            calc_is50 = round(effective_is * effective_f, 4)
            if is50_num is None:
                reg_err_comp("Is(50) (MPa)", rd.get("is50"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Is(50) (MPa)")
            elif abs(is50_num - calc_is50) > tolerance:
                reg_err_comp("Is(50) (MPa)", is50_num, "ERR_PLT_IS50_INCONGRUENTE", actual=is50_num, expected=calc_is50)
        elif valida_w_norm == "NO" and is50_num is not None:
            reg_err_comp("Is(50) (MPa)", is50_num, "ERR_PLT_IS50_INCONGRUENTE", actual=is50_num, expected="NULL (Muestra Inválida)")

        effective_is50 = is50_num if is50_num is not None else calc_is50

        # UCS y Resistencia ISRM
        ucs_num = sanitize_number(rd.get("ucs"))
        calc_ucs = None
        if effective_is50 is not None and exp_k is not None:
            calc_ucs = round(effective_is50 * exp_k, 3)
            if ucs_num is None:
                reg_err_comp("RCU (MPa)", rd.get("ucs"), "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="RCU (MPa)")
            elif abs(ucs_num - calc_ucs) > tolerance:
                reg_err_comp("RCU (MPa)", ucs_num, "ERR_PLT_UCS_INCONGRUENTE", actual=ucs_num, expected=calc_ucs)

        effective_ucs = ucs_num if ucs_num is not None else calc_ucs
        isrm_raw = _norm_str(rd.get("isrm"))
        if effective_ucs is not None:
            exp_isrm = get_expected_isrm(effective_ucs)
            if not isrm_raw:
                reg_err_comp("Resistencia ISRM", None, "ERR_PLT_CAMPO_OBLIGATORIO_VACIO", col_name="Resistencia ISRM")
            elif exp_isrm and isrm_raw != exp_isrm:
                reg_err_comp("Resistencia ISRM", isrm_raw, "ERR_PLT_RESISTENCIA_ISRM_INCONGRUENTE",
                             actual=isrm_raw, ucs_val=round(effective_ucs, 2), expected=exp_isrm)

    return {
        "formato_detectado": "FORMAT_COMPACT_FIELD",
        "total_filas_procesadas": total_filas,
        "incidencias": incidencias,
        "resumen_por_celda": resumen_celdas,
    }


# =============================================================================
# DESPACHADOR PRINCIPAL (SSOT)
# =============================================================================
def validate_plt_excel(
    file_path: str,
    output_json_path: Optional[str] = None,
    tolerance: float = 0.1,
) -> dict:
    """
    Audita un archivo Excel de Ensayos PLT (.xlsx / .xlsm) contra las reglas QAQC oficiales.
    Autodetecta el formato (Estándar 34 Columnas vs Formato de Campo Compacto) y ejecuta el validador correspondiente.
    """
    t_start = time.time()
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Archivo Excel no encontrado: {file_path}")

    formato_tipo, header_row_idx = detect_plt_file_format(file_path)

    if formato_tipo == "COMPACT_FIELD":
        diag = validate_plt_compact_field(file_path, header_row_idx=header_row_idx, tolerance=tolerance)
    else:
        diag = validate_plt_standard_34(file_path, header_row_idx=header_row_idx, tolerance=tolerance)

    elapsed_sec = time.time() - t_start
    total_filas = diag.get("total_filas_procesadas", 0)
    incidencias = diag.get("incidencias", [])
    num_alertas = sum(1 for i in incidencias if i.get("tipo_incidencia") == "ALERTA")
    num_advertencias = sum(1 for i in incidencias if i.get("tipo_incidencia") == "ADVERTENCIA")
    num_vacios = sum(1 for i in incidencias if i.get("tipo_incidencia") == "VACIO")

    output_json = {
        "nombre_archivo": os.path.basename(file_path),
        "formato_detectado": formato_tipo,
        "total_filas_procesadas": total_filas,
        "total_celdas_evaluadas": total_filas * PLT_MANDATORY_COLS_COUNT,
        "metricas_globales": {
            "total_registros": total_filas,
            "total_celdas": len(diag.get("resumen_por_celda", {})),
            "total_alertas": num_alertas,
            "total_advertencias": num_advertencias,
            "total_vacios": num_vacios,
        },
        "incidencias": incidencias,
        "resumen_por_celda": diag.get("resumen_por_celda", {}),
    }

    if output_json_path:
        os.makedirs(os.path.dirname(output_json_path), exist_ok=True)
        tmp_path = output_json_path + ".tmp"
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(output_json, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, output_json_path)

    print(
        f"[QAQC PLT] Formato: {formato_tipo} | Validacion en {elapsed_sec:.2f}s | "
        f"Alertas: {num_alertas} | Advertencias: {num_advertencias} | Vacios: {num_vacios}"
    )

    return output_json
