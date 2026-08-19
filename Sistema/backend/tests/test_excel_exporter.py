"""
tests/test_excel_exporter.py — Pruebas unitarias para el motor de exportación a Excel.
"""

import io
import openpyxl
import pytest
from app.services.excel_exporter import export_ventanas_to_excel, _sanitize_val


def test_sanitize_val():
    assert _sanitize_val(None) is None
    assert _sanitize_val(-1) is None
    assert _sanitize_val(-1.0) is None
    assert _sanitize_val("-1") is None
    assert _sanitize_val("None") is None
    assert _sanitize_val("null") is None
    assert _sanitize_val("undefined") is None
    assert _sanitize_val("MZQ") == "MZQ"
    assert _sanitize_val(45.5) == 45.5
    assert _sanitize_val(0) == 0


def test_export_single_ventana_basic():
    ventana_data = {
        "codigo": "VNT-001",
        "excel_data": {
            "codigo": "VNT-001",
            "este_ini": 100.0,
            "norte_ini": 200.0,
            "cota_ini": 300.0,
            "este_fin": 110.0,
            "norte_fin": 205.0,
            "cota_fin": 302.0,
            "altura_m": 12.0,
            "dip_talud": 65.0,
            "lito_3": "MZQ",
            "alteracion": "d",
            "intemperismo": "f",
            "sector": "NORTE",
            "fase": "FASE_1",
            "nivel": "NV_200",
            "mapeador": "GEOLOGO_1",
            "fecha": "2026-08-18",
            "condicion_agua_rmr76": "C",
            "dureza_rmr76": "R3",
            "gsi_superficie": "B",
            "gsi_estructura": "M",
            "gsi_visual_rmr76": 55.0,
            "comentarios": "Mapeo regular de prueba"
        },
        "estructuras": [
            {
                "familia_id": 1,
                "distancia_m": 1.5,
                "tipo_estructura": "JN",
                "dip": 45.0,
                "dip_dir": 120.0,
                "n_estructuras": 1,
                "abertura_mm": 1.0,
                "espesor_mm": 2.0,
                "continuidad_m": 3.0,
                "espaciamiento_m": 0.5,
                "n_extremos_visibles": 1,
                "terminacion": 1,
                "relleno_1_codigo": "c",
                "relleno_2_codigo": -1,
                "jrc": 8.0,
                "rugosidad_codigo": "3",
                "forma_estructura": "P",
                "alteracion_codigo": "d"
            },
            {
                "familia_id": 2,
                "distancia_m": 3.0,
                "tipo_estructura": "JN",
                "dip": 60.0,
                "dip_dir": 210.0,
                "n_estructuras": 2,
                "abertura_mm": 2.0,
                "espesor_mm": 3.0,
                "continuidad_m": 4.0,
                "espaciamiento_m": 0.8,
                "n_extremos_visibles": 2,
                "terminacion": 1,
                "relleno_1_codigo": "b",
                "relleno_2_codigo": -1,
                "jrc": 10.0,
                "rugosidad_codigo": "4",
                "forma_estructura": "O",
                "alteracion_codigo": "d"
            }
        ]
    }

    buf = export_ventanas_to_excel(ventana_data)
    wb = openpyxl.load_workbook(buf, data_only=False)

    # 1. Verificar Hoja 'ventana'
    assert "ventana" in wb.sheetnames
    ws_v = wb["ventana"]
    assert ws_v["A4"].value == "VNT-001"
    assert ws_v["B5"].value == 100.0
    assert ws_v["B6"].value == 110.0
    assert ws_v["K6"].value == 12.0
    assert ws_v["N5"].value == 65.0
    assert ws_v["P4"].value == "MZQ"
    assert ws_v["P7"].value == "MZQ"
    assert ws_v["U4"].value == "NORTE"
    assert ws_v["U7"].value == "NORTE"

    # Fórmulas en filas activas (15 y 16)
    assert ws_v["A15"].value == 1
    assert ws_v["B15"].value == 1.5
    assert str(ws_v["W15"].value).startswith("=VLOOKUP(V15,RMR!")
    assert str(ws_v["AB15"].value) == "=SUM(X15+Y15+Z15+AA15+W15)"
    assert str(ws_v["W16"].value).startswith("=VLOOKUP(V16,RMR!")

    # Filas inactivas sin fórmulas
    assert ws_v["W17"].value is None
    assert ws_v["AB17"].value is None

    # Fórmulas RMR adaptadas
    assert "J15*F15+J16*F16" in str(ws_v["AW11"].value)
    assert "SUM(F15:F16)" in str(ws_v["AW11"].value)

    # 2. Verificar Hoja 'BD'
    assert "BD" in wb.sheetnames
    ws_bd = wb["BD"]
    # Debe tener exactamente 2 filas de datos continuas (filas 2 y 3)
    assert ws_bd["A2"].value == 1
    assert ws_bd["A3"].value == 2
    assert ws_bd["A4"].value is None  # Sin filas vacías sobrantes

    # Fila Padre (Fila 2)
    assert ws_bd["B2"].value == 4
    assert "=INDIRECT(" in str(ws_bd["C2"].value)
    assert ws_bd["BJ2"].value == 15
    assert "=INDIRECT(" in str(ws_bd["BK2"].value)
    assert "=BK2*COS(BL2)+F2" in str(ws_bd["BN2"].value)

    # Fila Hija (Fila 3)
    assert str(ws_bd["C3"].value) == "=C2"
    assert str(ws_bd["F3"].value) == "=F2"
    assert ws_bd["BJ3"].value == 16
    assert "=INDIRECT(" in str(ws_bd["BK3"].value)


def test_export_ventana_expanded_structures_and_families():
    # Ventana con 16 estructuras (> 14) y 4 familias (> 3)
    estructuras = []
    for i in range(1, 17):
        estructuras.append({
            "familia_id": (i % 4) + 1,  # Familias 1, 2, 3, 4
            "distancia_m": round(i * 0.8, 2),
            "tipo_estructura": "JN",
            "dip": 40.0 + i,
            "dip_dir": 100.0 + (i * 5),
            "n_estructuras": 1,
            "abertura_mm": 1.0,
            "espesor_mm": 1.0,
            "continuidad_m": 2.0,
            "espaciamiento_m": 0.5,
            "n_extremos_visibles": 1,
            "terminacion": 1,
            "relleno_1_codigo": "c",
            "relleno_2_codigo": -1,
            "jrc": 8.0,
            "rugosidad_codigo": "3",
            "forma_estructura": "P",
            "alteracion_codigo": "d"
        })

    ventana_data = {
        "codigo": "VNT-EXPAND",
        "excel_data": {"codigo": "VNT-EXPAND", "este_ini": 500, "norte_ini": 600, "cota_ini": 700},
        "estructuras": estructuras
    }

    buf = export_ventanas_to_excel(ventana_data)
    wb = openpyxl.load_workbook(buf, data_only=False)

    ws_v = wb["ventana"]
    # Fila 15 a 30 (16 estructuras)
    assert ws_v["B15"].value == 0.8
    assert ws_v["B30"].value == 12.8
    assert str(ws_v["W30"].value).startswith("=VLOOKUP(V30,RMR!")

    # Hoja BD debe tener exactamente 16 filas continuas (filas 2 a 17)
    ws_bd = wb["BD"]
    assert ws_bd["A2"].value == 1
    assert ws_bd["A17"].value == 16
    assert ws_bd["A18"].value is None


def test_export_multiple_ventanas():
    v1 = {
        "codigo": "CELL-A",
        "excel_data": {"codigo": "CELL-A", "este_ini": 100, "norte_ini": 200, "cota_ini": 300},
        "estructuras": [
            {"familia_id": 1, "distancia_m": 1.0, "tipo_estructura": "JN", "dip": 50.0, "dip_dir": 100.0}
        ]
    }
    v2 = {
        "codigo": "CELL-B",
        "excel_data": {"codigo": "CELL-B", "este_ini": 400, "norte_ini": 500, "cota_ini": 600},
        "estructuras": [
            {"familia_id": 1, "distancia_m": 2.0, "tipo_estructura": "FL", "dip": 70.0, "dip_dir": 200.0},
            {"familia_id": 2, "distancia_m": 4.0, "tipo_estructura": "JN", "dip": 80.0, "dip_dir": 250.0}
        ]
    }

    buf = export_ventanas_to_excel([v1, v2])
    wb = openpyxl.load_workbook(buf, data_only=False)

    # Hoja 'ventana': Celda 1 en A4, Celda 2 después de separación
    ws_v = wb["ventana"]
    assert ws_v["A4"].value == "CELL-A"

    # Hoja 'BD': Exactamente 1 (Cell A) + 2 (Cell B) = 3 filas continuas (filas 2, 3, 4)
    ws_bd = wb["BD"]
    assert ws_bd["A2"].value == 1
    assert ws_bd["A3"].value == 2
    assert ws_bd["A4"].value == 3
    assert ws_bd["A5"].value is None

    # Celda 2 empieza como Fila Padre en fila 3 de BD
    assert "=INDIRECT(" in str(ws_bd["C3"].value)
    # Celda 2 Hija en fila 4 de BD
    assert str(ws_bd["C4"].value) == "=C3"
